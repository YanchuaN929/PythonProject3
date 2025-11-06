#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
任务指派模块
负责接口工程师和室主任的任务指派功能
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
import os
import sys


def get_resource_path(relative_path):
    """获取资源文件的绝对路径（支持打包后的exe）"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_responsible_column(file_type):
    """
    获取各文件类型的责任人列名
    
    参数:
        file_type: 文件类型(1-6)
    
    返回:
        str: Excel列名，如'R'、'AP'、'AM'等
    """
    column_map = {
        1: 'R',   # 文件1责任人列（索引17）
        2: 'AM',  # 文件2责任人列（索引38）
        3: 'AP',  # 文件3责任人列（索引41）
        4: 'AH',  # 文件4责任人列（索引33）
        5: 'K',   # 文件5责任人列（索引10）
        6: 'X',   # 文件6责任人列（索引23）
    }
    return column_map.get(file_type)


def get_name_list():
    """
    从姓名角色表读取姓名列表
    
    返回:
        list: 姓名列表（已去重排序）
    """
    try:
        xls_path = get_resource_path("excel_bin/姓名角色表.xlsx")
        if not os.path.exists(xls_path):
            print(f"姓名角色表不存在: {xls_path}")
            return []
        
        df = pd.read_excel(xls_path)
        
        # 第一列为姓名
        if len(df.columns) == 0:
            return []
        
        names = df.iloc[:, 0].dropna().astype(str).tolist()
        
        # 过滤空值和'nan'
        names = [name.strip() for name in names if name.strip() and name.strip().lower() != 'nan']
        
        # 去重并排序
        names = sorted(set(names))
        
        return names
        
    except Exception as e:
        print(f"读取姓名列表失败: {e}")
        import traceback
        traceback.print_exc()
        return []


def is_interface_engineer(user_roles):
    """
    判断是否是接口工程师
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        bool: 是否是接口工程师
    """
    if not user_roles:
        return False
    
    for role in user_roles:
        if '接口工程师' in str(role):
            return True
    
    return False


def is_director(user_roles):
    """
    判断是否是室主任
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        bool: 是否是室主任
    """
    if not user_roles:
        return False
    
    director_roles = ['一室主任', '二室主任', '建筑总图室主任']
    
    for role in user_roles:
        if role in director_roles:
            return True
    
    return False


def get_department(user_roles):
    """
    根据室主任角色获取科室名称
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        str: 科室名称
    """
    dept_map = {
        "一室主任": "结构一室",
        "二室主任": "结构二室",
        "建筑总图室主任": "建筑总图室"
    }
    
    for role in user_roles:
        if role in dept_map:
            return dept_map[role]
    
    return ""


def parse_interface_engineer_project(user_roles):
    """
    从接口工程师角色中提取项目号
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        str: 项目号，如"2016"
    """
    import re
    
    for role in user_roles:
        if '接口工程师' in str(role):
            # 提取项目号（4位数字）
            match = re.search(r'(\d{4})', role)
            if match:
                return match.group(1)
    
    return None


def get_interface_id_column_index(file_type):
    """
    获取各文件类型的接口号列索引（参考window.py的interface_column_index逻辑）
    
    参数:
        file_type: 文件类型(1-6)
    
    返回:
        int: DataFrame列索引
    """
    column_map = {
        1: 0,   # 文件1接口号列 - A列 = 索引0
        2: 17,  # 文件2接口号列 - R列 = 索引17
        3: 2,   # 文件3接口号列 - C列 = 索引2
        4: 4,   # 文件4接口号列 - E列 = 索引4
        5: 0,   # 文件5接口号列 - A列 = 索引0
        6: 4,   # 文件6接口号列 - E列 = 索引4
    }
    return column_map.get(file_type, 0)


def check_unassigned(processed_results, user_roles, project_id=None):
    """
    检测所有处理结果中没有责任人的数据
    
    参数:
        processed_results: 6个文件的处理结果字典 {file_type: DataFrame}
        user_roles: 当前用户角色列表
        project_id: 项目号（接口工程师需要）
    
    返回:
        list: 未指派任务列表，每个任务包含：
              - file_type: 文件类型
              - project_id: 项目号
              - interface_id: 接口号
              - file_path: 源文件路径
              - row_index: Excel行号
    """
    unassigned = []
    
    for file_type, df in processed_results.items():
        if df is None or df.empty:
            continue
        
        # 筛选责任人为空的数据
        mask = (df['责任人'].isna()) | (df['责任人'].astype(str).str.strip() == '') | (df['责任人'].astype(str).str.strip() == '无')
        df_unassigned = df[mask].copy()
        
        if df_unassigned.empty:
            continue
        
        # 角色权限过滤
        if is_interface_engineer(user_roles):
            # 接口工程师：只看自己负责的项目
            if project_id and '项目号' in df_unassigned.columns:
                df_unassigned = df_unassigned[df_unassigned['项目号'].astype(str) == str(project_id)]
        
        elif is_director(user_roles):
            # 室主任：只看自己科室的数据 + "请室主任确认"
            my_department = get_department(user_roles)
            if my_department and '科室' in df_unassigned.columns:
                df_unassigned = df_unassigned[
                    df_unassigned['科室'].isin([my_department, '请室主任确认'])
                ]
        
        # 转换为字典列表
        interface_col_idx = get_interface_id_column_index(file_type)
        
        for idx, row in df_unassigned.iterrows():
            # 优先尝试使用'接口号'列名（用于测试和有命名列的DataFrame）
            interface_id = ''
            if '接口号' in df_unassigned.columns:
                try:
                    interface_id = row['接口号']
                    if pd.isna(interface_id) or str(interface_id).strip() == '':
                        interface_id = ''
                except Exception as e:
                    print(f"从'接口号'列获取失败: {e}")
            
            # 备用方案：使用列索引（用于真实的Excel数据，列名是数字索引）
            if not interface_id or interface_id == '':
                try:
                    if interface_col_idx < len(row):
                        interface_id = row.iloc[interface_col_idx]
                        if pd.isna(interface_id) or str(interface_id).strip() == '':
                            interface_id = ''
                except Exception:
                    # 静默处理列索引越界（某些DataFrame列数较少）
                    pass
            
            task = {
                'file_type': file_type,
                'project_id': row.get('项目号', ''),
                'interface_id': str(interface_id) if interface_id and not pd.isna(interface_id) else '',
                'file_path': row.get('source_file', ''),
                'row_index': row.get('原始行号', 0),
                'interface_time': row.get('接口时间', ''),
                'department': row.get('科室', '')
            }
            
            # 确保有必要的字段
            if task['file_path'] and task['row_index']:
                unassigned.append(task)
    
    return unassigned


def save_assignment(file_type, file_path, row_index, assigned_name):
    """
    保存指派结果到Excel的责任人列（单个任务）
    
    注意：此函数已废弃，建议使用save_assignments_batch进行批量指派
    
    参数:
        file_type: 文件类型(1-6)
        file_path: Excel文件路径
        row_index: Excel行号
        assigned_name: 指派的责任人姓名
    
    返回:
        bool: 成功返回True，失败返回False
    """
    # 调用批量指派函数
    assignments = [{
        'file_type': file_type,
        'file_path': file_path,
        'row_index': row_index,
        'assigned_name': assigned_name
    }]
    
    results = save_assignments_batch(assignments)
    return results.get('success_count', 0) > 0


def save_assignments_batch(assignments):
    """
    批量保存指派结果到Excel（优化版，按文件分组）
    
    参数:
        assignments: 指派列表，每项包含:
            {
                'file_type': int,
                'file_path': str,
                'row_index': int,
                'assigned_name': str,
                'interface_id': str (可选，用于Registry)
                'project_id': str (可选，用于Registry)
            }
    
    返回:
        dict: {
            'success_count': int,  # 成功数量
            'failed_tasks': list,  # 失败的任务信息
            'registry_updates': int  # Registry更新数量
        }
    """
    import pandas as pd
    from collections import defaultdict
    
    # 按文件路径分组
    file_groups = defaultdict(list)
    for assignment in assignments:
        file_path = assignment['file_path']
        file_groups[file_path].append(assignment)
    
    success_count = 0
    failed_tasks = []
    registry_updates = 0
    
    # 按文件批量处理
    for file_path, file_assignments in file_groups.items():
        try:
            # 1. 检查文件是否存在
            if not os.path.exists(file_path):
                print(f"[指派] 文件不存在: {file_path}")
                for assignment in file_assignments:
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': '文件不存在'
                    })
                continue
            
            # 2. 文件锁定检测
            try:
                with open(file_path, 'r+b') as f:
                    pass
            except PermissionError:
                print(f"[指派] 文件被占用: {file_path}")
                for assignment in file_assignments:
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': '文件被占用'
                    })
                continue
            
            # 3. 打开Excel文件（只打开一次）
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 4. 读取DataFrame用于Registry（只读一次）
            df = None
            try:
                df = pd.read_excel(file_path, sheet_name=0)
            except Exception as e:
                print(f"[指派] 读取DataFrame失败: {e}")
            
            # 5. 批量写入责任人
            for assignment in file_assignments:
                try:
                    file_type = assignment['file_type']
                    row_index = assignment['row_index']
                    assigned_name = assignment['assigned_name']
                    
                    # 获取责任人列名
                    col_name = get_responsible_column(file_type)
                    if not col_name:
                        print(f"[指派] 无法确定责任人列: file_type={file_type}")
                        failed_tasks.append({
                            'interface_id': assignment.get('interface_id', '未知'),
                            'reason': '无法确定责任人列'
                        })
                        continue
                    
                    # 写入责任人
                    ws[f"{col_name}{row_index}"] = assigned_name
                    success_count += 1
                    
                    print(f"[指派] 成功: 行{row_index}, 责任人={assigned_name}")
                    
                except Exception as e:
                    print(f"[指派] 单个任务失败: {e}")
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': str(e)
                    })
            
            # 6. 保存Excel（只保存一次）
            wb.save(file_path)
            wb.close()
            print(f"[指派] 文件已保存: {file_path}")
            
            # 7. 批量调用Registry钩子
            if df is not None:
                try:
                    from registry import hooks as registry_hooks
                    from registry.util import extract_interface_id, extract_project_id
                    
                    for assignment in file_assignments:
                        try:
                            row_index = assignment['row_index']
                            # row_index是Excel行号（包含表头），DataFrame索引需要减2
                            df_row_idx = row_index - 2
                            
                            if 0 <= df_row_idx < len(df):
                                row_data = df.iloc[df_row_idx]
                                
                                interface_id = extract_interface_id(row_data, assignment['file_type'])
                                project_id = extract_project_id(row_data, assignment['file_type'])
                                
                                if interface_id and project_id:
                                    registry_hooks.on_assigned(
                                        file_type=assignment['file_type'],
                                        file_path=file_path,
                                        row_index=row_index,
                                        interface_id=interface_id,
                                        project_id=project_id,
                                        assigned_by="系统用户",  # TODO: 传递实际用户
                                        assigned_to=assignment['assigned_name']
                                    )
                                    registry_updates += 1
                        except Exception as e:
                            print(f"[Registry] 单个任务钩子失败: {e}")
                    
                    print(f"[Registry] 共更新 {registry_updates} 个任务")
                    
                except Exception as e:
                    print(f"[Registry] 批量钩子失败: {e}")
            
        except Exception as e:
            print(f"[指派] 文件处理失败: {file_path}, 错误: {e}")
            import traceback
            traceback.print_exc()
            for assignment in file_assignments:
                failed_tasks.append({
                    'interface_id': assignment.get('interface_id', '未知'),
                    'reason': str(e)
                })
    
    return {
        'success_count': success_count,
        'failed_tasks': failed_tasks,
        'registry_updates': registry_updates
    }


class AssignmentDialog(tk.Toplevel):
    """任务指派界面"""
    
    def __init__(self, parent, unassigned_tasks, name_list):
        """
        初始化任务指派对话框
        
        参数:
            parent: 父窗口
            unassigned_tasks: 未指派任务列表
            name_list: 姓名列表（从姓名角色表读取）
        """
        super().__init__(parent)
        
        self.unassigned_tasks = unassigned_tasks
        self.name_list = name_list
        self.assignment_entries = []  # 存储每行的输入控件
        self.batch_name_var = tk.StringVar()  # 批量指派的姓名
        self.assignment_successful = False  # 【新增】标记是否成功指派
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置界面"""
        self.title("任务指派")
        self.geometry("900x600")
        
        # 居中显示
        self.transient(self.master)
        self.grab_set()
        
        # 标题
        title_label = ttk.Label(
            self,
            text=f"任务指派（共{len(self.unassigned_tasks)}个未指派任务）",
            font=('Arial', 14, 'bold')
        )
        title_label.pack(pady=10)
        
        # 创建滚动区域
        canvas_frame = ttk.Frame(self)
        canvas_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(canvas_frame)  # 使用系统默认背景色，去掉白色
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 批量指派控制栏
        batch_frame = ttk.Frame(scrollable_frame)
        batch_frame.pack(fill='x', pady=5)
        
        # 全选/全不选按钮
        ttk.Button(batch_frame, text="全选", command=self.select_all, width=8).pack(side='left', padx=5)
        ttk.Button(batch_frame, text="全不选", command=self.deselect_all, width=8).pack(side='left', padx=5)
        
        # 批量指派输入框
        ttk.Label(batch_frame, text="批量指派给：").pack(side='left', padx=5)
        batch_combo = ttk.Combobox(batch_frame, textvariable=self.batch_name_var, values=self.name_list, width=15)
        batch_combo.pack(side='left', padx=5)
        batch_combo.bind('<KeyRelease>', lambda e: self.on_batch_search(e, batch_combo))
        
        # 批量应用按钮
        ttk.Button(batch_frame, text="应用到勾选项", command=self.apply_batch_assignment, width=12).pack(side='left', padx=5)
        
        # 分隔线
        ttk.Separator(scrollable_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # 表头
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill='x', pady=5)
        
        # 添加"选择"列
        ttk.Label(header_frame, text="选择", width=6, font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5)
        ttk.Label(header_frame, text="文件类型", width=12, font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5)
        ttk.Label(header_frame, text="项目号", width=10, font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5)
        ttk.Label(header_frame, text="接口号", width=25, font=('Arial', 10, 'bold')).grid(row=0, column=3, padx=5)
        ttk.Label(header_frame, text="接口时间", width=12, font=('Arial', 10, 'bold')).grid(row=0, column=4, padx=5)
        ttk.Label(header_frame, text="指派人", width=20, font=('Arial', 10, 'bold')).grid(row=0, column=5, padx=5)
        
        # 文件类型映射
        file_type_names = {
            1: "内部需打开",
            2: "内部需回复",
            3: "外部需打开",
            4: "外部需回复",
            5: "三维提资",
            6: "收发文函"
        }
        
        # 为每个未指派任务创建一行
        for i, task in enumerate(self.unassigned_tasks):
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', pady=2)
            
            # 选择复选框
            checkbox_var = tk.BooleanVar(value=False)
            checkbox = ttk.Checkbutton(row_frame, variable=checkbox_var)
            checkbox.grid(row=0, column=0, padx=5)
            
            # 文件类型
            file_type_name = file_type_names.get(task['file_type'], f"文件{task['file_type']}")
            ttk.Label(row_frame, text=file_type_name, width=12).grid(row=0, column=1, padx=5)
            
            # 项目号
            ttk.Label(row_frame, text=str(task['project_id']), width=10).grid(row=0, column=2, padx=5)
            
            # 接口号（处理空值）
            interface_id = str(task.get('interface_id', ''))
            if not interface_id or interface_id == 'None' or interface_id.strip() == '':
                interface_id = "（无接口号）"
                print(f"[警告] 任务{i+1}无接口号数据: {task}")
            elif len(interface_id) > 30:
                interface_id = interface_id[:27] + "..."
            ttk.Label(row_frame, text=interface_id, width=25).grid(row=0, column=3, padx=5)
            
            # 接口时间
            ttk.Label(row_frame, text=str(task['interface_time']), width=12).grid(row=0, column=4, padx=5)
            
            # 指派人下拉框（带实时搜索）
            combobox = ttk.Combobox(row_frame, values=self.name_list, width=18)
            combobox.grid(row=0, column=5, padx=5)
            
            # 绑定实时搜索 - 在输入时自动弹出下拉菜单
            combobox.bind('<KeyRelease>', lambda event, cb=combobox: self.on_search(event, cb))
            
            # 【修复】禁用鼠标滚轮改变值的行为
            combobox.bind('<MouseWheel>', lambda e: "break")
            combobox.bind('<Button-4>', lambda e: "break")  # Linux向上滚动
            combobox.bind('<Button-5>', lambda e: "break")  # Linux向下滚动
            
            # 保存引用
            self.assignment_entries.append({
                'task': task,
                'combobox': combobox,
                'checkbox_var': checkbox_var
            })
        
        # 布局Canvas和Scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 按钮区域
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="确认指派", command=self.on_confirm).pack(side='left', padx=10)
        ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)
    
    def select_all(self):
        """全选所有任务"""
        for entry in self.assignment_entries:
            entry['checkbox_var'].set(True)
    
    def deselect_all(self):
        """全不选所有任务"""
        for entry in self.assignment_entries:
            entry['checkbox_var'].set(False)
    
    def on_batch_search(self, event, combobox):
        """批量指派输入框的实时搜索"""
        search_text = combobox.get().strip()
        
        if not search_text:
            combobox['values'] = self.name_list
            return
        
        filtered = [name for name in self.name_list if search_text in name]
        combobox['values'] = filtered
    
    def apply_batch_assignment(self):
        """将批量指派的姓名应用到所有勾选的任务"""
        batch_name = self.batch_name_var.get().strip()
        
        if not batch_name:
            messagebox.showwarning("提示", "请先输入要批量指派的姓名", parent=self)
            return
        
        # 统计勾选数量
        checked_count = sum(1 for entry in self.assignment_entries if entry['checkbox_var'].get())
        
        if checked_count == 0:
            messagebox.showwarning("提示", "请至少勾选一个任务", parent=self)
            return
        
        # 应用到勾选的任务
        for entry in self.assignment_entries:
            if entry['checkbox_var'].get():
                entry['combobox'].set(batch_name)
        
        messagebox.showinfo("成功", f"已将 \"{batch_name}\" 应用到 {checked_count} 个任务", parent=self)
    
    def on_search(self, event, combobox):
        """实时搜索回调 - 只更新下拉列表，不自动弹出"""
        search_text = combobox.get().strip()
        
        if not search_text:
            # 恢复完整列表
            combobox['values'] = self.name_list
            return
        
        # 过滤姓名列表（包含搜索文本的）
        filtered = [name for name in self.name_list if search_text in name]
        combobox['values'] = filtered
        
        # 不自动弹出下拉菜单，让用户自己点击下拉按钮
    
    def on_confirm(self):
        """确认指派按钮回调（优化版，使用批量保存）"""
        assignments = []
        
        # 收集所有指派
        for entry in self.assignment_entries:
            task = entry['task']
            assigned_name = entry['combobox'].get().strip()
            
            # 跳过空值
            if not assigned_name:
                continue
            
            assignments.append({
                'file_type': task['file_type'],
                'file_path': task['file_path'],
                'row_index': task['row_index'],
                'assigned_name': assigned_name,
                'interface_id': task.get('interface_id', '未知'),
                'project_id': task.get('project_id', '')
            })
        
        if not assignments:
            messagebox.showwarning("提示", "请至少选择一个责任人进行指派", parent=self)
            return
        
        # 显示处理中提示
        processing_label = ttk.Label(self, text="正在批量指派，请稍候...", font=('Arial', 12))
        processing_label.pack(pady=10)
        self.update()
        
        # 执行批量指派
        try:
            results = save_assignments_batch(assignments)
            success_count = results['success_count']
            failed_tasks = results['failed_tasks']
            registry_updates = results['registry_updates']
            
            # 隐藏处理中提示
            processing_label.destroy()
            
            # 显示结果
            if success_count > 0:
                msg = f"成功指派 {success_count} 个任务"
                if registry_updates > 0:
                    msg += f"\nRegistry已更新 {registry_updates} 条记录"
                if failed_tasks:
                    msg += f"\n\n失败 {len(failed_tasks)} 个任务：\n"
                    msg += "\n".join([f"- {t['interface_id']}: {t['reason']}" for t in failed_tasks[:5]])
                    if len(failed_tasks) > 5:
                        msg += f"\n... 等共{len(failed_tasks)}个失败"
                
                messagebox.showinfo("指派结果", msg, parent=self)
                
                if not failed_tasks:
                    self.assignment_successful = True  # 【新增】标记成功
                    self.destroy()
            else:
                messagebox.showerror("失败", "所有任务指派失败，请检查文件是否被占用", parent=self)
        except Exception as e:
            processing_label.destroy()
            messagebox.showerror("错误", f"指派过程中发生错误：\n{str(e)}", parent=self)
            import traceback
            traceback.print_exc()


def show_assignment_dialog(parent, unassigned_tasks, name_list):
    """
    显示任务指派对话框
    
    参数:
        parent: 父窗口
        unassigned_tasks: 未指派任务列表
        name_list: 姓名列表
    """
    if not unassigned_tasks:
        messagebox.showinfo("提示", "当前没有需要指派的任务", parent=parent)
        return
    
    if not name_list:
        messagebox.showwarning("警告", "无法读取姓名列表，请检查姓名角色表.xlsx", parent=parent)
        return
    
    dialog = AssignmentDialog(parent, unassigned_tasks, name_list)
    dialog.wait_window()


# 测试代码
if __name__ == "__main__":
    # 测试get_responsible_column
    print("测试get_responsible_column:")
    for i in range(1, 7):
        col = get_responsible_column(i)
        print(f"  文件{i}: {col}")
    
    # 测试get_name_list
    print("\n测试get_name_list:")
    names = get_name_list()
    print(f"  读取到 {len(names)} 个姓名")
    if names:
        print(f"  前5个: {names[:5]}")
    
    # 测试角色判断
    print("\n测试角色判断:")
    test_roles = [
        ['一室主任'],
        ['2016接口工程师'],
        ['设计人员'],
        ['所长']
    ]
    
    for roles in test_roles:
        print(f"  角色: {roles}")
        print(f"    是接口工程师: {is_interface_engineer(roles)}")
        print(f"    是室主任: {is_director(roles)}")
        if is_director(roles):
            print(f"    科室: {get_department(roles)}")
        if is_interface_engineer(roles):
            print(f"    项目号: {parse_interface_engineer_project(roles)}")

