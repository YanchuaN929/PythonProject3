#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据处理模块
此文件包含所有Excel文件的数据处理逻辑
"""

import pandas as pd
import datetime
import os
import warnings
import re
import threading
import unicodedata
from copy import copy
from functools import lru_cache

# 忽略pandas警告
warnings.filterwarnings('ignore')


_REGISTRY_READ_SNAPSHOT_LOCK = threading.RLock()
_REGISTRY_READ_SNAPSHOT = None


def begin_registry_read_snapshot():
    """Enable one-run Registry read reuse across workbook worker threads."""
    global _REGISTRY_READ_SNAPSHOT
    with _REGISTRY_READ_SNAPSHOT_LOCK:
        _REGISTRY_READ_SNAPSHOT = {}


def end_registry_read_snapshot():
    global _REGISTRY_READ_SNAPSHOT
    with _REGISTRY_READ_SNAPSHOT_LOCK:
        _REGISTRY_READ_SNAPSHOT = None


def _registry_snapshot_load(key, loader):
    with _REGISTRY_READ_SNAPSHOT_LOCK:
        if _REGISTRY_READ_SNAPSHOT is None:
            return loader()
        if key not in _REGISTRY_READ_SNAPSHOT:
            _REGISTRY_READ_SNAPSHOT[key] = loader()
        return _REGISTRY_READ_SNAPSHOT[key]

# 导入科室参数化配置
try:
    from utils.dept_config import (
        get_department_codes,
        get_organization_filter,
        get_organization_filter_file6,
        get_projects_standard_filter,
        map_code_to_department,
        match_department_name,
        contains_department_code,
    )
except ImportError:
    # 后备：模块不可用时使用硬编码默认值（不应发生）
    def get_department_codes():
        return ["25C1", "25C2", "25C3"]
    def get_organization_filter():
        return "河北分公司-建筑结构所"
    def get_organization_filter_file6():
        return "河北分公司.建筑结构所"
    def map_code_to_department(s):
        for code, name in {"25C1": "结构一室", "25C2": "结构二室", "25C3": "建筑总图室"}.items():
            if code in s:
                return name
        return ""
    def match_department_name(s):
        for name in ["结构一室", "结构二室", "建筑总图室"]:
            if name in s:
                return name
        return s
    def contains_department_code(s):
        return any(c in s for c in ["25C1", "25C2", "25C3"])
    def get_projects_standard_filter():
        return ["1907", "2016", "2026"]

# 导入项目特殊调整模块（1818项目日期减6天等）
try:
    from utils.adjust import adjust_date_for_project
except ImportError:
    def adjust_date_for_project(cell_date, project_id):
        """兜底函数：无调整"""
        return cell_date


def apply_assignment_memory(result_df, file_type):
    """
    应用指派记忆：为责任人为空的行填充历史指派记忆

    参数:
        result_df: 处理结果DataFrame，必须包含'责任人'、'项目号'列
        file_type: 文件类型（1-6）

    返回:
        DataFrame: 应用记忆后的结果（原地修改）
    """
    if result_df is None or result_df.empty:
        return result_df

    if '责任人' not in result_df.columns or '项目号' not in result_df.columns:
        return result_df

    try:
        from services.assignment_memory import get_memory

        # 获取接口号列名（根据文件类型）
        interface_col = '接口号' if '接口号' in result_df.columns else None

        memory_applied_count = 0
        for idx in result_df.index:
            # 检查责任人是否为空
            responsible = str(result_df.at[idx, '责任人']).strip()
            if responsible and responsible.lower() not in ['', 'nan', 'none', '无']:
                continue  # 已有责任人，跳过

            # 获取项目号
            project_id = str(result_df.at[idx, '项目号']).strip()
            if not project_id or project_id.lower() in ['nan', 'none']:
                continue

            # 获取接口号
            interface_id = ''
            if interface_col and interface_col in result_df.columns:
                interface_id = str(result_df.at[idx, interface_col]).strip()
                # 去除角色后缀
                interface_id = re.sub(r'\([^)]*\)$', '', interface_id).strip()

            if not interface_id or interface_id.lower() in ['nan', 'none']:
                continue

            # 查询指派记忆
            memory_name = get_memory(file_type, project_id, interface_id)
            if memory_name:
                result_df.at[idx, '责任人'] = memory_name
                memory_applied_count += 1

        if memory_applied_count > 0:
            print(f"[AssignmentMemory] 文件{file_type}: 应用了 {memory_applied_count} 条指派记忆")

    except Exception as e:
        # 记忆功能失败不影响主流程
        print(f"[AssignmentMemory] 应用指派记忆失败: {e}")

    return result_df


def _get_version_rank(version_value):
    if version_value is None:
        return 0
    version_str = unicodedata.normalize("NFKC", str(version_value)).strip()
    if not version_str or version_str.lower() in ("nan", "none"):
        return 0
    match = re.search(r"[A-Za-z]", version_str)
    if not match:
        return 0
    letter = match.group(0).upper()
    return ord(letter) - ord("A") + 1


def _filter_rows_by_highest_version(df, file_type, rows, version_col_index):
    if not rows:
        return rows
    if version_col_index is None:
        return rows
    if len(df.columns) <= version_col_index:
        print(f"警告：文件列数不足，无法访问版次列(索引{version_col_index})，跳过版次筛选")
        return rows

    interface_col_map = {
        1: 0,
        2: 17,
        3: 2,
        4: 4,
        5: 0,
        6: 4,
    }

    interface_values = None
    try:
        if "接口号" in df.columns:
            interface_values = df["接口号"]
        else:
            interface_col_index = interface_col_map.get(file_type)
            if interface_col_index is not None and len(df.columns) > interface_col_index:
                interface_values = df.iloc[:, interface_col_index]
        version_values = df.iloc[:, version_col_index]
    except Exception:
        interface_values = None

    if interface_values is None:
        try:
            from registry.util import extract_interface_id
        except Exception:
            extract_interface_id = None
    else:
        extract_interface_id = None

    best_rank = {}
    best_rows = {}
    keep_rows = set()

    for idx in rows:
        if idx < 0 or idx >= len(df):
            continue
        if interface_values is not None:
            try:
                interface_id = str(interface_values.iat[idx]).strip()
                interface_id = re.sub(r'\([^)]*\)$', '', interface_id).strip()
            except Exception:
                interface_id = ""
        else:
            row_data = df.iloc[idx]
            interface_id = ""
            if extract_interface_id:
                try:
                    interface_id = extract_interface_id(row_data, file_type) or ""
                except Exception:
                    interface_id = ""
            elif hasattr(row_data, "get"):
                interface_id = str(row_data.get("接口号", "") or "").strip()

        interface_id = str(interface_id).strip()
        if not interface_id:
            keep_rows.add(idx)
            continue

        version_value = version_values.iat[idx] if interface_values is not None else df.iloc[idx, version_col_index]
        rank = _get_version_rank(version_value)
        current_rank = best_rank.get(interface_id)
        if current_rank is None or rank > current_rank:
            best_rank[interface_id] = rank
            best_rows[interface_id] = [idx]
        elif rank == current_rank:
            best_rows[interface_id].append(idx)

    for row_list in best_rows.values():
        keep_rows.update(row_list)

    removed = len(rows) - len(keep_rows)
    if removed > 0:
        print(f"[版本筛选] 文件{file_type}：剔除 {removed} 行低版本接口，仅保留最高版次")
        try:
            from core import Monitor
            Monitor.log_info(f"[版本筛选] 文件{file_type}：剔除 {removed} 行低版本接口，仅保留最高版次")
        except Exception:
            pass

    return keep_rows


def _extract_file_project_id(file_path):
    """从文件名提取四位项目号，失败时返回空字符串。"""
    filename = os.path.basename(file_path or "")
    match = re.search(r'(\d{4})', filename)
    return match.group(1) if match else ""


def _extract_file5_project_id(file_path):
    """三维提资文件名可能包含日期，优先提取靠近业务语义的项目号。"""
    filename = os.path.basename(file_path or "")
    match = re.search(r'(\d{4})项目', filename)
    if match:
        return match.group(1)
    match = re.search(r'三维接口提资清单[-_\s]*(\d{4})', filename)
    if match:
        return match.group(1)
    match = re.match(r'^(\d{4})接口提资清单', filename)
    if match:
        return match.group(1)
    return _extract_file_project_id(file_path)


def _load_latest_registry_pending_tasks_uncached(file_type, db_path, wal):
    """
    读取指定文件类型在 Registry 中最新 business 记录仍处于待审查的任务。

    统一口径：
    - 取非 archived 记录，以及未确认但被 missing_from_source 误归档的记录
    - 按 interface_id + project_id 只保留最新一条
    - 只有 status=completed 且 display_status 为待审查/待指派人审查才允许加回
    """
    from registry.db import get_connection, close_connection_after_use

    pending_tasks = []
    if not db_path or not os.path.exists(db_path):
        return pending_tasks

    conn = get_connection(db_path, wal)
    try:
        cursor = conn.execute("""
            SELECT current.interface_id, current.project_id, current.display_status,
                   CASE
                       WHEN current.status = 'archived'
                            AND current.archive_reason = 'missing_from_source'
                            AND current.confirmed_at IS NULL
                            AND current.completed_at IS NOT NULL
                       THEN 'completed'
                       WHEN current.status = 'archived'
                            AND current.archive_reason = 'missing_from_source'
                            AND current.confirmed_at IS NULL
                       THEN 'open'
                       ELSE current.status
                   END AS effective_status,
                   current.last_seen_at
            FROM tasks AS current
            WHERE current.file_type = ?
              AND (current.ignored = 0 OR current.ignored IS NULL)
              AND (
                  current.status != 'archived'
                  OR (
                      current.status = 'archived'
                      AND current.archive_reason = 'missing_from_source'
                      AND current.confirmed_at IS NULL
                  )
              )
              AND NOT EXISTS (
                  SELECT 1
                  FROM tasks AS confirmed
                  WHERE confirmed.business_id = current.business_id
                    AND confirmed.status = 'archived'
                    AND confirmed.archive_reason IN (
                        'confirmed_by_superior',
                        'completed_in_new_source'
                    )
                    AND confirmed.confirmed_at IS NOT NULL
                    AND COALESCE(
                            confirmed.archived_at,
                            confirmed.confirmed_at,
                            confirmed.last_seen_at,
                            ''
                        ) >= COALESCE(current.last_seen_at, current.first_seen_at, '')
              )
            ORDER BY current.last_seen_at DESC, current.rowid DESC
        """, (file_type,))
        rows = cursor.fetchall()
    finally:
        close_connection_after_use()

    latest_registry_tasks = {}
    for interface_id, project_id, display_status, status, last_seen_at in rows:
        key = (str(interface_id or "").strip(), str(project_id or "").strip())
        if key in latest_registry_tasks:
            continue
        latest_registry_tasks[key] = (
            str(display_status or "").strip(),
            str(status or "").strip(),
            last_seen_at,
        )

    for (interface_id, project_id), (display_status, status, last_seen_at) in latest_registry_tasks.items():
        if status == 'completed' and display_status in ('待审查', '待指派人审查'):
            pending_tasks.append(
                (interface_id, project_id, display_status, status, last_seen_at)
            )

    return pending_tasks


def _load_latest_registry_pending_tasks(file_type, db_path, wal):
    key = (
        "pending",
        int(file_type),
        os.path.normcase(os.path.abspath(os.fspath(db_path))) if db_path else "",
        bool(wal),
    )
    return _registry_snapshot_load(
        key,
        lambda: _load_latest_registry_pending_tasks_uncached(file_type, db_path, wal),
    )


def _load_latest_confirmed_archive_times_uncached(file_type, project_id, db_path, wal):
    """Return the latest confirmed archive plan time for each business interface."""
    from registry.db import get_connection, close_connection_after_use

    archive_times = {}
    if not db_path or not os.path.exists(db_path):
        return archive_times

    conn = get_connection(db_path, wal)
    try:
        rows = conn.execute(
            """
            SELECT interface_id, interface_time
            FROM tasks
            WHERE file_type = ?
              AND project_id = ?
              AND status = 'archived'
              AND archive_reason = 'confirmed_by_superior'
              AND confirmed_at IS NOT NULL
            ORDER BY archived_at DESC, last_seen_at DESC, rowid DESC
            """,
            (int(file_type), str(project_id or "").strip()),
        ).fetchall()
        for interface_id, interface_time in rows:
            key = str(interface_id or "").strip()
            if key and key not in archive_times:
                archive_times[key] = str(interface_time or "").strip()
    except Exception:
        return {}
    finally:
        close_connection_after_use()
    return archive_times


def _load_latest_confirmed_archive_times(file_type, project_id, db_path, wal):
    key = (
        "confirmed_archive_times",
        int(file_type),
        str(project_id or "").strip(),
        os.path.normcase(os.path.abspath(os.fspath(db_path))) if db_path else "",
        bool(wal),
    )
    return _registry_snapshot_load(
        key,
        lambda: _load_latest_confirmed_archive_times_uncached(
            file_type, project_id, db_path, wal
        ),
    )


def _build_registry_records_index(records, file_type, file_path, project_id=None, allow_interface_only_fallback=False):
    """为流式精简记录建立 Registry 加回索引，避免额外构造 DataFrame。"""
    if file_type == 5:
        file_project_id = project_id or _extract_file5_project_id(file_path)
    else:
        file_project_id = project_id or _extract_file_project_id(file_path)
    excel_index = {}
    excel_index_by_iface = {}
    spec = None
    if records:
        spec = records[0].get("_stream_spec")
    if not spec:
        spec = STREAM_FILE_SPECS.get(file_type) if "STREAM_FILE_SPECS" in globals() else None
    if not spec:
        return file_project_id, excel_index, excel_index_by_iface
    interface_alias = spec["columns"][spec["interface"]]

    for record in records:
        try:
            idx = int(record.get("_df_index", -1))
        except Exception:
            continue
        if idx < 0:
            continue
        raw = record.get("_raw") or {}
        interface_id = _normalize_interface_id(
            record.get("接口号") or raw.get(interface_alias)
        )
        if not interface_id:
            continue
        row_project_id = str(record.get("项目号") or "").strip()
        project = row_project_id or file_project_id
        if project:
            excel_index.setdefault((interface_id, project), []).append(idx)
        elif allow_interface_only_fallback:
            excel_index_by_iface.setdefault(interface_id, []).append(idx)

    return file_project_id, excel_index, excel_index_by_iface


def _build_registry_excel_index(df, file_type, file_path, allow_interface_only_fallback=False, project_id=None):
    """为 Registry 加回逻辑建立当前 Excel 的接口索引。"""
    if isinstance(df, list):
        return _build_registry_records_index(
            df,
            file_type,
            file_path,
            project_id=project_id,
            allow_interface_only_fallback=allow_interface_only_fallback,
        )

    file_project_id = project_id or (_extract_file5_project_id(file_path) if file_type == 5 else _extract_file_project_id(file_path))
    excel_index = {}
    excel_index_by_iface = {}
    interface_col_map = {
        1: 0,
        2: 17,
        3: 2,
        4: 4,
        5: 0,
        6: 4,
    }

    interface_values = None
    project_values = None
    source_file_values = None
    try:
        if "接口号" in df.columns:
            interface_values = df["接口号"]
        else:
            interface_col_index = interface_col_map.get(file_type)
            if interface_col_index is not None and len(df.columns) > interface_col_index:
                interface_values = df.iloc[:, interface_col_index]
        if "项目号" in df.columns:
            project_values = df["项目号"]
        if "source_file" in df.columns:
            source_file_values = df["source_file"]
    except Exception:
        interface_values = None

    if interface_values is None:
        from registry.util import extract_interface_id, extract_project_id
    else:
        extract_interface_id = None
        extract_project_id = None

    for idx in range(len(df)):
        if idx == 0:
            continue
        try:
            if interface_values is not None:
                interface_id = str(interface_values.iat[idx]).strip()
                interface_id = re.sub(r'\([^)]*\)$', '', interface_id).strip()
            else:
                row_data = df.iloc[idx]
                interface_id = extract_interface_id(row_data, file_type)
            if not interface_id:
                continue

            if project_values is not None:
                row_project_id = str(project_values.iat[idx]).strip()
            elif source_file_values is not None:
                source_basename = os.path.basename(str(source_file_values.iat[idx] or ""))
                match = re.search(r'(\d{4})', source_basename)
                row_project_id = match.group(1) if match else ""
            else:
                row_project_id = extract_project_id(row_data, file_type) if extract_project_id else ""
            row_project_id = row_project_id or ""
            project_id = row_project_id or file_project_id
            if project_id:
                key = (interface_id, project_id)
                if key not in excel_index:
                    excel_index[key] = []
                excel_index[key].append(idx)
            elif allow_interface_only_fallback:
                if interface_id not in excel_index_by_iface:
                    excel_index_by_iface[interface_id] = []
                excel_index_by_iface[interface_id].append(idx)
        except Exception:
            continue

    return file_project_id, excel_index, excel_index_by_iface


def _merge_registry_pending_rows(
    file_type,
    file_path,
    df,
    final_rows,
    allowed_rows,
    allow_interface_only_fallback=False,
    project_id=None,
):
    """
    将 Registry 中仍待审查的任务统一加回当前处理结果。

    参数说明：
    - allowed_rows: 各文件类型自己的业务基筛条件

    说明：
    - 这里统一负责“按最新 business 记录找回待审查任务”
    - “设计人员写回后应退出主列表、上级仍可见待审查”的角色差异，
      交给显示层按 Registry 状态过滤，避免在原始处理层直接切断上级审查链路
    """
    from registry.hooks import _cfg

    pending_rows = set()

    cfg = _cfg()
    db_path = cfg.get('registry_db_path')
    wal = bool(cfg.get("registry_wal", False))
    registry_tasks = _load_latest_registry_pending_tasks(file_type, db_path, wal)
    if not registry_tasks:
        return final_rows, pending_rows

    file_project_id, excel_index, excel_index_by_iface = _build_registry_excel_index(
        df,
        file_type,
        file_path,
        allow_interface_only_fallback=allow_interface_only_fallback,
        project_id=project_id,
    )

    print(f"[Registry] 文件类型{file_type}最新待审查任务数: {len(registry_tasks)}")
    print(
        f"[Registry] 文件类型{file_type} Excel索引建立完成"
        f"(项目={file_project_id or 'N/A'})，唯一接口{len(excel_index)}个"
    )

    for reg_interface_id, reg_project_id, reg_display_status, _reg_status, _reg_last_seen_at in registry_tasks:
        key = (reg_interface_id, reg_project_id)
        matched_indices = excel_index.get(key, []) or []
        if not matched_indices and allow_interface_only_fallback and not file_project_id:
            matched_indices = excel_index_by_iface.get(reg_interface_id, []) or []

        for idx in matched_indices:
            if idx not in allowed_rows:
                continue
            if idx in final_rows:
                continue
            pending_rows.add(idx)
            print(
                f"[Registry] 加回待审查任务: {reg_interface_id}, 行{idx + 2}, 状态:{reg_display_status}"
            )

    if pending_rows:
        final_rows = final_rows | pending_rows
        print(f"[Registry] 共加回{len(pending_rows)}条待审查任务")
    else:
        print("[Registry] 未找到需加回的待审查任务")

    return final_rows, pending_rows


# ============================================================
# 文件识别与项目号提取
# ============================================================

def find_target_file(excel_files):
    """
    查找符合特定格式的待处理文件1（兼容性函数，返回第一个匹配的文件）
    格式：四位数字+按项目导出IDI手册+日期
    例如：2016按项目导出IDI手册2025-08-01-17_55_52
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files1(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files1(excel_files):
    """
    查找所有符合特定格式的待处理文件1
    格式：四位数字+按项目导出IDI手册+日期
    例如：2016按项目导出IDI手册2025-08-01-17_55_52
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^(\d{4})按项目导出IDI手册\d{4}-\d{2}-\d{2}.*\.(xlsx|xlsm|xls)$'
    matched_files = []

    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件1...")
    except Exception:
        pass

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name, flags=re.IGNORECASE)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件1格式: {file_name}, 项目号: {project_id}")
            try:
                from core import Monitor
                Monitor.log_success(f"找到待处理文件1: 项目{project_id} - {file_name}")
            except Exception:
                pass

    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件1")
        try:
            from core import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件1，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except Exception:
            pass
    else:
        try:
            from core import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件1")
        except Exception:
            pass

    return matched_files


def find_target_file2(excel_files):
    """
    查找符合特定格式的待处理文件2（兼容性函数，返回第一个匹配的文件）
    格式：内部接口信息单报表+12位数字，前4位为项目号
    例如：内部接口信息单报表201612345678
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files2(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files2(excel_files):
    """
    查找所有符合特定格式的待处理文件2
    格式：内部接口信息单报表+12位数字，前4位为项目号
    例如：内部接口信息单报表201612345678
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^内部接口信息单报表(\d{4})\d{8}\.(xlsx|xlsm|xls)$'
    matched_files = []

    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件2...")
    except Exception:
        pass

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name, flags=re.IGNORECASE)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件2格式: {file_name}, 项目号: {project_id}")
            try:
                from core import Monitor
                Monitor.log_success(f"找到待处理文件2: 项目{project_id} - {file_name}")
            except Exception:
                pass

    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件2")
        try:
            from core import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件2，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except Exception:
            pass
    else:
        try:
            from core import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件2")
        except Exception:
            pass

    return matched_files


def find_target_file3(excel_files):
    """
    查找符合特定格式的待处理文件3（兼容性函数，返回第一个匹配的文件）
    格式：外部接口ICM报表+四位数字（项目号）+日期（8位）
    例如：外部接口ICM报表201620250801.xlsx
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files3(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files3(excel_files):
    """
    查找所有符合特定格式的待处理文件3
    格式：外部接口ICM报表+四位数字（项目号）+日期（8位）
    例如：外部接口ICM报表201620250801.xlsx
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^外部接口ICM报表(\d{4})\d{8}\.(xlsx|xlsm|xls)$'
    matched_files = []

    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件3...")
    except Exception:
        pass

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name, flags=re.IGNORECASE)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件3格式: {file_name}, 项目号: {project_id}")
            try:
                from core import Monitor
                Monitor.log_success(f"找到待处理文件3: 项目{project_id} - {file_name}")
            except Exception:
                pass
    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件3")
        try:
            from core import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件3，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except Exception:
            pass
    else:
        try:
            from core import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件3")
        except Exception:
            pass
    return matched_files


def find_target_file4(excel_files):
    """
    查找符合特定格式的待处理文件4（兼容性函数，返回第一个匹配的文件）
    格式：外部接口单报表+四位数字（项目号）+日期（8位）
    例如：外部接口单报表201620250801.xlsx
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files4(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files4(excel_files):
    """
    查找所有符合特定格式的待处理文件4
    格式：外部接口单报表+四位数字（项目号）+日期（8位）
    例如：外部接口单报表201620250801.xlsx
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^外部接口单报表(\d{4})\d{8}\.(xlsx|xlsm|xls)$'
    matched_files = []

    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件4...")
    except Exception:
        pass

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name, flags=re.IGNORECASE)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件4格式: {file_name}, 项目号: {project_id}")
            try:
                from core import Monitor
                Monitor.log_success(f"找到待处理文件4: 项目{project_id} - {file_name}")
            except Exception:
                pass

    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件4")
        try:
            from core import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件4，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except Exception:
            pass
    else:
        try:
            from core import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件4")
        except Exception:
            pass

    return matched_files


def find_target_file5(excel_files):
    """
    查找符合特定格式的待处理文件5（兼容性函数，返回第一个匹配的文件）
    格式：兼容“2016接口提资清单”和“三维接口提资清单1915...”命名
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files5(excel_files)
    if all_files:
        return all_files[0]
    return None, None


def find_all_target_files5(excel_files):
    """
    查找所有符合特定格式的待处理文件5
    格式：兼容“2016接口提资清单”和“三维接口提资清单1915...”命名
    返回：[(文件路径, 项目号), ...] 列表
    """
    matched_files = []
    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件5(三维提资接口)...")
    except Exception:
        pass
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        if not re.search(r'接口提资清单', file_name):
            continue
        if not re.search(r'\.(xlsx|xls|xlsm)$', file_name, flags=re.IGNORECASE):
            continue
        project_id = _extract_file5_project_id(file_path)
        if project_id:
            matched_files.append((file_path, project_id))
            try:
                from core import Monitor
                Monitor.log_success(f"找到待处理文件5: 项目{project_id} - {file_name}")
            except Exception:
                pass
    return matched_files


def find_target_file6(excel_files):
    """
    查找符合特定格式的待处理文件6（兼容性函数，返回第一个匹配的文件）
    规则：文件名中包含“收发文清单”，若其后紧随四位数字则作为项目号
         示例：收发文清单2016.xlsx → 项目号=2016；
         未紧随四位数字时，项目号为空字符串（兼容旧命名）
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files6(excel_files)
    if all_files:
        return all_files[0]
    return None, None


def find_all_target_files6(excel_files):
    """
    批量查找所有“收发文清单”文件。
    优先提取紧随其后的四位数字作为项目号；若未匹配，则项目号为空字符串。
    """
    matched_files = []
    try:
        from core import Monitor
        Monitor.log_process("开始批量识别待处理文件6(收发文函)...")
    except Exception:
        pass
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        if file_name.lower().endswith(('.xlsx', '.xlsm', '.xls')) and ("收发文清单" in file_name):
            # 优先匹配 紧随“收发文清单”的四位数字 作为项目号
            try:
                # 紧随“收发文清单”的四位数字作为项目号，例如：收发文清单2016.xlsx
                m = re.search(r"收发文清单(\d{4})", file_name)
                project_id = m.group(1) if m else ""
            except Exception:
                project_id = ""
            matched_files.append((file_path, project_id))
            try:
                from core import Monitor
                if project_id:
                    Monitor.log_success(f"找到待处理文件6: 项目{project_id} - {file_name}")
                else:
                    Monitor.log_success(f"找到待处理文件6(未识别项目号): {file_name}")
            except Exception:
                pass
    return matched_files


def find_target_file7(excel_files):
    """Find the first FU workbook named '<project>项目标准表格'."""
    matched_files = find_all_target_files7(excel_files)
    return matched_files[0] if matched_files else (None, None)


def find_all_target_files7(excel_files):
    """Find FU workbooks and return ``(path, project_id)`` pairs."""
    matched_files = []
    pattern = re.compile(r"^(\d{4})项目标准表格\.(?:xlsx|xlsm|xls)$", re.IGNORECASE)
    for file_path in excel_files:
        match = pattern.fullmatch(os.path.basename(file_path))
        if match:
            matched_files.append((file_path, match.group(1)))
    return matched_files


# ============================================================
# 当前精简列流式处理与导出实现
# ============================================================

STREAM_RESULT_SCHEMA_VERSION = "selected_columns_v5"

STREAM_EXPORT_COLUMNS = [
    "状态",
    "项目号",
    "接口号",
    "接口时间",
    "科室",
    "主办室",
    "责任人",
    "原始行号",
]


@lru_cache(maxsize=256)
def _col_to_index(col_letter):
    value = 0
    for ch in str(col_letter).strip().upper():
        if not ("A" <= ch <= "Z"):
            continue
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value - 1


@lru_cache(maxsize=256)
def _index_to_col(one_based_index):
    value = int(one_based_index)
    if value <= 0:
        raise ValueError(f"列序号必须为正整数: {one_based_index}")
    letters = []
    while value:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def _cell_to_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _is_empty_cell(value):
    text = _cell_to_text(value)
    return text == "" or text.lower() in ("nan", "none", "nat")


def _extract_chinese_text(value, default=""):
    text = _cell_to_text(value)
    if not text:
        return default
    found = re.findall(r"[\u4e00-\u9fa5]+", text)
    result = "".join(found)
    return result if result else default


def _normalize_interface_id(value):
    text = _cell_to_text(value)
    if not text:
        return ""
    return re.sub(r"\([^)]*\)$", "", text).strip()


def _parse_datetime_value(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)

    text = _cell_to_text(value)
    if not text:
        return None

    java_date = re.fullmatch(
        r"(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
        r"(\d{1,2})\s+(\d{2}):(\d{2}):(\d{2})\s+"
        r"(?:CST|GMT\+?8|Asia/Shanghai)\s+(\d{4})",
        text,
        flags=re.IGNORECASE,
    )
    if java_date:
        month_map = {
            name.lower(): month
            for month, name in enumerate(
                ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"),
                start=1,
            )
        }
        try:
            return datetime.datetime(
                int(java_date.group(6)),
                month_map[java_date.group(1).lower()],
                int(java_date.group(2)),
                int(java_date.group(3)),
                int(java_date.group(4)),
                int(java_date.group(5)),
            )
        except ValueError:
            return None

    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
    )
    for fmt in formats:
        try:
            return datetime.datetime.strptime(text, fmt)
        except Exception:
            continue

    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        if hasattr(parsed, "to_pydatetime"):
            return parsed.to_pydatetime()
        if isinstance(parsed, datetime.datetime):
            return parsed
    except Exception:
        return None
    return None


def _date_window(current_datetime):
    current_day = current_datetime.day
    current_year = current_datetime.year
    current_month = current_datetime.month
    start_date = datetime.datetime(current_year, 1, 1)
    if current_day <= 19:
        if current_month == 12:
            end_date = datetime.datetime(current_year, 12, 31)
        else:
            end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
    else:
        if current_month == 12:
            end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
        elif current_month == 11:
            end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    return start_date, end_date


def _in_month_window(value, current_datetime, project_id=None, exclude_4444=False):
    if exclude_4444 and _cell_to_text(value).startswith("4444"):
        return False
    parsed = _parse_datetime_value(value)
    if parsed is None:
        return False
    try:
        parsed = adjust_date_for_project(parsed, project_id)
    except Exception:
        pass
    start_date, end_date = _date_window(current_datetime)
    return start_date.date() <= parsed.date() <= end_date.date()


def _format_date_value(value):
    parsed = _parse_datetime_value(value)
    return parsed.strftime("%Y.%m.%d") if parsed is not None else ""


def _filter_valid_owner_names(names_text, valid_names_set):
    """仅保留角色表中存在的责任人姓名。"""
    if not names_text or not valid_names_set:
        return names_text or ""
    filtered = []
    for name in str(names_text).split(","):
        normalized = name.strip()
        if normalized and normalized in valid_names_set:
            filtered.append(normalized)
    return ",".join(filtered)


def _split_owner_names(value, valid_names_set=None):
    text = _cell_to_text(value)
    if not text:
        return ""
    for sep in [",", "，", ";", "；", "/", "、"]:
        text = text.replace(sep, ",")
    tokens = [item.strip() for item in text.split(",") if item.strip()]
    names = ",".join(tokens)
    if valid_names_set:
        try:
            filtered = _filter_valid_owner_names(names, valid_names_set)
            if filtered:
                names = filtered
        except Exception:
            pass
    return names


STREAM_FILE_SPECS = {
    1: {
        "columns": {"A": "a", "B": "discard", "H": "dept", "K": "time", "M": "completed", "R": "responsible"},
        "interface": "A",
        "time": "M",
        "completed": "M",
        "responsible": "R",
        "assign": "R",
        "sheet": "内部需打开接口",
        "filename": "内部需打开接口",
    },
    2: {
        "columns": {"A": "a", "E": "version", "F": "transfer", "I": "dept", "M": "time", "N": "completed", "R": "interface", "AB": "exclude", "AM": "responsible"},
        "interface": "R",
        "time": "N",
        "completed": "N",
        "responsible": "AM",
        "assign": "AM",
        "sheet": "内部需回复接口",
        "filename": "内部需回复接口",
    },
    3: {
        "columns": {"C": "interface", "I": "status", "L": "time_l", "M": "time_m", "Q": "completed_q", "T": "completed_t", "AC": "version", "AL": "org", "AO": "dept", "AP": "responsible"},
        "interface": "C",
        "time": None,
        "completed": None,
        "responsible": "AP",
        "assign": "AP",
        "sheet": "外部需打开接口",
        "filename": "外部需打开接口",
    },
    4: {
        "columns": {"E": "interface", "I": "version", "P": "status_p", "S": "time", "V": "completed", "AC": "status_ac", "AF": "org", "AG": "dept", "AH": "responsible"},
        "interface": "E",
        "time": "V",
        "completed": "V",
        "responsible": "AH",
        "assign": "AH",
        "sheet": "外部需回复接口",
        "filename": "外部需回复接口",
    },
    5: {
        "columns": {"A": "interface", "G": "dept", "K": "responsible", "L": "time", "N": "completed"},
        "interface": "A",
        "time": "N",
        "completed": "N",
        "responsible": "K",
        "assign": "K",
        "sheet": "三维提资接口",
        "filename": "三维提资接口",
    },
    6: {
        "columns": {"E": "interface", "I": "time", "J": "completed", "M": "reply_status", "V": "org", "W": "host_office", "X": "responsible", "AC": "version"},
        "interface": "E",
        "time": "J",
        "completed": "J",
        "responsible": "X",
        "assign": "X",
        "sheet": "收发文函",
        "filename": "收发文函",
    },
    7: {
        "columns": {"A": "file_code", "B": "internal_code", "C": "title", "D": "actual_fu", "E": "fu_plan", "F": "responsible"},
        "interface": "B",
        "time": "D",
        "completed": "D",
        "responsible": "F",
        "assign": "F",
        "sheet": "FU",
        "filename": "FU",
        "export_columns": ["原始行号", "状态", "项目号", "内部编码", "中文标题", "FU计划", "实际FU日期", "责任人"],
    },
}


def _normalize_stream_header(value):
    text = unicodedata.normalize("NFKC", _cell_to_text(value))
    text = re.sub(r"[\s\r\n]+", "", text)
    return text.replace("（", "(").replace("）", ")")


FILE6_HEADER_ALIASES = {
    "interface": ("收发文编号", "收文编号", "发文编号"),
    "time": ("要求回文期限", "要求回复期限", "计划回文日期"),
    "completed": ("我方回文日期", "实际回文日期", "回文日期"),
    "reply_status": ("回文状态", "回复状态"),
    "org": ("主办部门(所)", "主办部门", "主办所"),
    "host_office": ("主办室", "主办科室"),
    "responsible": ("主办人", "责任人"),
    "version": ("版次", "版本"),
}


def _resolve_stream_spec_from_headers(file_type, headers):
    """按真实表头解析易变模板；未识别到业务模板时保留旧列位兼容。"""
    base_spec = STREAM_FILE_SPECS[file_type]
    if file_type == 2:
        normalized_headers = {
            _normalize_stream_header(value): col_letter
            for col_letter, value in headers.items()
            if _normalize_stream_header(value)
        }
        responsible_column = normalized_headers.get(_normalize_stream_header("程序主办人"))
        if not responsible_column:
            return base_spec
        spec = copy(base_spec)
        spec["columns"] = {
            letter: alias
            for letter, alias in base_spec["columns"].items()
            if alias != "responsible"
        }
        spec["columns"][responsible_column] = "responsible"
        spec["responsible"] = responsible_column
        spec["assign"] = responsible_column
        return spec
    if file_type != 6:
        return base_spec

    normalized_headers = {
        _normalize_stream_header(value): col_letter
        for col_letter, value in headers.items()
        if _normalize_stream_header(value)
    }
    resolved = {}
    for alias, candidates in FILE6_HEADER_ALIASES.items():
        for candidate in candidates:
            col_letter = normalized_headers.get(_normalize_stream_header(candidate))
            if col_letter:
                resolved[alias] = col_letter
                break

    # 没有识别到文件6的锚点时，兼容历史模板及测试构造文件的固定列位。
    if "interface" not in resolved:
        return base_spec

    missing = [alias for alias in FILE6_HEADER_ALIASES if alias not in resolved]
    if missing:
        missing_text = "、".join(missing)
        raise ValueError(f"收发文函表头不完整，缺少字段: {missing_text}")

    spec = copy(base_spec)
    spec["columns"] = {
        resolved["interface"]: "interface",
        resolved["time"]: "time",
        resolved["completed"]: "completed",
        resolved["reply_status"]: "reply_status",
        resolved["org"]: "org",
        resolved["host_office"]: "host_office",
        resolved["responsible"]: "responsible",
        resolved["version"]: "version",
    }
    spec["interface"] = resolved["interface"]
    spec["time"] = resolved["time"]
    spec["completed"] = resolved["completed"]
    spec["responsible"] = resolved["responsible"]
    spec["assign"] = resolved["responsible"]
    return spec


def _build_selected_record(spec, row_number, value_getter):
    raw = {}
    for letter, alias in spec["columns"].items():
        raw[alias] = value_getter(_col_to_index(letter))
    return {
        "_df_index": row_number - 2,
        "原始行号": row_number,
        "_raw": raw,
        "_stream_spec": spec,
    }


def _read_xlsx_physical_records(workbook, worksheet, file_type, base_spec):
    """直接遍历工作表XML中的实际行；不信任dimension，也不构造缺失空单元格。"""
    from xml.etree.ElementTree import iterparse

    from openpyxl.utils.datetime import from_ISO8601, from_excel
    from openpyxl.worksheet._reader import _cast_number

    header_probe_max = 64 if file_type in (2, 6) else (
        max(_col_to_index(letter) for letter in base_spec["columns"]) + 1
    )
    records = []
    spec = None
    required_columns = []
    selected_columns = set()
    selected_column_numbers = {}

    def parse_value(cell):
        data_type = cell.get("t", "n")
        style_id = int(cell.get("s", 0) or 0)
        if data_type == "inlineStr":
            return "".join(
                node.text or ""
                for node in cell.iter()
                if node.tag.endswith("}t") or node.tag == "t"
            )
        value_node = next(
            (node for node in cell if node.tag.endswith("}v") or node.tag == "v"),
            None,
        )
        value = value_node.text if value_node is not None else None
        if value is None:
            return None
        if data_type == "s":
            return worksheet._shared_strings[int(value)]
        if data_type == "b":
            return bool(int(value))
        if data_type == "d":
            return from_ISO8601(value)
        if data_type in {"str", "e"}:
            return value
        number = _cast_number(value)
        if style_id in workbook._date_formats:
            try:
                return from_excel(
                    number,
                    workbook.epoch,
                    timedelta=style_id in workbook._timedelta_formats,
                )
            except (OverflowError, ValueError):
                return value
        return number

    with worksheet._get_source() as source:
        row_counter = 0
        for _event, element in iterparse(source, events=("end",)):
            if not (element.tag.endswith("}row") or element.tag == "row"):
                continue
            row_counter += 1
            try:
                row_number = int(element.get("r") or row_counter)
            except (TypeError, ValueError):
                row_number = row_counter
            values = {}
            fallback_column = 0
            for cell in element:
                if not (cell.tag.endswith("}c") or cell.tag == "c"):
                    continue
                fallback_column += 1
                coordinate = str(cell.get("r") or "")
                split_at = 0
                coordinate_length = len(coordinate)
                while split_at < coordinate_length and coordinate[split_at].isalpha():
                    split_at += 1
                column_letters = coordinate[:split_at].upper()
                if row_number != 1 and selected_column_numbers and column_letters:
                    column_number = selected_column_numbers.get(column_letters)
                    if column_number is None:
                        continue
                elif column_letters:
                    column_number = _col_to_index(column_letters) + 1
                else:
                    column_number = fallback_column
                if row_number == 1:
                    if column_number > header_probe_max:
                        continue
                elif selected_columns and column_number not in selected_columns:
                    continue
                values[column_number] = parse_value(cell)
            if row_number == 1:
                headers = {
                    _index_to_col(column_number): values.get(column_number)
                    for column_number in range(1, header_probe_max + 1)
                }
                spec = _resolve_stream_spec_from_headers(file_type, headers)
                required_columns = sorted(spec["columns"], key=_col_to_index)
                selected_columns = {
                    _col_to_index(letter) + 1
                    for letter in spec["columns"]
                }
                selected_column_numbers = {
                    str(letter).upper(): _col_to_index(letter) + 1
                    for letter in spec["columns"]
                }
            elif row_number >= 2:
                if spec is None:
                    spec = _resolve_stream_spec_from_headers(file_type, {})
                    required_columns = sorted(spec["columns"], key=_col_to_index)
                    selected_columns = {
                        _col_to_index(letter) + 1
                        for letter in spec["columns"]
                    }
                    selected_column_numbers = {
                        str(letter).upper(): _col_to_index(letter) + 1
                        for letter in spec["columns"]
                    }
                records.append(_build_selected_record(
                    spec,
                    row_number,
                    lambda zero_idx, values=values: values.get(zero_idx + 1),
                ))
            element.clear()

    if spec is None:
        spec = _resolve_stream_spec_from_headers(file_type, {})
        required_columns = sorted(spec["columns"], key=_col_to_index)
    return records, required_columns


def _read_sparse_fu_records_ooxml(file_path, base_spec=None):
    """Read only real A-F cells from the first FU worksheet.

    Some production FU workbooks contain more than one million empty ``row``
    elements.  Iterating actual ``c`` elements avoids materialising those rows
    while retaining Excel row numbers, scalar types and date conversion.
    """
    from utils.excel_io import OoxmlWorksheetSnapshot, worksheet_archive_path_by_index

    base_spec = base_spec or STREAM_FILE_SPECS[7]
    sheet_path = worksheet_archive_path_by_index(file_path, 0)
    snapshot = OoxmlWorksheetSnapshot(file_path, sheet_path=sheet_path)
    selected_letters = {str(letter).upper() for letter in base_spec["columns"]}
    header_values = {}
    row_values = {}

    for reference, cell_xml in snapshot.iter_cells():
        match = re.fullmatch(r"([A-Za-z]+)([1-9][0-9]*)", reference)
        if not match:
            continue
        column_letter = match.group(1).upper()
        if column_letter not in selected_letters:
            continue
        row_number = int(match.group(2))
        value = snapshot.typed_value_from_cell_xml(cell_xml)
        if row_number == 1:
            header_values[column_letter] = value
        elif row_number >= 2:
            row_values.setdefault(row_number, {})[column_letter] = value

    spec = _resolve_stream_spec_from_headers(7, header_values)
    required_columns = sorted(spec["columns"], key=_col_to_index)
    records = [
        _build_selected_record(
            spec,
            row_number,
            lambda zero_idx, values=values: values.get(_index_to_col(zero_idx + 1)),
        )
        for row_number, values in sorted(row_values.items())
    ]
    return records, required_columns


def _read_sparse_fu_records_openpyxl(file_path, base_spec=None):
    """Compatibility reader kept as a correctness fallback and test oracle."""
    from openpyxl import load_workbook

    base_spec = base_spec or STREAM_FILE_SPECS[7]
    records = []
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb.worksheets[0]
        header_probe_max = max(
            _col_to_index(letter) for letter in base_spec["columns"]
        ) + 1
        headers = {
            _index_to_col(col_idx): ws.cell(1, col_idx).value
            for col_idx in range(1, header_probe_max + 1)
        }
        spec = _resolve_stream_spec_from_headers(7, headers)
        required_columns = sorted(spec["columns"], key=_col_to_index)
        required_indices = {_col_to_index(letter) for letter in spec["columns"]}
        max_col = max(required_indices) + 1 if required_indices else 1
        physical_rows = sorted({
            row_idx
            for row_idx, col_idx in ws._cells
            if row_idx >= 2 and col_idx <= max_col
        })
        for row_number in physical_rows:
            records.append(_build_selected_record(
                spec,
                row_number,
                lambda zero_idx, row_number=row_number: ws.cell(
                    row_number, zero_idx + 1
                ).value,
            ))
    finally:
        wb.close()
    return records, required_columns


def _read_selected_excel_records(file_path, file_type):
    base_spec = STREAM_FILE_SPECS[file_type]
    records = []
    lower_path = str(file_path).lower()

    if lower_path.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        if file_type == 7:
            try:
                return _read_sparse_fu_records_ooxml(file_path, base_spec)
            except Exception as exc:
                # Unsupported producer-specific OOXML must never change the
                # business result: fall back to the proven workbook reader.
                print(f"FU快速读取失败，回退兼容模式: {exc}")
                return _read_sparse_fu_records_openpyxl(file_path, base_spec)

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            return _read_xlsx_physical_records(wb, ws, file_type, base_spec)
        finally:
            wb.close()

    if lower_path.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(0)
        header_probe_max = min(sheet.ncols, 64 if file_type in (2, 6) else sheet.ncols)
        headers = {
            _index_to_col(col_idx + 1): sheet.cell_value(0, col_idx)
            for col_idx in range(header_probe_max)
        }
        spec = _resolve_stream_spec_from_headers(file_type, headers)
        required_columns = sorted(spec["columns"], key=_col_to_index)
        required_indices = {
            _col_to_index(letter): (letter, alias)
            for letter, alias in spec["columns"].items()
        }
        for row_idx in range(1, sheet.nrows):
            raw = {}
            for col_idx, (letter, alias) in required_indices.items():
                if col_idx >= sheet.ncols:
                    raw[alias] = None
                    continue
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                raw[alias] = value
            records.append({
                "_df_index": row_idx - 1,
                "原始行号": row_idx + 1,
                "_raw": raw,
                "_stream_spec": spec,
            })
        return records, required_columns

    raise ValueError(f"不支持的Excel格式: {file_path}")


def _records_to_registry_index_df(records, file_type, file_path, project_id):
    spec = records[0].get("_stream_spec") if records else STREAM_FILE_SPECS[file_type]
    interface_alias = spec["columns"][spec["interface"]]
    fallback_project_id = _extract_file5_project_id(file_path) if file_type == 5 else _extract_file_project_id(file_path)
    rows = []
    for record in records:
        raw = record["_raw"]
        rows.append({
            "项目号": project_id or fallback_project_id,
            "接口号": _cell_to_text(raw.get(interface_alias)),
            "source_file": os.path.abspath(file_path),
            "原始行号": record["原始行号"],
        })
    return pd.DataFrame(rows)


def _highest_version_rows(records, interface_alias, version_alias):
    if not version_alias:
        return {record["_df_index"] for record in records if record["_df_index"] >= 0}

    best_rank = {}
    best_rows = {}
    keep_rows = set()
    for record in records:
        idx = record["_df_index"]
        if idx < 0:
            continue
        raw = record["_raw"]
        interface_id = _normalize_interface_id(raw.get(interface_alias))
        if not interface_id:
            keep_rows.add(idx)
            continue
        rank = _get_version_rank(raw.get(version_alias))
        if interface_id not in best_rank or rank > best_rank[interface_id]:
            best_rank[interface_id] = rank
            best_rows[interface_id] = [idx]
        elif rank == best_rank[interface_id]:
            best_rows[interface_id].append(idx)

    for row_list in best_rows.values():
        keep_rows.update(row_list)
    return keep_rows


def _standard_result_row(
    record,
    *,
    file_type,
    file_path,
    project_id,
    interface_time,
    department="",
    host_office="",
    responsible="",
    source_column="",
    completed_col=None,
    completed_value=None,
    interface_id=None,
):
    spec = record.get("_stream_spec") or STREAM_FILE_SPECS[file_type]
    raw = record["_raw"]
    interface_alias = spec["columns"][spec["interface"]]
    completed_letter = completed_col or spec["completed"] or ""
    completed_alias = spec["columns"].get(completed_letter) if completed_letter else None
    if completed_value is None and completed_alias:
        completed_value = raw.get(completed_alias)

    row = {
        "项目号": project_id or _extract_file_project_id(file_path),
        "接口号": _cell_to_text(interface_id if interface_id is not None else raw.get(interface_alias)),
        "接口时间": interface_time,
        "科室": department,
        "主办室": host_office,
        "责任人": responsible,
        "原始行号": record["原始行号"],
        "source_file": os.path.abspath(file_path),
        "_file_type": file_type,
        "_interface_col": spec["interface"],
        "_time_col": completed_letter,
        "_completed_col": completed_letter,
        "_responsible_col": spec["responsible"],
        "_assign_col": spec["assign"],
        "_source_column": source_column or "",
        "_completed_col_value": _cell_to_text(completed_value),
        "_stream_schema_version": STREAM_RESULT_SCHEMA_VERSION,
    }
    return row


def _finalize_stream_result(records_by_idx, final_rows, build_row, file_type, include_zero=True):
    minimum_index = 0 if include_zero else 1
    final_indices = sorted(idx for idx in final_rows if idx >= minimum_index and idx in records_by_idx)
    if not final_indices:
        result_df = pd.DataFrame(columns=STREAM_EXPORT_COLUMNS + [
            "source_file",
            "_file_type",
            "_interface_col",
            "_time_col",
            "_completed_col",
            "_responsible_col",
            "_assign_col",
            "_source_column",
            "_completed_col_value",
            "_stream_schema_version",
        ])
    else:
        rows = [build_row(records_by_idx[idx], idx) for idx in final_indices]
        result_df = pd.DataFrame(rows)
        result_df = apply_assignment_memory(result_df, file_type=file_type)
    result_df.attrs["_stream_schema_version"] = STREAM_RESULT_SCHEMA_VERSION
    return result_df


def _stream_process_file1(file_path, current_datetime):
    project_id = _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 1)
    records_by_idx = {record["_df_index"]: record for record in records}
    p1 = set()
    p2 = set()
    p3 = set()
    p4 = set()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if contains_department_code(_cell_to_text(raw.get("dept"))):
            p1.add(idx)
        if _in_month_window(raw.get("time"), current_datetime, project_id):
            p2.add(idx)
        if not _is_empty_cell(raw.get("a")) and _is_empty_cell(raw.get("completed")):
            p3.add(idx)
        if "作废" in _cell_to_text(raw.get("discard")):
            p4.add(idx)

    final_rows = p1 & p2 & p3 - p4
    final_rows, _ = _merge_registry_pending_rows(
        file_type=1,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=(p1 & p2) - p4,
        project_id=project_id,
    )

    def build(record, _idx):
        raw = record["_raw"]
        return _standard_result_row(
            record,
            file_type=1,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(raw.get("time")),
            department=map_code_to_department(_cell_to_text(raw.get("dept"))),
            responsible=_extract_chinese_text(raw.get("responsible")),
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 1)


def _stream_process_file2(file_path, current_datetime, project_id=None):
    project_id = project_id or _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 2)
    records_by_idx = {record["_df_index"]: record for record in records}
    version_allowed = _highest_version_rows(records, "interface", "version")
    standard_projects = set(get_projects_standard_filter())
    p1 = set()
    p2 = set()
    p3 = set()
    p4 = set()
    org_filter = get_organization_filter()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        dept_text = _cell_to_text(raw.get("dept"))
        if org_filter in dept_text or contains_department_code(dept_text):
            p1.add(idx)
        if _in_month_window(raw.get("time"), current_datetime, project_id):
            p2.add(idx)
        if _cell_to_text(raw.get("exclude")).startswith("4444") and _cell_to_text(raw.get("transfer")) == "传递":
            p3.add(idx)
        if not _is_empty_cell(raw.get("a")) and _is_empty_cell(raw.get("completed")):
            p4.add(idx)

    if str(project_id) in standard_projects:
        final_rows = p1 & p2 & p4
        allowed_rows = p1 & p2
    else:
        final_rows = p1 & p2 & p4 - p3
        allowed_rows = p1 & p2 - p3

    final_rows, _ = _merge_registry_pending_rows(
        file_type=2,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=allowed_rows,
        project_id=project_id,
    )
    final_rows = final_rows & version_allowed

    def build(record, _idx):
        raw = record["_raw"]
        dept_text = _cell_to_text(raw.get("dept"))
        matched = match_department_name(dept_text)
        department = matched if matched != dept_text else ""
        return _standard_result_row(
            record,
            file_type=2,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(raw.get("time")),
            department=department,
            responsible=_extract_chinese_text(raw.get("responsible"), default="无"),
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 2)


def _stream_process_file3(file_path, current_datetime):
    project_id = _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 3)
    records_by_idx = {record["_df_index"]: record for record in records}
    version_allowed = _highest_version_rows(records, "interface", "version")
    p1 = set()
    p2 = set()
    p3 = set()
    p4 = set()
    p5 = set()
    p6 = set()
    org_filter = get_organization_filter()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if _cell_to_text(raw.get("status")) == "B":
            p1.add(idx)
        if _cell_to_text(raw.get("org")).startswith(org_filter):
            p2.add(idx)
        if _in_month_window(raw.get("time_m"), current_datetime, project_id, exclude_4444=True):
            p3.add(idx)
        if _in_month_window(raw.get("time_l"), current_datetime, project_id, exclude_4444=True):
            p4.add(idx)
        if _is_empty_cell(raw.get("completed_q")):
            p5.add(idx)
        if _is_empty_cell(raw.get("completed_t")):
            p6.add(idx)

    source_m_rows = p1 & p2 & p3
    source_l_rows = p1 & p2 & p4
    group1 = source_m_rows & p6
    group2 = source_l_rows & p5
    final_rows = group1 | group2

    final_rows, _ = _merge_registry_pending_rows(
        file_type=3,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=p1 & p2 & (p3 | p4),
        project_id=project_id,
    )
    final_rows = final_rows & version_allowed

    def source_for_idx(idx):
        if idx in source_m_rows and idx not in source_l_rows:
            return "M"
        if idx in source_l_rows and idx not in source_m_rows:
            return "L"
        return "M"

    def build(record, idx):
        raw = record["_raw"]
        source_column = source_for_idx(idx)
        dept_text = _cell_to_text(raw.get("dept"))
        department = "请室主任确认" if not dept_text else match_department_name(dept_text)
        if source_column == "L":
            time_value = raw.get("time_l")
            completed_col = "Q"
            completed_value = raw.get("completed_q")
        else:
            time_value = raw.get("time_m")
            completed_col = "T"
            completed_value = raw.get("completed_t")
        return _standard_result_row(
            record,
            file_type=3,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(time_value),
            department=department,
            responsible=_extract_chinese_text(raw.get("responsible")),
            source_column=source_column,
            completed_col=completed_col,
            completed_value=completed_value,
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 3)


def _stream_process_file4(file_path, current_datetime):
    project_id = _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 4)
    records_by_idx = {record["_df_index"]: record for record in records}
    version_allowed = _highest_version_rows(records, "interface", "version")
    p1 = set()
    p2 = set()
    p3 = set()
    p4 = set()
    org_filter = get_organization_filter()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if _cell_to_text(raw.get("org")).startswith(org_filter):
            p1.add(idx)
        status_p = _cell_to_text(raw.get("status_p"))
        if status_p == "B" or (not status_p and _cell_to_text(raw.get("status_ac")) == "B"):
            p2.add(idx)
        if _in_month_window(raw.get("time"), current_datetime, project_id):
            p3.add(idx)
        if _is_empty_cell(raw.get("completed")):
            p4.add(idx)

    final_rows = p1 & p2 & p3 & p4
    final_rows, _ = _merge_registry_pending_rows(
        file_type=4,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=p1 & p2 & p3,
        allow_interface_only_fallback=True,
        project_id=project_id,
    )
    final_rows = final_rows & version_allowed

    def build(record, _idx):
        raw = record["_raw"]
        dept_text = _cell_to_text(raw.get("dept"))
        department = "请室主任确认" if not dept_text else match_department_name(dept_text)
        return _standard_result_row(
            record,
            file_type=4,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(raw.get("time")),
            department=department,
            responsible=_extract_chinese_text(raw.get("responsible")),
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 4)


def _stream_process_file5(file_path, current_datetime):
    project_id = _extract_file5_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 5)
    records_by_idx = {record["_df_index"]: record for record in records}
    p1 = set()
    p2 = set()
    p3 = set()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if contains_department_code(_cell_to_text(raw.get("dept"))):
            p1.add(idx)
        if _in_month_window(raw.get("time"), current_datetime, project_id):
            p2.add(idx)
        if _is_empty_cell(raw.get("completed")):
            p3.add(idx)

    final_rows = p1 & p2 & p3
    final_rows, _ = _merge_registry_pending_rows(
        file_type=5,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=p1 & p2,
        project_id=project_id,
    )

    def build(record, _idx):
        raw = record["_raw"]
        return _standard_result_row(
            record,
            file_type=5,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(raw.get("time")),
            department=map_code_to_department(_cell_to_text(raw.get("dept"))),
            responsible=_extract_chinese_text(raw.get("responsible")),
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 5)


def _stream_process_file6(file_path, current_datetime, skip_date_filter=False, valid_names_set=None):
    project_id = _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 6)
    records_by_idx = {record["_df_index"]: record for record in records}
    version_allowed = _highest_version_rows(records, "interface", "version")
    p1 = set()
    p_i = set()
    p3 = set()
    p4 = set()
    org_filter = get_organization_filter_file6()
    today = current_datetime.date()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if org_filter in _cell_to_text(raw.get("org")):
            p1.add(idx)
        parsed = _parse_datetime_value(raw.get("time"))
        if parsed is not None:
            p_i.add(idx)
            try:
                adjusted = adjust_date_for_project(parsed, project_id)
            except Exception:
                adjusted = parsed
            if (adjusted.date() - today).days <= 14:
                p3.add(idx)
        if _cell_to_text(raw.get("reply_status")) in ("尚未回复", "超期未回复"):
            p4.add(idx)

    if skip_date_filter:
        final_rows = p1 & p_i & p4
        allowed_rows = p1 & p_i
    else:
        final_rows = p1 & p_i & p3 & p4
        allowed_rows = p1 & p_i & p3

    final_rows, _ = _merge_registry_pending_rows(
        file_type=6,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=allowed_rows,
        project_id=project_id,
    )
    final_rows = final_rows & version_allowed

    def build(record, _idx):
        raw = record["_raw"]
        return _standard_result_row(
            record,
            file_type=6,
            file_path=file_path,
            project_id=project_id,
            interface_time=_format_date_value(raw.get("time")),
            department="",
            host_office=_cell_to_text(raw.get("host_office")),
            responsible=_split_owner_names(raw.get("responsible"), valid_names_set),
        )

    return _finalize_stream_result(records_by_idx, final_rows, build, 6)


def _assign_file7_interface_ids(records):
    """Bind duplicate internal codes to deterministic per-row interface IDs."""
    grouped = {}
    for record in records:
        raw = record.get("_raw") or {}
        internal_code = _cell_to_text(raw.get("internal_code")).strip()
        if internal_code:
            grouped.setdefault(internal_code, []).append(record)

    for internal_code, group in grouped.items():
        if len(group) == 1:
            group[0]["接口号"] = internal_code
            continue
        ordered = sorted(
            group,
            key=lambda item: (
                _cell_to_text(item["_raw"].get("file_code")),
                _cell_to_text(item["_raw"].get("title")),
                int(item.get("原始行号", 0)),
            ),
        )
        for sequence, record in enumerate(ordered, start=1):
            record["接口号"] = f"{internal_code}#{sequence:02d}"


def _stream_process_file7(file_path, current_datetime):
    project_id = _extract_file_project_id(file_path)
    records, _ = _read_selected_excel_records(file_path, 7)
    _assign_file7_interface_ids(records)
    records_by_idx = {record["_df_index"]: record for record in records}

    plan_rows = set()
    open_rows = set()
    for record in records:
        idx = record["_df_index"]
        raw = record["_raw"]
        if not _is_empty_cell(raw.get("fu_plan")) and _in_month_window(
            raw.get("fu_plan"), current_datetime, project_id
        ):
            plan_rows.add(idx)
            if _is_empty_cell(raw.get("actual_fu")):
                open_rows.add(idx)

    # A confirmed FU cycle stays closed if D is cleared accidentally. A changed
    # E plan with D cleared is a new cycle under the same business_id.
    try:
        from registry.hooks import _cfg

        cfg = _cfg()
        archived_plans = _load_latest_confirmed_archive_times(
            7,
            project_id,
            cfg.get("registry_db_path"),
            bool(cfg.get("registry_wal", False)),
        )
    except Exception:
        archived_plans = {}

    final_rows = set()
    for idx in open_rows:
        record = records_by_idx[idx]
        interface_id = str(record.get("接口号") or "").strip()
        current_plan = _format_date_value(record["_raw"].get("fu_plan"))
        if archived_plans.get(interface_id) == current_plan:
            continue
        final_rows.add(idx)

    final_rows, _ = _merge_registry_pending_rows(
        file_type=7,
        file_path=file_path,
        df=records,
        final_rows=final_rows,
        allowed_rows=plan_rows,
        project_id=project_id,
    )

    def build(record, _idx):
        raw = record["_raw"]
        interface_id = str(record.get("接口号") or "").strip()
        row = _standard_result_row(
            record,
            file_type=7,
            file_path=file_path,
            project_id=project_id,
            interface_id=interface_id,
            interface_time=_format_date_value(raw.get("fu_plan")),
            department="请室主任确认",
            responsible=_extract_chinese_text(raw.get("responsible")),
        )
        row.update({
            "内部编码": interface_id,
            "中文标题": _cell_to_text(raw.get("title")),
            "FU计划": _format_date_value(raw.get("fu_plan")),
            "实际FU日期": _format_date_value(raw.get("actual_fu")),
        })
        return row

    return _finalize_stream_result(
        records_by_idx, final_rows, build, 7, include_zero=True
    )


def _log_stream_process_start(file_type, file_path):
    print(f"开始流式处理文件{file_type}: {os.path.basename(file_path)}")
    try:
        from core import Monitor
        Monitor.log_process(f"开始流式处理文件{file_type}: {os.path.basename(file_path)}")
    except Exception:
        pass


def process_target_file(file_path, current_datetime):
    _log_stream_process_start(1, file_path)
    return _stream_process_file1(file_path, current_datetime)


def process_target_file2(file_path, current_datetime, project_id=None):
    _log_stream_process_start(2, file_path)
    return _stream_process_file2(file_path, current_datetime, project_id)


def process_target_file3(file_path, current_datetime):
    _log_stream_process_start(3, file_path)
    return _stream_process_file3(file_path, current_datetime)


def process_target_file4(file_path, current_datetime):
    _log_stream_process_start(4, file_path)
    return _stream_process_file4(file_path, current_datetime)


def process_target_file5(file_path, current_datetime):
    _log_stream_process_start(5, file_path)
    return _stream_process_file5(file_path, current_datetime)


def process_target_file6(file_path, current_datetime, skip_date_filter=False, valid_names_set=None):
    _log_stream_process_start(6, file_path)
    return _stream_process_file6(file_path, current_datetime, skip_date_filter, valid_names_set)


def process_target_file7(file_path, current_datetime):
    _log_stream_process_start(7, file_path)
    return _stream_process_file7(file_path, current_datetime)


def _export_stream_result(df, current_datetime, output_dir, project_id, file_type):
    from openpyxl import Workbook

    spec = STREAM_FILE_SPECS[file_type]
    final_output_dir = output_dir
    if project_id:
        final_output_dir = os.path.join(output_dir, f"{project_id}结果文件")
    os.makedirs(final_output_dir, exist_ok=True)

    date_str = current_datetime.strftime("%Y-%m-%d") if hasattr(current_datetime, "strftime") else str(current_datetime)[:10]
    base_filename = f"{spec['filename']}{date_str}"
    output_path = os.path.join(final_output_dir, f"{base_filename}.xlsx")
    counter = 1
    while os.path.exists(output_path):
        output_path = os.path.join(final_output_dir, f"{base_filename}({counter}).xlsx")
        counter += 1

    export_columns = spec.get("export_columns", STREAM_EXPORT_COLUMNS)
    export_df = pd.DataFrame()
    for col in export_columns:
        if df is not None and col in getattr(df, "columns", []):
            export_df[col] = df[col]
        else:
            export_df[col] = [""] * (0 if df is None else len(df))

    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet"]
    ws.append(export_columns)
    for _, row in export_df.iterrows():
        ws.append([row.get(col, "") for col in export_columns])

    for col_idx, col_name in enumerate(export_columns, start=1):
        max_width = max([len(str(col_name))] + [
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(2, ws.max_row + 1)
        ])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_width * 1.3, 10), 60)

    wb.save(output_path)
    wb.close()
    try:
        from core import Monitor
        Monitor.log_success(f"{spec['sheet']}导出完成: {output_path}")
    except Exception:
        pass
    return output_path


def export_result_to_excel(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 1)


def export_result_to_excel2(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 2)


def export_result_to_excel3(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 3)


def export_result_to_excel4(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 4)


def export_result_to_excel5(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 5)


def export_result_to_excel6(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 6)


def export_result_to_excel7(df, original_file_path, current_datetime, output_dir, project_id=None):
    return _export_stream_result(df, current_datetime, output_dir, project_id, 7)


if __name__ == "__main__":
    # 如果直接运行此文件，显示提示信息
    print("Excel数据处理模块已加载")
    print("请通过主程序(base.py)来使用此模块")
