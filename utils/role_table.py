"""姓名角色表读取工具。

角色表历史上同时存在“首行为表头”和“首行即人员数据”两种格式。
本模块统一按 A 列姓名、B 列角色读取，避免无表头文件漏掉第一位人员。
"""

import os

import pandas as pd
from openpyxl import load_workbook


def _is_header_row(name_value, role_value):
    """判断首行是否为角色表表头。"""
    name_text = "" if name_value is None else str(name_value).strip()
    role_text = "" if role_value is None else str(role_value).strip()
    return ("姓名" in name_text) or ("角色" in role_text)


def read_role_table(file_path):
    """读取姓名角色表并返回固定列 ``姓名``、``角色`` 的 DataFrame。"""
    lower_path = os.fspath(file_path).lower()
    rows = []

    if lower_path.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            for row in worksheet.iter_rows(min_col=1, max_col=2, values_only=True):
                rows.append((row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None))
        finally:
            workbook.close()
    elif lower_path.endswith(".xls"):
        raw = pd.read_excel(file_path, sheet_name=0, engine="xlrd", header=None, usecols=[0, 1])
        rows = list(raw.itertuples(index=False, name=None))
    else:
        raise ValueError("不支持的姓名角色表格式: {}".format(file_path))

    if rows and _is_header_row(rows[0][0], rows[0][1]):
        rows = rows[1:]

    return pd.DataFrame(rows, columns=["姓名", "角色"])
