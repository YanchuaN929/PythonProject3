#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工作簿安全读写辅助工具。
"""

from __future__ import annotations

import base64
import copy
import datetime
import html
import os
import posixpath
import re
import subprocess
import tempfile
import threading
import time
import zipfile
from xml.sax.saxutils import escape as xml_escape
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


REPLACE_RETRY_DELAYS = (0.5, 1.0, 2.0, 4.0, 8.0)
DEFAULT_WORKBOOK_LOCK_TIMEOUT_SECONDS = 12.0
WORKBOOK_LOCK_RETRY_INTERVAL_SECONDS = 0.08

_LOCAL_WORKBOOK_LOCKS = {}
_LOCAL_WORKBOOK_LOCKS_GUARD = threading.Lock()


class ExcelWriteError(RuntimeError):
    """携带失败阶段与是否已落盘信息的Excel写入异常。"""

    def __init__(
        self,
        code: str,
        stage: str,
        message: str,
        *,
        retryable: bool = False,
        committed: bool = False,
    ):
        self.code = str(code or "EXCEL_WRITE_FAILED")
        self.stage = str(stage or "UNKNOWN")
        self.detail = str(message or "Excel写入失败")
        self.retryable = bool(retryable)
        self.committed = bool(committed)
        super().__init__(
            f"[{self.code}][{self.stage}] {self.detail}"
            f"（已落盘：{'是' if self.committed else '否'}）"
        )


def _get_local_workbook_lock(file_path: str) -> threading.Lock:
    key = os.path.normcase(os.path.abspath(os.fspath(file_path)))
    with _LOCAL_WORKBOOK_LOCKS_GUARD:
        lock = _LOCAL_WORKBOOK_LOCKS.get(key)
        if lock is None:
            lock = threading.Lock()
            _LOCAL_WORKBOOK_LOCKS[key] = lock
        return lock


def _is_lock_contention(error: Exception) -> bool:
    return (
        isinstance(error, (BlockingIOError, PermissionError))
        or getattr(error, "errno", None) in {11, 13}
        or getattr(error, "winerror", None) in {32, 33, 36}
    )


class SharedWorkbookLock:
    """同一工作簿的跨线程、跨进程短时写锁，适用于Windows/SMB公共盘。"""

    def __init__(self, file_path: str, timeout: float = DEFAULT_WORKBOOK_LOCK_TIMEOUT_SECONDS):
        self.file_path = os.path.abspath(os.fspath(file_path))
        self.lock_path = f"{self.file_path}.interface.lock"
        self.timeout = max(0.0, float(timeout))
        self._local_lock = _get_local_workbook_lock(self.file_path)
        self._handle = None
        self._local_acquired = False

    def _try_lock_handle(self):
        handle = open(self.lock_path, "a+b")
        try:
            handle.seek(0, os.SEEK_END)
            if handle.tell() == 0:
                handle.write(b"0")
                handle.flush()
            handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:  # pragma: no cover - 发布环境为Windows
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            return handle
        except Exception:
            handle.close()
            raise

    def __enter__(self):
        deadline = time.monotonic() + self.timeout
        remaining = max(0.0, deadline - time.monotonic())
        self._local_acquired = self._local_lock.acquire(timeout=remaining)
        if not self._local_acquired:
            raise ExcelWriteError(
                "WORKBOOK_BUSY",
                "ACQUIRE_LOCK",
                "同一工作簿正在由本机另一个写入任务处理，请稍后重试。",
                retryable=True,
                committed=False,
            )

        try:
            while True:
                try:
                    self._handle = self._try_lock_handle()
                    return self
                except Exception as exc:
                    if not _is_lock_contention(exc):
                        raise ExcelWriteError(
                            "LOCK_CREATE_FAILED",
                            "ACQUIRE_LOCK",
                            f"无法在目标目录建立工作簿写锁：{exc}",
                            retryable=isinstance(exc, OSError),
                            committed=False,
                        ) from exc
                    if time.monotonic() >= deadline:
                        raise ExcelWriteError(
                            "WORKBOOK_BUSY",
                            "ACQUIRE_LOCK",
                            "该工作簿正在由其他客户端写入，请稍后重试。",
                            retryable=True,
                            committed=False,
                        ) from exc
                    time.sleep(WORKBOOK_LOCK_RETRY_INTERVAL_SECONDS)
        except Exception:
            self._local_lock.release()
            self._local_acquired = False
            raise

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            if self._handle is not None:
                try:
                    self._handle.seek(0)
                    if os.name == "nt":
                        import msvcrt

                        msvcrt.locking(self._handle.fileno(), msvcrt.LK_UNLCK, 1)
                    else:  # pragma: no cover - 发布环境为Windows
                        import fcntl

                        fcntl.flock(self._handle.fileno(), fcntl.LOCK_UN)
                finally:
                    self._handle.close()
                    self._handle = None
            try:
                os.remove(self.lock_path)
            except OSError:
                pass
        finally:
            if self._local_acquired:
                self._local_lock.release()
                self._local_acquired = False
        return False


def normalize_header_text(value) -> str:
    """将业务表头规范化为可稳定比较的文本。"""
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    return text.replace("(", "（").replace(")", "）")


def column_number_to_letter(column_number: int) -> str:
    value = int(column_number)
    if value <= 0:
        raise ValueError(f"无效Excel列序号: {column_number}")
    result = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def find_header_column(worksheet, expected_headers, *, max_columns: int = 256):
    """在首行按表头文本查找列；返回列字母，找不到返回None。"""
    expected = {
        normalize_header_text(item)
        for item in (expected_headers if isinstance(expected_headers, (list, tuple, set)) else [expected_headers])
        if normalize_header_text(item)
    }
    if not expected:
        return None
    upper = min(max(int(getattr(worksheet, "max_column", 1) or 1), 1), int(max_columns))
    for column_number in range(1, upper + 1):
        if normalize_header_text(worksheet.cell(row=1, column=column_number).value) in expected:
            return column_number_to_letter(column_number)
    return None


def _copy_header_style(worksheet, source_column: int, target_column: int) -> None:
    source = worksheet.cell(row=1, column=max(1, int(source_column)))
    target = worksheet.cell(row=1, column=int(target_column))
    try:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
    except Exception:
        pass


def ensure_program_column(worksheet, preferred_column: str, header_text: str) -> str:
    """取得或创建程序专用列，绝不覆盖已有的其他业务表头。"""
    existing = find_header_column(worksheet, header_text)
    if existing:
        return existing

    preferred_number = _column_letter_to_index(preferred_column) + 1
    preferred_cell = worksheet.cell(row=1, column=preferred_number)
    preferred_header = normalize_header_text(preferred_cell.value)
    if preferred_header and preferred_header != normalize_header_text(header_text):
        target_number = max(int(worksheet.max_column or 1), preferred_number) + 1
    else:
        target_number = preferred_number

    target = worksheet.cell(row=1, column=target_number)
    if normalize_header_text(target.value) not in {"", normalize_header_text(header_text)}:
        raise ExcelWriteError(
            "PROGRAM_COLUMN_CONFLICT",
            "RESOLVE_COLUMNS",
            f"无法建立程序专用列“{header_text}”，目标列已有业务表头“{target.value}”。",
            retryable=False,
            committed=False,
        )
    if not normalize_header_text(target.value):
        _copy_header_style(worksheet, max(1, target_number - 1), target_number)
        target.value = header_text
    return column_number_to_letter(target_number)


def _is_replace_retryable(error: Exception) -> bool:
    if isinstance(error, PermissionError):
        return True
    winerror = getattr(error, "winerror", None)
    errno_value = getattr(error, "errno", None)
    return winerror in {5, 32, 33} or errno_value in {13, 16}


def replace_file_with_retry(
    temp_path: str,
    file_path: str,
    retry_delays=REPLACE_RETRY_DELAYS,
) -> None:
    """在共享盘短暂占用时有限重试正式文件替换。"""
    delays = tuple(retry_delays or ())
    last_error = None
    for attempt in range(len(delays) + 1):
        try:
            os.replace(temp_path, file_path)
            return
        except Exception as exc:
            last_error = exc
            if not _is_replace_retryable(exc) or attempt >= len(delays):
                break
            time.sleep(float(delays[attempt]))

    retryable = bool(last_error and _is_replace_retryable(last_error))
    raise ExcelWriteError(
        "REPLACE_BLOCKED" if retryable else "REPLACE_FAILED",
        "REPLACE_TARGET",
        f"无法用临时文件替换正式文件，文件可能被占用或当前账户缺少替换权限：{last_error}",
        retryable=retryable,
        committed=False,
    ) from last_error


def is_legacy_xls(file_path: str) -> bool:
    """Return whether a workbook uses the legacy binary Excel format."""
    return str(file_path).lower().endswith(".xls")


def _column_letter_to_index(column: str) -> int:
    index = 0
    for char in str(column).upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"无效Excel列名: {column}")
        index = index * 26 + ord(char) - ord("A") + 1
    return index - 1


def read_legacy_xls_cell(file_path: str, cell_reference: str):
    """Read one cell from an .xls workbook without loading the full sheet."""
    import xlrd

    match = re.fullmatch(r"([A-Za-z]+)(\d+)", str(cell_reference).strip())
    if not match:
        raise ValueError(f"无效Excel单元格: {cell_reference}")
    row_index = int(match.group(2)) - 1
    column_index = _column_letter_to_index(match.group(1))
    book = xlrd.open_workbook(file_path, on_demand=True)
    try:
        sheet = book.sheet_by_index(0)
        if row_index >= sheet.nrows or column_index >= sheet.ncols:
            return None
        cell = sheet.cell(row_index, column_index)
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, book.datemode)
        return cell.value
    finally:
        book.release_resources()


def _powershell_literal(value) -> str:
    if value is None:
        return "$null"
    if isinstance(value, bool):
        return "$true" if value else "$false"
    if isinstance(value, (int, float)):
        return repr(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def write_legacy_xls_cells(file_path: str, updates) -> None:
    """Atomically update cells in an .xls workbook through installed Excel."""
    if not is_legacy_xls(file_path):
        raise ValueError(f"不是.xls文件: {file_path}")
    file_path = os.path.abspath(file_path)
    parent_dir = os.path.dirname(file_path) or "."
    try:
        fd, temp_path = tempfile.mkstemp(
            prefix=".__excel_write_",
            suffix=".xls",
            dir=parent_dir,
        )
        os.close(fd)
        os.remove(temp_path)
    except PermissionError as exc:
        raise ExcelWriteError(
            "NO_DIRECTORY_PERMISSION",
            "CREATE_TEMP",
            f"无法在目标目录创建写入临时文件：{parent_dir}；{exc}",
            retryable=False,
            committed=False,
        ) from exc
    except OSError as exc:
        raise ExcelWriteError(
            "TEMP_CREATE_FAILED",
            "CREATE_TEMP",
            f"创建写入临时文件失败：{parent_dir}；{exc}",
            retryable=False,
            committed=False,
        ) from exc

    source_literal = _powershell_literal(file_path)
    temp_literal = _powershell_literal(temp_path)
    update_lines = []
    for update in updates:
        cell = str(update["cell"]).strip().upper()
        if not re.fullmatch(r"[A-Z]+\d+", cell):
            raise ValueError(f"无效Excel单元格: {cell}")
        range_expr = f"$sheet.Range({_powershell_literal(cell)})"
        value = update.get("value")
        if isinstance(value, datetime.datetime):
            value = value.date()
        if isinstance(value, datetime.date):
            iso_date = value.isoformat()
            update_lines.append(
                f"{range_expr}.Value = [datetime]::ParseExact('{iso_date}', 'yyyy-MM-dd', "
                "[Globalization.CultureInfo]::InvariantCulture)"
            )
        else:
            update_lines.append(f"{range_expr}.Value2 = {_powershell_literal(value)}")
        number_format = update.get("number_format")
        if number_format:
            update_lines.append(f"{range_expr}.NumberFormat = {_powershell_literal(number_format)}")

    script = "\n".join([
        "$ErrorActionPreference = 'Stop'",
        "$excel = $null",
        "$workbook = $null",
        "try {",
        "  $excel = New-Object -ComObject Excel.Application",
        "  $excel.Visible = $false",
        "  $excel.DisplayAlerts = $false",
        "  $excel.AskToUpdateLinks = $false",
        f"  $workbook = $excel.Workbooks.Open({source_literal}, 0, $false)",
        "  $sheet = $workbook.Worksheets.Item(1)",
        *[f"  {line}" for line in update_lines],
        f"  $workbook.SaveCopyAs({temp_literal})",
        "  $workbook.Close($false)",
        "  [void][Runtime.InteropServices.Marshal]::ReleaseComObject($sheet)",
        "  [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)",
        "  $workbook = $null",
        "  $excel.Quit()",
        "  [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)",
        "  $excel = $null",
        "} finally {",
        "  if ($null -ne $workbook) { $workbook.Close($false) }",
        "  if ($null -ne $excel) { $excel.Quit() }",
        "  [GC]::Collect()",
        "  [GC]::WaitForPendingFinalizers()",
        "}",
    ])
    encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
    creationflags = int(getattr(subprocess, "CREATE_NO_WINDOW", 0))

    try:
        result = subprocess.run(
            [
                "powershell.exe",
                "-NoLogo",
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-EncodedCommand",
                encoded,
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            creationflags=creationflags,
            timeout=120,
            check=False,
        )
        if result.returncode != 0:
            error_bytes = result.stderr or result.stdout
            try:
                error_text = error_bytes.decode("utf-8", errors="replace").strip()
            except Exception:
                error_text = str(error_bytes)
            raise ExcelWriteError(
                "SAVE_FAILED",
                "SAVE_TEMP",
                "无法通过Microsoft Excel写入.xls文件；请确认本机已安装Excel且文件未被占用。"
                f" 详细信息: {error_text or 'Excel COM调用失败'}",
                retryable=False,
                committed=False,
            )
        if not os.path.exists(temp_path):
            raise ExcelWriteError(
                "TEMP_NOT_CREATED",
                "SAVE_TEMP",
                "Excel未生成写入后的临时文件",
                retryable=False,
                committed=False,
            )

        import xlrd

        verify_book = xlrd.open_workbook(temp_path, on_demand=True)
        verify_book.release_resources()
        replace_file_with_retry(temp_path, file_path)
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def _needs_keep_vba(file_path: str) -> bool:
    """xlsm 文件需要保留 VBA 内容。"""
    return str(file_path).lower().endswith(".xlsm")


def open_workbook_for_edit(file_path: str):
    """以可编辑方式打开工作簿，并尽量保留原有宏内容。"""
    return load_workbook(file_path, keep_vba=_needs_keep_vba(file_path))


def atomic_save_workbook(workbook, file_path: str) -> None:
    """
    先保存到同目录临时文件，再原子替换目标文件，避免中途失败把原文件写坏。
    """
    file_path = os.path.abspath(file_path)
    parent_dir = os.path.dirname(file_path) or "."
    suffix = os.path.splitext(file_path)[1] or ".xlsx"
    fd = None
    temp_path = None

    try:
        try:
            fd, temp_path = tempfile.mkstemp(
                prefix=".__excel_write_",
                suffix=suffix,
                dir=parent_dir,
            )
            os.close(fd)
            fd = None
        except PermissionError as exc:
            raise ExcelWriteError(
                "NO_DIRECTORY_PERMISSION",
                "CREATE_TEMP",
                f"无法在目标目录创建写入临时文件：{parent_dir}；{exc}",
                retryable=False,
                committed=False,
            ) from exc
        except OSError as exc:
            raise ExcelWriteError(
                "TEMP_CREATE_FAILED",
                "CREATE_TEMP",
                f"创建写入临时文件失败：{parent_dir}；{exc}",
                retryable=False,
                committed=False,
            ) from exc

        try:
            workbook.save(temp_path)
        except Exception as exc:
            raise ExcelWriteError(
                "SAVE_FAILED",
                "SAVE_TEMP",
                f"保存Excel临时文件失败：{exc}",
                retryable=False,
                committed=False,
            ) from exc

        # 先验证临时文件可正常打开，再替换正式文件。
        try:
            verify_wb = open_workbook_for_edit(temp_path)
            verify_wb.close()
        except Exception as exc:
            raise ExcelWriteError(
                "TEMP_VERIFY_FAILED",
                "VERIFY_TEMP",
                f"临时文件无法重新打开，未替换正式文件：{exc}",
                retryable=False,
                committed=False,
            ) from exc

        replace_file_with_retry(temp_path, file_path)
    finally:
        if fd is not None:
            try:
                os.close(fd)
            except Exception:
                pass
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


_OOXML_CELL_REF_RE = re.compile(r"^([A-Za-z]+)([1-9][0-9]*)$")


def worksheet_archive_path(file_path: str, worksheet) -> str:
    """Resolve an openpyxl worksheet to its OOXML archive member path."""
    sheet_name = str(getattr(worksheet, "title", "") or "")
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            relationships_root = ET.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )

        relationship_id = None
        for element in workbook_root.iter():
            if element.tag.rsplit("}", 1)[-1] != "sheet":
                continue
            if str(element.attrib.get("name", "")) != sheet_name:
                continue
            relationship_id = next(
                (
                    value
                    for key, value in element.attrib.items()
                    if key.rsplit("}", 1)[-1] == "id"
                ),
                None,
            )
            break

        if relationship_id:
            for relationship in relationships_root:
                if relationship.attrib.get("Id") != relationship_id:
                    continue
                target = str(relationship.attrib.get("Target", "") or "").replace("\\", "/")
                if target.startswith("/"):
                    path = target.lstrip("/")
                else:
                    path = posixpath.normpath(posixpath.join("xl", target))
                if path:
                    return path
    except Exception as exc:
        raise ExcelWriteError(
            "SHEET_PATH_UNKNOWN",
            "RESOLVE_SHEET",
            f"无法确定目标工作表在Excel文件中的XML路径：{exc}",
            retryable=False,
            committed=False,
        ) from exc

    raise ExcelWriteError(
        "SHEET_PATH_UNKNOWN",
        "RESOLVE_SHEET",
        f"无法找到工作表“{sheet_name}”对应的XML路径。",
        retryable=False,
        committed=False,
    )


def _cell_reference_parts(cell_reference: str):
    reference = str(cell_reference or "").strip().upper()
    match = _OOXML_CELL_REF_RE.fullmatch(reference)
    if not match:
        raise ValueError(f"无效Excel单元格: {cell_reference}")
    return reference, match.group(1), int(match.group(2))


def _ooxml_column_number(column: str) -> int:
    number = 0
    for char in str(column or "").upper():
        number = number * 26 + ord(char) - ord("A") + 1
    return number


def _ooxml_row_pattern(row_number: int):
    row_value = str(int(row_number)).encode("ascii")
    lookahead = rb'(?=[^>]*\br\s*=\s*["\']' + row_value + rb'["\'])'
    return re.compile(
        rb'(?:<row\b'
        + lookahead
        + rb'[^>]*/\s*>|<row\b'
        + lookahead
        + rb'[^>]*>.*?</row\s*>)',
        re.DOTALL,
    )


def _ooxml_cell_pattern(cell_reference: str):
    escaped = re.escape(str(cell_reference).upper().encode("ascii"))
    lookahead = rb'(?=[^>]*\br\s*=\s*["\']' + escaped + rb'["\'])'
    return re.compile(
        rb'(?:<c\b'
        + lookahead
        + rb'[^>]*/\s*>|<c\b'
        + lookahead
        + rb'[^>]*>.*?</c\s*>)',
        re.DOTALL,
    )


def _ooxml_cell_reference_from_xml(cell_xml: bytes):
    match = re.search(rb'\br\s*=\s*["\']([A-Za-z]+[1-9][0-9]*)["\']', cell_xml[:512])
    if not match:
        return None
    try:
        return match.group(1).decode("ascii").upper()
    except Exception:
        return None


def _ooxml_inline_string_cell(cell_reference: str, value, existing_cell=None, style_id=None) -> bytes:
    reference = str(cell_reference).upper()
    attributes = b' r="' + reference.encode("ascii") + b'"'
    if existing_cell:
        opening = re.match(rb'<c\b([^>]*)/?>', existing_cell, flags=re.DOTALL)
        if opening:
            attributes = opening.group(1)
            attributes = re.sub(rb'\s*/\s*$', b'', attributes)
            attributes = re.sub(
                rb'\s+t\s*=\s*(["\']).*?\1',
                b'',
                attributes,
                flags=re.DOTALL,
            )
    if not re.search(rb'\br\s*=', attributes):
        attributes += b' r="' + reference.encode("ascii") + b'"'
    if style_id not in (None, "", 0, "0") and not re.search(rb'\bs\s*=', attributes):
        attributes += b' s="' + str(int(style_id)).encode("ascii") + b'"'

    text = "" if value is None else str(value)
    escaped_text = xml_escape(text, {'"': "&quot;", "'": "&apos;"})
    escaped_text = escaped_text.replace("\r", "&#13;")
    body = escaped_text.encode("utf-8")
    return (
        b"<c"
        + attributes
        + b' t="inlineStr"><is><t xml:space="preserve">'
        + body
        + b"</t></is></c>"
    )


def _ooxml_number_cell(cell_reference: str, value, existing_cell=None, style_id=None) -> bytes:
    """Build a numeric cell while preserving the target cell's other attributes."""
    reference = str(cell_reference).upper()
    attributes = b' r="' + reference.encode("ascii") + b'"'
    if existing_cell:
        opening = re.match(rb'<c\b([^>]*)/?>', existing_cell, flags=re.DOTALL)
        if opening:
            attributes = re.sub(rb'\s*/\s*$', b'', opening.group(1))
            attributes = re.sub(
                rb'\s+t\s*=\s*(["\']).*?\1',
                b'',
                attributes,
                flags=re.DOTALL,
            )
    if not re.search(rb'\br\s*=', attributes):
        attributes += b' r="' + reference.encode("ascii") + b'"'
    if style_id not in (None, ""):
        encoded_style = str(int(style_id)).encode("ascii")
        if re.search(rb'\bs\s*=', attributes):
            attributes = re.sub(
                rb'(\bs\s*=\s*["\'])[^"\']*(["\'])',
                rb'\g<1>' + encoded_style + rb'\g<2>',
                attributes,
                count=1,
            )
        elif int(style_id) != 0:
            attributes += b' s="' + encoded_style + b'"'
    number_text = str(value).encode("ascii")
    return b"<c" + attributes + b"><v>" + number_text + b"</v></c>"


def _patch_ooxml_row(row_xml: bytes, update) -> bytes:
    reference, target_column, _row_number = _cell_reference_parts(update["cell"])
    pattern = _ooxml_cell_pattern(reference)
    existing_match = pattern.search(row_xml)
    existing_xml = existing_match.group(0) if existing_match else None
    if update.get("value_type") == "number":
        replacement = _ooxml_number_cell(
            reference,
            update.get("value"),
            existing_cell=existing_xml,
            style_id=update.get("style_id"),
        )
    else:
        replacement = _ooxml_inline_string_cell(
            reference,
            update.get("value"),
            existing_cell=existing_xml,
            style_id=update.get("style_id"),
        )
    if existing_match:
        return row_xml[:existing_match.start()] + replacement + row_xml[existing_match.end():]

    if row_xml.rstrip().endswith(b"/>"):
        open_tag = row_xml.rstrip()[:-2] + b">"
        return open_tag + replacement + b"</row>"

    target_number = _ooxml_column_number(target_column)
    candidate_pattern = re.compile(
        rb'(?:<c\b[^>]*/\s*>|<c\b[^>]*>.*?</c\s*>)',
        flags=re.DOTALL,
    )
    for candidate in candidate_pattern.finditer(row_xml):
        candidate_reference = _ooxml_cell_reference_from_xml(candidate.group(0))
        if not candidate_reference:
            continue
        candidate_column = _OOXML_CELL_REF_RE.fullmatch(candidate_reference).group(1)
        if _ooxml_column_number(candidate_column) > target_number:
            return row_xml[:candidate.start()] + replacement + row_xml[candidate.start():]

    closing = row_xml.rfind(b"</row")
    if closing < 0:
        raise ExcelWriteError(
            "ROW_XML_INVALID",
            "PATCH_SHEET",
            f"目标行{_row_number}的XML结构不完整。",
            retryable=False,
            committed=False,
        )
    return row_xml[:closing] + replacement + row_xml[closing:]


def _patch_ooxml_sheet_xml(sheet_xml: bytes, updates) -> bytes:
    patched = sheet_xml
    grouped = {}
    for update in updates:
        reference, _column, row_number = _cell_reference_parts(update["cell"])
        normalized = dict(update)
        normalized["cell"] = reference
        grouped.setdefault(row_number, []).append(normalized)

    for row_number in sorted(grouped):
        row_pattern = _ooxml_row_pattern(row_number)
        row_match = row_pattern.search(patched)
        if not row_match:
            raise ExcelWriteError(
                "ROW_NOT_FOUND",
                "PATCH_SHEET",
                f"目标工作表中不存在第{row_number}行，未修改Excel。",
                retryable=False,
                committed=False,
            )
        row_xml = row_match.group(0)
        for update in sorted(
            grouped[row_number],
            key=lambda item: _ooxml_column_number(_cell_reference_parts(item["cell"])[1]),
        ):
            row_xml = _patch_ooxml_row(row_xml, update)
        patched = patched[:row_match.start()] + row_xml + patched[row_match.end():]
    return patched


def atomic_patch_ooxml_cells(file_path: str, sheet_path: str, updates) -> None:
    """Atomically patch only selected cells in one OOXML worksheet.

    Other archive members are copied with their original uncompressed bytes and
    are never parsed or regenerated by openpyxl.  A temporary workbook is built
    in the target directory and replaces the source only after the patched
    worksheet can be read back successfully.
    """
    file_path = os.path.abspath(os.fspath(file_path))
    normalized_sheet_path = str(sheet_path or "").replace("\\", "/").lstrip("/")
    if not normalized_sheet_path:
        raise ValueError("目标工作表XML路径不能为空")
    if not updates:
        return

    parent_dir = os.path.dirname(file_path) or "."
    suffix = os.path.splitext(file_path)[1] or ".xlsx"
    fd = None
    temp_path = None
    try:
        try:
            fd, temp_path = tempfile.mkstemp(
                prefix=".__excel_xml_write_",
                suffix=suffix,
                dir=parent_dir,
            )
            os.close(fd)
            fd = None
        except PermissionError as exc:
            raise ExcelWriteError(
                "NO_DIRECTORY_PERMISSION",
                "CREATE_TEMP",
                f"无法在目标目录创建写入临时文件：{parent_dir}；{exc}",
                retryable=False,
                committed=False,
            ) from exc
        except OSError as exc:
            raise ExcelWriteError(
                "TEMP_CREATE_FAILED",
                "CREATE_TEMP",
                f"创建写入临时文件失败：{parent_dir}；{exc}",
                retryable=False,
                committed=False,
            ) from exc

        patched_count = 0
        try:
            with zipfile.ZipFile(file_path, "r") as source, zipfile.ZipFile(
                temp_path,
                "w",
                allowZip64=True,
            ) as target:
                target.comment = source.comment
                for original_info in source.infolist():
                    data = source.read(original_info.filename)
                    if original_info.filename.replace("\\", "/") == normalized_sheet_path:
                        data = _patch_ooxml_sheet_xml(data, updates)
                        patched_count += 1
                    info = copy.copy(original_info)
                    target.writestr(info, data)
        except ExcelWriteError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise ExcelWriteError(
                "OOXML_PATCH_FAILED",
                "SAVE_TEMP",
                f"无法生成定点写入后的Excel临时文件：{exc}",
                retryable=isinstance(exc, OSError),
                committed=False,
            ) from exc

        if patched_count != 1:
            raise ExcelWriteError(
                "SHEET_XML_NOT_FOUND",
                "PATCH_SHEET",
                f"目标工作表XML不存在或重复：{normalized_sheet_path}",
                retryable=False,
                committed=False,
            )

        try:
            with zipfile.ZipFile(temp_path, "r") as verify_archive:
                verify_xml = verify_archive.read(normalized_sheet_path)
            for update in updates:
                cell_xml = _ooxml_cell_pattern(update["cell"]).search(verify_xml)
                if not cell_xml:
                    raise ValueError(f"找不到写入后的单元格：{update['cell']}")
        except Exception as exc:
            raise ExcelWriteError(
                "TEMP_VERIFY_FAILED",
                "VERIFY_TEMP",
                f"定点写入临时文件验证失败，未替换正式文件：{exc}",
                retryable=False,
                committed=False,
            ) from exc

        replace_file_with_retry(temp_path, file_path)
    finally:
        if fd is not None:
            try:
                os.close(fd)
            except Exception:
                pass
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def read_ooxml_inline_cell(file_path: str, sheet_path: str, cell_reference: str):
    """Read a text-like cell directly from one OOXML worksheet member."""
    normalized_sheet_path = str(sheet_path or "").replace("\\", "/").lstrip("/")
    reference, _column, _row = _cell_reference_parts(cell_reference)
    with zipfile.ZipFile(file_path, "r") as archive:
        sheet_xml = archive.read(normalized_sheet_path)
        match = _ooxml_cell_pattern(reference).search(sheet_xml)
        if not match:
            return None
        cell_xml = match.group(0)
        opening = re.match(rb'<c\b([^>]*)/?>', cell_xml, flags=re.DOTALL)
        attributes = opening.group(1) if opening else b""
        type_match = re.search(rb'\bt\s*=\s*["\']([^"\']+)["\']', attributes)
        cell_type = type_match.group(1).decode("ascii", errors="ignore") if type_match else ""

        if cell_type == "inlineStr":
            parts = re.findall(rb'<t\b[^>]*>(.*?)</t\s*>', cell_xml, flags=re.DOTALL)
            return "".join(html.unescape(part.decode("utf-8")) for part in parts)

        value_match = re.search(rb'<v\b[^>]*>(.*?)</v\s*>', cell_xml, flags=re.DOTALL)
        if not value_match:
            return None
        raw_value = html.unescape(value_match.group(1).decode("utf-8"))
        if cell_type == "s":
            shared_xml = archive.read("xl/sharedStrings.xml")
            shared_items = re.findall(rb'<si\b[^>]*>(.*?)</si\s*>', shared_xml, flags=re.DOTALL)
            index = int(raw_value)
            if index < 0 or index >= len(shared_items):
                return None
            text_parts = re.findall(rb'<t\b[^>]*>(.*?)</t\s*>', shared_items[index], flags=re.DOTALL)
            return "".join(html.unescape(part.decode("utf-8")) for part in text_parts)
        return raw_value


def active_worksheet_archive_path(file_path: str) -> str:
    """Resolve the active worksheet without loading the workbook object model."""
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            relationships_root = ET.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )

        sheets = [
            element
            for element in workbook_root.iter()
            if element.tag.rsplit("}", 1)[-1] == "sheet"
        ]
        if not sheets:
            raise ValueError("工作簿没有工作表")

        active_index = 0
        for element in workbook_root.iter():
            if element.tag.rsplit("}", 1)[-1] != "workbookView":
                continue
            try:
                active_index = int(element.attrib.get("activeTab", 0) or 0)
            except (TypeError, ValueError):
                active_index = 0
            break
        if active_index < 0 or active_index >= len(sheets):
            active_index = 0

        relationship_id = next(
            (
                value
                for key, value in sheets[active_index].attrib.items()
                if key.rsplit("}", 1)[-1] == "id"
            ),
            None,
        )
        if not relationship_id:
            raise ValueError("活动工作表缺少关系标识")
        for relationship in relationships_root:
            if relationship.attrib.get("Id") != relationship_id:
                continue
            target = str(relationship.attrib.get("Target", "") or "").replace("\\", "/")
            if target.startswith("/"):
                return target.lstrip("/")
            return posixpath.normpath(posixpath.join("xl", target))
        raise ValueError("活动工作表关系不存在")
    except ExcelWriteError:
        raise
    except Exception as exc:
        raise ExcelWriteError(
            "SHEET_PATH_UNKNOWN",
            "RESOLVE_SHEET",
            f"无法确定活动工作表XML路径：{exc}",
            retryable=False,
            committed=False,
        ) from exc


def worksheet_archive_path_by_index(file_path: str, sheet_index: int = 0) -> str:
    """Resolve a worksheet archive member by workbook order."""
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            relationships_root = ET.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )

        sheets = [
            element
            for element in workbook_root.iter()
            if element.tag.rsplit("}", 1)[-1] == "sheet"
        ]
        index = int(sheet_index)
        if index < 0 or index >= len(sheets):
            raise IndexError(f"工作表序号超出范围: {sheet_index}")
        relationship_id = next(
            (
                value
                for key, value in sheets[index].attrib.items()
                if key.rsplit("}", 1)[-1] == "id"
            ),
            None,
        )
        if not relationship_id:
            raise ValueError("工作表缺少关系标识")
        for relationship in relationships_root:
            if relationship.attrib.get("Id") != relationship_id:
                continue
            target = str(relationship.attrib.get("Target", "") or "").replace("\\", "/")
            if target.startswith("/"):
                return target.lstrip("/")
            return posixpath.normpath(posixpath.join("xl", target))
        raise ValueError("工作表关系不存在")
    except ExcelWriteError:
        raise
    except Exception as exc:
        raise ExcelWriteError(
            "SHEET_PATH_UNKNOWN",
            "RESOLVE_SHEET",
            f"无法确定第{sheet_index + 1}个工作表XML路径：{exc}",
            retryable=False,
            committed=False,
        ) from exc


_OOXML_ANY_CELL_PATTERN = re.compile(
    rb'(?:<c\b[^>]*/\s*>|<c\b[^>]*>.*?</c\s*>)',
    flags=re.DOTALL,
)


class OoxmlWorksheetSnapshot:
    """Read-only, lightweight view of one OOXML worksheet."""

    def __init__(self, file_path: str, sheet_path: str = None):
        self.file_path = os.path.abspath(os.fspath(file_path))
        self.sheet_path = sheet_path or active_worksheet_archive_path(self.file_path)
        self.sheet_path = self.sheet_path.replace("\\", "/").lstrip("/")
        self._shared_items = None
        with zipfile.ZipFile(self.file_path, "r") as archive:
            self.sheet_xml = archive.read(self.sheet_path)
            try:
                self.shared_strings_xml = archive.read("xl/sharedStrings.xml")
            except KeyError:
                self.shared_strings_xml = None
            try:
                workbook_xml = archive.read("xl/workbook.xml")
            except KeyError:
                workbook_xml = b""
            try:
                self.styles_xml = archive.read("xl/styles.xml")
            except KeyError:
                self.styles_xml = None
        self.date_1904 = bool(re.search(rb'<workbookPr\b[^>]*\bdate1904\s*=\s*["\'](?:1|true)["\']', workbook_xml))
        self._style_formats = None

    def cell_xml(self, cell_reference: str):
        reference, _column, _row = _cell_reference_parts(cell_reference)
        match = _ooxml_cell_pattern(reference).search(self.sheet_xml)
        return match.group(0) if match else None

    @staticmethod
    def _cell_type_and_raw_value(cell_xml):
        if not cell_xml:
            return "", None
        opening = re.match(rb'<c\b([^>]*)/?>', cell_xml, flags=re.DOTALL)
        attributes = opening.group(1) if opening else b""
        type_match = re.search(rb'\bt\s*=\s*["\']([^"\']+)["\']', attributes)
        cell_type = type_match.group(1).decode("ascii", errors="ignore") if type_match else ""
        if cell_type == "inlineStr":
            parts = re.findall(rb'<t\b[^>]*>(.*?)</t\s*>', cell_xml, flags=re.DOTALL)
            return cell_type, "".join(html.unescape(part.decode("utf-8")) for part in parts)
        value_match = re.search(rb'<v\b[^>]*>(.*?)</v\s*>', cell_xml, flags=re.DOTALL)
        if not value_match:
            return cell_type, None
        return cell_type, html.unescape(value_match.group(1).decode("utf-8"))

    def _shared_strings(self):
        if self._shared_items is None:
            if not self.shared_strings_xml:
                self._shared_items = []
            else:
                self._shared_items = []
                for item in re.findall(
                    rb'<si\b[^>]*>(.*?)</si\s*>',
                    self.shared_strings_xml,
                    flags=re.DOTALL,
                ):
                    parts = re.findall(rb'<t\b[^>]*>(.*?)</t\s*>', item, flags=re.DOTALL)
                    self._shared_items.append(
                        "".join(html.unescape(part.decode("utf-8")) for part in parts)
                    )
        return self._shared_items

    def value_from_cell_xml(self, cell_xml):
        cell_type, raw_value = self._cell_type_and_raw_value(cell_xml)
        if raw_value is None:
            return None
        if cell_type == "s":
            try:
                index = int(raw_value)
                values = self._shared_strings()
                return values[index] if 0 <= index < len(values) else None
            except (TypeError, ValueError):
                return None
        if cell_type == "b":
            return str(raw_value).strip() == "1"
        return raw_value

    def _number_format_for_style(self, style_id: int):
        if self._style_formats is None:
            from openpyxl.styles.numbers import BUILTIN_FORMATS

            custom_formats = {}
            style_formats = []
            if self.styles_xml:
                root = ET.fromstring(self.styles_xml)
                for element in root.iter():
                    if element.tag.rsplit("}", 1)[-1] == "numFmt":
                        try:
                            custom_formats[int(element.attrib.get("numFmtId"))] = element.attrib.get(
                                "formatCode", ""
                            )
                        except (TypeError, ValueError):
                            continue
                cell_xfs = next(
                    (
                        element
                        for element in root.iter()
                        if element.tag.rsplit("}", 1)[-1] == "cellXfs"
                    ),
                    None,
                )
                if cell_xfs is not None:
                    for xf in cell_xfs:
                        try:
                            num_fmt_id = int(xf.attrib.get("numFmtId", 0) or 0)
                        except (TypeError, ValueError):
                            num_fmt_id = 0
                        style_formats.append(
                            custom_formats.get(num_fmt_id, BUILTIN_FORMATS.get(num_fmt_id, ""))
                        )
            self._style_formats = style_formats
        try:
            return self._style_formats[int(style_id)]
        except (IndexError, TypeError, ValueError):
            return ""

    def typed_value_from_cell_xml(self, cell_xml):
        """Return an openpyxl-compatible scalar without loading worksheet rows."""
        from openpyxl.styles.numbers import is_date_format, is_timedelta_format
        from openpyxl.utils.datetime import (
            CALENDAR_MAC_1904,
            CALENDAR_WINDOWS_1900,
            from_ISO8601,
            from_excel,
        )
        from openpyxl.worksheet._reader import _cast_number

        cell_type, raw_value = self._cell_type_and_raw_value(cell_xml)
        if raw_value is None:
            return None
        if cell_type == "s":
            try:
                index = int(raw_value)
                values = self._shared_strings()
                return values[index] if 0 <= index < len(values) else None
            except (TypeError, ValueError):
                return None
        if cell_type == "b":
            return str(raw_value).strip() == "1"
        if cell_type == "d":
            return from_ISO8601(raw_value)
        if cell_type in {"inlineStr", "str", "e"}:
            return raw_value

        number = _cast_number(raw_value)
        style_id = self.style_id_from_cell_xml(cell_xml)
        number_format = self._number_format_for_style(style_id)
        if number_format and is_date_format(number_format):
            try:
                epoch = CALENDAR_MAC_1904 if self.date_1904 else CALENDAR_WINDOWS_1900
                return from_excel(
                    number,
                    epoch,
                    timedelta=is_timedelta_format(number_format),
                )
            except (OverflowError, ValueError):
                return raw_value
        return number

    @staticmethod
    def style_id_from_cell_xml(cell_xml):
        if not cell_xml:
            return 0
        opening = re.match(rb'<c\b([^>]*)/?>', cell_xml, flags=re.DOTALL)
        attributes = opening.group(1) if opening else b""
        style_match = re.search(rb'\bs\s*=\s*["\']([0-9]+)["\']', attributes)
        return int(style_match.group(1)) if style_match else 0

    def value(self, cell_reference: str):
        return self.value_from_cell_xml(self.cell_xml(cell_reference))

    def style_id(self, cell_reference: str):
        cell_xml = self.cell_xml(cell_reference)
        if not cell_xml:
            return None
        opening = re.match(rb'<c\b([^>]*)/?>', cell_xml, flags=re.DOTALL)
        attributes = opening.group(1) if opening else b""
        style_match = re.search(rb'\bs\s*=\s*["\']([0-9]+)["\']', attributes)
        return int(style_match.group(1)) if style_match else None

    def iter_cells(self):
        for match in _OOXML_ANY_CELL_PATTERN.finditer(self.sheet_xml):
            cell_xml = match.group(0)
            reference = _ooxml_cell_reference_from_xml(cell_xml)
            if reference:
                yield reference, cell_xml

    def header_values(self):
        values = {}
        for reference, cell_xml in self.iter_cells():
            match = _OOXML_CELL_REF_RE.fullmatch(reference)
            if match and int(match.group(2)) == 1:
                values[match.group(1)] = self.value_from_cell_xml(cell_xml)
        return values

    def max_column_number(self):
        maximum = 1
        for reference, _cell_xml in self.iter_cells():
            match = _OOXML_CELL_REF_RE.fullmatch(reference)
            if match:
                maximum = max(maximum, _ooxml_column_number(match.group(1)))
        return maximum

    def rows_matching(self, column: str, normalized_expected: str, normalizer):
        column = str(column or "").upper()
        matches = []
        for reference, cell_xml in self.iter_cells():
            parsed = _OOXML_CELL_REF_RE.fullmatch(reference)
            if not parsed or parsed.group(1) != column:
                continue
            if normalizer(self.value_from_cell_xml(cell_xml)) == normalized_expected:
                matches.append(int(parsed.group(2)))
        return matches
