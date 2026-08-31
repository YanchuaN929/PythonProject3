#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回文单号输入处理模块
"""

import tkinter as tk
from tkinter import ttk, messagebox
from dataclasses import dataclass
from datetime import date
import os
import re
import sys
import time

from write_tasks import get_write_task_manager, get_pending_cache
from utils.excel_io import (
    ExcelWriteError,
    OoxmlWorksheetSnapshot,
    SharedWorkbookLock,
    atomic_patch_ooxml_cells,
    atomic_save_workbook,
    column_number_to_letter,
    ensure_program_column,
    is_legacy_xls,
    normalize_header_text,
    open_workbook_for_edit,
    read_ooxml_inline_cell,
    read_legacy_xls_cell,
    write_legacy_xls_cells,
)


RESPONSE_INTERFACE_COLUMN_MAP = {
    1: "A",
    2: "R",
    3: "C",
    4: "E",
    5: "A",
    6: "E",
}
VERIFY_RETRY_DELAYS = (0.5, 1.0, 2.0, 4.0)


@dataclass(frozen=True)
class ResponseWriteResult:
    success: bool
    row_index: int
    already_present: bool = False

    def __bool__(self):
        return self.success


def _normalize_response_interface_id(value) -> str:
    text = "" if value is None else str(value).strip()
    if not text or text.lower() in {"none", "nan"}:
        return ""
    text = re.sub(r"\([^)]*\)$", "", text).strip()
    return re.sub(r"\s+", "", text).upper()


def _column_letter_to_number(column: str) -> int:
    value = 0
    for char in str(column or "").upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"无效Excel列名: {column}")
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _resolve_xlsx_response_row(worksheet, file_type, row_index, interface_id) -> int:
    requested_row = int(row_index)
    expected = _normalize_response_interface_id(interface_id)
    if not expected:
        return requested_row

    interface_column = RESPONSE_INTERFACE_COLUMN_MAP.get(int(file_type))
    if not interface_column:
        raise ExcelWriteError(
            "INTERFACE_COLUMN_UNKNOWN",
            "VALIDATE_ROW",
            f"无法确定文件类型{file_type}的接口号列",
            retryable=False,
            committed=False,
        )

    actual_at_requested = ""
    if requested_row >= 2:
        actual_at_requested = _normalize_response_interface_id(
            worksheet[f"{interface_column}{requested_row}"].value
        )
        if actual_at_requested == expected:
            return requested_row

    column_number = _column_letter_to_number(interface_column)
    matches = []
    for candidate_row, row in enumerate(worksheet.iter_rows(
        min_row=2,
        max_row=max(2, int(worksheet.max_row or 2)),
        min_col=column_number,
        max_col=column_number,
        values_only=True,
    ), start=2):
        if _normalize_response_interface_id(row[0] if row else None) == expected:
            matches.append(candidate_row)

    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ExcelWriteError(
            "ROW_NOT_FOUND",
            "VALIDATE_ROW",
            f"原行号{requested_row}的接口号已变化，且工作簿中找不到接口号：{interface_id}",
            retryable=False,
            committed=False,
        )
    raise ExcelWriteError(
        "ROW_AMBIGUOUS",
        "VALIDATE_ROW",
        f"原行号{requested_row}的接口号已变化，工作簿中存在{len(matches)}条相同接口号，已拒绝自动写入：{interface_id}",
        retryable=False,
        committed=False,
    )


def _resolve_ooxml_response_row(snapshot, file_type, row_index, interface_id) -> int:
    requested_row = int(row_index)
    expected = _normalize_response_interface_id(interface_id)
    if not expected:
        return requested_row

    interface_column = RESPONSE_INTERFACE_COLUMN_MAP.get(int(file_type))
    if not interface_column:
        raise ExcelWriteError(
            "INTERFACE_COLUMN_UNKNOWN",
            "VALIDATE_ROW",
            f"无法确定文件类型{file_type}的接口号列",
            retryable=False,
            committed=False,
        )
    if requested_row >= 2:
        current = _normalize_response_interface_id(
            snapshot.value(f"{interface_column}{requested_row}")
        )
        if current == expected:
            return requested_row

    matches = snapshot.rows_matching(
        interface_column,
        expected,
        _normalize_response_interface_id,
    )
    matches = [row for row in matches if row >= 2]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ExcelWriteError(
            "ROW_NOT_FOUND",
            "VALIDATE_ROW",
            f"原行号{requested_row}的接口号已变化，且工作簿中找不到接口号：{interface_id}",
            retryable=False,
            committed=False,
        )
    raise ExcelWriteError(
        "ROW_AMBIGUOUS",
        "VALIDATE_ROW",
        f"原行号{requested_row}的接口号已变化，工作簿中存在{len(matches)}条相同接口号，已拒绝自动写入：{interface_id}",
        retryable=False,
        committed=False,
    )


def _resolve_xls_response_row(file_path, file_type, row_index, interface_id) -> int:
    requested_row = int(row_index)
    expected = _normalize_response_interface_id(interface_id)
    if not expected:
        return requested_row

    interface_column = RESPONSE_INTERFACE_COLUMN_MAP.get(int(file_type))
    if not interface_column:
        raise ExcelWriteError(
            "INTERFACE_COLUMN_UNKNOWN",
            "VALIDATE_ROW",
            f"无法确定文件类型{file_type}的接口号列",
            retryable=False,
            committed=False,
        )

    import xlrd

    book = xlrd.open_workbook(file_path, on_demand=True)
    try:
        sheet = book.sheet_by_index(0)
        column_index = _column_letter_to_number(interface_column) - 1
        if 2 <= requested_row <= sheet.nrows:
            if _normalize_response_interface_id(
                sheet.cell_value(requested_row - 1, column_index)
            ) == expected:
                return requested_row
        matches = [
            excel_row
            for excel_row in range(2, sheet.nrows + 1)
            if _normalize_response_interface_id(
                sheet.cell_value(excel_row - 1, column_index)
            ) == expected
        ]
    finally:
        book.release_resources()

    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ExcelWriteError(
            "ROW_NOT_FOUND",
            "VALIDATE_ROW",
            f"原行号{requested_row}的接口号已变化，且工作簿中找不到接口号：{interface_id}",
            retryable=False,
            committed=False,
        )
    raise ExcelWriteError(
        "ROW_AMBIGUOUS",
        "VALIDATE_ROW",
        f"原行号{requested_row}的接口号已变化，工作簿中存在{len(matches)}条相同接口号，已拒绝自动写入：{interface_id}",
        retryable=False,
        committed=False,
    )


def _response_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _ensure_response_cell_writable(current_value, response_number, cell_reference: str) -> bool:
    current_text = _response_text(current_value)
    expected_text = _response_text(response_number)
    if current_text == expected_text and expected_text:
        return False
    if current_text:
        raise ExcelWriteError(
            "RESPONSE_CONFLICT",
            "CHECK_EXISTING",
            f"{cell_reference}已有不同回文单号“{current_text}”，本次“{expected_text}”未覆盖",
            retryable=False,
            committed=False,
        )
    return True


def _verify_xlsx_response(
    file_path,
    row_index,
    response_column,
    response_number,
    sheet_path=None,
) -> None:
    expected_text = _response_text(response_number)
    last_error = None
    last_value = None
    delays = tuple(VERIFY_RETRY_DELAYS)
    for attempt in range(len(delays) + 1):
        try:
            if sheet_path:
                last_value = read_ooxml_inline_cell(
                    file_path,
                    sheet_path,
                    f"{response_column}{row_index}",
                )
            else:
                workbook = open_workbook_for_edit(file_path)
                try:
                    last_value = workbook.active[f"{response_column}{row_index}"].value
                finally:
                    workbook.close()
            if _response_text(last_value) == expected_text:
                return
            last_error = RuntimeError(
                f"期望:{response_number}, 实际:{last_value}"
            )
        except Exception as exc:
            last_error = exc
        if attempt < len(delays):
            time.sleep(float(delays[attempt]))

    raise ExcelWriteError(
        "VERIFY_FAILED",
        "VERIFY_FINAL",
        f"正式文件已替换，但多次重新打开核验仍失败：{last_error}",
        retryable=True,
        committed=True,
    ) from last_error


def _verify_xls_response(file_path, row_index, response_column, response_number) -> None:
    expected_text = _response_text(response_number)
    last_error = None
    delays = tuple(VERIFY_RETRY_DELAYS)
    for attempt in range(len(delays) + 1):
        try:
            actual = read_legacy_xls_cell(
                file_path,
                f"{response_column}{row_index}",
            )
            if _response_text(actual) == expected_text:
                return
            last_error = RuntimeError(
                f"期望:{response_number}, 实际:{actual}"
            )
        except Exception as exc:
            last_error = exc
        if attempt < len(delays):
            time.sleep(float(delays[attempt]))

    raise ExcelWriteError(
        "VERIFY_FAILED",
        "VERIFY_FINAL",
        f"正式文件已替换，但多次重新打开核验仍失败：{last_error}",
        retryable=True,
        committed=True,
    ) from last_error


def _make_stdio_unicode_safe():
    """Prevent diagnostic prints from breaking write workflows on GBK consoles."""
    for stream in (getattr(sys, "stdout", None), getattr(sys, "stderr", None)):
        if stream is None or not hasattr(stream, "reconfigure"):
            continue
        try:
            stream.reconfigure(errors="replace")
        except Exception:
            pass


def _safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        stream = kwargs.get("file") or getattr(sys, "stdout", None)
        if stream is None:
            return
        sep = kwargs.get("sep", " ")
        end = kwargs.get("end", "\n")
        text = sep.join(str(arg) for arg in args) + end
        encoding = getattr(stream, "encoding", None) or "utf-8"
        try:
            safe_text = text.encode(encoding, errors="replace").decode(encoding, errors="replace")
            stream.write(safe_text)
            stream.flush()
        except Exception:
            pass


_make_stdio_unicode_safe()

# 导入Registry模块
try:
    from registry import hooks as registry_hooks
except ImportError:
    _safe_print("警告: 未找到registry模块")
    registry_hooks = None


def get_excel_lock_owner(file_path: str) -> str:
    """
    获取Excel文件的占用者用户名
    
    Excel打开文件时会创建 ~$文件名 的临时锁定文件，
    其中包含打开文件的用户名信息。
    
    参数:
        file_path: Excel文件的完整路径
        
    返回:
        占用者用户名，如果无法获取则返回空字符串
    """
    try:
        dir_path = os.path.dirname(file_path)
        file_name = os.path.basename(file_path)
        
        # Excel临时文件格式: ~$文件名
        # 对于长文件名，Excel会截取前面部分
        lock_file_name = "~$" + file_name
        lock_file_path = os.path.join(dir_path, lock_file_name)
        
        # 如果标准锁定文件不存在，尝试查找以 ~$ 开头的文件
        if not os.path.exists(lock_file_path):
            # 搜索目录中所有以 ~$ 开头的文件
            for f in os.listdir(dir_path):
                if f.startswith("~$"):
                    # 检查是否与目标文件相关（去掉 ~$ 后是否匹配）
                    potential_name = f[2:]  # 去掉 ~$
                    if file_name.startswith(potential_name) or potential_name in file_name:
                        lock_file_path = os.path.join(dir_path, f)
                        break
        
        if not os.path.exists(lock_file_path):
            return ""
        
        # 读取锁定文件内容获取用户名
        # Excel锁定文件是二进制格式，用户名通常在文件开头以Unicode编码存储
        with open(lock_file_path, 'rb') as f:
            content = f.read()
        
        # 尝试解码用户名
        # 方法1: 直接解码 UTF-16-LE（Windows常用编码）
        try:
            # 跳过前面可能的字节，用户名通常在前128字节内
            # 查找连续的可打印Unicode字符
            decoded = content[:128].decode('utf-16-le', errors='ignore')
            # 过滤出有效字符（字母、数字、中文等）
            user_name = ""
            for char in decoded:
                if char.isprintable() and char not in '\x00\x01\x02\x03\x04\x05\x06\x07\x08':
                    user_name += char
                elif user_name:  # 遇到非法字符且已有用户名，停止
                    break
            
            if user_name and len(user_name) >= 2:
                return user_name.strip()
        except Exception:
            pass
        
        # 方法2: 尝试 GBK 解码（中文Windows系统）
        try:
            decoded = content[:64].decode('gbk', errors='ignore')
            user_name = ""
            for char in decoded:
                if char.isprintable() and ord(char) > 31:
                    user_name += char
                elif user_name:
                    break
            
            if user_name and len(user_name) >= 2:
                return user_name.strip()
        except Exception:
            pass
        
        return ""
        
    except Exception as e:
        print(f"[文件锁定] 获取占用者信息失败: {e}")
        return ""


class InterfaceInputDialog(tk.Toplevel):
    """回文单号输入弹窗"""
    
    def __init__(self, parent, interface_id, file_type, file_path, row_index, 
                 user_name, project_id, source_column=None, file_manager=None, 
                 viewer=None, item_id=None, columns=None, on_success=None, has_assignor=False,
                 user_roles=None):
        """
        参数:
            parent: 父窗口
            interface_id: 接口号
            file_type: 文件类型(1-6)
            file_path: 原始Excel文件路径
            row_index: Excel行号
            user_name: 当前用户姓名
            project_id: 项目号
            source_column: 文件3专用，'M'或'L'，表示筛选来源
            file_manager: 文件管理器实例（用于自动勾选）
            viewer: Treeview控件（用于立即刷新显示）
            item_id: Treeview中的行ID（用于立即刷新显示）
            columns: 列名列表（用于查找"是否已完成"列索引）
        """
        super().__init__(parent)
        
        self.interface_id = interface_id
        self.file_type = file_type
        self.file_path = file_path
        self.row_index = row_index
        self.user_name = user_name
        self.user_roles = user_roles or []
        self.project_id = project_id
        self.source_column = source_column
        self.file_manager = file_manager
        self.viewer = viewer  # 保存Treeview引用
        self.item_id = item_id  # 保存行ID
        self.columns = columns  # 保存列名
        self.on_success = on_success
        self.has_assignor = has_assignor
        
        # 【新增】存储已填写的回文单号信息
        self.existing_response = None  # 存储已填写的回文单号
        self.completed_info = None     # 存储完成信息

        # 优先从主程序配置中获取 data_folder，并同步到 registry hooks
        data_folder = self._resolve_data_folder_from_app()
        if data_folder:
            try:
                from registry import hooks as registry_hooks
                registry_hooks.set_data_folder(data_folder)
            except Exception:
                pass
        else:
            # 兜底：从当前文件路径推导（寻找 .registry）
            try:
                from registry import hooks as registry_hooks
                registry_hooks._ensure_data_folder_from_path(self.file_path)
            except Exception:
                pass
        
        # 查询Registry中是否已填写回文单号
        self._load_existing_response()
        
        self.setup_ui()

    def _resolve_data_folder_from_app(self) -> str:
        """从主程序配置中解析数据文件夹路径。"""
        # 优先从顶层窗口获取 app 引用
        try:
            top = self.winfo_toplevel()
            app = getattr(top, "app", None)
            if app and isinstance(getattr(app, "config", None), dict):
                folder = str(app.config.get("folder_path", "") or "").strip()
                if folder:
                    return folder
        except Exception:
            pass

        # 兜底：沿 master 链查找 app 引用
        try:
            node = self
            for _ in range(10):
                app = getattr(node, "app", None)
                if app and isinstance(getattr(app, "config", None), dict):
                    folder = str(app.config.get("folder_path", "") or "").strip()
                    if folder:
                        return folder
                node = getattr(node, "master", None)
                if node is None:
                    break
        except Exception:
            pass
        return ""
    
    def _load_existing_response(self):
        """从Registry查询已填写的回文单号"""
        try:
            from registry.hooks import _cfg
            from registry.db import get_connection, close_connection_after_use
            from registry.util import make_task_id
            
            cfg = _cfg()
            if not cfg.get('registry_enabled'):
                return
            
            db_path = cfg.get('registry_db_path')
            if not db_path or not os.path.exists(db_path):
                return
            
            # 【修复】去除接口号的角色后缀，与extract_interface_id保持一致
            # 例如 "S-YA---1ZJ-02-25C3-25C3(建筑总图室主任)" -> "S-YA---1ZJ-02-25C3-25C3"
            import re
            clean_interface_id = re.sub(r'\([^)]*\)$', '', self.interface_id).strip() if self.interface_id else self.interface_id
            
            task_id = make_task_id(
                self.file_type,
                self.project_id,
                clean_interface_id,  # 使用清理后的接口号
                os.path.basename(self.file_path),
                self.row_index
            )
            
            conn = get_connection(db_path, bool(cfg.get('registry_wal', False)))
            try:
                cursor = conn.execute("""
                    SELECT response_number, completed_at, completed_by
                    FROM tasks
                    WHERE id = ? AND status IN ('completed', 'confirmed')
                """, (task_id,))
                
                row = cursor.fetchone()
                if row and row[0]:  # 确保response_number不为空
                    self.existing_response = row[0]
                    self.completed_info = {
                        'completed_at': row[1],
                        'completed_by': row[2]
                    }
            finally:
                close_connection_after_use()
        except Exception as e:
            print(f"[Registry] 查询已填写回文单号失败: {e}")
    
    def setup_ui(self):
        """设置界面"""
        # 居中显示
        self.transient(self.master)
        self.grab_set()
        
        if self.existing_response:
            # 【已填写回文单号】显示只读信息
            self.title("回文单号（已填写）")
            self.geometry("450x280")
            self.resizable(False, False)
            
            # 标题
            title_label = ttk.Label(self, text=f"接口号: {self.interface_id}",
                                    font=('Arial', 12, 'bold'))
            title_label.pack(pady=10)
            
            # 显示已填写的回文单号（只读）
            info_frame = ttk.LabelFrame(self, text="已填写信息", padding=15)
            info_frame.pack(pady=10, padx=20, fill='both', expand=True)
            
            ttk.Label(info_frame, text="回文单号:").grid(row=0, column=0, sticky='w', padx=5, pady=8)
            response_label = ttk.Label(info_frame, text=self.existing_response,
                                       font=('Arial', 11, 'bold'), foreground='blue')
            response_label.grid(row=0, column=1, sticky='w', padx=5, pady=8)
            
            if self.completed_info:
                if self.completed_info.get('completed_by'):
                    ttk.Label(info_frame, text="填写人:").grid(row=1, column=0, sticky='w', padx=5, pady=8)
                    ttk.Label(info_frame, text=self.completed_info['completed_by']).grid(row=1, column=1, sticky='w', padx=5, pady=8)
                
                if self.completed_info.get('completed_at'):
                    ttk.Label(info_frame, text="填写时间:").grid(row=2, column=0, sticky='w', padx=5, pady=8)
                    completed_time = str(self.completed_info['completed_at'])[:19]  # 截断到秒
                    ttk.Label(info_frame, text=completed_time).grid(row=2, column=1, sticky='w', padx=5, pady=8)
            
            # 关闭按钮（优化样式）
            button_frame = ttk.Frame(self)
            button_frame.pack(pady=15)
            close_btn = ttk.Button(button_frame, text="关闭", command=self.destroy, width=12)
            close_btn.pack()
            
        else:
            # 【未填写回文单号】显示输入界面（原有逻辑）
            self.title("回文单号输入")
            self.geometry("400x200")
            self.resizable(False, False)
            
            # 标题
            title_label = ttk.Label(self, text=f"接口号: {self.interface_id}", 
                                    font=('Arial', 12, 'bold'))
            title_label.pack(pady=10)
            
            # 输入框
            input_frame = ttk.Frame(self)
            input_frame.pack(pady=10, padx=20, fill='x')
            
            ttk.Label(input_frame, text="回文单号:").pack(side='left', padx=5)
            
            self.entry = ttk.Entry(input_frame, width=30)
            self.entry.pack(side='left', padx=5, fill='x', expand=True)
            self.entry.focus_set()
            
            # 按钮
            button_frame = ttk.Frame(self)
            button_frame.pack(pady=20)
            
            ttk.Button(button_frame, text="确认", command=self.on_confirm).pack(side='left', padx=10)
            ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)
            
            # 绑定Enter键
            self.entry.bind('<Return>', lambda e: self.on_confirm())
    
    def on_confirm(self):
        """确认按钮回调"""
        response_number = self.entry.get().strip()
        
        if not response_number:
            messagebox.showwarning("警告", "请输入回文单号", parent=self)
            return
        
        try:
            manager = get_write_task_manager()
            role = ""
            try:
                if isinstance(self.user_roles, (list, tuple)) and self.user_roles:
                    role = " ".join(str(x) for x in self.user_roles if x)
                elif self.user_roles:
                    role = str(self.user_roles)
            except Exception:
                role = ""
            data_folder = self._resolve_data_folder_from_app() or None
            if not data_folder:
                try:
                    from registry import hooks as registry_hooks
                    data_folder = registry_hooks.get_data_folder()
                except Exception:
                    data_folder = None
            task = manager.submit_response_task(
                file_path=self.file_path,
                file_type=self.file_type,
                row_index=self.row_index,
                interface_id=self.interface_id,
                response_number=response_number,
                user_name=self.user_name,
                project_id=self.project_id,
                source_column=self.source_column,
                role=role,
                data_folder=data_folder,
                description=f"{self.user_name} 填写回文单号 {self.interface_id}",
            )
            try:
                cache = get_pending_cache()
                cache.add_response_entry(
                    task.task_id,
                    {
                        "file_path": self.file_path,
                        "file_type": self.file_type,
                        "row_index": self.row_index,
                        "response_number": response_number,
                        "user_name": self.user_name,
                        "project_id": self.project_id,
                        "has_assignor": self.has_assignor,
                    },
                )
            except Exception as cache_error:
                print(f"[PendingCache] 记录回文单号任务失败: {cache_error}")
            messagebox.showinfo("已提交", "回文单号写入任务已提交，后台将自动执行。", parent=self)
            if callable(self.on_success):
                try:
                    self.on_success(self.file_path, self.row_index, self.file_type)
                except Exception as cb_error:
                    print(f"[PendingCache] 回调失败: {cb_error}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("错误", f"提交写入任务失败: {str(e)}", parent=self)


def _write_response_to_excel_unlocked(
    file_path,
    file_type,
    row_index,
    response_number,
    user_name,
    project_id,
    source_column=None,
    *,
    interface_id=None,
    return_details=False,
):
    """
    写入回文单号到Excel文件
    
    参数:
        file_path: Excel文件路径
        file_type: 文件类型(1-6)
        row_index: Excel行号（从2开始，因为第1行是标题）
        response_number: 回文单号
        user_name: 用户姓名
        project_id: 项目号
        source_column: 文件3专用，'M'或'L'
        interface_id: 用于确认原行号仍对应同一接口；未传时保持旧调用兼容
        return_details: True时返回包含实际写入行号的ResponseWriteResult
    
    返回:
        bool | ResponseWriteResult: 成功返回True；详细模式返回实际写入结果
    """
    wb = None
    stage = "PRECHECK"
    committed = False
    resolved_row = int(row_index)
    _make_stdio_unicode_safe()

    def _success(*, already_present=False):
        result = ResponseWriteResult(
            success=True,
            row_index=int(resolved_row),
            already_present=bool(already_present),
        )
        return result if return_details else True

    try:
        if not os.path.exists(file_path):
            raise ExcelWriteError(
                "FILE_NOT_FOUND",
                stage,
                f"目标Excel文件不存在：{file_path}",
                retryable=False,
                committed=False,
            )
        
        try:
            with open(file_path, 'r+b'):
                pass
        except PermissionError as exc:
            lock_owner = get_excel_lock_owner(file_path)
            owner_text = f"，当前占用者可能为【{lock_owner}】" if lock_owner else ""
            raise ExcelWriteError(
                "FILE_LOCKED" if lock_owner else "FILE_ACCESS_DENIED",
                stage,
                f"无法以读写方式打开目标文件，文件可能被占用或当前Windows账户无写权限{owner_text}：{exc}",
                retryable=True,
                committed=False,
            ) from exc
        
        if is_legacy_xls(file_path):
            stage = "VALIDATE_ROW"
            resolved_row = _resolve_xls_response_row(
                file_path,
                file_type,
                row_index,
                interface_id,
            )
            legacy_source_column = source_column
            if file_type == 3 and legacy_source_column not in {"M", "L"}:
                m_value = read_legacy_xls_cell(file_path, f"M{resolved_row}")
                l_value = read_legacy_xls_cell(file_path, f"L{resolved_row}")
                t_value = read_legacy_xls_cell(file_path, f"T{resolved_row}")
                q_value = read_legacy_xls_cell(file_path, f"Q{resolved_row}")
                if m_value not in (None, "") and t_value in (None, ""):
                    legacy_source_column = "M"
                elif l_value not in (None, "") and q_value in (None, ""):
                    legacy_source_column = "L"
                else:
                    legacy_source_column = "M"

            columns = get_write_columns(file_type, resolved_row, None, legacy_source_column)
            if not columns:
                raise ExcelWriteError(
                    "WRITE_COLUMNS_UNKNOWN",
                    "RESOLVE_COLUMNS",
                    f"无法确定写入列位置: file_type={file_type}",
                    retryable=False,
                    committed=False,
                )

            response_cell = f"{columns['response_col']}{resolved_row}"
            stage = "CHECK_EXISTING"
            should_write = _ensure_response_cell_writable(
                read_legacy_xls_cell(file_path, response_cell),
                response_number,
                response_cell,
            )
            if not should_write:
                _safe_print(f"[幂等] 回文单号已存在，无需重复写入: {response_cell}")
                return _success(already_present=True)

            updates = [
                {"cell": response_cell, "value": response_number},
                {"cell": f"{columns['time_col']}{resolved_row}", "value": date.today().isoformat()},
                {"cell": f"{columns['name_col']}{resolved_row}", "value": user_name},
            ]
            for header_update in columns.get("header_updates", []):
                header_cell = str(header_update["cell"])
                expected_header = str(header_update["value"])
                current_header = read_legacy_xls_cell(file_path, header_cell)
                if current_header in (None, ""):
                    updates.append({"cell": header_cell, "value": expected_header})
                elif normalize_header_text(current_header) != normalize_header_text(expected_header):
                    raise ExcelWriteError(
                        "PROGRAM_COLUMN_CONFLICT",
                        "RESOLVE_COLUMNS",
                        f"{header_cell}已有业务表头“{current_header}”，未写入程序字段“{expected_header}”。",
                        retryable=False,
                        committed=False,
                    )
            if file_type == 6:
                expected_time = read_legacy_xls_cell(file_path, f"I{resolved_row}")
                if expected_time:
                    import pandas as pd

                    parsed = pd.to_datetime(expected_time, errors="coerce")
                    if pd.notna(parsed):
                        reply_status = "按时回复" if date.today() <= parsed.date() else "延期回复"
                        updates.append({"cell": f"M{resolved_row}", "value": reply_status})

            stage = "SAVE_TEMP"
            write_legacy_xls_cells(file_path, updates)
            committed = True
            stage = "VERIFY_FINAL"
            _verify_xls_response(
                file_path,
                resolved_row,
                columns['response_col'],
                response_number,
            )
            _safe_print(f"成功写入.xls: {file_path}, 行{resolved_row}, 回文单号={response_number}")
            return _success()

        stage = "OPEN_WORKBOOK"
        try:
            snapshot = OoxmlWorksheetSnapshot(file_path)
        except Exception as exc:
            raise ExcelWriteError(
                "OPEN_FAILED",
                stage,
                f"无法打开目标Excel文件：{exc}",
                retryable=isinstance(exc, (PermissionError, OSError)),
                committed=False,
            ) from exc
        sheet_path = snapshot.sheet_path

        stage = "VALIDATE_ROW"
        resolved_row = _resolve_ooxml_response_row(
            snapshot,
            file_type,
            row_index,
            interface_id,
        )
        columns = get_ooxml_write_columns(
            file_type,
            resolved_row,
            snapshot,
            source_column,
        )
        if not columns:
            raise ExcelWriteError(
                "WRITE_COLUMNS_UNKNOWN",
                "RESOLVE_COLUMNS",
                f"无法确定写入列位置: file_type={file_type}",
                retryable=False,
                committed=False,
            )
        
        response_col = columns['response_col']
        time_col = columns['time_col']
        name_col = columns['name_col']

        response_cell = f"{response_col}{resolved_row}"
        stage = "CHECK_EXISTING"
        should_write = _ensure_response_cell_writable(
            snapshot.value(response_cell),
            response_number,
            response_cell,
        )
        if not should_write:
            _safe_print(f"[幂等] 回文单号已存在，无需重复写入: {response_cell}")
            return _success(already_present=True)

        updates = [
            {"cell": response_cell, "value": response_number},
            {"cell": f"{time_col}{resolved_row}", "value": date.today().strftime('%Y-%m-%d')},
            {"cell": f"{name_col}{resolved_row}", "value": user_name},
        ]
        updates.extend(columns.get("header_updates", []))
        
        if file_type == 6:
            try:
                expected_time = snapshot.value(f"I{resolved_row}")
                
                # 比较当前日期和预期时间
                from datetime import datetime
                today = date.today()
                
                # 解析预期时间
                if expected_time:
                    try:
                        # 尝试解析为日期对象
                        if isinstance(expected_time, datetime):
                            expected_date = expected_time.date()
                        elif isinstance(expected_time, date):
                            expected_date = expected_time
                        else:
                            # 尝试字符串解析
                            import pandas as pd
                            expected_text = str(expected_time).strip()
                            parsed = None
                            if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", expected_text):
                                try:
                                    numeric_date = float(expected_text)
                                    if 1 <= numeric_date <= 100000:
                                        from openpyxl.utils.datetime import from_excel
                                        from openpyxl.utils.datetime import CALENDAR_MAC_1904, WINDOWS_EPOCH

                                        epoch = CALENDAR_MAC_1904 if snapshot.date_1904 else WINDOWS_EPOCH
                                        parsed = from_excel(numeric_date, epoch=epoch)
                                except Exception:
                                    parsed = None
                            if parsed is None:
                                parsed = pd.to_datetime(expected_time, errors='coerce')
                            if pd.notna(parsed):
                                expected_date = parsed.date() if hasattr(parsed, "date") else None
                            else:
                                expected_date = None
                        
                        # 根据对比结果写入M列（第13列）
                        if expected_date:
                            if today <= expected_date:
                                reply_status = "按时回复"
                            else:
                                reply_status = "延期回复"
                            
                            updates.append({"cell": f"M{resolved_row}", "value": reply_status})
                            _safe_print(f"[文件6] 自动更新M列: {reply_status} (预期:{expected_date}, 实际:{today})")
                        else:
                            _safe_print("[文件6] 无法解析预期时间，跳过M列更新")
                    except Exception as parse_error:
                        _safe_print(f"[文件6] 解析预期时间失败: {parse_error}")
                else:
                    _safe_print("[文件6] I列预期时间为空，跳过M列更新")
            except Exception as e:
                _safe_print(f"[文件6] 更新M列失败: {e}")
                # 即使M列更新失败，也不影响回文单号写入
        
        stage = "SAVE_TEMP"
        atomic_patch_ooxml_cells(file_path, sheet_path, updates)
        committed = True

        stage = "VERIFY_FINAL"
        _safe_print("[验证] 开始验证Excel写入...")
        _verify_xlsx_response(
            file_path,
            resolved_row,
            response_col,
            response_number,
            sheet_path=sheet_path,
        )
        _safe_print("[验证] ✓ Excel写入验证成功")
        _safe_print(f"成功写入: {file_path}, 行{resolved_row}, 回文单号={response_number}")
        return _success()

    except ExcelWriteError as e:
        _safe_print("[ERROR] 写入回文单号失败!")
        _safe_print(f"  文件路径: {file_path}")
        _safe_print(f"  文件类型: {file_type}")
        _safe_print(f"  原行号: {row_index}")
        _safe_print(f"  实际行号: {resolved_row}")
        _safe_print(f"  回文单号: {response_number}")
        _safe_print(f"  错误信息: {e}")
        raise
    except Exception as e:
        wrapped = ExcelWriteError(
            "EXCEL_WRITE_FAILED",
            stage,
            f"回文单号写入发生未分类异常：{e}",
            retryable=isinstance(e, (PermissionError, OSError)),
            committed=committed,
        )
        _safe_print(f"[ERROR] {wrapped}")
        raise wrapped from e
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def write_response_to_excel(
    file_path,
    file_type,
    row_index,
    response_number,
    user_name,
    project_id,
    source_column=None,
    *,
    interface_id=None,
    return_details=False,
):
    """在单个工作簿短时共享锁内执行回文写入。"""
    if not os.path.exists(file_path):
        return _write_response_to_excel_unlocked(
            file_path,
            file_type,
            row_index,
            response_number,
            user_name,
            project_id,
            source_column,
            interface_id=interface_id,
            return_details=return_details,
        )
    with SharedWorkbookLock(file_path):
        return _write_response_to_excel_unlocked(
            file_path,
            file_type,
            row_index,
            response_number,
            user_name,
            project_id,
            source_column,
            interface_id=interface_id,
            return_details=return_details,
        )


def _fu_base_interface_id(value) -> str:
    return re.sub(r"#\d{2}$", "", _normalize_response_interface_id(value))


def _resolve_xlsx_fu_row(worksheet, row_index, interface_id) -> int:
    requested_row = int(row_index)
    expected = _normalize_response_interface_id(interface_id)
    base_expected = _fu_base_interface_id(expected)
    if not base_expected:
        return requested_row
    if requested_row >= 2 and _fu_base_interface_id(worksheet[f"B{requested_row}"].value) == base_expected:
        return requested_row

    populated_rows = None
    try:
        populated_rows = sorted({
            int(row)
            for (row, column), cell in worksheet._cells.items()
            if int(row) >= 2 and int(column) == 2 and cell.value not in (None, "")
        })
    except Exception:
        populated_rows = None
    candidate_rows = populated_rows
    if candidate_rows is None:
        candidate_rows = range(2, max(2, int(worksheet.max_row or 2)) + 1)

    matches = []
    for candidate_row in candidate_rows:
        if _fu_base_interface_id(worksheet[f"B{candidate_row}"].value) == base_expected:
            matches.append(candidate_row)
    if len(matches) == 1:
        return matches[0]
    sequence_match = re.search(r"#(\d{2})$", expected)
    if sequence_match and matches:
        ordered = sorted(
            matches,
            key=lambda row: (
                _response_text(worksheet[f"A{row}"].value),
                _response_text(worksheet[f"C{row}"].value),
                row,
            ),
        )
        sequence = int(sequence_match.group(1))
        if 1 <= sequence <= len(ordered):
            return ordered[sequence - 1]
    if not matches:
        raise ExcelWriteError(
            "ROW_NOT_FOUND",
            "VALIDATE_ROW",
            f"原行号{requested_row}已变化，且找不到FU内部编码：{interface_id}",
            retryable=False,
            committed=False,
        )
    raise ExcelWriteError(
        "ROW_AMBIGUOUS",
        "VALIDATE_ROW",
        f"FU内部编码“{interface_id}”存在{len(matches)}条匹配，已拒绝自动写入。",
        retryable=False,
        committed=False,
    )


def _resolve_xls_fu_row(file_path, row_index, interface_id) -> int:
    requested_row = int(row_index)
    expected = _normalize_response_interface_id(interface_id)
    base_expected = _fu_base_interface_id(expected)
    if not base_expected:
        return requested_row
    if _fu_base_interface_id(read_legacy_xls_cell(file_path, f"B{requested_row}")) == base_expected:
        return requested_row

    import xlrd

    book = xlrd.open_workbook(file_path, on_demand=True)
    try:
        sheet = book.sheet_by_index(0)
        matches = [
            excel_row
            for excel_row in range(2, sheet.nrows + 1)
            if _fu_base_interface_id(sheet.cell_value(excel_row - 1, 1)) == base_expected
        ]
        sequence_match = re.search(r"#(\d{2})$", expected)
        if len(matches) == 1:
            return matches[0]
        if sequence_match and matches:
            ordered = sorted(
                matches,
                key=lambda row: (
                    _response_text(sheet.cell_value(row - 1, 0)),
                    _response_text(sheet.cell_value(row - 1, 2)),
                    row,
                ),
            )
            sequence = int(sequence_match.group(1))
            if 1 <= sequence <= len(ordered):
                return ordered[sequence - 1]
    finally:
        book.release_resources()
    if not matches:
        raise ExcelWriteError(
            "ROW_NOT_FOUND",
            "VALIDATE_ROW",
            f"原行号{requested_row}已变化，且找不到FU内部编码：{interface_id}",
            retryable=False,
            committed=False,
        )
    raise ExcelWriteError(
        "ROW_AMBIGUOUS",
        "VALIDATE_ROW",
        f"FU内部编码“{interface_id}”存在{len(matches)}条匹配，已拒绝自动写入。",
        retryable=False,
        committed=False,
    )


def _write_fu_completion_unlocked(
    file_path,
    row_index,
    completion_date=None,
    *,
    interface_id=None,
    return_details=False,
):
    """Write and verify the actual FU date in column D for a type-7 task."""
    wb = None
    verify_wb = None
    resolved_row = int(row_index)
    _make_stdio_unicode_safe()

    def _success():
        result = ResponseWriteResult(success=True, row_index=resolved_row)
        return result if return_details else True

    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(file_path)
        try:
            with open(file_path, "r+b"):
                pass
        except PermissionError:
            lock_owner = get_excel_lock_owner(file_path)
            owner_text = f"【{lock_owner}】" if lock_owner else "其他用户"
            messagebox.showerror("文件占用", f"文件正被{owner_text}占用，请稍后再试")
            return False

        target_date = completion_date or date.today().isoformat()
        if isinstance(target_date, str):
            target_date = date.fromisoformat(target_date[:10])

        if is_legacy_xls(file_path):
            resolved_row = _resolve_xls_fu_row(file_path, row_index, interface_id)
            current_value = read_legacy_xls_cell(file_path, f"D{resolved_row}")
            if current_value not in (None, ""):
                raise ValueError(f"D{resolved_row} 已有实际FU日期，未覆盖原值")
            write_legacy_xls_cells(
                file_path,
                [{
                    "cell": f"D{resolved_row}",
                    "value": target_date,
                    "number_format": "yyyy/m/d",
                }],
            )
            verify_value = read_legacy_xls_cell(file_path, f"D{resolved_row}")
            if hasattr(verify_value, "date"):
                verify_value = verify_value.date()
            if verify_value != target_date:
                raise RuntimeError(
                    f"验证失败：D{resolved_row} 期望 {target_date}，实际 {verify_value}"
                )
            return _success()

        wb = open_workbook_for_edit(file_path)
        ws = wb.active
        resolved_row = _resolve_xlsx_fu_row(ws, row_index, interface_id)
        target_cell = ws[f"D{resolved_row}"]
        if target_cell.value not in (None, ""):
            raise ValueError(f"D{resolved_row} 已有实际FU日期，未覆盖原值")
        target_cell.value = target_date
        target_cell.number_format = "yyyy/m/d"
        atomic_save_workbook(wb, file_path)
        wb.close()
        wb = None

        verify_wb = open_workbook_for_edit(file_path)
        verify_value = verify_wb.active[f"D{resolved_row}"].value
        if hasattr(verify_value, "date"):
            verify_value = verify_value.date()
        if verify_value != target_date:
            raise RuntimeError(
                f"验证失败：D{resolved_row} 期望 {target_date}，实际 {verify_value}"
            )
        verify_wb.close()
        verify_wb = None
        return _success()
    except Exception as e:
        _safe_print(f"[ERROR] 写入FU日期失败: {e}")
        try:
            messagebox.showerror("写入失败", f"无法写入实际FU日期\n\n错误：{e}")
        except Exception:
            pass
        return False
    finally:
        for workbook in (wb, verify_wb):
            try:
                if workbook is not None:
                    workbook.close()
            except Exception:
                pass
        try:
            if verify_wb is not None:
                verify_wb.close()
        except Exception:
            pass


def write_fu_completion_to_excel(
    file_path,
    row_index,
    completion_date=None,
    *,
    interface_id=None,
    return_details=False,
):
    """在单个工作簿短时共享锁内写入FU实际完成日期。"""
    if not os.path.exists(file_path):
        return _write_fu_completion_unlocked(
            file_path,
            row_index,
            completion_date,
            interface_id=interface_id,
            return_details=return_details,
        )
    with SharedWorkbookLock(file_path):
        return _write_fu_completion_unlocked(
            file_path,
            row_index,
            completion_date,
            interface_id=interface_id,
            return_details=return_details,
        )


def get_write_columns(file_type, row_index, worksheet, source_column=None):
    """
    获取各文件类型的写入列位置
    
    参数:
        file_type: 文件类型(1-6)
        row_index: Excel行号
        worksheet: openpyxl工作表对象
        source_column: 文件3专用，'M'或'L'
    
    返回:
        dict: {'response_col': 'S', 'time_col': 'N', 'name_col': 'V'}
        或 None（如果无法确定）
    """
    # 文件1和文件4使用业务区之外的程序专用列，避免覆盖“备注/重新打开编号”。
    column_map = {
        1: {'response_col': 'W', 'time_col': 'M', 'name_col': 'V'},
        2: {'response_col': 'P', 'time_col': 'N', 'name_col': 'AL'},
        4: {'response_col': 'U', 'time_col': 'V', 'name_col': 'AX'},
        5: {'response_col': 'V', 'time_col': 'N', 'name_col': 'W'},
        6: {'response_col': 'L', 'time_col': 'J', 'name_col': 'N'},
    }
    
    if file_type in column_map:
        columns = dict(column_map[file_type])
        program_headers = {
            1: [('V', '程序填写人', 'name_col'), ('W', '程序回文单号', 'response_col')],
            2: [('AL', '程序填写人', 'name_col')],
            4: [('AX', '程序填写人', 'name_col')],
        }
        if worksheet is not None:
            for preferred, header, key in program_headers.get(file_type, []):
                columns[key] = ensure_program_column(worksheet, preferred, header)
        else:
            columns['header_updates'] = [
                {'cell': f'{preferred}1', 'value': header}
                for preferred, header, _key in program_headers.get(file_type, [])
            ]
        return columns
    
    # 文件3特殊逻辑：根据source_column判断
    if file_type == 3:
        if source_column == 'M':
            # M列筛选：V/T/BM
            result = {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif source_column == 'L':
            # L列筛选：S/Q/BM
            result = {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 如果未指定，尝试自动判断
            result = determine_file3_source_and_columns(row_index, worksheet)
        if worksheet is not None:
            result['name_col'] = ensure_program_column(worksheet, 'BM', '程序填写人')
        else:
            result['header_updates'] = [{'cell': 'BM1', 'value': '程序填写人'}]
        return result
    
    return None


def _ensure_ooxml_program_column(snapshot, headers, max_column_state, preferred, header_text):
    expected = normalize_header_text(header_text)
    for column, value in headers.items():
        if normalize_header_text(value) == expected:
            return column, False

    preferred_number = _column_letter_to_number(preferred)
    preferred_value = normalize_header_text(headers.get(preferred, ""))
    if preferred_value and preferred_value != expected:
        target_number = max(max_column_state[0], preferred_number) + 1
    else:
        target_number = preferred_number
    target_column = column_number_to_letter(target_number)
    target_value = normalize_header_text(headers.get(target_column, ""))
    if target_value not in {"", expected}:
        raise ExcelWriteError(
            "PROGRAM_COLUMN_CONFLICT",
            "RESOLVE_COLUMNS",
            f"无法建立程序专用列“{header_text}”，{target_column}1已有业务表头“{headers.get(target_column)}”。",
            retryable=False,
            committed=False,
        )
    headers[target_column] = header_text
    max_column_state[0] = max(max_column_state[0], target_number)
    return target_column, not bool(target_value)


def get_ooxml_write_columns(file_type, row_index, snapshot, source_column=None):
    """Resolve response columns from OOXML metadata without loading openpyxl."""
    column_map = {
        1: {'response_col': 'W', 'time_col': 'M', 'name_col': 'V'},
        2: {'response_col': 'P', 'time_col': 'N', 'name_col': 'AL'},
        4: {'response_col': 'U', 'time_col': 'V', 'name_col': 'AX'},
        5: {'response_col': 'V', 'time_col': 'N', 'name_col': 'W'},
        6: {'response_col': 'L', 'time_col': 'J', 'name_col': 'N'},
    }
    if file_type == 3:
        if source_column not in {'M', 'L'}:
            m_value = snapshot.value(f"M{row_index}")
            l_value = snapshot.value(f"L{row_index}")
            t_value = snapshot.value(f"T{row_index}")
            q_value = snapshot.value(f"Q{row_index}")
            if _response_text(m_value) and not _response_text(t_value):
                source_column = 'M'
            elif _response_text(l_value) and not _response_text(q_value):
                source_column = 'L'
            else:
                source_column = 'M'
        columns = (
            {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
            if source_column == 'M'
            else {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        )
    else:
        columns = dict(column_map.get(file_type) or {})
    if not columns:
        return None

    header_specs = {
        1: [('name_col', 'V', '程序填写人'), ('response_col', 'W', '程序回文单号')],
        2: [('name_col', 'AL', '程序填写人')],
        3: [('name_col', 'BM', '程序填写人')],
        4: [('name_col', 'AX', '程序填写人')],
    }
    headers = snapshot.header_values()
    max_column_state = [snapshot.max_column_number()]
    header_updates = []
    for key, preferred, header_text in header_specs.get(file_type, []):
        column, should_create = _ensure_ooxml_program_column(
            snapshot,
            headers,
            max_column_state,
            preferred,
            header_text,
        )
        columns[key] = column
        if should_create:
            target_style = snapshot.style_id(f"{column}1")
            if target_style is None:
                previous = column_number_to_letter(max(1, _column_letter_to_number(column) - 1))
                target_style = snapshot.style_id(f"{previous}1")
            header_updates.append({
                'cell': f'{column}1',
                'value': header_text,
                'style_id': target_style,
            })
    columns['header_updates'] = header_updates
    return columns


def determine_file3_source_and_columns(row_index, worksheet):
    """
    判断文件3某行是因M列还是L列被筛选出
    
    参数:
        row_index: Excel行号
        worksheet: openpyxl工作表对象
    
    返回:
        dict: 写入列位置
    """
    try:
        # 读取M列和L列的值
        m_val = worksheet[f"M{row_index}"].value
        l_val = worksheet[f"L{row_index}"].value
        
        # 读取T列和Q列的值（回复时间列）
        t_val = worksheet[f"T{row_index}"].value
        q_val = worksheet[f"Q{row_index}"].value
        
        # 简化判断逻辑：
        # 如果M列有时间数据且T列为空，判断为M列来源
        # 如果L列有时间数据且Q列为空，判断为L列来源
        # 优先M列
        
        m_has_time = m_val is not None and str(m_val).strip() != ''
        t_is_empty = t_val is None or str(t_val).strip() == ''
        
        l_has_time = l_val is not None and str(l_val).strip() != ''
        q_is_empty = q_val is None or str(q_val).strip() == ''
        
        if m_has_time and t_is_empty:
            # M列来源
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif l_has_time and q_is_empty:
            # L列来源
            return {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 默认M列
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
    
    except Exception as e:
        print(f"判断文件3来源失败: {e}")
        # 默认返回M列
        return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}


# 测试代码
if __name__ == "__main__":
    # 测试get_write_columns
    columns_1 = get_write_columns(1, 5, None)
    print(f"文件1写入列: {columns_1}")
    
    columns_3_m = get_write_columns(3, 5, None, 'M')
    print(f"文件3(M列)写入列: {columns_3_m}")
    
    columns_3_l = get_write_columns(3, 5, None, 'L')
    print(f"文件3(L列)写入列: {columns_3_l}")
