#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
import threading

import pandas as pd
import pytest
from openpyxl import Workbook


pytestmark = pytest.mark.allow_empty_name


def test_local_cache_get_read_connection_does_not_deadlock(tmp_path):
    from registry.local_cache import LocalCacheManager

    network_db_path = tmp_path / "network" / "registry.db"
    network_db_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(network_db_path))
    try:
        conn.execute("CREATE TABLE demo (id INTEGER PRIMARY KEY, name TEXT)")
        conn.execute("INSERT INTO demo(name) VALUES ('ok')")
        conn.commit()
    finally:
        conn.close()

    manager = LocalCacheManager(
        str(network_db_path),
        local_cache_dir=str(tmp_path / "cache"),
        sync_interval=300,
    )

    result = {"done": False, "conn": None, "error": None}

    def worker():
        try:
            result["conn"] = manager.get_read_connection()
        except Exception as exc:  # pragma: no cover
            result["error"] = str(exc)
        finally:
            result["done"] = True

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    thread.join(timeout=2.0)

    try:
        assert result["done"] is True, "get_read_connection 出现死锁"
        assert result["error"] is None
        assert result["conn"] is not None
        assert result["conn"].execute("SELECT name FROM demo").fetchone()[0] == "ok"
    finally:
        manager.cleanup()


def test_local_cache_rebuilds_when_local_db_is_malformed(tmp_path):
    from registry.local_cache import LocalCacheManager

    network_db_path = tmp_path / "network" / "registry.db"
    network_db_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(network_db_path))
    try:
        conn.execute("CREATE TABLE demo (id INTEGER PRIMARY KEY, name TEXT)")
        conn.execute("INSERT INTO demo(name) VALUES ('healthy')")
        conn.commit()
    finally:
        conn.close()

    cache_dir = tmp_path / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_db_path = cache_dir / "registry_local.db"
    local_db_path.write_bytes(b"not-a-real-sqlite-db")

    manager = LocalCacheManager(
        str(network_db_path),
        local_cache_dir=str(cache_dir),
        sync_interval=300,
    )

    try:
        local_conn = manager.get_read_connection()
        assert local_conn is not None
        assert local_conn.execute("SELECT name FROM demo").fetchone()[0] == "healthy"
    finally:
        manager.cleanup()


def test_get_read_connection_logs_fallback_reason(monkeypatch):
    from registry import db as registry_db

    captured = []
    sentinel = object()

    class BrokenCache:
        network_db_path = "//server/share/.registry/registry.db"

        @staticmethod
        def get_read_connection():
            raise RuntimeError("cache init failed")

    original_cache = registry_db._local_cache_manager
    original_enabled = registry_db._local_cache_enabled
    try:
        registry_db._local_cache_enabled = True
        registry_db._local_cache_manager = BrokenCache()
        monkeypatch.setattr(registry_db, "_is_network_path", lambda _path: True)
        monkeypatch.setattr(registry_db, "get_connection", lambda *_args, **_kwargs: sentinel)
        monkeypatch.setattr(
            registry_db,
            "_diag_log",
            lambda event, **fields: captured.append((event, fields)),
        )

        result = registry_db.get_read_connection("//server/share/.registry/registry.db")

        assert result is sentinel
        assert any(
            event == "read_conn_fallback"
            and fields.get("reason") == "local_cache_init_failed"
            for event, fields in captured
        )
    finally:
        registry_db._local_cache_manager = original_cache
        registry_db._local_cache_enabled = original_enabled


def test_get_display_status_uses_read_connection(monkeypatch):
    from registry import service as registry_service

    class FakeCursor:
        def __init__(self, row=None, rows=None):
            self._row = row
            self._rows = rows or []

        def fetchone(self):
            return self._row

        def fetchall(self):
            return self._rows

    class FakeConn:
        def execute(self, sql, _params=()):
            if "PRAGMA table_info" in sql:
                return FakeCursor(
                    rows=[
                        (0, "status", "TEXT", 0, None, 0),
                        (1, "display_status", "TEXT", 0, None, 0),
                        (2, "assigned_by", "TEXT", 0, None, 0),
                        (3, "role", "TEXT", 0, None, 0),
                        (4, "confirmed_at", "TEXT", 0, None, 0),
                        (5, "responsible_person", "TEXT", 0, None, 0),
                        (6, "ignored", "INTEGER", 0, None, 0),
                    ]
                )
            return FakeCursor(("open", "待完成", None, None, None, "张三", 0))

    monkeypatch.setattr(registry_service, "get_read_connection", lambda _db_path: FakeConn())
    monkeypatch.setattr(
        registry_service,
        "get_connection",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("不应走写连接")),
    )
    monkeypatch.setattr(registry_service, "close_connection_after_use", lambda: None)

    result = registry_service.get_display_status(
        "E:/dummy/.registry/registry.db",
        False,
        [
            {
                "file_type": 1,
                "project_id": "2016",
                "interface_id": "S-TEST-01",
                "source_file": "source.xlsx",
                "row_index": 2,
                "interface_time": "2026-02-10",
            }
        ],
    )

    assert len(result) == 1
    assert next(iter(result.values())).startswith("📌")


def test_atomic_save_workbook_keeps_file_openable(tmp_path):
    from utils.excel_io import atomic_save_workbook, open_workbook_for_edit

    file_path = tmp_path / "sample.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "before"
    wb.save(str(file_path))
    wb.close()

    wb = open_workbook_for_edit(str(file_path))
    try:
        wb.active["A1"] = "after"
        atomic_save_workbook(wb, str(file_path))
    finally:
        wb.close()

    verify_wb = open_workbook_for_edit(str(file_path))
    try:
        assert verify_wb.active["A1"].value == "after"
    finally:
        verify_wb.close()

    leftover = [p.name for p in tmp_path.iterdir() if p.name.startswith(".__excel_write_")]
    assert leftover == []


def test_write_response_to_excel_keeps_workbook_valid(tmp_path, monkeypatch):
    from ui.input_handler import write_response_to_excel
    from utils.excel_io import open_workbook_for_edit

    file_path = tmp_path / "response.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "标题"
    ws["L2"] = ""
    ws["J2"] = ""
    ws["N2"] = ""
    wb.save(str(file_path))
    wb.close()

    monkeypatch.setattr("tkinter.messagebox.showerror", lambda *_args, **_kwargs: None)

    ok = write_response_to_excel(
        str(file_path),
        file_type=6,
        row_index=2,
        response_number="HF-001",
        user_name="测试用户",
        project_id="2026",
        source_column=None,
    )

    assert ok is True

    verify_wb = open_workbook_for_edit(str(file_path))
    try:
        verify_ws = verify_wb.active
        assert verify_ws["L2"].value == "HF-001"
        assert verify_ws["N2"].value == "测试用户"
    finally:
        verify_wb.close()


def test_save_assignments_batch_keeps_workbook_valid(tmp_path, monkeypatch):
    from services.distribution import save_assignments_batch
    from utils.excel_io import open_workbook_for_edit

    file_path = tmp_path / "assignment.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "标题"
    ws["R2"] = ""
    wb.save(str(file_path))
    wb.close()

    monkeypatch.setattr("services.distribution.log_info", lambda *_args, **_kwargs: None)
    monkeypatch.setattr("services.distribution.log_success", lambda *_args, **_kwargs: None)
    monkeypatch.setattr("services.distribution.log_error", lambda *_args, **_kwargs: None)
    monkeypatch.setattr("registry.hooks.on_assigned", lambda *_args, **_kwargs: None)

    result = save_assignments_batch(
        [
            {
                "file_type": 1,
                "file_path": str(file_path),
                "row_index": 2,
                "assigned_name": "张三",
                "interface_id": "S-TEST-01",
                "project_id": "2016",
                "assigned_by": "测试接口工程师",
            }
        ]
    )

    assert result["success_count"] == 1

    verify_wb = open_workbook_for_edit(str(file_path))
    try:
        assert verify_wb.active["R2"].value == "张三"
    finally:
        verify_wb.close()


def test_registry_accepts_2416_without_touching_existing_project(tmp_path, monkeypatch):
    from registry import db as registry_db
    from registry import hooks

    data_folder = tmp_path / "data"
    data_folder.mkdir()
    monkeypatch.setattr(hooks, "_DATA_FOLDER", str(data_folder))
    monkeypatch.setattr(hooks, "_RUNTIME_DISABLED_REASON", "")
    monkeypatch.setattr(hooks, "_RUNTIME_DISABLE_NOTIFIED", False)

    old_df = pd.DataFrame([{
        "项目号": "2026",
        "接口号": "S-OLD-01",
        "接口时间": "2026.05.27",
        "科室": "结构一室",
        "责任人": "张三",
        "原始行号": 2,
        "_completed_col_value": "",
    }])
    new_df = pd.DataFrame([{
        "项目号": "2416",
        "接口号": "S-NEW-01",
        "接口时间": "2026.05.27",
        "科室": "结构一室",
        "责任人": "李四",
        "原始行号": 2,
        "_completed_col_value": "",
    }])

    hooks.on_process_done(1, "2026", str(tmp_path / "2026按项目导出IDI手册2026-05-27.xlsx"), old_df)
    hooks.on_process_done(1, "2416", str(tmp_path / "2416按项目导出IDI手册2026-05-27.xlsx"), new_df)
    registry_db.close_connection_after_use()

    db_path = data_folder / ".registry" / "registry.db"
    conn = sqlite3.connect(str(db_path))
    try:
        task_rows = conn.execute(
            "SELECT project_id, interface_id, business_id FROM tasks ORDER BY project_id"
        ).fetchall()
        event_rows = conn.execute(
            "SELECT event, project_id FROM events WHERE event = 'process_done' ORDER BY project_id"
        ).fetchall()
    finally:
        conn.close()
        registry_db.close_connection_after_use()

    assert task_rows == [
        ("2026", "S-OLD-01", "1|2026|S-OLD-01"),
        ("2416", "S-NEW-01", "1|2416|S-NEW-01"),
    ]
    assert event_rows == [("process_done", "2026"), ("process_done", "2416")]
