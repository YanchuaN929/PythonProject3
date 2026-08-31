#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime
import multiprocessing
import os
import pickle
import re
import threading
import time
import zipfile

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from core import main
from utils.dept_config import get_organization_filter_file6


pytestmark = pytest.mark.allow_empty_name


def _force_worksheet_dimension_to_a1(path):
    """模拟真实 POI 文件错误声明 A1:A1，但保留工作表中的业务单元格。"""
    temp_path = path.with_suffix(".dimension-fix.tmp")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(
                    br'<dimension ref="[^"]+"\s*/>',
                    b'<dimension ref="A1:A1"/>',
                    data,
                    count=1,
                )
            target.writestr(item, data)
    os.replace(temp_path, path)


def _response_process_worker(path, row_index, interface_id, response_number, start_event, result_queue):
    try:
        from ui.input_handler import write_response_to_excel

        if not start_event.wait(10):
            result_queue.put((False, "start timeout"))
            return
        result = write_response_to_excel(
            path,
            file_type=1,
            row_index=row_index,
            response_number=response_number,
            user_name=f"用户{row_index}",
            project_id="2026",
            interface_id=interface_id,
            return_details=True,
        )
        result_queue.put((bool(result.success), str(result.row_index)))
    except Exception as exc:  # pragma: no cover - 失败信息由父进程断言显示
        result_queue.put((False, repr(exc)))


@pytest.fixture
def no_registry_merge(monkeypatch):
    monkeypatch.setattr(
        main,
        "_merge_registry_pending_rows",
        lambda **kwargs: (kwargs["final_rows"], set()),
    )


@pytest.mark.parametrize(
    ("file_type", "filename", "cells", "expected_alias", "expected_value"),
    [
        (2, "内部接口信息单报表202620260831.xlsx", {"A2": "row", "R2": "I-2"}, "interface", "I-2"),
        (3, "外部接口ICM报表202620260831.xlsx", {"C2": "I-3", "BL2": "tail"}, "interface", "I-3"),
        (4, "外部接口单报表202620260831.xlsx", {"E2": "I-4", "AW2": "tail"}, "interface", "I-4"),
    ],
)
def test_selected_reader_recovers_poi_a1_dimension(
    tmp_path, file_type, filename, cells, expected_alias, expected_value
):
    path = tmp_path / filename
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "header"
    for cell, value in cells.items():
        worksheet[cell] = value
    workbook.save(path)
    workbook.close()
    _force_worksheet_dimension_to_a1(path)

    records, _columns = main._read_selected_excel_records(str(path), file_type)

    assert len(records) == 1
    assert records[0]["原始行号"] == 2
    assert records[0]["_raw"][expected_alias] == expected_value


def test_file6_current_headers_drive_display_and_assignment(tmp_path, monkeypatch, no_registry_merge):
    from registry import hooks
    from services.distribution import save_assignments_batch

    path = tmp_path / "收发文清单2026.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    headers = {
        "E": "收发文编号",
        "I": "要求回文期限",
        "J": "我方回文日期",
        "M": "回文状态",
        "W": "主办部门（所）",
        "X": "主办室",
        "Y": "主办人",
        "AD": "版次",
    }
    for column, header in headers.items():
        worksheet[f"{column}1"] = header
    worksheet["E2"] = "FH-2026-001"
    worksheet["I2"] = datetime.datetime(2026, 8, 31)
    worksheet["M2"] = "尚未回复"
    worksheet["W2"] = get_organization_filter_file6()
    worksheet["X2"] = "结构一室"
    worksheet["Y2"] = "张三,李四"
    worksheet["AD2"] = "Ａ"
    workbook.save(path)
    workbook.close()

    result = main.process_target_file6(
        str(path), datetime.datetime(2026, 8, 31), valid_names_set={"张三", "李四"}
    )

    assert len(result) == 1
    row = result.iloc[0]
    assert row["主办室"] == "结构一室"
    assert row["责任人"] == "张三,李四"
    assert row["_responsible_col"] == "Y"
    assert row["_assign_col"] == "Y"

    monkeypatch.setattr(hooks, "on_assigned", lambda **_kwargs: True)
    saved = save_assignments_batch([{
        "file_type": 6,
        "file_path": str(path),
        "row_index": 2,
        "assigned_name": "王五",
        "interface_id": "FH-2026-001",
        "project_id": "2026",
        "assigned_by": "室主任",
    }])
    assert saved["success_count"] == 1

    verify = load_workbook(path, read_only=True, data_only=True)
    try:
        assert verify.active["X2"].value == "结构一室"
        assert verify.active["Y2"].value == "王五"
    finally:
        verify.close()


def test_response_writer_preserves_file1_remark_and_file4_reopen_number(tmp_path):
    from ui.input_handler import write_response_to_excel

    file1 = tmp_path / "2026按项目导出IDI手册2026-08-31.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "IDI编号"
    worksheet["M1"] = "实际答复时间"
    worksheet["S1"] = "备注"
    worksheet["A2"] = "IDI-001"
    worksheet["S2"] = "不得覆盖的备注"
    workbook.save(file1)
    workbook.close()

    assert write_response_to_excel(
        str(file1), 1, 2, "HF-001", "张三", "2026", interface_id="IDI-001"
    ) is True
    verify = load_workbook(file1, read_only=True, data_only=True)
    try:
        assert verify.active["S2"].value == "不得覆盖的备注"
        assert verify.active["V1"].value == "程序填写人"
        assert verify.active["V2"].value == "张三"
        assert verify.active["W1"].value == "程序回文单号"
        assert verify.active["W2"].value == "HF-001"
    finally:
        verify.close()

    file4 = tmp_path / "外部接口单报表202620260831.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["E1"] = "接口单编号"
    worksheet["U1"] = "回文编号"
    worksheet["V1"] = "回文日期"
    worksheet["AT1"] = "重新打开编号"
    worksheet["E2"] = "ICM-001"
    worksheet["AT2"] = "REOPEN-BUSINESS-001"
    workbook.save(file4)
    workbook.close()

    assert write_response_to_excel(
        str(file4), 4, 2, "HF-004", "李四", "2026", interface_id="ICM-001"
    ) is True
    verify = load_workbook(file4, read_only=True, data_only=True)
    try:
        assert verify.active["AT2"].value == "REOPEN-BUSINESS-001"
        assert verify.active["AX1"].value == "程序填写人"
        assert verify.active["AX2"].value == "李四"
    finally:
        verify.close()


def test_file2_assignment_uses_program_column_without_overwriting_business(tmp_path, monkeypatch):
    from registry import hooks
    from services.distribution import save_assignments_batch

    path = tmp_path / "内部接口信息单报表202620260831.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["R1"] = "IDI编号"
    worksheet["R2"] = "IDI-002"
    worksheet["AM1"] = "已有业务字段"
    worksheet["AM2"] = "业务值"
    workbook.save(path)
    workbook.close()
    monkeypatch.setattr(hooks, "on_assigned", lambda **_kwargs: True)

    saved = save_assignments_batch([{
        "file_type": 2,
        "file_path": str(path),
        "row_index": 2,
        "assigned_name": "王五",
        "interface_id": "IDI-002",
        "project_id": "2026",
        "assigned_by": "接口工程师",
    }])

    assert saved["success_count"] == 1
    verify = load_workbook(path, read_only=True, data_only=True)
    try:
        assert verify.active["AM1"].value == "已有业务字段"
        assert verify.active["AM2"].value == "业务值"
        assert verify.active["AN1"].value == "程序主办人"
        assert verify.active["AN2"].value == "王五"
    finally:
        verify.close()

    records, _columns = main._read_selected_excel_records(str(path), 2)
    assert records[0]["_raw"]["responsible"] == "王五"
    assert records[0]["_stream_spec"]["assign"] == "AN"


def test_fu_sparse_reader_visits_only_physical_rows(tmp_path):
    path = tmp_path / "2026项目标准表格.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    for column, header in {
        "A": "文件编码",
        "B": "内部编码",
        "C": "中文标题",
        "D": "实际FU日期",
        "E": "FU计划",
        "F": "责任人",
    }.items():
        worksheet[f"{column}1"] = header
    row_number = 1_048_412
    worksheet[f"A{row_number}"] = "FILE-LAST"
    worksheet[f"B{row_number}"] = "FU-LAST"
    worksheet[f"E{row_number}"] = datetime.datetime(2026, 8, 31)
    workbook.save(path)
    workbook.close()

    records, _columns = main._read_selected_excel_records(str(path), 7)

    assert len(records) == 1
    assert records[0]["原始行号"] == row_number
    assert records[0]["_raw"]["internal_code"] == "FU-LAST"


def test_empty_result_cache_requires_current_schema_envelope(tmp_path):
    from services.file_manager import FileIdentityManager

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    cache_dir = tmp_path / "result-cache"
    cache_dir.mkdir()
    manager = FileIdentityManager(
        cache_file=str(tmp_path / "file-cache.json"),
        result_cache_dir=str(cache_dir),
    )
    empty = pd.DataFrame()
    empty.attrs["_stream_schema_version"] = main.STREAM_RESULT_SCHEMA_VERSION

    assert manager.save_cached_result(str(source), "2026", "file2", empty) is True
    loaded = manager.load_cached_result(str(source), "2026", "file2")
    assert isinstance(loaded, pd.DataFrame)
    assert loaded.empty
    assert loaded.attrs["_stream_schema_version"] == main.STREAM_RESULT_SCHEMA_VERSION

    cache_file = manager._get_cache_filename(str(source), "2026", "file2")
    with open(cache_file, "wb") as handle:
        pickle.dump(pd.DataFrame(), handle)
    assert manager.load_cached_result(str(source), "2026", "file2") is None
    assert not os.path.exists(cache_file)


def test_role_tokens_and_multi_owner_matching():
    from base import ExcelProcessorApp

    assert ExcelProcessorApp._parse_role_tokens(
        "设计人员  1818接口工程师2026 接口工程师"
    ) == ["设计人员", "1818接口工程师", "2026接口工程师"]

    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app.user_name = "张三"
    source = pd.DataFrame({
        "责任人": ["张三", "李四,张三", "张三丰", "李四"],
        "原始行号": [2, 3, 4, 5],
    })
    filtered = app._filter_by_single_role(source, "设计人员")
    assert filtered["原始行号"].tolist() == [2, 3]


def test_fullwidth_version_and_java_cst_date_are_supported():
    assert main._get_version_rank("Ａ") == main._get_version_rank("A") == 1
    assert main._get_version_rank("!") == 0
    parsed = main._parse_datetime_value("Mon Aug 31 14:05:06 CST 2026")
    assert parsed == datetime.datetime(2026, 8, 31, 14, 5, 6)
    assert main._parse_datetime_value("Mon Feb 30 00:00:00 CST 2026") is None


def test_two_processes_write_different_rows_without_lost_update(tmp_path):
    path = tmp_path / "2026按项目导出IDI手册2026-08-31.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "IDI编号"
    worksheet["M1"] = "实际答复时间"
    worksheet["S1"] = "备注"
    worksheet["A2"] = "IDI-P1"
    worksheet["A3"] = "IDI-P2"
    workbook.save(path)
    workbook.close()

    context = multiprocessing.get_context("spawn")
    start_event = context.Event()
    result_queue = context.Queue()
    processes = [
        context.Process(
            target=_response_process_worker,
            args=(str(path), 2, "IDI-P1", "HF-P1", start_event, result_queue),
        ),
        context.Process(
            target=_response_process_worker,
            args=(str(path), 3, "IDI-P2", "HF-P2", start_event, result_queue),
        ),
    ]
    for process in processes:
        process.start()
    start_event.set()
    for process in processes:
        process.join(20)
        assert process.exitcode == 0
    outcomes = [result_queue.get(timeout=3) for _process in processes]
    assert all(success for success, _detail in outcomes), outcomes

    verify = load_workbook(path, read_only=True, data_only=True)
    try:
        assert verify.active["W2"].value == "HF-P1"
        assert verify.active["W3"].value == "HF-P2"
    finally:
        verify.close()


def test_cold_cache_prewarm_reads_different_files_in_parallel():
    from base import ExcelProcessorApp

    class Cache:
        def __init__(self):
            self.saved = []

        def load_cached_result(self, *_args):
            return None

        def save_cached_result(self, file_path, project_id, file_type, result):
            self.saved.append((file_path, project_id, file_type, len(result)))
            return True

    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app.file_manager = Cache()
    guard = threading.Lock()
    active = 0
    peak = 0

    def process(file_path, _current_datetime):
        nonlocal active, peak
        with guard:
            active += 1
            peak = max(peak, active)
        time.sleep(0.05)
        with guard:
            active -= 1
        return pd.DataFrame({"source": [file_path]})

    jobs = [
        {
            "file_path": f"file-{index}.xlsx",
            "project_id": str(2000 + index),
            "file_type": "file1",
            "process_func": process,
            "args": (datetime.datetime(2026, 8, 31),),
        }
        for index in range(4)
    ]

    result = app._prewarm_result_caches_parallel(jobs, max_workers=4)

    assert result == {"processed": 4, "failed": 0}
    assert peak >= 2
    assert len(app.file_manager.saved) == 4
