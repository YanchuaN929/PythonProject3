from openpyxl import Workbook

from utils.role_table import read_role_table


def _write_role_book(path, rows):
    workbook = Workbook()
    worksheet = workbook.active
    for row_index, (name, role) in enumerate(rows, start=1):
        worksheet.cell(row=row_index, column=1, value=name)
        worksheet.cell(row=row_index, column=2, value=role)
    workbook.save(path)
    workbook.close()


def test_role_table_without_header_keeps_first_person(tmp_path):
    path = tmp_path / "roles.xlsx"
    _write_role_book(path, [("首位人员", "设计人员"), ("第二位人员", "一室主任")])

    result = read_role_table(path)

    assert result.to_dict("records") == [
        {"姓名": "首位人员", "角色": "设计人员"},
        {"姓名": "第二位人员", "角色": "一室主任"},
    ]


def test_role_table_with_header_skips_only_header(tmp_path):
    path = tmp_path / "roles-with-header.xlsx"
    _write_role_book(path, [("姓名", "角色"), ("首位人员", "设计人员")])

    result = read_role_table(path)

    assert result.to_dict("records") == [{"姓名": "首位人员", "角色": "设计人员"}]


def test_current_role_table_keeps_first_person_and_li_zixiang_role():
    result = read_role_table("excel_bin/姓名角色表.xlsx")

    assert result.iloc[0].to_dict() == {"姓名": "张翼飞", "角色": "设计人员"}
    role = result.loc[result["姓名"] == "李子香", "角色"].iloc[0]
    assert role == "设计人员、1818接口工程师、2026接口工程师"


def test_nuclear_process_role_table_is_readable():
    result = read_role_table("excel_bin/姓名角色表-核电工艺所.xlsx")

    assert result.iloc[0].to_dict() == {"姓名": "贾雅永", "角色": "设计人员"}
    assert len(result) == 169
    duplicated = result[result.duplicated("姓名", keep=False)]
    for _name, rows in duplicated.groupby("姓名"):
        assert rows["角色"].astype(str).str.strip().nunique() == 1
