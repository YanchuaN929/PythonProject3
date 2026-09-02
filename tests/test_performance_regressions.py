import datetime
import sqlite3
import threading
import zipfile
from types import SimpleNamespace

import pandas as pd
from openpyxl import Workbook, load_workbook


def _make_fu_workbook(tmp_path, filename="1916项目标准表格.xlsx"):
    path = tmp_path / filename
    workbook = Workbook()
    sheet = workbook.active
    for column, header in {
        "A": "文件编码",
        "B": "内部编码",
        "C": "中文标题",
        "D": "实际FU日期",
        "E": "FU计划",
        "F": "责任人",
    }.items():
        sheet[f"{column}1"] = header
    sheet["A2"] = "FILE-A"
    sheet["B2"] = "FU-001"
    sheet["C2"] = "标题"
    sheet["E2"] = datetime.datetime(2026, 9, 3)
    sheet["F2"] = "张三"
    workbook.save(path)
    workbook.close()
    return path


def test_sparse_fu_reader_matches_legacy_scalar_types(tmp_path):
    from core import main

    path = _make_fu_workbook(tmp_path)
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        sheet["D5"] = datetime.datetime(2026, 9, 1)
        sheet["E5"] = datetime.datetime(2026, 9, 2)
        sheet["B5"] = "FU-005"
        sheet["A8"] = "styled-empty-row"
        sheet["D8"].number_format = "yyyy/m/d"
        workbook.save(path)
    finally:
        workbook.close()

    fast_records, fast_columns = main._read_sparse_fu_records_ooxml(str(path))
    legacy_records, legacy_columns = main._read_sparse_fu_records_openpyxl(str(path))

    def normalized(records):
        return [
            (
                item["_df_index"],
                item["原始行号"],
                [
                    (key, type(value).__name__, repr(value))
                    for key, value in sorted(item["_raw"].items())
                ],
            )
            for item in records
        ]

    assert fast_columns == legacy_columns
    assert normalized(fast_records) == normalized(legacy_records)


def test_fu_completion_changes_only_target_worksheet_member(tmp_path, monkeypatch):
    from ui.input_handler import write_fu_completion_to_excel
    from utils.excel_io import active_worksheet_archive_path

    path = _make_fu_workbook(tmp_path)
    target_sheet = active_worksheet_archive_path(str(path))
    with zipfile.ZipFile(path, "r") as archive:
        before = {name: archive.read(name) for name in archive.namelist()}

    monkeypatch.setattr("ui.input_handler.messagebox.showerror", lambda *_a, **_k: None)
    assert write_fu_completion_to_excel(
        str(path),
        2,
        "2026-09-04",
        interface_id="FU-001",
    ) is True

    with zipfile.ZipFile(path, "r") as archive:
        after = {name: archive.read(name) for name in archive.namelist()}
    assert before.keys() == after.keys()
    assert [name for name in before if before[name] != after[name]] == [target_sheet]

    workbook = load_workbook(path, data_only=False)
    try:
        actual = workbook.active["D2"].value
        assert actual.date() == datetime.date(2026, 9, 4)
        assert workbook.active["E2"].value == datetime.datetime(2026, 9, 3)
        assert workbook.active["F2"].value == "张三"
    finally:
        workbook.close()


def test_prewarm_result_is_consumed_without_second_cache_read():
    from base import ExcelProcessorApp

    class Cache:
        def __init__(self):
            self.load_count = 0

        def load_cached_result(self, *_args):
            self.load_count += 1
            return None

        def save_cached_result(self, *_args):
            return True

    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app.file_manager = Cache()
    expected = pd.DataFrame({"责任人": ["张三"], "接口号": ["A"]})
    jobs = [{
        "file_path": "one.xlsx",
        "project_id": "2026",
        "file_type": "file1",
        "process_func": lambda *_args: expected,
        "args": (datetime.datetime(2026, 9, 2),),
    }]

    prewarmed = app._prewarm_result_caches_parallel(jobs, max_workers=4)
    app._run_preloaded_results = prewarmed["results"]
    actual = app._process_with_cache(
        "one.xlsx",
        "2026",
        "file1",
        lambda *_args: (_ for _ in ()).throw(AssertionError("must not recompute")),
        datetime.datetime(2026, 9, 2),
    )

    assert actual is expected
    assert app.file_manager.load_count == 1
    assert app._last_cache_hit_info["hit"] is False


def test_registry_pending_snapshot_loads_once_per_type(monkeypatch):
    from core import main

    calls = []
    expected = [("I-1", "2026", "待审查", "completed", "2026-09-02")]
    monkeypatch.setattr(
        main,
        "_load_latest_registry_pending_tasks_uncached",
        lambda *args: calls.append(args) or expected,
    )

    main.begin_registry_read_snapshot()
    try:
        assert main._load_latest_registry_pending_tasks(2, "registry.db", False) is expected
        assert main._load_latest_registry_pending_tasks(2, "registry.db", False) is expected
    finally:
        main.end_registry_read_snapshot()
    assert len(calls) == 1

    main._load_latest_registry_pending_tasks(2, "registry.db", False)
    assert len(calls) == 2


def test_file_types_5_to_7_processing_is_worker_safe_and_keeps_raw_registry_rows(monkeypatch):
    import base
    from base import ExcelProcessorApp

    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app.current_datetime = datetime.datetime(2026, 9, 2, 10, 0, 0)
    app.target_files5 = [("five.xlsx", "1818")]
    app.target_files6 = [("six.xlsx", "2026")]
    app.target_files7 = [("seven.xlsx", "1916")]

    process_calls = []

    def process_with_cache(file_path, project_id, file_type, _process_func, *args):
        process_calls.append((file_path, project_id, file_type, args))
        return pd.DataFrame({"接口号": [f"I-{project_id}"], "责任人": ["原始责任人"]})

    app._process_with_cache = process_with_cache
    app.apply_role_based_filter = lambda frame, project_id: frame.assign(责任人="筛选后责任人")
    registry_rows = []
    monkeypatch.setattr(
        base,
        "registry_hooks",
        SimpleNamespace(
            on_process_done=lambda **kwargs: registry_rows.append(kwargs) or True
        ),
    )
    flags = {"count": 0}
    main_module = SimpleNamespace(
        process_target_file5=object(),
        process_target_file6=object(),
        process_target_file7=object(),
    )

    results = app._process_file_types_5_to_7_background(
        main_module=main_module,
        enabled_types={5, 6, 7},
        can_reuse_refresh_cache=False,
        all_file_paths=[],
        changed_files=set(),
        registry_bootstrap_needed=True,
        registry_write_flags=flags,
        file6_context=({"张三"}, False),
    )

    assert [call[2] for call in process_calls] == ["file5", "file6", "file7"]
    assert process_calls[1][3][1:] == (False, {"张三"})
    assert flags["count"] == 3
    assert [item["file_type"] for item in registry_rows] == [5, 6, 7]
    assert all(item["result_df"].iloc[0]["责任人"] == "原始责任人" for item in registry_rows)
    assert all(results[file_type].iloc[0]["责任人"] == "筛选后责任人" for file_type in (5, 6, 7))
    assert all(getattr(app, f"has_processed_results{file_type}") for file_type in (5, 6, 7))


def test_task_panel_coalesces_overlapping_refreshes(monkeypatch):
    from write_tasks import task_panel

    started = []

    class FakeThread:
        def __init__(self, **kwargs):
            started.append(kwargs)

        def start(self):
            return None

    class Var:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    panel = task_panel.TaskRecordPanel.__new__(task_panel.TaskRecordPanel)
    panel._destroyed = False
    panel._refresh_job = None
    panel._refresh_lock = threading.Lock()
    panel._refresh_in_progress = False
    panel._refresh_requested = False
    panel._last_task_snapshot = []
    panel.only_mine_var = Var(False)
    panel.status_var = Var("")
    panel.get_current_user = lambda: "张三"
    monkeypatch.setattr(task_panel.threading, "Thread", FakeThread)

    panel.refresh_tasks()
    panel.refresh_tasks()

    assert len(started) == 1
    assert panel._refresh_in_progress is True
    assert panel._refresh_requested is True


def test_task_panel_initializes_shared_schema_only_once(monkeypatch):
    from write_tasks import task_panel

    ensure_values = []

    class Connection:
        def close(self):
            return None

    panel = task_panel.TaskRecordPanel.__new__(task_panel.TaskRecordPanel)
    panel._shared_schema_paths = set()
    monkeypatch.setattr(
        task_panel,
        "registry_hooks",
        SimpleNamespace(_cfg=lambda: {
            "registry_enabled": True,
            "registry_db_path": "shared-registry.db",
            "registry_wal": False,
        }),
    )
    monkeypatch.setattr(
        task_panel,
        "shared_list_tasks",
        lambda _conn, **kwargs: ensure_values.append(kwargs["ensure_schema_first"]) or [],
    )
    monkeypatch.setattr("registry.db.open_isolated_connection", lambda *_args: Connection())

    assert panel._collect_shared_tasks(only_mine=False, current_user="") == ([], "")
    assert panel._collect_shared_tasks(only_mine=False, current_user="") == ([], "")
    assert ensure_values == [True, False]


def test_selective_pending_refresh_only_redraws_requested_tab():
    from base import ExcelProcessorApp

    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app._tab_render_signatures = {0: "one", 2: "three", 6: "seven"}
    app.has_processed_results3 = True
    app.processing_results3 = pd.DataFrame({"接口号": ["X"]})
    calls = []
    app.display_results3 = lambda *_args, **_kwargs: calls.append(3)

    app._refresh_views_with_pending_cache({3})

    assert calls == [3]
    assert app._tab_render_signatures == {0: "one", 6: "seven"}


def test_registry_batch_unique_business_ids_do_not_use_row_queries(tmp_path, monkeypatch):
    from registry import service
    from registry.db import close_connection_after_use

    db_path = str(tmp_path / "registry.db")
    key = {
        "file_type": 2,
        "project_id": "2026",
        "interface_id": "BATCH-PREFETCH",
        "source_file": "source.xlsx",
        "row_index": 2,
    }
    service.upsert_task(
        db_path,
        False,
        key,
        {"display_status": "待完成", "responsible_person": "张三"},
        datetime.datetime(2026, 9, 2, 8, 0, 0),
    )
    close_connection_after_use()

    monkeypatch.setattr(
        service,
        "find_task_by_business_id",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("row query used")),
    )
    monkeypatch.setattr(
        service,
        "find_latest_confirmed_archive_by_business_id",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("archive row query used")),
    )
    count = service.batch_upsert_tasks(
        db_path,
        False,
        [{
            "key": key,
            "fields": {
                "display_status": "待完成",
                "responsible_person": "张三",
                "interface_time": "2026.09.03",
                "_completed_col_value": "",
            },
        }],
        datetime.datetime(2026, 9, 2, 9, 0, 0),
    )
    close_connection_after_use()
    assert count == 1

    connection = sqlite3.connect(db_path)
    try:
        assert connection.execute(
            "SELECT responsible_person FROM tasks WHERE business_id = ?",
            ("2|2026|BATCH-PREFETCH",),
        ).fetchone() == ("张三",)
    finally:
        connection.close()
