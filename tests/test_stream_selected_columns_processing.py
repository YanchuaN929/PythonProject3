#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime

import pytest
from openpyxl import Workbook, load_workbook

from core import main
from utils.dept_config import get_organization_filter, get_organization_filter_file6


def _write_row(ws, row, values):
    for col, value in values.items():
        ws[f"{col}{row}"] = value


def _make_workbook(tmp_path, filename, rows):
    path = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "header"
    ws["A2"] = "skip-row"
    for row_no, values in rows:
        _write_row(ws, row_no, values)
    wb.save(path)
    wb.close()
    return str(path)


@pytest.fixture
def no_registry_merge(monkeypatch):
    monkeypatch.setattr(
        main,
        "_merge_registry_pending_rows",
        lambda **kwargs: (kwargs["final_rows"], set()),
    )


def test_stream_processors_do_not_call_full_table_read(monkeypatch, tmp_path, no_registry_merge):
    def fail_read_excel(*args, **kwargs):
        raise AssertionError("pd.read_excel should not be used by stream processors")

    monkeypatch.setattr(main.pd, "read_excel", fail_read_excel)
    now = datetime.datetime(2026, 5, 14)
    org = get_organization_filter()
    org6 = get_organization_filter_file6()

    files = {
        1: _make_workbook(tmp_path, "2026按项目导出IDI手册2026-05-14.xlsx", [
            (3, {"A": "S-A-25C1", "B": "", "H": "结构25C1", "K": now, "M": "", "R": "张三"})
        ]),
        2: _make_workbook(tmp_path, "内部接口信息单报表202620260514.xlsx", [
            (3, {"A": "x", "E": "A", "F": "", "I": org, "M": now, "N": "", "R": "S-B-25C2", "AB": "", "AM": "李四"})
        ]),
        3: _make_workbook(tmp_path, "外部接口ICM报表202620260514.xlsx", [
            (3, {"C": "S-C-M", "I": "B", "M": now, "T": "", "AC": "A", "AL": org, "AO": "结构一室", "AP": "王五"}),
            (4, {"C": "S-C-L", "I": "B", "L": now, "Q": "", "AC": "A", "AL": org, "AO": "结构二室", "AP": "赵六"}),
        ]),
        4: _make_workbook(tmp_path, "外部接口单报表202620260514.xlsx", [
            (3, {"E": "S-D-25C1", "I": "A", "P": "B", "S": now, "V": "", "AF": org, "AG": "", "AH": "周七"})
        ]),
        5: _make_workbook(tmp_path, "2026接口提资清单.xlsx", [
            (3, {"A": "S-E-25C2", "G": "25C2", "K": "吴八", "L": now, "N": ""})
        ]),
        6: _make_workbook(tmp_path, "收发文清单2026.xlsx", [
            (3, {"E": "S-F-25C3", "I": now, "J": "", "M": "尚未回复", "V": org6, "W": "结构一室", "X": "郑九", "AC": "A"})
        ]),
    }

    results = {
        1: main.process_target_file(files[1], now),
        2: main.process_target_file2(files[2], now, "2026"),
        3: main.process_target_file3(files[3], now),
        4: main.process_target_file4(files[4], now),
        5: main.process_target_file5(files[5], now),
        6: main.process_target_file6(files[6], now),
    }

    for file_type, df in results.items():
        assert not df.empty, f"file{file_type} should produce one or more rows"
        assert set(["项目号", "接口号", "接口时间", "科室", "主办室", "责任人", "原始行号"]).issubset(df.columns)
        assert set(["source_file", "_file_type", "_interface_col", "_time_col", "_completed_col", "_responsible_col", "_assign_col", "_completed_col_value", "_stream_schema_version"]).issubset(df.columns)
        assert df["_stream_schema_version"].eq(main.STREAM_RESULT_SCHEMA_VERSION).all()

    file3 = results[3].sort_values("接口号").reset_index(drop=True)
    assert file3.loc[0, "_source_column"] == "L"
    assert file3.loc[0, "_time_col"] == "Q"
    assert file3.loc[0, "_completed_col"] == "Q"
    assert file3.loc[1, "_source_column"] == "M"
    assert file3.loc[1, "_time_col"] == "T"
    assert file3.loc[1, "_completed_col"] == "T"

    file6 = results[6].iloc[0]
    assert results[1].iloc[0]["_time_col"] == "M"
    assert results[2].iloc[0]["_time_col"] == "N"
    assert results[4].iloc[0]["_time_col"] == "V"
    assert results[5].iloc[0]["_time_col"] == "N"
    assert file6["_time_col"] == "J"
    assert file6["_completed_col"] == "J"
    assert file6["主办室"] == "结构一室"


def test_stream_processors_include_excel_row_two(tmp_path, no_registry_merge):
    now = datetime.datetime(2026, 5, 14)
    org = get_organization_filter()
    org6 = get_organization_filter_file6()
    files = {
        1: _make_workbook(tmp_path, "2026按项目导出IDI手册2026-05-14.xlsx", [
            (2, {"A": "S-A-ROW2", "B": "", "H": "结构25C1", "K": now, "M": "", "R": "张三"})
        ]),
        2: _make_workbook(tmp_path, "内部接口信息单报表202620260514.xlsx", [
            (2, {"A": "x", "E": "A", "F": "", "I": org, "M": now, "N": "", "R": "S-B-ROW2", "AB": "", "AM": "李四"})
        ]),
        3: _make_workbook(tmp_path, "外部接口ICM报表202620260514.xlsx", [
            (2, {"C": "S-C-ROW2", "I": "B", "M": now, "T": "", "AC": "A", "AL": org, "AO": "结构一室", "AP": "王五"})
        ]),
        4: _make_workbook(tmp_path, "外部接口单报表202620260514.xlsx", [
            (2, {"E": "S-D-ROW2", "I": "A", "P": "B", "S": now, "V": "", "AF": org, "AG": "", "AH": "周七"})
        ]),
        5: _make_workbook(tmp_path, "2026接口提资清单.xlsx", [
            (2, {"A": "S-E-ROW2", "G": "25C2", "K": "吴八", "L": now, "N": ""})
        ]),
        6: _make_workbook(tmp_path, "收发文清单2026.xlsx", [
            (2, {"E": "S-F-ROW2", "I": now, "J": "", "M": "尚未回复", "V": org6, "W": "结构一室", "X": "郑九", "AC": "A"})
        ]),
    }
    results = {
        1: main.process_target_file(files[1], now),
        2: main.process_target_file2(files[2], now, "2026"),
        3: main.process_target_file3(files[3], now),
        4: main.process_target_file4(files[4], now),
        5: main.process_target_file5(files[5], now),
        6: main.process_target_file6(files[6], now),
    }

    for file_type, result in results.items():
        assert result["原始行号"].tolist() == [2], f"file{file_type} skipped Excel row 2"


@pytest.mark.parametrize("project_id", ["2016", "2026"])
def test_file2_2026_uses_same_standard_rule_as_2016(
    tmp_path,
    no_registry_merge,
    project_id,
):
    now = datetime.datetime(2026, 9, 4)
    org = get_organization_filter()
    source = _make_workbook(
        tmp_path,
        f"内部接口信息单报表{project_id}20260904.xlsx",
        [
            (
                3,
                {
                    "A": "x",
                    "E": "A",
                    "F": "传递",
                    "I": org,
                    "M": now,
                    "N": "",
                    "R": f"S-{project_id}-STANDARD",
                    "AB": "4444-计划关闭",
                    "AM": "测试责任人",
                },
            )
        ],
    )

    result = main.process_target_file2(source, now, project_id)

    assert result["接口号"].tolist() == [f"S-{project_id}-STANDARD"]


def test_stream_export_only_contains_business_columns(tmp_path):
    df = main.pd.DataFrame([{
        "状态": "待完成",
        "项目号": "2026",
        "接口号": "S-A",
        "接口时间": "2026.05.14",
        "科室": "结构一室",
        "主办室": "",
        "责任人": "张三",
        "原始行号": 3,
        "_completed_col": "M",
        "_stream_schema_version": main.STREAM_RESULT_SCHEMA_VERSION,
    }])
    out = main.export_result_to_excel(df, "unused.xlsx", datetime.datetime(2026, 5, 14), str(tmp_path), "2026")
    wb = load_workbook(out, read_only=True)
    try:
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == main.STREAM_EXPORT_COLUMNS
        assert "_completed_col" not in headers
        assert "_stream_schema_version" not in headers
    finally:
        wb.close()


def test_file5_finder_accepts_three_dimensional_naming(tmp_path):
    legacy = tmp_path / "2026接口提资清单.xlsx"
    compact = tmp_path / "三维接口提资清单191520260211.xlsx"
    dated = tmp_path / "三维接口提资清单-2026-3-10-1915项目.xls"
    ignored = tmp_path / "2026其他清单.xlsx"

    files = [str(legacy), str(compact), str(dated), str(ignored)]
    matched = {path: project_id for path, project_id in main.find_all_target_files5(files)}

    assert matched[str(legacy)] == "2026"
    assert matched[str(compact)] == "1915"
    assert matched[str(dated)] == "1915"
    assert str(ignored) not in matched
    assert main._extract_file5_project_id(str(dated)) == "1915"


def test_file5_stream_uses_three_dimensional_project_id(tmp_path, no_registry_merge):
    now = datetime.datetime(2026, 5, 14)
    path = _make_workbook(tmp_path, "三维接口提资清单191520260211.xlsx", [
        (3, {"A": "S-E-25C2", "G": "25C2", "K": "吴八", "L": now, "N": ""})
    ])

    df = main.process_target_file5(path, now)

    assert not df.empty
    assert df.iloc[0]["项目号"] == "1915"


def test_registry_merge_can_index_stream_records(monkeypatch, tmp_path):
    monkeypatch.setattr(
        main,
        "_load_latest_registry_pending_tasks",
        lambda file_type, db_path, wal: [("S-A-1", "2026", "待审查", "completed", "2026-05-14T00:00:00")],
    )
    path = tmp_path / "2026按项目导出IDI手册2026-05-14.xlsx"
    records = [
        {"_df_index": 0, "原始行号": 2, "_raw": {"a": "header"}},
        {"_df_index": 1, "原始行号": 3, "_raw": {"a": "S-A-1"}},
    ]

    final_rows, pending_rows = main._merge_registry_pending_rows(
        file_type=1,
        file_path=str(path),
        df=records,
        final_rows=set(),
        allowed_rows={1},
        project_id="2026",
    )

    assert final_rows == {1}
    assert pending_rows == {1}


def test_all_file_finders_accept_2416_project():
    files = [
        r"C:\tmp\2416按项目导出IDI手册2026-05-27.xlsx",
        r"C:\tmp\内部接口信息单报表241620260527.xlsx",
        r"C:\tmp\外部接口ICM报表241620260527.xlsx",
        r"C:\tmp\外部接口单报表241620260527.xlsx",
        r"C:\tmp\2416接口提资清单.xlsx",
        r"C:\tmp\三维接口提资清单241620260527.xlsx",
        r"C:\tmp\收发文清单2416.xlsx",
        r"C:\tmp\2416项目标准表格.xlsx",
    ]

    assert main.find_all_target_files1(files) == [(files[0], "2416")]
    assert main.find_all_target_files2(files) == [(files[1], "2416")]
    assert main.find_all_target_files3(files) == [(files[2], "2416")]
    assert main.find_all_target_files4(files) == [(files[3], "2416")]
    assert main.find_all_target_files5(files) == [(files[4], "2416"), (files[5], "2416")]
    assert main.find_all_target_files6(files) == [(files[6], "2416")]
    assert main.find_all_target_files7(files) == [(files[7], "2416")]


def test_all_file_finders_accept_xlsm_and_case_insensitive_extensions():
    files = [
        r"C:\tmp\2416按项目导出IDI手册2026-05-27.XLSM",
        r"C:\tmp\内部接口信息单报表241620260527.XLSM",
        r"C:\tmp\外部接口ICM报表241620260527.XLSM",
        r"C:\tmp\外部接口单报表241620260527.XLSM",
        r"C:\tmp\2416接口提资清单.XLSM",
        r"C:\tmp\收发文清单2416.XLSM",
        r"C:\tmp\2416项目标准表格.XLSM",
    ]

    assert main.find_all_target_files1(files) == [(files[0], "2416")]
    assert main.find_all_target_files2(files) == [(files[1], "2416")]
    assert main.find_all_target_files3(files) == [(files[2], "2416")]
    assert main.find_all_target_files4(files) == [(files[3], "2416")]
    assert main.find_all_target_files5(files) == [(files[4], "2416")]
    assert main.find_all_target_files6(files) == [(files[5], "2416")]
    assert main.find_all_target_files7(files) == [(files[6], "2416")]
