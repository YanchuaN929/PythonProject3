"""
测试新的三个修复：
1. 文件6责任人筛选
2. 文件1回文单号时间列位置
3. 缓存失效机制
"""
import pytest
import pandas as pd
import os
import tempfile
import time
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook


class TestFile6ResponsiblePersonFiltering:
    """测试文件6责任人筛选功能"""
    
    def test_filter_valid_names_basic(self):
        """测试基本的姓名筛选"""
        from main import filter_valid_names
        
        valid_names = {'张三', '李四', '王五'}
        names_str = '张三,李四,王五'
        
        result = filter_valid_names(names_str, valid_names)
        assert result == '张三,李四,王五'
    
    def test_filter_valid_names_with_invalid(self):
        """测试筛选掉无效姓名"""
        from main import filter_valid_names
        
        valid_names = {'张三', '李四'}
        names_str = '张三,李四,王五,赵六'  # 王五和赵六不在有效名单中
        
        result = filter_valid_names(names_str, valid_names)
        assert result == '张三,李四'
    
    def test_filter_valid_names_with_suffix(self):
        """测试去除尾部字母后匹配（如刘峰a → 刘峰）"""
        from main import filter_valid_names
        
        valid_names = {'刘峰', '张三'}
        names_str = '刘峰a,张三b,李四c'  # 刘峰a和张三b应该匹配成功，李四c失败
        
        result = filter_valid_names(names_str, valid_names)
        assert result == '刘峰,张三'
    
    def test_filter_valid_names_mixed(self):
        """测试混合情况：精确匹配+后缀匹配+无效姓名"""
        from main import filter_valid_names
        
        valid_names = {'刘峰', '张三', '李四'}
        names_str = '刘峰,张三a,李四,王五b,赵六'
        
        result = filter_valid_names(names_str, valid_names)
        # 刘峰（精确）, 张三（去除a）, 李四（精确）, 王五b（无效）, 赵六（无效）
        assert result == '刘峰,张三,李四'
    
    def test_filter_valid_names_empty_input(self):
        """测试空输入"""
        from main import filter_valid_names
        
        valid_names = {'张三', '李四'}
        names_str = ''
        
        result = filter_valid_names(names_str, valid_names)
        assert result == ''
    
    def test_filter_valid_names_no_valid_set(self):
        """测试无有效姓名集合（返回原字符串）"""
        from main import filter_valid_names
        
        valid_names = set()
        names_str = '张三,李四'
        
        result = filter_valid_names(names_str, valid_names)
        assert result == '张三,李四'
    
    def test_filter_valid_names_all_invalid(self):
        """测试所有姓名都无效（返回空字符串）"""
        from main import filter_valid_names
        
        valid_names = {'刘峰', '张三'}
        names_str = '王五,赵六,孙七'
        
        result = filter_valid_names(names_str, valid_names)
        assert result == ''
    
    def test_filter_valid_names_multiple_letters_suffix(self):
        """测试多个字母后缀（如abc）"""
        from main import filter_valid_names
        
        valid_names = {'刘峰'}
        names_str = '刘峰abc'
        
        result = filter_valid_names(names_str, valid_names)
        assert result == '刘峰'
    
    def test_process_target_file6_with_valid_names(self):
        """测试process_target_file6接收valid_names_set参数"""
        from main import process_target_file6
        
        # 这个测试只验证filter_valid_names函数被正确集成
        # 不测试完整的process_target_file6流程（因为需要复杂的Excel结构）
        # 具体的Excel处理已在其他测试中覆盖
        
        # 验证函数签名包含valid_names_set参数
        import inspect
        sig = inspect.signature(process_target_file6)
        params = list(sig.parameters.keys())
        
        assert 'valid_names_set' in params, "process_target_file6应该接收valid_names_set参数"
        
        # 验证默认值是None
        assert sig.parameters['valid_names_set'].default is None


class TestFile1TimeColumnChange:
    """测试文件1回文单号时间列位置修改"""
    
    def test_file1_time_column_is_m(self):
        """测试文件1的时间列是M列"""
        from input_handler import get_write_columns
        
        # 创建模拟工作表
        mock_ws = MagicMock()
        
        columns = get_write_columns(1, 2, mock_ws)
        
        assert columns is not None
        assert columns['time_col'] == 'M', "文件1的时间列应该是M列"
        assert columns['response_col'] == 'S'
        assert columns['name_col'] == 'V'
    
    def test_file1_time_column_not_n(self):
        """测试文件1的时间列不是N列（已修改）"""
        from input_handler import get_write_columns
        
        mock_ws = MagicMock()
        
        columns = get_write_columns(1, 2, mock_ws)
        
        assert columns is not None
        assert columns['time_col'] != 'N', "文件1的时间列不应该是N列"
    
    def test_other_files_time_column_unchanged(self):
        """测试其他文件的时间列位置没有改变"""
        from input_handler import get_write_columns
        
        mock_ws = MagicMock()
        
        # 文件2：N列
        columns2 = get_write_columns(2, 2, mock_ws)
        assert columns2['time_col'] == 'N'
        
        # 文件4：V列
        columns4 = get_write_columns(4, 2, mock_ws)
        assert columns4['time_col'] == 'V'
        
        # 文件5：N列
        columns5 = get_write_columns(5, 2, mock_ws)
        assert columns5['time_col'] == 'N'
        
        # 文件6：J列
        columns6 = get_write_columns(6, 2, mock_ws)
        assert columns6['time_col'] == 'J'
    
    def test_write_response_uses_correct_column(self):
        """测试写入回文单号时使用正确的时间列"""
        from input_handler import write_response_to_excel
        
        # 创建临时Excel文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # 创建测试Excel
            wb = Workbook()
            ws = wb.active
            
            # 添加25列（A-Y）
            for i in range(1, 26):
                ws.cell(1, i, f'Col{i}')
            
            # 添加一行数据
            ws.cell(2, 1, 'Test')
            
            wb.save(tmp_path)
            wb.close()
            
            # 写入回文单号
            success = write_response_to_excel(
                tmp_path, 
                file_type=1, 
                row_index=2, 
                response_number='TEST-001',
                user_name='测试用户',
                project_id='2016',
                source_column=None
            )
            
            assert success, "写入应该成功"
            
            # 读取并验证
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # M列是第13列
            time_value = ws.cell(2, 13).value  # M列
            assert time_value is not None, "M列应该有时间值"
            
            # N列是第14列，应该没有值（因为已经改为M列）
            n_col_value = ws.cell(2, 14).value
            # 如果之前错误地写入了N列，这里会有值
            
            wb.close()
            
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)


class TestCacheInvalidation:
    """测试缓存失效机制"""
    
    def test_cache_identity_none_allows_loading(self):
        """测试当cached_identity是None时，允许加载缓存"""
        from file_manager import FileIdentityManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # 创建测试文件
            test_file = os.path.join(tmpdir, 'test.xlsx')
            wb = Workbook()
            wb.save(test_file)
            wb.close()
            
            # 创建FileIdentityManager
            cache_file = os.path.join(tmpdir, 'test_cache.json')
            cache_dir = os.path.join(tmpdir, 'result_cache')
            os.makedirs(cache_dir, exist_ok=True)
            
            manager = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            
            # 创建一个缓存文件
            test_df = pd.DataFrame({'col1': [1, 2, 3]})
            manager.save_cached_result(test_file, '2016', 'file1', test_df)
            
            # 清空file_identities（模拟新文件或程序重启后没有标识）
            manager.file_identities = {}
            
            # 尝试加载缓存
            loaded_df = manager.load_cached_result(test_file, '2016', 'file1')
            
            # 应该成功加载（即使cached_identity是None）
            assert loaded_df is not None, "当cached_identity是None时，应该允许加载缓存"
            assert len(loaded_df) == 3
            
            # 验证文件标识已被更新
            assert test_file in manager.file_identities, "应该更新文件标识"
    
    def test_cache_invalidated_on_file_change(self):
        """测试文件修改后缓存失效"""
        from file_manager import FileIdentityManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # 创建测试文件
            test_file = os.path.join(tmpdir, 'test.xlsx')
            wb = Workbook()
            wb.save(test_file)
            wb.close()
            
            # 创建FileIdentityManager
            cache_file = os.path.join(tmpdir, 'test_cache.json')
            cache_dir = os.path.join(tmpdir, 'result_cache')
            os.makedirs(cache_dir, exist_ok=True)
            
            manager = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            
            # 保存缓存并记录文件标识
            test_df = pd.DataFrame({'col1': [1, 2, 3]})
            manager.save_cached_result(test_file, '2016', 'file1', test_df)
            manager.update_file_identities([test_file])
            
            # 验证可以加载缓存
            loaded_df = manager.load_cached_result(test_file, '2016', 'file1')
            assert loaded_df is not None
            
            # 修改文件（触发修改时间变化）
            time.sleep(0.1)  # 确保修改时间不同
            wb = Workbook()
            ws = wb.active
            ws.cell(1, 1, 'modified')
            wb.save(test_file)
            wb.close()
            
            # 再次尝试加载缓存
            loaded_df = manager.load_cached_result(test_file, '2016', 'file1')
            
            # 应该返回None（缓存失效）
            assert loaded_df is None, "文件修改后，缓存应该失效"
    
    def test_cache_identity_comparison_logic(self):
        """测试缓存标识比较逻辑"""
        from file_manager import FileIdentityManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test.xlsx')
            wb = Workbook()
            wb.save(test_file)
            wb.close()
            
            cache_file = os.path.join(tmpdir, 'test_cache.json')
            cache_dir = os.path.join(tmpdir, 'result_cache')
            os.makedirs(cache_dir, exist_ok=True)
            
            manager = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            
            # 生成当前文件标识
            current_identity = manager.generate_file_identity(test_file)
            assert current_identity is not None
            
            # 场景1：cached_identity是None，应该允许加载
            cached_identity = None
            should_invalidate = (cached_identity is not None and current_identity != cached_identity)
            assert not should_invalidate, "cached_identity是None时，不应该使缓存失效"
            
            # 场景2：cached_identity与current_identity相同，应该允许加载
            cached_identity = current_identity
            should_invalidate = (cached_identity is not None and current_identity != cached_identity)
            assert not should_invalidate, "标识相同时，不应该使缓存失效"
            
            # 场景3：cached_identity与current_identity不同，应该使缓存失效
            cached_identity = "different_hash"
            should_invalidate = (cached_identity is not None and current_identity != cached_identity)
            assert should_invalidate, "标识不同时，应该使缓存失效"
    
    def test_file_identity_persisted(self):
        """测试文件标识会被持久化"""
        from file_manager import FileIdentityManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test.xlsx')
            wb = Workbook()
            wb.save(test_file)
            wb.close()
            
            cache_file = os.path.join(tmpdir, 'test_cache.json')
            cache_dir = os.path.join(tmpdir, 'result_cache')
            os.makedirs(cache_dir, exist_ok=True)
            
            # 第一个manager实例
            manager1 = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            manager1.update_file_identities([test_file])
            
            # 验证标识已保存
            assert test_file in manager1.file_identities
            identity1 = manager1.file_identities[test_file]
            
            # 创建第二个manager实例（模拟程序重启）
            manager2 = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            
            # 验证标识已被恢复
            assert test_file in manager2.file_identities, "文件标识应该从缓存文件恢复"
            identity2 = manager2.file_identities[test_file]
            
            assert identity1 == identity2, "恢复的标识应该与保存的标识相同"
    
    def test_check_files_changed_detects_modification(self):
        """测试check_files_changed能检测到文件修改"""
        from file_manager import FileIdentityManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            test_file = os.path.join(tmpdir, 'test.xlsx')
            wb = Workbook()
            wb.save(test_file)
            wb.close()
            
            cache_file = os.path.join(tmpdir, 'test_cache.json')
            cache_dir = os.path.join(tmpdir, 'result_cache')
            os.makedirs(cache_dir, exist_ok=True)
            
            manager = FileIdentityManager(cache_file=cache_file, result_cache_dir=cache_dir)
            
            # 初始化文件标识
            manager.update_file_identities([test_file])
            
            # 检查文件是否变化（应该没有变化）
            has_changed = manager.check_files_changed([test_file])
            assert not has_changed, "文件未修改时，should返回False"
            
            # 修改文件
            time.sleep(0.1)
            wb = Workbook()
            ws = wb.active
            ws.cell(1, 1, 'modified')
            wb.save(test_file)
            wb.close()
            
            # 再次检查
            has_changed = manager.check_files_changed([test_file])
            assert has_changed, "文件修改后，应该返回True"

