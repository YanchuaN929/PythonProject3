from concurrent.futures import ThreadPoolExecutor

import pytest
from openpyxl import Workbook, load_workbook

from services.account_service import (
    AccountBusyError,
    AuthenticationError,
    _SharedFileLock,
    change_password,
    list_accounts,
    verify_password,
)


def _make_role_table(path, rows, *, with_header=False):
    workbook = Workbook()
    worksheet = workbook.active
    if with_header:
        worksheet.append(["姓名", "角色", "密码"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _password_cells(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return {
            str(row[0]).strip(): row[2]
            for row in worksheet.iter_rows(min_col=1, max_col=3, values_only=True)
            if row[0] and str(row[0]).strip() != "姓名"
        }
    finally:
        workbook.close()


def test_list_and_verify_accounts_support_headerless_default_password(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(
        table,
        [
            ["张三", "设计人员", None],
            ["李四", "所领导", "custom"],
            ["张三", "2026接口工程师", None],
        ],
    )

    assert list_accounts(table) == ["张三", "李四"]
    assert verify_password(table, "张三", "password") is True
    assert verify_password(table, "张三", "wrong") is False
    assert verify_password(table, "李四", "custom") is True


def test_change_password_updates_third_column_only(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(
        table,
        [["张三", "设计人员", "password"], ["李四", "所领导", "password"]],
        with_header=True,
    )

    change_password(table, "张三", "password", "new-secret")

    assert verify_password(table, "张三", "new-secret") is True
    assert verify_password(table, "李四", "password") is True
    assert _password_cells(table) == {"张三": "new-secret", "李四": "password"}


def test_wrong_current_password_does_not_modify_workbook(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(table, [["张三", "设计人员", "password"]])
    original = table.read_bytes()

    with pytest.raises(AuthenticationError, match="当前密码不正确"):
        change_password(table, "张三", "wrong", "new-secret")

    assert table.read_bytes() == original


def test_concurrent_changes_for_different_users_do_not_overwrite(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(
        table,
        [["张三", "设计人员", "password"], ["李四", "所领导", "password"]],
    )

    with ThreadPoolExecutor(max_workers=2) as pool:
        first = pool.submit(change_password, table, "张三", "password", "zhang-new")
        second = pool.submit(change_password, table, "李四", "password", "li-new")
        first.result(timeout=10)
        second.result(timeout=10)

    assert _password_cells(table) == {"张三": "zhang-new", "李四": "li-new"}


def test_concurrent_changes_for_same_user_cannot_reuse_old_password(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(table, [["张三", "设计人员", "password"]])

    def attempt(new_password):
        try:
            change_password(table, "张三", "password", new_password)
            return "ok"
        except AuthenticationError:
            return "rejected"

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(attempt, ["first-new", "second-new"]))

    assert sorted(results) == ["ok", "rejected"]
    assert _password_cells(table)["张三"] in {"first-new", "second-new"}


def test_lock_contention_reports_busy_without_touching_workbook(tmp_path):
    table = tmp_path / "姓名角色表.xlsx"
    _make_role_table(table, [["张三", "设计人员", "password"]])

    with _SharedFileLock(str(table), timeout=0.1, stale_seconds=300):
        with pytest.raises(AccountBusyError, match="其他用户修改"):
            change_password(
                table,
                "张三",
                "password",
                "new-secret",
                lock_timeout=0.15,
            )

    assert verify_password(table, "张三", "password") is True
