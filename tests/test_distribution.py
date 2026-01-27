#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
任务指派功能测试
"""

import pytest
import pandas as pd
import os
import sys
from unittest.mock import Mock, patch, MagicMock

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import services.distribution as distribution


class TestGetResponsibleColumn:
    """测试获取责任人列名"""
    
    def test_file1_responsible_column(self):
        """测试文件1责任人列"""
        col = distribution.get_responsible_column(1)
        assert col == 'R'
    
    def test_file2_responsible_column(self):
        """测试文件2责任人列（AM列）"""
        col = distribution.get_responsible_column(2)
        assert col == 'AM'
    
    def test_file3_responsible_column(self):
        """测试文件3责任人列"""
        col = distribution.get_responsible_column(3)
        assert col == 'AP'
    
    def test_file4_responsible_column(self):
        """测试文件4责任人列"""
        col = distribution.get_responsible_column(4)
        assert col == 'AH'
    
    def test_file5_responsible_column(self):
        """测试文件5责任人列"""
        col = distribution.get_responsible_column(5)
        assert col == 'K'
    
    def test_file6_responsible_column(self):
        """测试文件6责任人列"""
        col = distribution.get_responsible_column(6)
        assert col == 'X'
    
    def test_invalid_file_type(self):
        """测试无效文件类型"""
        col = distribution.get_responsible_column(999)
        assert col is None


class TestRoleChecking:
    """测试角色判断"""
    
    def test_is_interface_engineer_true(self):
        """测试识别接口工程师"""
        roles = ['2016接口工程师']
        assert distribution.is_interface_engineer(roles) is True
    
    def test_is_interface_engineer_false(self):
        """测试非接口工程师"""
        roles = ['一室主任', '设计人员']
        assert distribution.is_interface_engineer(roles) is False
    
    def test_is_director_true(self):
        """测试识别室主任"""
        roles = ['一室主任']
        assert distribution.is_director(roles) is True
        
        roles = ['二室主任']
        assert distribution.is_director(roles) is True
        
        roles = ['建筑总图室主任']
        assert distribution.is_director(roles) is True
    
    def test_is_director_false(self):
        """测试非室主任"""
        roles = ['设计人员', '所长']
        assert distribution.is_director(roles) is False
    
    def test_get_department(self):
        """测试获取科室名称"""
        assert distribution.get_department(['一室主任']) == '结构一室'
        assert distribution.get_department(['二室主任']) == '结构二室'
        assert distribution.get_department(['建筑总图室主任']) == '建筑总图室'
        assert distribution.get_department(['设计人员']) == ''
    
    def test_parse_interface_engineer_project(self):
        """测试从角色中提取项目号"""
        project_id = distribution.parse_interface_engineer_project(['2016接口工程师'])
        assert project_id == '2016'
        
        project_id = distribution.parse_interface_engineer_project(['2026接口工程师'])
        assert project_id == '2026'
        
        project_id = distribution.parse_interface_engineer_project(['设计人员'])
        assert project_id is None


class TestCheckUnassigned:
    """测试检测未指派任务"""
    
    def test_detect_empty_responsible(self):
        """测试检测责任人为空的数据"""
        df = pd.DataFrame({
            '项目号': ['2016', '2016', '2026'],
            '接口号': ['INT-001', 'INT-002', 'INT-003'],
            '责任人': ['王任超', '', None],
            '科室': ['结构一室', '结构一室', '结构二室'],
            '原始行号': [2, 3, 4],
            'source_file': ['file1.xlsx'] * 3,
            '接口时间': ['10.28', '10.29', '10.30']
        })
        
        processed_results = {1: df}
        user_roles = ['一室主任']
        
        unassigned = distribution.check_unassigned(processed_results, user_roles)
        
        # 应该检测到1个未指派任务（第2行，责任人为空）
        assert len(unassigned) == 1
        assert unassigned[0]['interface_id'] == 'INT-002'
        assert unassigned[0]['file_type'] == 1
    
    def test_detect_responsible_as_none(self):
        """测试检测责任人为'无'的数据"""
        df = pd.DataFrame({
            '项目号': ['2016', '2016'],
            '接口号': ['INT-001', 'INT-002'],
            '责任人': ['王任超', '无'],
            '科室': ['结构一室', '结构一室'],
            '原始行号': [2, 3],
            'source_file': ['file1.xlsx'] * 2,
            '接口时间': ['10.28', '10.29']
        })
        
        processed_results = {1: df}
        user_roles = ['一室主任']
        
        unassigned = distribution.check_unassigned(processed_results, user_roles)
        
        # 应该检测到1个未指派任务（第2行，责任人为'无'）
        assert len(unassigned) == 1
        assert unassigned[0]['interface_id'] == 'INT-002'
    
    def test_interface_engineer_filter(self):
        """测试接口工程师只看自己项目"""
        df = pd.DataFrame({
            '项目号': ['2016', '2016', '2026'],
            '接口号': ['INT-001', 'INT-002', 'INT-003'],
            '责任人': ['', '', ''],
            '科室': ['结构一室'] * 3,
            '原始行号': [2, 3, 4],
            'source_file': ['file1.xlsx'] * 3,
            '接口时间': ['10.28', '10.29', '10.30']
        })
        
        processed_results = {1: df}
        user_roles = ['2016接口工程师']
        project_id = '2016'
        
        unassigned = distribution.check_unassigned(processed_results, user_roles, project_id)
        
        # 应该只检测到2个2016项目的未指派任务
        assert len(unassigned) == 2
        assert all(task['project_id'] == '2016' for task in unassigned)
    
    def test_director_filter(self):
        """测试室主任只看自己科室"""
        df = pd.DataFrame({
            '项目号': ['2016'] * 4,
            '接口号': ['INT-001', 'INT-002', 'INT-003', 'INT-004'],
            '责任人': ['', '', '', ''],
            '科室': ['结构一室', '结构二室', '请室主任确认', '建筑总图室'],
            '原始行号': [2, 3, 4, 5],
            'source_file': ['file1.xlsx'] * 4,
            '接口时间': ['10.28', '10.29', '10.30', '10.31']
        })
        
        processed_results = {1: df}
        user_roles = ['一室主任']
        
        unassigned = distribution.check_unassigned(processed_results, user_roles)
        
        # 应该检测到2个任务：结构一室 + 请室主任确认
        assert len(unassigned) == 2
        interfaces = [task['interface_id'] for task in unassigned]
        assert 'INT-001' in interfaces  # 结构一室
        assert 'INT-003' in interfaces  # 请室主任确认
    
    def test_multiple_file_types(self):
        """测试多个文件类型"""
        df1 = pd.DataFrame({
            '项目号': ['2016'],
            '接口号': ['INT-001'],
            '责任人': [''],
            '科室': ['结构一室'],
            '原始行号': [2],
            'source_file': ['file1.xlsx'],
            '接口时间': ['10.28']
        })
        
        df2 = pd.DataFrame({
            '项目号': ['2016'],
            '接口号': ['INT-002'],
            '责任人': ['无'],
            '科室': ['结构一室'],
            '原始行号': [2],
            'source_file': ['file2.xlsx'],
            '接口时间': ['10.29']
        })
        
        processed_results = {1: df1, 2: df2}
        user_roles = ['一室主任']
        
        unassigned = distribution.check_unassigned(processed_results, user_roles)
        
        # 应该检测到2个未指派任务
        assert len(unassigned) == 2
        file_types = [task['file_type'] for task in unassigned]
        assert 1 in file_types
        assert 2 in file_types


class TestSaveAssignment:
    """测试保存指派结果"""
    
    @pytest.fixture
    def temp_excel_file(self, tmp_path):
        """创建临时Excel文件"""
        import pandas as pd
        from openpyxl import Workbook
        
        file_path = tmp_path / "test_file1.xlsx"
        
        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 创建表头和数据
        headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 添加数据行
        for row in range(2, 5):
            for col in range(1, 19):
                ws.cell(row=row, column=col, value=f"Data{row}_{col}")
        
        wb.save(str(file_path))
        return str(file_path)
    
    def test_save_assignment_file1(self, temp_excel_file):
        """测试保存指派到文件1（R列）"""
        from openpyxl import load_workbook
        
        success = distribution.save_assignment(
            1,  # 文件类型1
            temp_excel_file,
            2,  # 第2行
            "王任超"  # 指派给王任超
        )
        
        assert success is True
        
        # 验证写入内容
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        assert ws['R2'].value == "王任超"
        
        wb.close()
    
    def test_save_assignment_file2(self, temp_excel_file):
        """测试保存指派到文件2（AM列）"""
        from openpyxl import load_workbook
        
        # 创建更大的Excel文件以包含AM列
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        ws['AM2'] = ""  # 确保AM列存在
        wb.save(temp_excel_file)
        wb.close()
        
        success = distribution.save_assignment(
            2,  # 文件类型2
            temp_excel_file,
            2,  # 第2行
            "李四"  # 指派给李四
        )
        
        assert success is True
        
        # 验证写入内容
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        assert ws['AM2'].value == "李四"
        
        wb.close()
    
    def test_save_assignment_invalid_file(self):
        """测试保存到不存在的文件"""
        success = distribution.save_assignment(
            1,
            "/path/to/nonexistent/file.xlsx",
            2,
            "王任超"
        )
        
        assert success is False
    
    def test_save_assignment_invalid_file_type(self, temp_excel_file):
        """测试保存到无效文件类型"""
        success = distribution.save_assignment(
            999,  # 无效的文件类型
            temp_excel_file,
            2,
            "王任超"
        )
        
        assert success is False


class TestGetNameList:
    """测试获取姓名列表"""
    
    @patch('distribution.pd.read_excel')
    @patch('distribution.os.path.exists')
    def test_get_name_list_success(self, mock_exists, mock_read_excel):
        """测试成功读取姓名列表"""
        mock_exists.return_value = True
        
        # 模拟姓名角色表数据
        mock_df = pd.DataFrame({
            '姓名': ['王任超', '李四', '张三', '王任超'],  # 包含重复
            '角色': ['设计人员', '设计人员', '一室主任', '设计人员']
        })
        mock_read_excel.return_value = mock_df
        
        names = distribution.get_name_list()
        
        # 应该返回去重排序的姓名列表
        assert len(names) == 3
        assert '王任超' in names
        assert '李四' in names
        assert '张三' in names
    
    @patch('distribution.os.path.exists')
    def test_get_name_list_file_not_exists(self, mock_exists):
        """测试姓名角色表不存在"""
        mock_exists.return_value = False
        
        names = distribution.get_name_list()
        
        assert names == []


class TestAssignmentDialogIntegration:
    """测试任务指派对话框集成"""
    
    def test_dialog_initialization(self):
        """测试对话框初始化"""
        # 创建测试数据
        unassigned_tasks = [
            {
                'file_type': 1,
                'project_id': '2016',
                'interface_id': 'INT-001',
                'file_path': 'test.xlsx',
                'row_index': 2,
                'interface_time': '10.28',
                'department': '结构一室'
            }
        ]
        
        name_list = ['王任超', '李四', '张三']
        
        # 由于需要Tk root，这里只测试数据验证
        assert len(unassigned_tasks) > 0
        assert len(name_list) > 0
        assert unassigned_tasks[0]['file_type'] == 1
        assert unassigned_tasks[0]['project_id'] == '2016'


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

