#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""文件1数据库只读模式测试（离线SQL模拟）。"""

import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from core import main
from core.sql import file1_db_source
from scripts.db_tools.sql_explorer.validate_cims_sql_dump import (
    iter_insert_rows,
    parse_create_columns,
    parse_department_map,
    parse_user_map,
)
from write_tasks.executors import execute_response_task


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _load_offline_file1_rows(limit: int = 600):
    root = _project_root()
    table_sql = root / "example" / "CIMS-sql" / "IDIACP1000.sql"
    if not table_sql.exists():
        pytest.skip("离线 SQL 文件不存在：example/CIMS-sql/IDIACP1000.sql")

    columns = parse_create_columns(table_sql)
    mapping = {name.lower(): idx for idx, name in enumerate(columns)}
    required = {
        "proj_num",
        "item_number",
        "release_party",
        "swap_start_date",
        "actual_open_date",
        "depart_user",
        "created_by_id",
    }
    if not required.issubset(set(mapping)):
        pytest.skip("离线 IDIACP1000.sql 缺少必需列，跳过测试")

    project_id = ""
    rows = []
    for row in iter_insert_rows(table_sql):
        proj_val = str(row[mapping["proj_num"]] or "").strip()
        if not proj_val:
            continue
        if not project_id:
            project_id = proj_val
        if proj_val != project_id:
            continue
        rows.append(
            (
                row[mapping["item_number"]],
                row[mapping["release_party"]],
                row[mapping["swap_start_date"]],
                row[mapping["actual_open_date"]],
                row[mapping["depart_user"]],
                row[mapping["created_by_id"]],
            )
        )
        if len(rows) >= limit:
            break

    if not rows:
        pytest.skip("离线 SQL 未读取到可用 IDIACP1000 样本行")
    return project_id, rows


def _load_offline_user_name_map():
    root = _project_root()
    user_sql = root / "example" / "CIMS-sql" / "USER.sql"
    dept_sql = root / "example" / "CIMS-sql" / "DEPARTMENT.sql"
    if (not user_sql.exists()) or (not dept_sql.exists()):
        return {}

    dept_map = parse_department_map(dept_sql)
    user_map = parse_user_map(user_sql, dept_map)
    return {uid: str(info.get("user_name", "") or "") for uid, info in user_map.items()}


def test_file1_virtual_source_helpers():
    source = file1_db_source.build_file1_virtual_source("1818")
    assert source == "db://file1/1818"
    assert file1_db_source.is_file1_db_virtual_source(source)
    assert file1_db_source.extract_project_id_from_virtual_source(source) == "1818"
    assert file1_db_source.is_file1_db_source_list([source, "db://file1/1907"])


def test_get_file1_db_connection_status_failure(monkeypatch):
    def _raise_error():
        raise RuntimeError("mock connect failure")

    monkeypatch.setattr(file1_db_source, "create_connection_from_saved_profile", _raise_error)
    status = file1_db_source.get_file1_db_connection_status()
    assert status["connected"] == "0"
    assert "mock connect failure" in status["message"]


def test_fetch_file1_db_dataframe_with_offline_sql(monkeypatch):
    project_id, offline_rows = _load_offline_file1_rows()
    user_name_map = _load_offline_user_name_map()

    # 尽量复用离线样本；若样本不满足筛选条件，兜底修正一条用于验证链路
    candidate_row = None
    dept_codes = ["25B1", "25B2", "25B3", "25C1", "25C2", "25C3", "25E5", "25E6"]
    for row in offline_rows:
        release_party = str(row[1] or "")
        actual_open_date = row[3]
        if any(code in release_party for code in dept_codes) and str(actual_open_date or "").strip() == "":
            candidate_row = row
            break
    if candidate_row is None:
        seed = list(offline_rows[0])
        seed[1] = f"{seed[1] or ''} 25C1"
        seed[3] = None
        candidate_row = tuple(seed)

    class _DummyConn:
        def close(self):
            return None

    monkeypatch.setattr(file1_db_source, "create_connection_from_saved_profile", lambda: (_DummyConn(), "mock"))
    monkeypatch.setattr(file1_db_source, "_query_file1_latest_rows", lambda _conn, _pid: [candidate_row])
    monkeypatch.setattr(file1_db_source, "_query_user_name_map", lambda _conn: user_name_map)
    monkeypatch.setattr(file1_db_source, "get_department_codes", lambda: dept_codes)
    monkeypatch.setattr(file1_db_source, "_in_file1_time_window", lambda *_args, **_kwargs: True)
    monkeypatch.setattr(
        file1_db_source,
        "map_code_to_department",
        lambda text: "结构一室" if "25C1" in str(text or "") else "电气室",
    )

    df = file1_db_source.fetch_file1_db_dataframe(
        project_id=project_id,
        current_datetime=datetime.datetime(2026, 2, 13, 10, 0, 0),
    )
    assert not df.empty
    assert list(df.columns) == file1_db_source.RESULT_COLUMNS
    assert df.iloc[0]["source_file"].startswith("db://file1/")
    assert df.iloc[0]["项目号"] == project_id


def test_export_result_to_excel_db_mode_only_four_columns(tmp_path):
    project_id = "1818"
    source = file1_db_source.build_file1_virtual_source(project_id)
    input_df = pd.DataFrame(
        {
            "接口号": ["S-TEST-001"],
            "接口时间": ["2026.02.20"],
            "责任人": ["张三"],
            "科室": ["结构一室"],
            "项目号": [project_id],
            "原始行号": [2],
            "source_file": [source],
            "额外列": ["不会被导出"],
        }
    )

    output_path = main.export_result_to_excel(
        input_df,
        source,
        datetime.datetime(2026, 2, 13, 10, 0, 0),
        str(tmp_path),
        project_id=project_id,
    )
    wb = load_workbook(output_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    values = [ws.cell(row=2, column=idx).value for idx in range(1, ws.max_column + 1)]
    wb.close()

    assert headers == ["接口号", "接口日期", "责任人", "所属科室"]
    assert values == ["S-TEST-001", "2026.02.20", "张三", "结构一室"]


def test_execute_response_task_block_file1_readonly():
    ok = execute_response_task(
        {
            "file_type": 1,
            "file_path": "db://file1/1818",
            "row_index": 2,
            "response_number": "HF-001",
            "user_name": "测试",
            "project_id": "1818",
        }
    )
    assert ok is False

