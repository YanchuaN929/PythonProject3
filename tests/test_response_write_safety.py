import time
import zipfile
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from registry import hooks as registry_hooks
from utils.excel_io import ExcelWriteError
from write_tasks.manager import WriteTaskManager


pytestmark = pytest.mark.allow_empty_name


def _make_file4(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet["E1"] = "接口号"
    sheet["U1"] = "回文号"
    sheet["V1"] = "回文日期"
    sheet["AT1"] = "填写人"
    for row_index, interface_id, response_number in rows:
        sheet[f"E{row_index}"] = interface_id
        sheet[f"U{row_index}"] = response_number
    workbook.save(path)
    workbook.close()


def _wait_for_task_done(manager, task_id, timeout=5.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        status = manager.tasks[task_id].status
        if status in {"completed", "failed"}:
            return status
        time.sleep(0.02)
    return manager.tasks[task_id].status


def test_same_response_is_idempotent_without_saving_again(tmp_path, monkeypatch):
    from ui import input_handler

    file_path = tmp_path / "response.xlsx"
    _make_file4(file_path, [(2, "C-MX-001", "HF-001")])
    monkeypatch.setattr(
        input_handler,
        "atomic_save_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("相同回文单号不应再次保存Excel")
        ),
    )

    result = input_handler.write_response_to_excel(
        str(file_path),
        file_type=4,
        row_index=2,
        response_number="HF-001",
        user_name="测试用户",
        project_id="2026",
        interface_id="C-MX-001(设计人员)",
        return_details=True,
    )

    assert result.success is True
    assert result.already_present is True
    assert result.row_index == 2


def test_different_existing_response_is_rejected_without_overwrite(tmp_path):
    from ui import input_handler
    from utils.excel_io import open_workbook_for_edit

    file_path = tmp_path / "response_conflict.xlsx"
    _make_file4(file_path, [(2, "C-MX-001", "HF-OLD")])

    with pytest.raises(ExcelWriteError) as exc_info:
        input_handler.write_response_to_excel(
            str(file_path),
            file_type=4,
            row_index=2,
            response_number="HF-NEW",
            user_name="测试用户",
            project_id="2026",
            interface_id="C-MX-001",
        )

    assert exc_info.value.code == "RESPONSE_CONFLICT"
    workbook = open_workbook_for_edit(str(file_path))
    try:
        assert workbook.active["U2"].value == "HF-OLD"
    finally:
        workbook.close()


def test_changed_row_is_relocated_only_when_interface_is_unique(tmp_path):
    from ui import input_handler
    from utils.excel_io import open_workbook_for_edit

    file_path = tmp_path / "response_relocated.xlsx"
    _make_file4(
        file_path,
        [
            (2, "C-OTHER-001", ""),
            (3, "C-MX-001", ""),
        ],
    )

    result = input_handler.write_response_to_excel(
        str(file_path),
        file_type=4,
        row_index=2,
        response_number="HF-001",
        user_name="测试用户",
        project_id="2026",
        interface_id=" C-MX-001 (设计人员)",
        return_details=True,
    )

    assert result.row_index == 3
    workbook = open_workbook_for_edit(str(file_path))
    try:
        assert workbook.active["U2"].value in (None, "")
        assert workbook.active["U3"].value == "HF-001"
    finally:
        workbook.close()


def test_changed_row_with_duplicate_interfaces_is_rejected(tmp_path):
    from ui import input_handler

    file_path = tmp_path / "response_ambiguous.xlsx"
    _make_file4(
        file_path,
        [
            (2, "C-MX-001", ""),
            (3, "C-MX-001", ""),
            (4, "C-OTHER-001", ""),
        ],
    )

    with pytest.raises(ExcelWriteError) as exc_info:
        input_handler.write_response_to_excel(
            str(file_path),
            file_type=4,
            row_index=4,
            response_number="HF-001",
            user_name="测试用户",
            project_id="2026",
            interface_id="C-MX-001",
        )

    assert exc_info.value.code == "ROW_AMBIGUOUS"


def test_atomic_replace_retries_transient_share_block(tmp_path, monkeypatch):
    from utils import excel_io

    file_path = tmp_path / "replace_retry.xlsx"
    _make_file4(file_path, [(2, "C-MX-001", "")])
    workbook = excel_io.open_workbook_for_edit(str(file_path))
    workbook.active["U2"] = "HF-001"

    real_replace = excel_io.os.replace
    attempts = []
    slept = []

    def flaky_replace(source, target):
        attempts.append((source, target))
        if len(attempts) < 3:
            raise PermissionError("sharing violation")
        return real_replace(source, target)

    monkeypatch.setattr(excel_io.os, "replace", flaky_replace)
    monkeypatch.setattr(excel_io.time, "sleep", lambda delay: slept.append(delay))
    try:
        excel_io.atomic_save_workbook(workbook, str(file_path))
    finally:
        workbook.close()

    assert len(attempts) == 3
    assert slept == [0.5, 1.0]


def test_final_verification_retries_without_rewriting_excel(tmp_path, monkeypatch):
    from ui import input_handler

    file_path = tmp_path / "verify_retry.xlsx"
    _make_file4(file_path, [(2, "C-MX-001", "")])

    real_patch = input_handler.atomic_patch_ooxml_cells
    real_read = input_handler.read_ooxml_inline_cell
    patch_calls = []
    read_calls = []

    def counted_patch(path, sheet_path, updates):
        patch_calls.append(path)
        return real_patch(path, sheet_path, updates)

    def flaky_read(path, sheet_path, cell_reference):
        read_calls.append((path, sheet_path, cell_reference))
        if len(read_calls) < 3:
            raise PermissionError("temporary SMB read lock")
        return real_read(path, sheet_path, cell_reference)

    monkeypatch.setattr(input_handler, "atomic_patch_ooxml_cells", counted_patch)
    monkeypatch.setattr(input_handler, "read_ooxml_inline_cell", flaky_read)
    monkeypatch.setattr(input_handler.time, "sleep", lambda _delay: None)

    result = input_handler.write_response_to_excel(
        str(file_path),
        file_type=4,
        row_index=2,
        response_number="HF-001",
        user_name="测试用户",
        project_id="2026",
        interface_id="C-MX-001",
        return_details=True,
    )

    assert result.success is True
    assert len(patch_calls) == 1
    assert len(read_calls) == 3


def test_response_xml_patch_preserves_every_non_target_archive_member_and_cell(tmp_path):
    from ui.input_handler import write_response_to_excel
    from utils.excel_io import _ooxml_cell_pattern, open_workbook_for_edit

    file_path = tmp_path / "external_response.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["E1"] = "接口号"
    sheet["U1"] = "回文号"
    sheet["V1"] = "回文日期"
    sheet["AD1"] = "业务说明"
    sheet["E2"] = "C-MX-XML-001"
    sheet["U2"] = ""
    sheet["AD2"] = "第一行\r\n第二行"
    sheet["AD3"] = "仅回车\r保持"
    other = workbook.create_sheet("其他工作表")
    other["A1"] = "必须原样保留"
    workbook.save(file_path)
    workbook.close()

    baseline = open_workbook_for_edit(str(file_path))
    try:
        baseline_ad2 = baseline.active["AD2"].value
        baseline_ad3 = baseline.active["AD3"].value
    finally:
        baseline.close()

    with zipfile.ZipFile(file_path, "r") as archive:
        before_names = archive.namelist()
        before_members = {name: archive.read(name) for name in before_names}

    assert write_response_to_excel(
        str(file_path),
        4,
        2,
        "HF-XML-001",
        "测试用户",
        "2026",
        interface_id="C-MX-XML-001",
    ) is True

    with zipfile.ZipFile(file_path, "r") as archive:
        after_names = archive.namelist()
        after_members = {name: archive.read(name) for name in after_names}

    assert after_names == before_names
    target_sheet = "xl/worksheets/sheet1.xml"
    for member_name in before_names:
        if member_name != target_sheet:
            assert after_members[member_name] == before_members[member_name]

    def without_target_cells(xml_bytes):
        result = xml_bytes
        for reference in ("U2", "V2", "AX1", "AX2"):
            result = _ooxml_cell_pattern(reference).sub(b"", result)
        return result

    assert without_target_cells(after_members[target_sheet]) == without_target_cells(
        before_members[target_sheet]
    )

    verify = open_workbook_for_edit(str(file_path))
    try:
        assert verify.active["U2"].value == "HF-XML-001"
        assert verify.active["V2"].value is not None
        assert verify.active["AX1"].value == "程序填写人"
        assert verify.active["AX2"].value == "测试用户"
        assert verify.active["AD2"].value == baseline_ad2
        assert verify.active["AD3"].value == baseline_ad3
        assert verify["其他工作表"]["A1"].value == "必须原样保留"
    finally:
        verify.close()


def test_ooxml_patch_keeps_macro_and_custom_archive_members_byte_identical(tmp_path):
    from utils.excel_io import atomic_patch_ooxml_cells, read_ooxml_inline_cell

    file_path = tmp_path / "macro_response.xlsm"
    _make_file4(file_path, [(2, "C-MX-MACRO-001", "")])
    fake_vba = b"FAKE-VBA-PROJECT\x00\x01\xff"
    fake_custom = b"custom-member-must-not-change"
    with zipfile.ZipFile(file_path, "a") as archive:
        archive.writestr("xl/vbaProject.bin", fake_vba)
        archive.writestr("customXml/item99.xml", fake_custom)

    atomic_patch_ooxml_cells(
        str(file_path),
        "xl/worksheets/sheet1.xml",
        [{"cell": "U2", "value": "HF-MACRO-001"}],
    )

    with zipfile.ZipFile(file_path, "r") as archive:
        assert archive.read("xl/vbaProject.bin") == fake_vba
        assert archive.read("customXml/item99.xml") == fake_custom
        assert archive.testzip() is None
    assert (
        read_ooxml_inline_cell(str(file_path), "xl/worksheets/sheet1.xml", "U2")
        == "HF-MACRO-001"
    )


def test_ooxml_replace_failure_leaves_original_file_unchanged(tmp_path, monkeypatch):
    from utils import excel_io

    file_path = tmp_path / "replace_failure.xlsx"
    _make_file4(file_path, [(2, "C-MX-ROLLBACK-001", "")])
    original_bytes = file_path.read_bytes()

    def fail_replace(_temp_path, _file_path):
        raise ExcelWriteError(
            "REPLACE_BLOCKED",
            "REPLACE_TARGET",
            "模拟共享盘占用",
            retryable=True,
            committed=False,
        )

    monkeypatch.setattr(excel_io, "replace_file_with_retry", fail_replace)
    with pytest.raises(ExcelWriteError) as exc_info:
        excel_io.atomic_patch_ooxml_cells(
            str(file_path),
            "xl/worksheets/sheet1.xml",
            [{"cell": "U2", "value": "HF-ROLLBACK-001"}],
        )

    assert exc_info.value.code == "REPLACE_BLOCKED"
    assert file_path.read_bytes() == original_bytes
    assert list(tmp_path.glob(".__excel_xml_write_*")) == []


def test_real_error_is_preserved_in_task_record(tmp_path, monkeypatch):
    from ui import input_handler

    def fail_write(**_kwargs):
        raise ExcelWriteError(
            "FILE_LOCKED",
            "PRECHECK",
            "目标文件被测试用户占用",
            retryable=True,
            committed=False,
        )

    monkeypatch.setattr(input_handler, "write_response_to_excel", fail_write)
    manager = WriteTaskManager(state_path=tmp_path / "tasks.json")
    try:
        manager._sync_to_shared_log = MagicMock()
        task = manager.submit_response_task(
            file_path=str(tmp_path / "dummy.xlsx"),
            file_type=4,
            row_index=2,
            interface_id="C-MX-001",
            response_number="HF-001",
            user_name="测试用户",
            project_id="2026",
            source_column=None,
            description="测试真实错误保留",
        )

        assert _wait_for_task_done(manager, task.task_id) == "failed"
        assert "FILE_LOCKED" in manager.tasks[task.task_id].error
        assert "目标文件被测试用户占用" in manager.tasks[task.task_id].error
    finally:
        manager.shutdown()


def test_executor_persists_relocated_row_for_registry(tmp_path, monkeypatch):
    file_path = tmp_path / "response_executor_relocated.xlsx"
    _make_file4(
        file_path,
        [
            (2, "C-OTHER-001", ""),
            (3, "C-MX-001", ""),
        ],
    )
    registry_write = MagicMock()
    monkeypatch.setattr(registry_hooks, "on_response_written", registry_write)

    manager = WriteTaskManager(state_path=tmp_path / "tasks.json")
    try:
        manager._sync_to_shared_log = MagicMock()
        task = manager.submit_response_task(
            file_path=str(file_path),
            file_type=4,
            row_index=2,
            interface_id="C-MX-001",
            response_number="HF-001",
            user_name="测试用户",
            project_id="2026",
            source_column=None,
            data_folder=str(tmp_path),
            description="测试实际行号同步",
        )

        assert _wait_for_task_done(manager, task.task_id) == "completed"
        assert manager.tasks[task.task_id].payload["row_index"] == 3
        assert registry_write.call_args.kwargs["row_index"] == 3
    finally:
        manager.shutdown()
