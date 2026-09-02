from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook


FILE_CASES = {
    1: ("A", "W", None),
    2: ("R", "P", None),
    3: ("C", "V", "M"),
    4: ("E", "U", None),
    5: ("A", "V", None),
    6: ("E", "L", None),
}


def _make_response_book(path: Path, file_type: int):
    interface_col, _response_col, source_column = FILE_CASES[file_type]
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "测试表头"
    for row, interface_id in ((2, f"IF-{file_type}-A"), (3, f"IF-{file_type}-B")):
        ws[f"{interface_col}{row}"] = interface_id
        if file_type == 3:
            ws[f"{source_column}{row}"] = date.today().isoformat()
        if file_type == 6:
            ws[f"I{row}"] = date.today() + timedelta(days=1)
    wb.save(path)
    wb.close()


@pytest.mark.parametrize("file_type", [1, 2, 3, 4, 5, 6])
def test_batch_response_allows_one_number_for_multiple_interfaces_and_saves_once(
    tmp_path, monkeypatch, file_type
):
    from services import batch_response

    file_path = tmp_path / f"type{file_type}.xlsx"
    _make_response_book(file_path, file_type)
    original_patch = batch_response.atomic_patch_ooxml_cells
    patch_calls = []

    def counted_patch(*args, **kwargs):
        patch_calls.append((args, kwargs))
        return original_patch(*args, **kwargs)

    monkeypatch.setattr(batch_response, "atomic_patch_ooxml_cells", counted_patch)
    source_column = FILE_CASES[file_type][2]
    items = [
        {
            "file_path": str(file_path),
            "file_type": file_type,
            "row_index": row,
            "interface_id": f"IF-{file_type}-{suffix}",
            "response_number": "HF-SAME-001",
            "user_name": "测试人员",
            "project_id": "2026",
            "source_column": source_column,
        }
        for row, suffix in ((2, "A"), (3, "B"))
    ]

    result = batch_response.write_responses_batch(items)

    assert result["success_count"] == 2
    assert result["failed_tasks"] == []
    assert len(patch_calls) == 1
    wb = load_workbook(file_path, data_only=False)
    try:
        response_col = FILE_CASES[file_type][1]
        assert str(wb.active[f"{response_col}2"].value) == "HF-SAME-001"
        assert str(wb.active[f"{response_col}3"].value) == "HF-SAME-001"
    finally:
        wb.close()


def test_batch_response_workbook_group_is_atomic_on_conflict(tmp_path):
    from services.batch_response import write_responses_batch

    file_path = tmp_path / "conflict.xlsx"
    _make_response_book(file_path, 4)
    wb = load_workbook(file_path)
    wb.active["U3"] = "HF-OLD"
    wb.save(file_path)
    wb.close()
    items = [
        {
            "file_path": str(file_path),
            "file_type": 4,
            "row_index": 2,
            "interface_id": "IF-4-A",
            "response_number": "HF-NEW",
            "user_name": "测试人员",
            "project_id": "2026",
        },
        {
            "file_path": str(file_path),
            "file_type": 4,
            "row_index": 3,
            "interface_id": "IF-4-B",
            "response_number": "HF-NEW",
            "user_name": "测试人员",
            "project_id": "2026",
        },
    ]

    result = write_responses_batch(items)

    assert result["success_count"] == 0
    assert len(result["failed_tasks"]) == 2
    wb = load_workbook(file_path)
    try:
        assert wb.active["U2"].value in (None, "")
        assert wb.active["U3"].value == "HF-OLD"
    finally:
        wb.close()


def test_repeated_batch_is_idempotent_and_does_not_save_again(tmp_path, monkeypatch):
    from services import batch_response

    file_path = tmp_path / "idempotent.xlsx"
    _make_response_book(file_path, 1)
    items = [
        {
            "file_path": str(file_path),
            "file_type": 1,
            "row_index": row,
            "interface_id": f"IF-1-{suffix}",
            "response_number": "HF-SAME-001",
            "user_name": "测试人员",
            "project_id": "2026",
        }
        for row, suffix in ((2, "A"), (3, "B"))
    ]
    assert batch_response.write_responses_batch(items)["success_count"] == 2
    save_mock = MagicMock(side_effect=AssertionError("幂等重试不应再次保存Excel"))
    monkeypatch.setattr(batch_response, "atomic_patch_ooxml_cells", save_mock)

    result = batch_response.write_responses_batch(items)

    assert result["success_count"] == 2
    assert result["already_present_count"] == 2
    save_mock.assert_not_called()


def test_parse_mapping_allows_duplicate_response_numbers():
    from ui.batch_response_dialog import parse_response_paste

    items = [
        {"interface_id": "IF-A", "enabled": True},
        {"interface_id": "IF-B", "enabled": True},
    ]
    updates = parse_response_paste(
        "接口号\t回文单号\nIF-A\tHF-SAME\nIF-B\tHF-SAME\n",
        items,
        "mapping",
    )
    assert updates == {0: "HF-SAME", 1: "HF-SAME"}


def test_parse_sequential_requires_exact_count():
    from ui.batch_response_dialog import parse_response_paste

    items = [
        {"interface_id": "IF-A", "enabled": True},
        {"interface_id": "IF-B", "enabled": True},
    ]
    with pytest.raises(ValueError, match="数量必须完全一致"):
        parse_response_paste("HF-ONLY-ONE", items, "sequential")


def test_fu_batch_saves_one_workbook_once(tmp_path, monkeypatch):
    from services import batch_response

    file_path = tmp_path / "fu.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "项目号"
    ws["B1"] = "内部编码"
    ws["B2"] = "FU-A"
    ws["B3"] = "FU-B"
    wb.save(file_path)
    wb.close()
    original_save = batch_response.atomic_save_workbook
    save_calls = []

    def counted_save(*args, **kwargs):
        save_calls.append((args, kwargs))
        return original_save(*args, **kwargs)

    monkeypatch.setattr(batch_response, "atomic_save_workbook", counted_save)
    items = [
        {
            "file_path": str(file_path),
            "file_type": 7,
            "row_index": row,
            "interface_id": interface_id,
            "completion_date": "2026-09-02",
            "user_name": "测试人员",
            "project_id": "2026",
        }
        for row, interface_id in ((2, "FU-A"), (3, "FU-B"))
    ]

    result = batch_response.write_fu_completions_batch(items)

    assert result["success_count"] == 2
    assert len(save_calls) == 1
    wb = load_workbook(file_path)
    try:
        assert wb.active["D2"].value.date() == date(2026, 9, 2)
        assert wb.active["D3"].value.date() == date(2026, 9, 2)
    finally:
        wb.close()


def test_response_batch_executor_registry_failure_only_returns_compensation(monkeypatch):
    from registry import hooks as registry_hooks
    from services import batch_response
    from write_tasks import executors

    successful = [
        {
            "file_path": "response.xlsx",
            "file_type": 4,
            "row_index": 2,
            "requested_row_index": 2,
            "interface_id": "IF-A",
            "response_number": "HF-SAME",
            "user_name": "测试人员",
            "project_id": "2026",
        },
        {
            "file_path": "response.xlsx",
            "file_type": 4,
            "row_index": 3,
            "requested_row_index": 3,
            "interface_id": "IF-B",
            "response_number": "HF-SAME",
            "user_name": "测试人员",
            "project_id": "2026",
        },
    ]
    write_mock = MagicMock(return_value={
        "success_count": 2,
        "already_present_count": 0,
        "successful_items": successful,
        "failed_tasks": [],
    })
    monkeypatch.setattr(batch_response, "write_responses_batch", write_mock)
    monkeypatch.setattr(registry_hooks, "on_response_written", MagicMock(return_value=False))
    payload = {"items": successful, "user_name": "测试人员", "data_folder": "E:/data"}

    result = executors.execute_response_batch_task(payload)

    write_mock.assert_called_once()
    assert result["success_count"] == 2
    assert len(result["registry_compensations"]) == 2
    assert all(item["operation"] == "response_written" for item in result["registry_compensations"])


def test_manager_submits_batch_task(tmp_path):
    from write_tasks.manager import WriteTaskManager

    manager = WriteTaskManager(state_path=tmp_path / "state.json")
    try:
        task = manager.submit_response_batch_task(
            items=[{
                "file_path": "missing.xlsx",
                "file_type": 1,
                "row_index": 2,
                "interface_id": "IF-A",
                "response_number": "HF-001",
            }],
            user_name="测试人员",
            data_folder="E:/data",
            description="批量回文测试",
        )
        assert task.task_type == "response_batch"
        assert task.payload["items"][0]["response_number"] == "HF-001"
        assert task.payload["batch_id"]
    finally:
        manager.shutdown()


def test_pending_cache_removes_only_failed_rows_from_partial_batch():
    from write_tasks.models import WriteTask
    from write_tasks.pending_cache import PendingCache

    cache = PendingCache()
    infos = [
        {"file_path": "book.xlsx", "file_type": 1, "row_index": 2, "response_number": "HF", "user_name": "张三"},
        {"file_path": "book.xlsx", "file_type": 1, "row_index": 3, "response_number": "HF", "user_name": "张三"},
    ]
    cache.add_response_entries("batch-1", infos)
    task = WriteTask(
        task_id="batch-1",
        task_type="response_batch",
        payload={
            "_result": {
                "successful_items": [{
                    "file_path": "book.xlsx",
                    "file_type": 1,
                    "row_index": 2,
                    "requested_row_index": 2,
                }],
                "failed_tasks": [{"file_path": "book.xlsx", "file_type": 1, "row_index": 3}],
            }
        },
        submitted_by="张三",
        description="test",
        status="completed",
    )

    cache.on_task_status_changed(task)

    rows = sorted(item["row_index"] for item in cache.get_summary(only_user="张三"))
    assert rows == [2]
