import datetime
import sqlite3

from openpyxl import Workbook, load_workbook

from core import main
from registry import service
from ui.input_handler import write_fu_completion_to_excel


def _make_fu_workbook(tmp_path, rows, filename="1916项目标准表格.xlsx"):
    path = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    for column, header in {
        "A": "文件编码",
        "B": "内部编码",
        "C": "中文标题",
        "D": "实际FU日期",
        "E": "FU计划",
        "F": "责任人",
    }.items():
        ws[f"{column}1"] = header
    for row_number, values in rows:
        for column, value in values.items():
            ws[f"{column}{row_number}"] = value
    wb.save(path)
    wb.close()
    return path


def test_file7_finder_requires_exact_project_standard_name(tmp_path):
    valid = tmp_path / "1916项目标准表格.xlsx"
    valid.touch()
    invalid = tmp_path / "1916项目标准表格-副本.xlsx"
    invalid.touch()

    assert main.find_all_target_files7([str(valid), str(invalid)]) == [
        (str(valid), "1916")
    ]
    assert main.find_target_file7([str(valid)]) == (str(valid), "1916")


def test_file7_stream_keeps_row2_and_stabilizes_duplicate_codes(monkeypatch, tmp_path):
    now = datetime.datetime(2026, 5, 13)
    path = _make_fu_workbook(tmp_path, [
        (2, {"A": "Z-FILE", "B": "FU-DUP", "C": "标题乙", "E": now}),
        (3, {"A": "A-FILE", "B": "FU-DUP", "C": "标题甲", "E": now}),
        (4, {"A": "C-FILE", "B": "FU-DONE", "C": "已完成", "D": now, "E": now}),
        (5, {"A": "D-FILE", "B": "FU-NOPLAN", "C": "无计划", "E": ""}),
    ])
    monkeypatch.setattr(
        main,
        "_merge_registry_pending_rows",
        lambda **kwargs: (kwargs["final_rows"], set()),
    )
    monkeypatch.setattr(main, "_load_latest_confirmed_archive_times", lambda *_args: {})

    df = main.process_target_file7(str(path), now)

    assert df["原始行号"].tolist() == [2, 3]
    assert df["项目号"].tolist() == ["1916", "1916"]
    by_row = df.set_index("原始行号")
    assert by_row.loc[2, "内部编码"] == "FU-DUP#02"
    assert by_row.loc[3, "内部编码"] == "FU-DUP#01"
    assert by_row.loc[2, "接口号"] == "FU-DUP#02"
    assert by_row.loc[2, "科室"] == "请室主任确认"
    assert by_row.loc[2, "_completed_col"] == "D"
    assert by_row.loc[2, "_assign_col"] == "F"


def test_file7_same_confirmed_plan_is_hidden_but_changed_plan_reopens(monkeypatch, tmp_path):
    now = datetime.datetime(2026, 5, 13)
    path = _make_fu_workbook(tmp_path, [
        (2, {"A": "A", "B": "FU-01", "C": "标题", "E": now}),
    ])
    monkeypatch.setattr(
        main,
        "_merge_registry_pending_rows",
        lambda **kwargs: (kwargs["final_rows"], set()),
    )
    monkeypatch.setattr(
        main,
        "_load_latest_confirmed_archive_times",
        lambda *_args: {"FU-01": "2026.05.13"},
    )
    assert main.process_target_file7(str(path), now).empty

    monkeypatch.setattr(
        main,
        "_load_latest_confirmed_archive_times",
        lambda *_args: {"FU-01": "2026.05.12"},
    )
    assert len(main.process_target_file7(str(path), now)) == 1

    archived = {
        "status": "archived",
        "archive_reason": "confirmed_by_superior",
        "completed_at": "2026-05-13T10:00:00",
        "confirmed_at": "2026-05-13T11:00:00",
        "interface_time": "2026.05.13",
    }
    assert service._should_suppress_rescan_after_confirmed_archive(
        archived,
        {"interface_time": "2026.05.13", "_completed_col_value": ""},
        7,
    ) is True
    assert service._should_suppress_rescan_after_confirmed_archive(
        archived,
        {"interface_time": "2026.05.14", "_completed_col_value": ""},
        7,
    ) is False


def test_file7_completion_writes_only_actual_date(tmp_path, monkeypatch):
    path = _make_fu_workbook(tmp_path, [
        (2, {"A": "A", "B": "FU-01", "C": "标题", "E": datetime.datetime(2026, 5, 13), "F": "张三"}),
    ])
    monkeypatch.setattr("ui.input_handler.messagebox.showerror", lambda *_args, **_kwargs: None)

    assert write_fu_completion_to_excel(str(path), 2, "2026-05-14") is True

    wb = load_workbook(path, data_only=False)
    try:
        ws = wb.active
        assert ws["D2"].value.date() == datetime.date(2026, 5, 14)
        assert ws["E2"].value == datetime.datetime(2026, 5, 13)
        assert ws["F2"].value == "张三"
    finally:
        wb.close()


def test_file7_assignment_writes_column_f(tmp_path, monkeypatch):
    from registry import hooks
    from services.distribution import save_assignments_batch

    path = _make_fu_workbook(tmp_path, [
        (2, {"A": "A", "B": "FU-01", "C": "标题", "E": datetime.datetime(2026, 5, 13)}),
    ])
    monkeypatch.setattr(hooks, "on_assigned", lambda **_kwargs: None)

    result = save_assignments_batch([{
        "file_type": 7,
        "project_id": "1916",
        "interface_id": "FU-01",
        "file_path": str(path),
        "row_index": 2,
        "assigned_name": "张三",
        "assigned_by": "一室主任",
    }])

    assert result["success_count"] == 1
    wb = load_workbook(path, read_only=True)
    try:
        assert wb.active["F2"].value == "张三"
        assert wb.active["D2"].value is None
    finally:
        wb.close()


def test_file7_export_contains_only_fu_business_columns(tmp_path):
    df = main.pd.DataFrame([{
        "原始行号": 2,
        "状态": "待完成",
        "项目号": "1916",
        "内部编码": "FU-01",
        "中文标题": "标题",
        "FU计划": "2026.05.13",
        "实际FU日期": "",
        "责任人": "张三",
        "source_file": "ignored.xlsx",
        "_file_type": 7,
    }])

    output = main.export_result_to_excel7(
        df,
        "unused.xlsx",
        datetime.datetime(2026, 5, 13),
        str(tmp_path),
        "1916",
    )
    wb = load_workbook(output, read_only=True)
    try:
        headers = [cell.value for cell in next(wb.active.iter_rows(min_row=1, max_row=1))]
        assert headers == main.STREAM_FILE_SPECS[7]["export_columns"]
        assert "source_file" not in headers
        assert "接口号" not in headers
    finally:
        wb.close()


def test_file7_gui_display_uses_internal_code_and_fu_columns():
    from ui.window import WindowManager

    manager = WindowManager.__new__(WindowManager)
    source = main.pd.DataFrame([{
        "状态": "待完成",
        "项目号": "1916",
        "接口号": "FU-01",
        "内部编码": "FU-01",
        "中文标题": "标题",
        "FU计划": "2026.05.13",
        "实际FU日期": "",
        "责任人": "张三",
        "原始行号": 2,
    }])

    display = manager._create_optimized_display(source, "FU", completed_rows=set())

    assert list(display.columns) == [
        "状态",
        "项目号",
        "内部编码",
        "中文标题",
        "FU计划",
        "实际FU日期",
        "责任人",
        "是否已完成",
    ]
    assert "接口号" not in display.columns


def test_file7_registry_completion_confirmation_and_new_cycle(monkeypatch, tmp_path):
    from registry import hooks

    now = datetime.datetime(2026, 5, 13, 10, 0, 0)
    path = _make_fu_workbook(tmp_path, [
        (2, {"A": "A", "B": "FU-CYCLE", "C": "标题", "E": now}),
    ])
    db_path = tmp_path / "registry.db"
    cfg = {
        "registry_enabled": True,
        "registry_db_path": str(db_path),
        "registry_wal": False,
    }
    monkeypatch.setattr(hooks, "_cfg", lambda: cfg)
    monkeypatch.setattr(hooks, "_ensure_data_folder_from_path", lambda _path: None)

    first_result = main.process_target_file7(str(path), now)
    hooks.on_process_done(7, "1916", str(path), first_result, now)
    assert write_fu_completion_to_excel(str(path), 2, "2026-05-13") is True
    assert hooks.on_fu_completed(
        file_path=str(path),
        row_index=2,
        interface_id="FU-CYCLE",
        actual_date="2026-05-13",
        user_name="张三",
        project_id="1916",
        role="设计人员",
        now=now + datetime.timedelta(minutes=1),
    ) is True

    conn = sqlite3.connect(db_path)
    try:
        status, display_status = conn.execute(
            "SELECT status, display_status FROM tasks WHERE file_type=7 AND interface_id='FU-CYCLE'"
        ).fetchone()
        assert (status, display_status) == ("completed", "待审查")
    finally:
        conn.close()

    assert hooks.on_confirmed_by_superior(
        file_type=7,
        file_path=str(path),
        row_index=2,
        user_name="一室主任",
        project_id="1916",
        interface_id="FU-CYCLE",
        role="一室主任",
        now=now + datetime.timedelta(minutes=2),
    ) is True

    conn = sqlite3.connect(db_path)
    try:
        task = conn.execute(
            "SELECT status, display_status, confirmed_by, archive_reason FROM tasks WHERE file_type=7"
        ).fetchone()
        assert task == ("archived", "已审查", "一室主任", "confirmed_by_superior")
        event_types = {row[0] for row in conn.execute("SELECT event FROM events WHERE file_type=7")}
        assert {"fu_completed", "confirmed", "archived"}.issubset(event_types)
    finally:
        conn.close()

    wb = load_workbook(path)
    ws = wb.active
    ws["D2"] = None
    wb.save(path)
    wb.close()
    assert main.process_target_file7(str(path), now).empty

    wb = load_workbook(path)
    ws = wb.active
    ws["E2"] = datetime.datetime(2026, 5, 14)
    wb.save(path)
    wb.close()
    next_cycle = main.process_target_file7(str(path), now)
    assert len(next_cycle) == 1
    assert next_cycle.iloc[0]["接口号"] == "FU-CYCLE"
