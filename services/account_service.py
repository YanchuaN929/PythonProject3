#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""基于姓名角色表的账户认证与并发安全密码写入。"""

from __future__ import annotations

import hmac
import os
import time
from dataclasses import dataclass
from typing import List

from openpyxl import load_workbook

from utils.excel_io import atomic_save_workbook, open_workbook_for_edit


DEFAULT_PASSWORD = "password"
DEFAULT_LOCK_TIMEOUT_SECONDS = 12.0


class AccountError(RuntimeError):
    """账户功能的可预期业务异常。"""


class AccountBusyError(AccountError):
    """共享账户表正由其他客户端修改。"""


class AccountDataError(AccountError):
    """账户表缺失、损坏或数据不一致。"""


class AuthenticationError(AccountError):
    """用户名或密码验证失败。"""


@dataclass(frozen=True)
class AccountRecord:
    name: str
    password: str


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_header_row(name_value, role_value, password_value=None) -> bool:
    name = _clean_cell(name_value)
    role = _clean_cell(role_value)
    password = _clean_cell(password_value)
    return "姓名" in name or "角色" in role or password in {"密码", "口令"}


def _effective_password(value) -> str:
    text = "" if value is None else str(value)
    return text if text else DEFAULT_PASSWORD


def _validate_table_path(file_path: str) -> str:
    path = os.path.abspath(os.fspath(file_path))
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise AccountDataError("账户表必须是 .xlsx 或 .xlsm 文件。")
    if not os.path.isfile(path):
        raise AccountDataError(f"账户表不存在：{path}")
    return path


def _read_records(file_path: str) -> List[AccountRecord]:
    path = _validate_table_path(file_path)
    last_error = None
    for attempt in range(3):
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.worksheets[0]
            records: List[AccountRecord] = []
            for row_index, row in enumerate(
                worksheet.iter_rows(min_col=1, max_col=3, values_only=True),
                start=1,
            ):
                name = _clean_cell(row[0] if len(row) > 0 else None)
                role = row[1] if len(row) > 1 else None
                password = row[2] if len(row) > 2 else None
                if row_index == 1 and _is_header_row(name, role, password):
                    continue
                if name and name.lower() not in {"nan", "none"}:
                    records.append(AccountRecord(name, _effective_password(password)))
            return records
        except Exception as exc:
            last_error = exc
            if attempt < 2:
                time.sleep(0.08 * (attempt + 1))
        finally:
            if workbook is not None:
                workbook.close()
    raise AccountDataError(f"无法读取账户表，请检查公共盘连接和文件完整性：{last_error}")


def list_accounts(file_path: str) -> List[str]:
    """按角色表原顺序返回去重后的账户姓名。"""
    result: List[str] = []
    seen = set()
    for record in _read_records(file_path):
        if record.name not in seen:
            seen.add(record.name)
            result.append(record.name)
    return result


def _password_for_user(records: List[AccountRecord], user_name: str) -> str:
    name = _clean_cell(user_name)
    passwords = {record.password for record in records if record.name == name}
    if not passwords:
        raise AuthenticationError("账户不存在，请联系管理员检查姓名角色表。")
    if len(passwords) > 1:
        raise AccountDataError(f"账户“{name}”存在多条不一致的密码记录，请联系管理员处理。")
    return next(iter(passwords))


def verify_password(file_path: str, user_name: str, password: str) -> bool:
    """验证指定账户密码；空白第三列按默认密码 ``password`` 处理。"""
    expected = _password_for_user(_read_records(file_path), user_name)
    supplied = "" if password is None else str(password)
    return hmac.compare_digest(expected, supplied)


class _SharedFileLock:
    """适用于本地磁盘及 SMB 公共盘的系统级字节范围锁。"""

    def __init__(self, target_path: str, timeout: float, stale_seconds: float = 300.0):
        self.lock_path = f"{target_path}.account.lock"
        self.timeout = max(0.0, float(timeout))
        self.acquired = False
        self._handle = None

    def _try_lock_handle(self):
        handle = open(self.lock_path, "a+b")
        try:
            handle.seek(0, os.SEEK_END)
            if handle.tell() == 0:
                handle.write(b"0")
                handle.flush()
            handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:  # pragma: no cover - 发布平台为 Windows
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            return handle
        except Exception:
            handle.close()
            raise

    def __enter__(self):
        deadline = time.monotonic() + self.timeout
        while True:
            try:
                self._handle = self._try_lock_handle()
                self.acquired = True
                return self
            except (BlockingIOError, PermissionError):
                if time.monotonic() >= deadline:
                    raise AccountBusyError(
                        "账户表正由其他用户修改，请稍后重试。"
                    )
                time.sleep(0.12)
            except OSError as exc:
                raise AccountDataError(f"无法在公共盘创建账户写入锁：{exc}") from exc

    def __exit__(self, exc_type, exc_value, traceback):
        if not self.acquired:
            return False
        try:
            self._handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self._handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:  # pragma: no cover - 发布平台为 Windows
                import fcntl

                fcntl.flock(self._handle.fileno(), fcntl.LOCK_UN)
        except Exception:
            pass
        finally:
            if self._handle is not None:
                try:
                    self._handle.close()
                except Exception:
                    pass
                self._handle = None
        try:
            os.remove(self.lock_path)
        except OSError:
            # 另一个客户端可能已经打开同一锁文件；保留空闲锁文件不影响后续使用。
            pass
        self.acquired = False
        return False


def change_password(
    file_path: str,
    user_name: str,
    current_password: str,
    new_password: str,
    *,
    lock_timeout: float = DEFAULT_LOCK_TIMEOUT_SECONDS,
) -> None:
    """在跨客户端锁内重新读取并原子更新账户表第三列。"""
    path = _validate_table_path(file_path)
    name = _clean_cell(user_name)
    if not name:
        raise AccountDataError("当前账户为空，无法修改密码。")
    if new_password is None or not str(new_password).strip():
        raise AccountDataError("新密码不能为空。")
    if len(str(new_password)) > 128:
        raise AccountDataError("新密码不能超过 128 个字符。")

    with _SharedFileLock(path, lock_timeout):
        # 必须在取得锁后重新验证，以防另一客户端刚刚修改过同一账户。
        records = _read_records(path)
        expected = _password_for_user(records, name)
        supplied = "" if current_password is None else str(current_password)
        if not hmac.compare_digest(expected, supplied):
            raise AuthenticationError("当前密码不正确。")

        workbook = None
        try:
            workbook = open_workbook_for_edit(path)
            worksheet = workbook.worksheets[0]
            matched_rows = []
            for row_index in range(1, worksheet.max_row + 1):
                row_name = _clean_cell(worksheet.cell(row=row_index, column=1).value)
                if row_index == 1 and _is_header_row(
                    row_name,
                    worksheet.cell(row=row_index, column=2).value,
                    worksheet.cell(row=row_index, column=3).value,
                ):
                    continue
                if row_name == name:
                    matched_rows.append(row_index)

            if not matched_rows:
                raise AccountDataError("当前账户已不在姓名角色表中，请刷新后重试。")
            for row_index in matched_rows:
                worksheet.cell(row=row_index, column=3).value = str(new_password)
            atomic_save_workbook(workbook, path)
        except AccountError:
            raise
        except PermissionError as exc:
            raise AccountDataError(
                "账户表正在被 Excel 或其他程序占用，请关闭文件后重试。"
            ) from exc
        except Exception as exc:
            raise AccountDataError(f"密码写入失败，账户表未被替换：{exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()
