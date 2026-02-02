#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
任务指派模块
负责接口工程师和室主任的任务指派功能
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
import os
import sys

from write_tasks import get_write_task_manager, get_pending_cache
from ui.ui_copy import copy_text, normalize_interface_id

# 导入文件锁定检测函数
try:
    from ui.input_handler import get_excel_lock_owner
except ImportError:
    get_excel_lock_owner = None


def get_resource_path(relative_path):
    """获取资源文件的绝对路径（支持打包后的exe）"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_responsible_column(file_type):
    """
    获取各文件类型的责任人列名
    
    参数:
        file_type: 文件类型(1-6)
    
    返回:
        str: Excel列名，如'R'、'AP'、'AM'等
    """
    column_map = {
        1: 'R',   # 文件1责任人列（索引17）
        2: 'AM',  # 文件2责任人列（索引38）
        3: 'AP',  # 文件3责任人列（索引41）
        4: 'AH',  # 文件4责任人列（索引33）
        5: 'K',   # 文件5责任人列（索引10）
        6: 'X',   # 文件6责任人列（索引23）
    }
    return column_map.get(file_type)


def get_name_list():
    """
    从姓名角色表读取姓名列表
    
    返回:
        list: 姓名列表（已去重排序）
    """
    try:
        xls_path = get_resource_path("excel_bin/姓名角色表.xlsx")
        if not os.path.exists(xls_path):
            print(f"姓名角色表不存在: {xls_path}")
            return []
        
        df = pd.read_excel(xls_path)
        
        # 第一列为姓名
        if len(df.columns) == 0:
            return []
        
        names = df.iloc[:, 0].dropna().astype(str).tolist()
        
        # 过滤空值和'nan'
        names = [name.strip() for name in names if name.strip() and name.strip().lower() != 'nan']
        
        # 去重并排序
        names = sorted(set(names))
        
        return names
        
    except Exception as e:
        print(f"读取姓名列表失败: {e}")
        import traceback
        traceback.print_exc()
        return []


def is_interface_engineer(user_roles):
    """
    判断是否是接口工程师
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        bool: 是否是接口工程师
    """
    if not user_roles:
        return False
    
    for role in user_roles:
        if '接口工程师' in str(role):
            return True
    
    return False


def is_director(user_roles):
    """
    判断是否是室主任
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        bool: 是否是室主任
    """
    if not user_roles:
        return False
    
    director_roles = ['一室主任', '二室主任', '建筑总图室主任']
    
    for role in user_roles:
        if role in director_roles:
            return True
    
    return False


def get_department(user_roles):
    """
    根据室主任角色获取科室名称
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        str: 科室名称
    """
    dept_map = {
        "一室主任": "结构一室",
        "二室主任": "结构二室",
        "建筑总图室主任": "建筑总图室"
    }
    
    for role in user_roles:
        if role in dept_map:
            return dept_map[role]
    
    return ""


def parse_interface_engineer_project(user_roles):
    """
    从接口工程师角色中提取项目号
    
    参数:
        user_roles: 用户角色列表
    
    返回:
        str: 项目号，如"2016"
    """
    import re
    
    for role in user_roles:
        if '接口工程师' in str(role):
            # 提取项目号（4位数字）
            match = re.search(r'(\d{4})', role)
            if match:
                return match.group(1)
    
    return None


def get_interface_id_column_index(file_type):
    """
    获取各文件类型的接口号列索引（参考window.py的interface_column_index逻辑）
    
    参数:
        file_type: 文件类型(1-6)
    
    返回:
        int: DataFrame列索引
    """
    column_map = {
        1: 0,   # 文件1接口号列 - A列 = 索引0
        2: 17,  # 文件2接口号列 - R列 = 索引17
        3: 2,   # 文件3接口号列 - C列 = 索引2
        4: 4,   # 文件4接口号列 - E列 = 索引4
        5: 0,   # 文件5接口号列 - A列 = 索引0
        6: 4,   # 文件6接口号列 - E列 = 索引4
    }
    return column_map.get(file_type, 0)


def check_unassigned(processed_results, user_roles, project_id=None, config=None):
    """
    检测所有处理结果中没有责任人的数据
    
    参数:
        processed_results: 6个文件的处理结果字典 {file_type: DataFrame}
        user_roles: 当前用户角色列表
        project_id: 项目号（接口工程师需要）
        config: 配置字典（用于自动过滤超期任务）
    
    返回:
        list: 未指派任务列表，每个任务包含：
              - file_type: 文件类型
              - project_id: 项目号
              - interface_id: 接口号
              - file_path: 源文件路径
              - row_index: Excel行号
    """
    unassigned = []
    
    # 【新增】获取超期过滤配置
    auto_hide_enabled = False
    threshold_days = 30
    if config:
        auto_hide_enabled = config.get("auto_hide_overdue_enabled", True)
        threshold_days = config.get("auto_hide_overdue_days", 30)
    
    pending_cache = None
    try:
        pending_cache = get_pending_cache()
    except Exception:
        pending_cache = None

    for file_type, df in processed_results.items():
        if df is None or df.empty:
            continue
        
        # 筛选责任人为空的数据
        mask = (df['责任人'].isna()) | (df['责任人'].astype(str).str.strip() == '') | (df['责任人'].astype(str).str.strip() == '无')
        df_unassigned = df[mask].copy()
        
        if df_unassigned.empty:
            continue
        
        # 角色权限过滤
        if is_interface_engineer(user_roles):
            # 接口工程师：只看自己负责的项目
            if project_id and '项目号' in df_unassigned.columns:
                df_unassigned = df_unassigned[df_unassigned['项目号'].astype(str) == str(project_id)]
        
        elif is_director(user_roles):
            # 室主任：只看自己科室的数据 + "请室主任确认"
            my_department = get_department(user_roles)
            if my_department and '科室' in df_unassigned.columns:
                df_unassigned = df_unassigned[
                    df_unassigned['科室'].isin([my_department, '请室主任确认'])
                ]
        
        # 转换为字典列表
        interface_col_idx = get_interface_id_column_index(file_type)
        
        for idx, row in df_unassigned.iterrows():
            # 优先尝试使用'接口号'列名（用于测试和有命名列的DataFrame）
            interface_id = ''
            if '接口号' in df_unassigned.columns:
                try:
                    interface_id = row['接口号']
                    if pd.isna(interface_id) or str(interface_id).strip() == '':
                        interface_id = ''
                except Exception as e:
                    print(f"从'接口号'列获取失败: {e}")
            
            # 备用方案：使用列索引（用于真实的Excel数据，列名是数字索引）
            if not interface_id or interface_id == '':
                try:
                    if interface_col_idx < len(row):
                        interface_id = row.iloc[interface_col_idx]
                        if pd.isna(interface_id) or str(interface_id).strip() == '':
                            interface_id = ''
                except Exception:
                    # 静默处理列索引越界（某些DataFrame列数较少）
                    pass
            
            file_path = row.get('source_file', '') or row.get('源文件', '')
            task = {
                'file_type': file_type,
                'project_id': row.get('项目号', ''),
                'interface_id': str(interface_id) if interface_id and not pd.isna(interface_id) else '',
                'file_path': file_path,
                'row_index': row.get('原始行号', 0),
                'interface_time': row.get('接口时间', ''),
                'department': row.get('科室', '')
            }
            
            # 确保有必要的字段
            if task['file_path'] and task['row_index']:
                if pending_cache and pending_cache.is_assignment_pending(task['file_path'], task['row_index'], file_type):
                    continue
                # 【新增】超期过滤：如果开启自动隐藏，过滤超期太久的任务
                if auto_hide_enabled and threshold_days > 0:
                    interface_time = task.get('interface_time', '')
                    if interface_time and str(interface_time).strip() not in ['', '-', 'nan', 'None', '未知']:
                        try:
                            from utils.date_utils import parse_mmdd_to_date, get_workday_difference
                            from datetime import date
                            today = date.today()
                            due_date = parse_mmdd_to_date(str(interface_time).strip(), today)
                            if due_date:
                                workday_diff = get_workday_difference(due_date, today)
                                if workday_diff < 0 and abs(workday_diff) > threshold_days:
                                    # 超期太久，跳过此任务
                                    continue
                        except Exception:
                            pass
                
                unassigned.append(task)
    
    return unassigned


def save_assignment(file_type, file_path, row_index, assigned_name):
    """
    保存指派结果到Excel的责任人列（单个任务）
    
    注意：此函数已废弃，建议使用save_assignments_batch进行批量指派
    
    参数:
        file_type: 文件类型(1-6)
        file_path: Excel文件路径
        row_index: Excel行号
        assigned_name: 指派的责任人姓名
    
    返回:
        bool: 成功返回True，失败返回False
    """
    # 调用批量指派函数
    assignments = [{
        'file_type': file_type,
        'file_path': file_path,
        'row_index': row_index,
        'assigned_name': assigned_name
    }]
    
    results = save_assignments_batch(assignments)
    return results.get('success_count', 0) > 0


def save_assignments_batch(assignments):
    """
    批量保存指派结果到Excel（优化版，按文件分组）
    
    参数:
        assignments: 指派列表，每项包含:
            {
                'file_type': int,
                'file_path': str,
                'row_index': int,
                'assigned_name': str,
                'interface_id': str (可选，用于Registry)
                'project_id': str (可选，用于Registry)
            }
    
    返回:
        dict: {
            'success_count': int,  # 成功数量
            'failed_tasks': list,  # 失败的任务信息
            'registry_updates': int  # Registry更新数量
        }
    """
    import pandas as pd
    from collections import defaultdict
    
    # 按文件路径分组
    file_groups = defaultdict(list)
    for assignment in assignments:
        file_path = assignment['file_path']
        file_groups[file_path].append(assignment)
    
    success_count = 0
    failed_tasks = []
    registry_updates = 0
    
    # 按文件批量处理
    for file_path, file_assignments in file_groups.items():
        try:
            # 1. 检查文件是否存在
            if not os.path.exists(file_path):
                print(f"[指派] 文件不存在: {file_path}")
                for assignment in file_assignments:
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': '文件不存在'
                    })
                continue
            
            # 2. 文件锁定检测
            try:
                with open(file_path, 'r+b'):
                    pass
            except PermissionError:
                # 尝试获取占用者信息
                lock_owner = ""
                if get_excel_lock_owner:
                    lock_owner = get_excel_lock_owner(file_path)
                
                if lock_owner:
                    print(f"[指派] 文件被占用: {file_path} (占用者: {lock_owner})")
                    reason = f'文件被 {lock_owner} 占用'
                else:
                    print(f"[指派] 文件被占用: {file_path}")
                    reason = '文件被占用'
                
                for assignment in file_assignments:
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': reason
                    })
                continue
            
            # 3. 打开Excel文件（只打开一次）
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 4. 读取DataFrame用于Registry（只读一次，可失败；Registry 将优先使用 payload 兜底）
            df = None
            try:
                df = pd.read_excel(file_path, sheet_name=0)
            except Exception as e:
                print(f"[指派] 读取DataFrame失败: {e}")
            
            # 5. 批量写入责任人
            for assignment in file_assignments:
                try:
                    file_type = assignment['file_type']
                    row_index = assignment['row_index']
                    assigned_name = assignment['assigned_name']
                    
                    # 获取责任人列名
                    col_name = get_responsible_column(file_type)
                    if not col_name:
                        print(f"[指派] 无法确定责任人列: file_type={file_type}")
                        failed_tasks.append({
                            'interface_id': assignment.get('interface_id', '未知'),
                            'reason': '无法确定责任人列'
                        })
                        continue
                    
                    # 写入责任人
                    ws[f"{col_name}{row_index}"] = assigned_name
                    success_count += 1
                    
                    print(f"[指派] 成功: 行{row_index}, 责任人={assigned_name}")
                    
                except Exception as e:
                    print(f"[指派] 单个任务失败: {e}")
                    failed_tasks.append({
                        'interface_id': assignment.get('interface_id', '未知'),
                        'reason': str(e)
                    })
            
            # 6. 保存Excel（只保存一次）
            wb.save(file_path)
            wb.close()
            print(f"[指派] 文件已保存: {file_path}")
            
            # 7. 批量调用Registry钩子（不依赖 DataFrame 一定成功；优先使用 assignment payload 的接口号/项目号兜底）
            try:
                from registry import hooks as registry_hooks
                from registry.util import extract_interface_id, extract_project_id

                for assignment in file_assignments:
                    try:
                        row_index = assignment['row_index']

                        # 兜底：先用 payload（更稳定，也避免行号映射不准导致“更新 0”）
                        interface_id = str(assignment.get("interface_id", "") or "").strip()
                        project_id = str(assignment.get("project_id", "") or "").strip()

                        # 若 df 可用，且能正确映射到行，则以 df 提取结果为准（更贴近真实Excel内容）
                        if df is not None:
                            try:
                                df_row_idx = row_index - 2  # Excel行号（含表头） -> df 行索引
                                if 0 <= df_row_idx < len(df):
                                    row_data = df.iloc[df_row_idx]
                                    df_interface_id = extract_interface_id(row_data, assignment['file_type'])
                                    df_project_id = extract_project_id(row_data, assignment['file_type'])
                                    if df_interface_id and df_project_id:
                                        interface_id = str(df_interface_id or "").strip()
                                        project_id = str(df_project_id or "").strip()
                            except Exception:
                                pass

                        if interface_id and project_id:
                            assigned_by = assignment.get('assigned_by', '系统用户')
                            registry_hooks.on_assigned(
                                file_type=assignment['file_type'],
                                file_path=file_path,
                                row_index=row_index,
                                interface_id=interface_id,
                                project_id=project_id,
                                assigned_by=assigned_by,
                                assigned_to=assignment['assigned_name']
                            )
                            registry_updates += 1
                    except Exception as e:
                        print(f"[Registry] 单个任务钩子失败: {e}")

                print(f"[Registry] 共更新 {registry_updates} 个任务")
            except Exception as e:
                print(f"[Registry] 批量钩子失败: {e}")
            
        except Exception as e:
            print(f"[指派] 文件处理失败: {file_path}, 错误: {e}")
            import traceback
            traceback.print_exc()
            for assignment in file_assignments:
                failed_tasks.append({
                    'interface_id': assignment.get('interface_id', '未知'),
                    'reason': str(e)
                })

    # 【新增】保存指派记忆（只记录成功的指派）
    if success_count > 0:
        try:
            from services.assignment_memory import batch_save_memories
            # 过滤出成功的指派（不在failed_tasks中的）
            failed_interface_ids = {t.get('interface_id') for t in failed_tasks}
            successful_assignments = [
                a for a in assignments
                if a.get('interface_id') not in failed_interface_ids
            ]
            memory_count = batch_save_memories(successful_assignments)
            if memory_count > 0:
                print(f"[AssignmentMemory] 已保存 {memory_count} 条指派记忆")
        except Exception as mem_error:
            print(f"[AssignmentMemory] 保存指派记忆失败: {mem_error}")

    return {
        'success_count': success_count,
        'failed_tasks': failed_tasks,
        'registry_updates': registry_updates
    }


class AssignmentDialog(tk.Toplevel):
    """任务指派界面"""
    
    def __init__(self, parent, unassigned_tasks, name_list, user_name=None, user_roles=None):
        """
        初始化任务指派对话框
        
        参数:
            parent: 父窗口
            unassigned_tasks: 未指派任务列表
            name_list: 姓名列表（从姓名角色表读取）
            user_name: 当前用户姓名（可选）
            user_roles: 当前用户角色列表（可选）
        """
        super().__init__(parent)
        
        self.unassigned_tasks = unassigned_tasks
        self.name_list = name_list
        # 优先使用显式传入的 user_name/user_roles；
        # 若调用方未传（历史调用点/外部脚本），尝试从 parent.app 回溯（base.py 会挂载 root.app = app）。
        inferred_name = (user_name or "").strip()
        inferred_roles = list(user_roles or [])
        if not inferred_name:
            try:
                app = getattr(parent, "app", None)
                if app and getattr(app, "config", None):
                    inferred_name = str(app.config.get("user_name", "") or "").strip()
                if not inferred_roles and app and hasattr(app, "user_roles"):
                    inferred_roles = list(getattr(app, "user_roles") or [])
            except Exception:
                pass
        self.user_name = inferred_name or "未知用户"
        self.user_roles = inferred_roles
        self.assignment_entries = []  # 存储每行的输入控件
        self.batch_name_var = tk.StringVar()  # 批量指派的姓名
        self.assignment_successful = False  # 【新增】标记是否成功指派
        self.assignment_payload = []
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置界面"""
        self.title("任务指派")
        self.geometry("900x600")
        
        # 居中显示
        self.transient(self.master)
        self.grab_set()
        
        # 标题区域（包含标题和重新指派按钮）
        title_frame = ttk.Frame(self)
        title_frame.pack(fill='x', pady=10)
        
        # 标题
        title_label = ttk.Label(
            title_frame,
            text=f"任务指派（共{len(self.unassigned_tasks)}个未指派任务）",
            font=('Arial', 14, 'bold')
        )
        title_label.pack(side='left', padx=20)
        
        # 重新指派按钮（强制指派功能）
        force_assign_btn = ttk.Button(
            title_frame,
            text="重新指派",
            command=self._show_force_assign_dialog
        )
        force_assign_btn.pack(side='right', padx=20)
        
        # 创建滚动区域
        canvas_frame = ttk.Frame(self)
        canvas_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(canvas_frame)  # 使用系统默认背景色，去掉白色
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 批量指派控制栏
        batch_frame = ttk.Frame(scrollable_frame)
        batch_frame.pack(fill='x', pady=5)
        
        # 全选/全不选按钮
        ttk.Button(batch_frame, text="全选", command=self.select_all, width=8).pack(side='left', padx=5)
        ttk.Button(batch_frame, text="全不选", command=self.deselect_all, width=8).pack(side='left', padx=5)
        
        # 批量指派输入框
        ttk.Label(batch_frame, text="批量指派给：").pack(side='left', padx=5)
        batch_combo = ttk.Combobox(batch_frame, textvariable=self.batch_name_var, values=self.name_list, width=15)
        batch_combo.pack(side='left', padx=5)
        batch_combo.bind('<KeyRelease>', lambda e: self.on_batch_search(e, batch_combo))
        
        # 批量应用按钮
        ttk.Button(batch_frame, text="应用到勾选项", command=self.apply_batch_assignment, width=12).pack(side='left', padx=5)

        # 复制勾选项（便于粘贴到Excel/聊天）
        ttk.Button(batch_frame, text="复制勾选项", command=self.copy_checked_tasks, width=10).pack(side='left', padx=5)

        # Ctrl+C 复制勾选项
        self.bind("<Control-c>", lambda e: self._on_copy_shortcut())
        self.bind("<Control-C>", lambda e: self._on_copy_shortcut())
        
        # 分隔线
        ttk.Separator(scrollable_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # 表头
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill='x', pady=5)
        
        # 添加"选择"列
        ttk.Label(header_frame, text="选择", width=6, font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5)
        ttk.Label(header_frame, text="文件类型", width=12, font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5)
        ttk.Label(header_frame, text="项目号", width=10, font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5)
        ttk.Label(header_frame, text="接口号", width=25, font=('Arial', 10, 'bold')).grid(row=0, column=3, padx=5)
        ttk.Label(header_frame, text="接口时间", width=12, font=('Arial', 10, 'bold')).grid(row=0, column=4, padx=5)
        ttk.Label(header_frame, text="指派人", width=20, font=('Arial', 10, 'bold')).grid(row=0, column=5, padx=5)
        
        # 文件类型映射
        file_type_names = {
            1: "内部需打开",
            2: "内部需回复",
            3: "外部需打开",
            4: "外部需回复",
            5: "三维提资",
            6: "收发文函"
        }
        
        # 为每个未指派任务创建一行
        for i, task in enumerate(self.unassigned_tasks):
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill='x', pady=2)
            
            # 选择复选框
            checkbox_var = tk.BooleanVar(value=False)
            checkbox = ttk.Checkbutton(row_frame, variable=checkbox_var)
            checkbox.grid(row=0, column=0, padx=5)
            
            # 文件类型
            file_type_name = file_type_names.get(task['file_type'], f"文件{task['file_type']}")
            ttk.Label(row_frame, text=file_type_name, width=12).grid(row=0, column=1, padx=5)
            
            # 项目号
            ttk.Label(row_frame, text=str(task['project_id']), width=10).grid(row=0, column=2, padx=5)
            
            # 接口号（处理空值）
            interface_id = str(task.get('interface_id', ''))
            if not interface_id or interface_id == 'None' or interface_id.strip() == '':
                interface_id = "（无接口号）"
                print(f"[警告] 任务{i+1}无接口号数据: {task}")
            elif len(interface_id) > 30:
                interface_id = interface_id[:27] + "..."
            ttk.Label(row_frame, text=interface_id, width=25).grid(row=0, column=3, padx=5)
            
            # 接口时间
            ttk.Label(row_frame, text=str(task['interface_time']), width=12).grid(row=0, column=4, padx=5)
            
            # 指派人下拉框（带实时搜索）
            combobox = ttk.Combobox(row_frame, values=self.name_list, width=18)
            combobox.grid(row=0, column=5, padx=5)
            
            # 绑定实时搜索 - 在输入时自动弹出下拉菜单
            combobox.bind('<KeyRelease>', lambda event, cb=combobox: self.on_search(event, cb))
            
            # 【修复】禁用鼠标滚轮改变值的行为
            combobox.bind('<MouseWheel>', lambda e: "break")
            combobox.bind('<Button-4>', lambda e: "break")  # Linux向上滚动
            combobox.bind('<Button-5>', lambda e: "break")  # Linux向下滚动
            
            # 保存引用
            self.assignment_entries.append({
                'task': task,
                'combobox': combobox,
                'checkbox_var': checkbox_var
            })
        
        # 布局Canvas和Scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 按钮区域
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="确认指派", command=self.on_confirm).pack(side='left', padx=10)
        ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)

    def _on_copy_shortcut(self):
        self.copy_checked_tasks()
        return "break"

    def copy_checked_tasks(self):
        """
        只复制当前勾选的“接口号”（按用户要求）。
        """
        interface_ids = []
        for entry in self.assignment_entries:
            if not entry['checkbox_var'].get():
                continue
            task = entry['task']
            interface_id = normalize_interface_id(task.get('interface_id', ''))
            if interface_id and interface_id not in ("（无接口号）", "未知"):
                interface_ids.append(interface_id)

        if not interface_ids:
            messagebox.showinfo("提示", "请先勾选要复制的任务", parent=self)
            return

        text = "\n".join(interface_ids).strip()
        ok = copy_text(self, text)
        if ok:
            messagebox.showinfo("已复制", f"已复制 {len(interface_ids)} 个接口号到剪贴板", parent=self)
    
    def select_all(self):
        """全选所有任务"""
        for entry in self.assignment_entries:
            entry['checkbox_var'].set(True)
    
    def deselect_all(self):
        """全不选所有任务"""
        for entry in self.assignment_entries:
            entry['checkbox_var'].set(False)
    
    def on_batch_search(self, event, combobox):
        """批量指派输入框的实时搜索"""
        search_text = combobox.get().strip()
        
        if not search_text:
            combobox['values'] = self.name_list
            return
        
        filtered = [name for name in self.name_list if search_text in name]
        combobox['values'] = filtered
    
    def apply_batch_assignment(self):
        """将批量指派的姓名应用到所有勾选的任务"""
        batch_name = self.batch_name_var.get().strip()
        
        if not batch_name:
            messagebox.showwarning("提示", "请先输入要批量指派的姓名", parent=self)
            return
        
        # 统计勾选数量
        checked_count = sum(1 for entry in self.assignment_entries if entry['checkbox_var'].get())
        
        if checked_count == 0:
            messagebox.showwarning("提示", "请至少勾选一个任务", parent=self)
            return
        
        # 应用到勾选的任务
        for entry in self.assignment_entries:
            if entry['checkbox_var'].get():
                entry['combobox'].set(batch_name)
        
        messagebox.showinfo("成功", f"已将 \"{batch_name}\" 应用到 {checked_count} 个任务", parent=self)
    
    def on_search(self, event, combobox):
        """实时搜索回调 - 只更新下拉列表，不自动弹出"""
        search_text = combobox.get().strip()
        
        if not search_text:
            # 恢复完整列表
            combobox['values'] = self.name_list
            return
        
        # 过滤姓名列表（包含搜索文本的）
        filtered = [name for name in self.name_list if search_text in name]
        combobox['values'] = filtered
        
        # 不自动弹出下拉菜单，让用户自己点击下拉按钮
    
    def on_confirm(self):
        """确认指派按钮回调（优化版，使用批量保存）"""
        assignments = []
        
        # 构建指派人信息（姓名+角色）
        assigned_by = self.user_name
        if self.user_roles:
            # 如果有多个角色，只取第一个主要角色
            role = self.user_roles[0] if self.user_roles else ""
            assigned_by = f"{self.user_name}（{role}）"
        
        # 收集所有指派
        for entry in self.assignment_entries:
            task = entry['task']
            assigned_name = entry['combobox'].get().strip()
            
            # 跳过空值
            if not assigned_name:
                continue
            
            assignments.append({
                'file_type': task['file_type'],
                'file_path': task['file_path'],
                'row_index': task['row_index'],
                'assigned_name': assigned_name,
                'assigned_by': assigned_by,  # 【新增】指派人信息
                'interface_id': task.get('interface_id', '未知'),
                'project_id': task.get('project_id', ''),
                'status_text': '待完成',
            })
        
        if not assignments:
            messagebox.showwarning("提示", "请至少选择一个责任人进行指派", parent=self)
            return
        
        # 显示处理中提示
        processing_label = ttk.Label(self, text="正在提交写入任务...", font=('Arial', 12))
        processing_label.pack(pady=10)
        self.update()
        
        try:
            manager = get_write_task_manager()
            # 描述补全：按“指派给谁”聚合计数，便于在写入任务记录窗快速看懂
            try:
                from collections import Counter

                cnt = Counter(a.get("assigned_name", "") for a in assignments if a.get("assigned_name"))
                parts = [f"{name}({num})" for name, num in cnt.most_common()]
                summary = "，".join(parts[:6])
                if len(parts) > 6:
                    summary += "…"
                desc = f"{self.user_name} 指派 {len(assignments)} 条：{summary}" if summary else f"{self.user_name} 指派 {len(assignments)} 条"
            except Exception:
                desc = f"{self.user_name} 指派 {len(assignments)} 条"
            task = manager.submit_assignment_task(
                assignments=assignments,
                submitted_by=self.user_name,
                description=desc,
            )
            try:
                pending_cache = get_pending_cache()
                pending_cache.add_assignment_entries(task.task_id, assignments)
            except Exception as cache_error:
                print(f"[PendingCache] 记录指派任务失败: {cache_error}")
            processing_label.destroy()
            messagebox.showinfo(
                "已提交",
                f"已提交 {len(assignments)} 条指派任务。\n后台写入完成后可在写入任务记录窗查看状态。",
                parent=self,
            )
            self.assignment_successful = True
            self.assignment_payload = assignments
            self.destroy()
        except Exception as e:
            processing_label.destroy()
            messagebox.showerror("错误", f"指派任务提交失败：\n{str(e)}", parent=self)
            import traceback
            traceback.print_exc()

    def _show_force_assign_dialog(self):
        """打开强制指派弹窗"""
        dialog = ForceAssignDialog(
            self,
            self.name_list,
            self.user_name,
            self.user_roles
        )
        dialog.wait_window()


class ForceAssignDialog(tk.Toplevel):
    """强制指派弹窗 - 手动输入业务标识进行指派"""

    # 文件类型映射
    FILE_TYPE_NAMES = {
        1: "内部需打开",
        2: "内部需回复",
        3: "外部需打开",
        4: "外部需回复",
        5: "三维提资",
        6: "收发文函"
    }

    def __init__(self, parent, name_list, user_name, user_roles):
        """
        初始化强制指派对话框

        参数:
            parent: 父窗口
            name_list: 姓名列表
            user_name: 当前用户姓名
            user_roles: 当前用户角色列表
        """
        super().__init__(parent)

        self.name_list = name_list
        self.user_name = user_name
        self.user_roles = user_roles

        self.setup_ui()

    def setup_ui(self):
        """设置界面"""
        self.title("重新指派")
        self.geometry("450x350")

        # 居中显示
        self.transient(self.master)
        self.grab_set()

        # 标题
        title_label = ttk.Label(
            self,
            text="手动指定任务进行重新指派",
            font=('Arial', 12, 'bold')
        )
        title_label.pack(pady=15)

        # 表单区域
        form_frame = ttk.Frame(self)
        form_frame.pack(fill='x', padx=30, pady=10)

        # 文件类型
        ttk.Label(form_frame, text="文件类型：").grid(row=0, column=0, sticky='e', pady=8)
        self.file_type_var = tk.StringVar()
        file_type_combo = ttk.Combobox(
            form_frame,
            textvariable=self.file_type_var,
            values=[f"{k} - {v}" for k, v in self.FILE_TYPE_NAMES.items()],
            width=25,
            state='readonly'
        )
        file_type_combo.grid(row=0, column=1, sticky='w', pady=8, padx=5)
        file_type_combo.current(0)  # 默认选中第一项

        # 项目号
        ttk.Label(form_frame, text="项目号：").grid(row=1, column=0, sticky='e', pady=8)
        self.project_id_var = tk.StringVar()
        project_entry = ttk.Entry(form_frame, textvariable=self.project_id_var, width=28)
        project_entry.grid(row=1, column=1, sticky='w', pady=8, padx=5)

        # 接口号
        ttk.Label(form_frame, text="接口号：").grid(row=2, column=0, sticky='e', pady=8)
        self.interface_id_var = tk.StringVar()
        interface_entry = ttk.Entry(form_frame, textvariable=self.interface_id_var, width=28)
        interface_entry.grid(row=2, column=1, sticky='w', pady=8, padx=5)

        # 指派对象
        ttk.Label(form_frame, text="指派给：").grid(row=3, column=0, sticky='e', pady=8)
        self.assigned_name_var = tk.StringVar()
        assigned_combo = ttk.Combobox(
            form_frame,
            textvariable=self.assigned_name_var,
            values=self.name_list,
            width=25
        )
        assigned_combo.grid(row=3, column=1, sticky='w', pady=8, padx=5)
        # 绑定实时搜索
        assigned_combo.bind('<KeyRelease>', lambda e: self._on_name_search(e, assigned_combo))

        # 提示信息
        hint_label = ttk.Label(
            self,
            text="提示：将根据输入的业务标识查找任务并覆盖指派人",
            foreground='gray'
        )
        hint_label.pack(pady=10)

        # 按钮区域
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=20)

        ttk.Button(button_frame, text="确认指派", command=self._on_confirm).pack(side='left', padx=10)
        ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)

    def _on_name_search(self, event, combobox):
        """姓名下拉框实时搜索"""
        search_text = combobox.get().strip()

        if not search_text:
            combobox['values'] = self.name_list
            return

        filtered = [name for name in self.name_list if search_text in name]
        combobox['values'] = filtered

    def _on_confirm(self):
        """确认指派按钮回调"""
        # 获取输入值
        file_type_str = self.file_type_var.get()
        project_id = self.project_id_var.get().strip()
        interface_id = self.interface_id_var.get().strip()
        assigned_name = self.assigned_name_var.get().strip()

        # 验证输入
        if not file_type_str:
            messagebox.showwarning("提示", "请选择文件类型", parent=self)
            return

        if not project_id:
            messagebox.showwarning("提示", "请输入项目号", parent=self)
            return

        if not interface_id:
            messagebox.showwarning("提示", "请输入接口号", parent=self)
            return

        if not assigned_name:
            messagebox.showwarning("提示", "请选择指派对象", parent=self)
            return

        # 解析文件类型
        try:
            file_type = int(file_type_str.split(' - ')[0])
        except (ValueError, IndexError):
            messagebox.showerror("错误", "文件类型格式错误", parent=self)
            return

        # 从Registry数据库查找匹配的任务
        try:
            from registry.service import find_tasks_for_force_assign
            from registry.config import get_config

            cfg = get_config()
            db_path = cfg.get('registry_db_path')
            wal = cfg.get('registry_wal', False)

            if not db_path:
                messagebox.showerror("错误", "未配置Registry数据库路径", parent=self)
                return

            tasks = find_tasks_for_force_assign(db_path, wal, file_type, project_id, interface_id)

            if not tasks:
                messagebox.showwarning(
                    "未找到任务",
                    f"未找到匹配的任务：\n"
                    f"文件类型：{self.FILE_TYPE_NAMES.get(file_type, file_type)}\n"
                    f"项目号：{project_id}\n"
                    f"接口号：{interface_id}\n\n"
                    f"请检查输入是否正确。",
                    parent=self
                )
                return

            # 构建指派人信息
            assigned_by = self.user_name
            if self.user_roles:
                role = self.user_roles[0] if self.user_roles else ""
                assigned_by = f"{self.user_name}（{role}）"

            # 构建指派列表（可能有多个源文件中的相同任务）
            assignments = []
            for task in tasks:
                # 需要根据 source_file（文件名）找到完整的文件路径
                # 由于数据库中只存储文件名，需要从配置中获取数据文件夹路径
                source_file = task['source_file']
                data_folder = cfg.get('data_folder', '')

                # 尝试在数据文件夹中查找匹配的文件
                file_path = self._find_source_file(data_folder, source_file, file_type)

                if file_path:
                    assignments.append({
                        'file_type': file_type,
                        'file_path': file_path,
                        'row_index': task['row_index'],
                        'assigned_name': assigned_name,
                        'assigned_by': assigned_by,
                        'interface_id': interface_id,
                        'project_id': project_id,
                        'status_text': '待完成',
                    })

            if not assignments:
                messagebox.showwarning(
                    "无法定位源文件",
                    f"找到了 {len(tasks)} 个匹配的任务记录，但无法定位源Excel文件。\n"
                    f"源文件名：{tasks[0]['source_file']}\n\n"
                    f"请确保数据文件夹配置正确且文件存在。",
                    parent=self
                )
                return

            # 提交指派任务
            manager = get_write_task_manager()
            desc = f"{self.user_name} 强制指派 {interface_id} -> {assigned_name}"
            write_task = manager.submit_assignment_task(
                assignments=assignments,
                submitted_by=self.user_name,
                description=desc,
            )

            # 记录到待处理缓存
            try:
                pending_cache = get_pending_cache()
                pending_cache.add_assignment_entries(write_task.task_id, assignments)
            except Exception as cache_error:
                print(f"[PendingCache] 记录强制指派任务失败: {cache_error}")

            # 保存到指派记忆
            try:
                from services.assignment_memory import save_memory
                save_memory(file_type, project_id, interface_id, assigned_name)
            except Exception as mem_error:
                print(f"[AssignmentMemory] 保存指派记忆失败: {mem_error}")

            messagebox.showinfo(
                "已提交",
                f"已提交强制指派任务：\n"
                f"接口号：{interface_id}\n"
                f"指派给：{assigned_name}\n"
                f"涉及 {len(assignments)} 个源文件。",
                parent=self
            )
            self.destroy()

        except Exception as e:
            messagebox.showerror("错误", f"强制指派失败：\n{str(e)}", parent=self)
            import traceback
            traceback.print_exc()

    def _find_source_file(self, data_folder, source_file, file_type):
        """
        在数据文件夹中查找源文件

        参数:
            data_folder: 数据文件夹路径
            source_file: 源文件名（数据库中存储的）
            file_type: 文件类型

        返回:
            str: 完整文件路径，如果找不到返回None
        """
        if not data_folder or not os.path.exists(data_folder):
            return None

        # 文件类型对应的子文件夹名称模式
        folder_patterns = {
            1: ['待处理文件1', '内部需打开'],
            2: ['待处理文件2', '内部需回复'],
            3: ['待处理文件3', '外部需打开'],
            4: ['待处理文件4', '外部需回复'],
            5: ['待处理文件5', '三维提资'],
            6: ['待处理文件6', '收发文函'],
        }

        # 尝试在对应的子文件夹中查找
        patterns = folder_patterns.get(file_type, [])
        for pattern in patterns:
            folder_path = os.path.join(data_folder, pattern)
            if os.path.exists(folder_path):
                file_path = os.path.join(folder_path, source_file)
                if os.path.exists(file_path):
                    return file_path

        # 如果在子文件夹中找不到，尝试直接在数据文件夹中查找
        file_path = os.path.join(data_folder, source_file)
        if os.path.exists(file_path):
            return file_path

        # 递归搜索
        for root, _dirs, files in os.walk(data_folder):
            if source_file in files:
                return os.path.join(root, source_file)

        return None


def show_assignment_dialog(parent, unassigned_tasks, name_list):
    """
    显示任务指派对话框
    
    参数:
        parent: 父窗口
        unassigned_tasks: 未指派任务列表
        name_list: 姓名列表
    """
    if not unassigned_tasks:
        messagebox.showinfo("提示", "当前没有需要指派的任务", parent=parent)
        return
    
    if not name_list:
        messagebox.showwarning("警告", "无法读取姓名列表，请检查姓名角色表.xlsx", parent=parent)
        return
    
    dialog = AssignmentDialog(parent, unassigned_tasks, name_list)
    dialog.wait_window()


# 测试代码
if __name__ == "__main__":
    # 测试get_responsible_column
    print("测试get_responsible_column:")
    for i in range(1, 7):
        col = get_responsible_column(i)
        print(f"  文件{i}: {col}")
    
    # 测试get_name_list
    print("\n测试get_name_list:")
    names = get_name_list()
    print(f"  读取到 {len(names)} 个姓名")
    if names:
        print(f"  前5个: {names[:5]}")
    
    # 测试角色判断
    print("\n测试角色判断:")
    test_roles = [
        ['一室主任'],
        ['2016接口工程师'],
        ['设计人员'],
        ['所长']
    ]
    
    for roles in test_roles:
        print(f"  角色: {roles}")
        print(f"    是接口工程师: {is_interface_engineer(roles)}")
        print(f"    是室主任: {is_director(roles)}")
        if is_director(roles):
            print(f"    科室: {get_department(roles)}")
        if is_interface_engineer(roles):
            print(f"    项目号: {parse_interface_engineer_project(roles)}")

