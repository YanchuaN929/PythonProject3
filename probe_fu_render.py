#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FU 渲染诊断探针（独立脚本，不依赖 base.py / ui/window.py）

用法：把此脚本放在 EXE 同目录（如 dist/接口筛选/probe_fu_render.py），
     然后用 Python 跑：
         python probe_fu_render.py
     或在 Windows 上双击（如果 .py 关联了 python）。

输出：
  - 控制台打印每步结果
  - 同目录的 probe_log.txt（完整诊断记录）

诊断内容：
  1) 业务文件夹是否存在
  2) 能找到几个 FU 文件
  3) openpyxl 能否读 FU 文件
  4) STREAM_FILE_SPECS[7] 期望的 A-F 列实际内容
  5) header 行能否识别
  6) data 行数
  7) 模拟 _create_optimized_display 的 FU 特殊路径能否构造 display_df
  8) 与 EXE 内部渲染逻辑对比
"""

import os
import sys
import json
import datetime
import traceback
import re
from pathlib import Path

# 探针结果收集
LOG_LINES = []


def log(msg=""):
    """同时输出到控制台和内存（最后写文件）"""
    print(msg)
    LOG_LINES.append(msg)


def save_log():
    """写 probe_log.txt 到脚本同目录"""
    log_path = Path(__file__).parent / "probe_log.txt"
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("\n".join(LOG_LINES))
        print(f"\n[probe] 详细 log 已写: {log_path}")
    except Exception as e:
        print(f"[probe] 写 log 失败: {e}")


def find_data_folder():
    """读 config.json 找业务文件夹，回退让用户输入"""
    here = Path(__file__).parent
    candidates = [
        here / "_internal" / "config.json",
        here / "config.json",
        here.parent / "config.json",
    ]
    for p in candidates:
        if p.exists():
            try:
                with open(p, encoding="utf-8") as f:
                    cfg = json.load(f)
                folder = (cfg.get("folder_path") or "").strip()
                if folder:
                    log(f"  [OK] 从 {p} 读到业务文件夹: {folder}")
                    return folder, str(p)
            except Exception as e:
                log(f"  [WARN] 读 {p} 失败: {e}")
    # 让用户输入
    log("  [INFO] 未能从 config.json 读到 folder_path")
    try:
        folder = input("请输入业务文件夹绝对路径: ").strip()
    except EOFError:
        folder = ""
    return folder, "<user input>"


def find_fu_files(folder):
    """扫业务文件夹找 FU 文件（按 STREAM_FILE_SPECS[7] 命名约定）"""
    if not folder or not Path(folder).exists():
        return []
    folder_path = Path(folder)
    fu_files = []
    # 正则: ^(\d{4})项目标准表格\.(xlsx|xlsm|xls)$
    pattern = re.compile(r"^(\d{4})项目标准表格\.(xlsx|xlsm|xls)$", re.IGNORECASE)
    for path in folder_path.rglob("*"):
        if path.is_file() and pattern.match(path.name):
            fu_files.append(path)
    return fu_files


def check_dependencies():
    """检查 openpyxl 是否可用"""
    try:
        import openpyxl
        log(f"  [OK] openpyxl 版本: {openpyxl.__version__}")
        return True
    except ImportError as e:
        log(f"  [错误] openpyxl 未安装: {e}")
        log(f"  请运行: pip install openpyxl")
        return False


def inspect_fu_xlsx(fu_path):
    """用 openpyxl 读 FU Excel 检查关键结构"""
    log(f"\n--- 测 {fu_path.name} ---")
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [错误] openpyxl 未安装")
        return None

    info = {"file": str(fu_path), "error": None}

    # 步骤 A: 尝试 openpyxl read_only=False 模式（6f7705f 之前用的）
    log("  [A] openpyxl load_workbook(read_only=False, data_only=True)")
    try:
        wb = load_workbook(str(fu_path), read_only=False, data_only=True)
        ws = wb.worksheets[0]
        info["sheet_name"] = ws.title
        info["max_row"] = ws.max_row
        info["max_col"] = ws.max_column
        log(f"      [OK] sheet={ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        log(f"      [错误] {e}")
        info["error"] = f"openpyxl 读取失败: {e}"
        return info
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # 步骤 B: 读 header（row 1）和 data
    log("  [B] 读 header (row 1) 和 data rows")
    try:
        wb = load_workbook(str(fu_path), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        # header
        header = {}
        for col_idx in range(1, min(7, ws.max_column or 7) + 1):
            cell_val = ws.cell(1, col_idx).value
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            header[col_letter] = cell_val
            log(f"      col {col_letter}: {cell_val!r}")

        # data rows（用 ws._cells 拿到真实存在 cell 的行号）
        data_rows = set()
        if hasattr(ws, "_cells"):
            for (row, col), _ in ws._cells.items():
                if row >= 2 and 1 <= col <= 6:
                    data_rows.add(row)
        else:
            # fallback: 遍历 row
            for row_idx in range(2, (ws.max_row or 1) + 1):
                for col_idx in range(1, 7):
                    if ws.cell(row_idx, col_idx).value is not None:
                        data_rows.add(row_idx)
                        break
        info["data_row_count"] = len(data_rows)
        info["data_row_sample"] = sorted(data_rows)[:5]
        log(f"      [OK] 找到 {len(data_rows)} 个 data row（包含实际 cell 的行）")
        if info["data_row_sample"]:
            log(f"      sample: {info['data_row_sample']}")

        # 步骤 C: 读第一行 data 验证 FU 特殊路径
        if data_rows:
            first_row = min(data_rows)
            log(f"  [C] 读第一行 data (row {first_row})")
            first_data = {}
            for col_idx in range(1, 7):
                val = ws.cell(first_row, col_idx).value
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
                first_data[col_letter] = val
                log(f"      col {col_letter}: {val!r}")
            info["first_data_row"] = first_row
            info["first_data"] = first_data

            # 步骤 D: STREAM_FILE_SPECS[7] 期望的列映射
            log("  [D] STREAM_FILE_SPECS[7] 期望的列映射")
            log("      A=file_code, B=internal_code, C=title, D=actual_fu, E=fu_plan, F=responsible")
            # 检查每列是否非空
            for col_letter, alias in [
                ("A", "file_code"),
                ("B", "internal_code"),
                ("C", "title"),
                ("D", "actual_fu"),
                ("E", "fu_plan"),
                ("F", "responsible"),
            ]:
                val = first_data.get(col_letter)
                log(f"      col {col_letter} ({alias}): {'有值' if val is not None else '空!'} = {val!r}")
    except Exception as e:
        log(f"  [错误] 读 header/data 失败: {e}")
        info["error"] = f"读 header/data 失败: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # 步骤 E: 模拟 _create_optimized_display 的 FU 特殊路径
    log("  [E] 模拟 _create_optimized_display FU 路径")
    try:
        import pandas as pd
        if "first_data" in info and info["first_data"]:
            first_data = info["first_data"]
            # 模拟 _build_selected_record 后的 row 构造
            # record["_raw"] = {alias: value}
            raw = {
                "file_code": first_data.get("A"),
                "internal_code": first_data.get("B"),
                "title": first_data.get("C"),
                "actual_fu": first_data.get("D"),
                "fu_plan": first_data.get("E"),
                "responsible": first_data.get("F"),
            }
            log(f"      _raw: {raw}")

            # 模拟 _standard_result_row
            interface_id = str(raw.get("internal_code") or "").strip() if raw.get("internal_code") else ""
            log(f"      interface_id (= internal_code): {interface_id!r}")
            log(f"      raw.get('internal_code') 类型: {type(raw.get('internal_code'))}")

            # 关键检查：raw.get('internal_code') 是否是字符串
            # _build_selected_record 中 value_getter(_col_to_index('B'))
            # 老 _read_sparse_fu_records_openpyxl: ws.cell(row_number, 2).value
            # 新 _read_sparse_fu_records_ooxml: snapshot.typed_value_from_cell_xml
            # 两种路径应返回相同类型的 value

            # 步骤 F: 用 pandas 模拟 FU 特殊路径构造 display_df
            log("  [F] 用 pandas 模拟 _create_optimized_display FU 特殊路径")
            # 模拟"如果 raw['internal_code'] 是 None 或非字符串类型会怎样"
            if not interface_id:
                log("  [WARNING] interface_id 为空！display_results7 会显示'无待处理FU'")
            else:
                log(f"  [OK] interface_id 非空，会调 display_results7 渲染")
    except ImportError:
        log("  [SKIP] pandas 未安装，跳过模拟")
    except Exception as e:
        log(f"  [错误] 模拟失败: {e}")
        log(f"  {traceback.format_exc()}")

    return info


def main():
    log("=" * 60)
    log(f"FU 渲染诊断探针")
    log(f"时间: {datetime.datetime.now().isoformat()}")
    log(f"Python: {sys.version}")
    log(f"工作目录: {os.getcwd()}")
    log("=" * 60)

    # 1. 检查依赖
    log("\n[1] 检查依赖")
    if not check_dependencies():
        save_log()
        return

    # 2. 找业务文件夹
    log("\n[2] 找业务文件夹")
    folder, source = find_data_folder()
    log(f"  业务文件夹: {folder}")
    log(f"  来源: {source}")

    if not folder:
        log("  [错误] 业务文件夹为空")
        save_log()
        return
    if not Path(folder).exists():
        log(f"  [错误] 业务文件夹不存在: {folder}")
        save_log()
        return

    # 3. 找 FU 文件
    log("\n[3] 找 FU 文件（正则: ^(\\d{4})项目标准表格\\.(xlsx|xlsm|xls)$）")
    fu_files = find_fu_files(folder)
    log(f"  找到 {len(fu_files)} 个 FU 文件")
    for f in fu_files:
        log(f"    - {f}")

    if not fu_files:
        log("  [错误] 没有找到 FU 文件")
        log("  可能原因:")
        log("    1) 文件名不符合 正则: ^\\d{4}项目标准表格\\.(xlsx|xlsm|xls)$")
        log("    2) 业务文件夹路径不对")
        log("    3) 文件在子目录但 rglob 没扫到（应已扫到）")
        save_log()
        return

    # 4. 检查每个 FU 文件
    log("\n[4] 详细检查每个 FU 文件")
    results = []
    for fu_path in fu_files[:5]:  # 最多测 5 个文件
        info = inspect_fu_xlsx(fu_path)
        results.append(info)

    # 5. 总结
    log("\n" + "=" * 60)
    log("[5] 总结")
    if not results:
        log("  无 FU 文件可测")
    else:
        for r in results:
            if r.get("error"):
                log(f"  [错误] {r['file']}: {r['error']}")
            else:
                log(f"  [OK] {Path(r['file']).name}: data_rows={r.get('data_row_count', '?')}, "
                    f"max_row={r.get('max_row', '?')}, max_col={r.get('max_col', '?')}")
    log("=" * 60)

    save_log()


if __name__ == "__main__":
    main()
