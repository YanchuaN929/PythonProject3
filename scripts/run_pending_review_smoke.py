#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Dict, List, Optional

import tkinter.messagebox as messagebox
from openpyxl import load_workbook

NOW = datetime(2026, 4, 24, 9, 0, 0)
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
EXAMPLE_DIR = ROOT / "example"
RUNTIME_DIR = ROOT / ".smoke_runtime" / "pending_review"

from base import ExcelProcessorApp
from core.main import (
    process_target_file,
    process_target_file2,
    process_target_file3,
    process_target_file4,
    process_target_file6,
)
from registry import hooks as registry_hooks
from registry.service import resolve_task_record
from registry.util import extract_interface_id, make_business_id, make_task_id
from ui.input_handler import InterfaceInputDialog, get_write_columns, write_response_to_excel


@dataclass
class SmokeCase:
    name: str
    file_type: int
    source_name: str
    project_id: str
    process_fn: Callable[[str], object]
    response_number: str
    designer_name: str
    superior_role: str
    superior_name: str
    target_row: int
    prepare_copy: Optional[Callable[[Path], None]] = None
    source_column: Optional[str] = None
    extra_verify: Optional[Callable[[Path, int], Dict[str, object]]] = None


def _noop_messagebox(*_args, **_kwargs):
    return None


messagebox.showerror = _noop_messagebox
messagebox.showwarning = _noop_messagebox
messagebox.showinfo = _noop_messagebox


def _assert(condition: bool, message: str):
    if not condition:
        raise AssertionError(message)


def _edit_workbook(path: Path, updates: Dict[str, object]):
    wb = load_workbook(path)
    try:
        ws = wb.active
        for cell, value in updates.items():
            ws[cell] = value
        wb.save(path)
    finally:
        wb.close()


def _prepare_file3_m(path: Path):
    _edit_workbook(
        path,
        {
            "M1069": datetime(2026, 5, 1),
            "AP1069": "张亮@zhangliangc",
        },
    )


def _prepare_file3_l(path: Path):
    _edit_workbook(
        path,
        {
            "L2067": datetime(2026, 5, 1),
            "AP2067": "张亮@zhangliangc",
        },
    )


def _prepare_file4(path: Path):
    _edit_workbook(
        path,
        {
            "S9506": datetime(2026, 5, 1),
        },
    )


def _verify_file6_extra(path: Path, row_index: int) -> Dict[str, object]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        expected_due = ws[f"I{row_index}"].value
        reply_status = ws[f"M{row_index}"].value
    finally:
        wb.close()

    if isinstance(expected_due, datetime):
        expected_due = expected_due.date()
    elif hasattr(expected_due, "year") and hasattr(expected_due, "month") and hasattr(expected_due, "day"):
        expected_due = date(expected_due.year, expected_due.month, expected_due.day)
    else:
        expected_due = None

    expected_status = None
    if expected_due:
        expected_status = "按时回复" if date.today() <= expected_due else "延期回复"
    return {
        "expected_reply_status": expected_status,
        "actual_reply_status": reply_status,
    }


def _make_view_df(df, source_file: Path, file_type: int, project_id: str, user_name: str, user_roles: List[str]):
    app = ExcelProcessorApp.__new__(ExcelProcessorApp)
    app.user_name = user_name
    app.user_roles = list(user_roles)
    app.user_role = user_roles[0] if user_roles else ""
    app.config = {"role_export_days": {}}
    app.auto_mode = False
    app._apply_overdue_filter = lambda frame, _file_type: frame

    role_df = ExcelProcessorApp.apply_role_based_filter(app, df.copy(), project_id=project_id)
    return ExcelProcessorApp._exclude_pending_confirmation_rows(
        app,
        role_df,
        str(source_file),
        file_type,
        project_id,
    )


def _load_existing_response(file_path: Path, file_type: int, row_index: int, interface_id: str, project_id: str):
    class DummyDialog:
        pass

    dialog = DummyDialog()
    dialog.file_path = str(file_path)
    dialog.file_type = file_type
    dialog.row_index = row_index
    dialog.interface_id = interface_id
    dialog.project_id = project_id
    dialog.existing_response = None
    dialog.completed_info = None
    InterfaceInputDialog._load_existing_response(dialog)
    return dialog.existing_response, dialog.completed_info


def _get_db_path() -> Path:
    cfg = registry_hooks._cfg()
    return Path(cfg["registry_db_path"])


def _query_exact_and_business_rows(key: Dict[str, object]) -> Dict[str, object]:
    db_path = _get_db_path()
    business_id = make_business_id(key["file_type"], key["project_id"], key["interface_id"])
    exact_tid = make_task_id(
        key["file_type"],
        key["project_id"],
        key["interface_id"],
        key["source_file"],
        key["row_index"],
    )

    conn = sqlite3.connect(str(db_path))
    try:
        exact_row = conn.execute(
            """
            SELECT id, status, display_status, response_number, confirmed_by, confirmed_at, archived_at
            FROM tasks
            WHERE id = ?
            """,
            (exact_tid,),
        ).fetchone()
        business_rows = conn.execute(
            """
            SELECT id, status, display_status, response_number, archived_at
            FROM tasks
            WHERE business_id = ?
            ORDER BY last_seen_at DESC, rowid DESC
            """,
            (business_id,),
        ).fetchall()
    finally:
        conn.close()

    latest_row = resolve_task_record(str(db_path), False, key)
    return {
        "exact_tid": exact_tid,
        "business_id": business_id,
        "exact_row": exact_row,
        "business_rows": business_rows,
        "latest_row": latest_row,
    }


def _read_written_cells(path: Path, columns: Dict[str, str], row_index: int) -> Dict[str, object]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        return {
            "response": ws[f"{columns['response_col']}{row_index}"].value,
            "time": ws[f"{columns['time_col']}{row_index}"].value,
            "name": ws[f"{columns['name_col']}{row_index}"].value,
        }
    finally:
        wb.close()


def _process_file(case: SmokeCase, path: Path):
    return case.process_fn(str(path))


def _find_result_row(result_df, target_row: int):
    matched = result_df[result_df["原始行号"] == target_row]
    _assert(not matched.empty, f"未找到目标行: {target_row}")
    return matched.iloc[0]


def _row_numbers(frame) -> set:
    if frame is None or getattr(frame, "empty", True):
        return set()
    if "原始行号" not in frame.columns:
        return set()
    return set(frame["原始行号"].tolist())


def _run_single_case(case: SmokeCase) -> Dict[str, object]:
    case_dir = RUNTIME_DIR / case.name
    if case_dir.exists():
        shutil.rmtree(case_dir)
    case_dir.mkdir(parents=True, exist_ok=True)

    source_path = EXAMPLE_DIR / case.source_name
    work_path = case_dir / source_path.name
    shutil.copy2(source_path, work_path)
    if case.prepare_copy:
        case.prepare_copy(work_path)

    registry_hooks.set_data_folder(str(case_dir))

    initial_df = _process_file(case, work_path)
    registry_hooks.on_process_done(case.file_type, case.project_id, str(work_path), initial_df, now=NOW)
    initial_row = _find_result_row(initial_df, case.target_row)
    interface_id = extract_interface_id(initial_row, case.file_type)
    key = {
        "file_type": case.file_type,
        "project_id": case.project_id,
        "interface_id": interface_id,
        "source_file": work_path.name,
        "row_index": case.target_row,
        "interface_time": str(initial_row.get("接口时间", "") or ""),
    }

    columns = get_write_columns(case.file_type, case.target_row, None, case.source_column)
    _assert(columns is not None, f"{case.name}: 无法解析写回列")

    write_ok = write_response_to_excel(
        str(work_path),
        case.file_type,
        case.target_row,
        case.response_number,
        case.designer_name,
        case.project_id,
        source_column=case.source_column,
        interface_id=interface_id,
    )
    _assert(write_ok, f"{case.name}: Excel 写回失败")

    registry_hooks.on_response_written(
        case.file_type,
        str(work_path),
        case.target_row,
        interface_id,
        case.response_number,
        case.designer_name,
        case.project_id,
        source_column=case.source_column,
        role="设计人员",
        now=NOW,
    )

    written_cells = _read_written_cells(work_path, columns, case.target_row)
    _assert(str(written_cells["response"]).strip() == case.response_number, f"{case.name}: 回文单号写入错误")
    _assert(str(written_cells["name"]).strip() == case.designer_name, f"{case.name}: 写回人列错误")
    _assert(str(written_cells["time"]).strip().startswith(date.today().isoformat()), f"{case.name}: 回文时间列错误")

    existing_response, completed_info = _load_existing_response(
        work_path,
        case.file_type,
        case.target_row,
        interface_id,
        case.project_id,
    )
    _assert(existing_response == case.response_number, f"{case.name}: 再次点击未进入已填写只读态")
    _assert(completed_info and completed_info.get("completed_by") == case.designer_name, f"{case.name}: 已填写信息缺少 completed_by")

    after_response_df = _process_file(case, work_path)
    _assert(case.target_row in set(after_response_df["原始行号"].tolist()), f"{case.name}: 待审查任务未被 Registry 加回")

    designer_view = _make_view_df(
        after_response_df,
        work_path,
        case.file_type,
        case.project_id,
        case.designer_name,
        ["设计人员"],
    )
    superior_view = _make_view_df(
        after_response_df,
        work_path,
        case.file_type,
        case.project_id,
        case.superior_name,
        [case.superior_role],
    )
    _assert(case.target_row not in _row_numbers(designer_view), f"{case.name}: 设计人员写回后仍看到提醒")
    _assert(case.target_row in _row_numbers(superior_view), f"{case.name}: 上级未看到待审查任务")

    status_after_response = registry_hooks.get_display_status([key], "设计人员")
    exact_after_response = _query_exact_and_business_rows(key)
    latest_after_response = exact_after_response["latest_row"]
    _assert(latest_after_response is not None, f"{case.name}: Registry 未找到最新任务")
    _assert(latest_after_response["status"] == "completed", f"{case.name}: 写回后状态不是 completed")
    _assert(latest_after_response["display_status"] == "待审查", f"{case.name}: 写回后 display_status 不是 待审查")
    non_archived_after_response = [row for row in exact_after_response["business_rows"] if row[4] is None]
    _assert(len(non_archived_after_response) == 1, f"{case.name}: business_id 出现非归档冲突记录")
    _assert(next(iter(status_after_response.values())).endswith("待审查"), f"{case.name}: 设计人员状态文本异常")

    registry_hooks.on_confirmed_by_superior(
        case.file_type,
        str(work_path),
        case.target_row,
        case.superior_name,
        case.project_id,
        interface_id=interface_id,
        role=case.superior_role,
        now=NOW + timedelta(minutes=5),
    )

    after_confirm_df = _process_file(case, work_path)
    designer_view_after_confirm = _make_view_df(
        after_confirm_df,
        work_path,
        case.file_type,
        case.project_id,
        case.designer_name,
        ["设计人员"],
    )
    superior_view_after_confirm = _make_view_df(
        after_confirm_df,
        work_path,
        case.file_type,
        case.project_id,
        case.superior_name,
        [case.superior_role],
    )
    _assert(case.target_row not in _row_numbers(designer_view_after_confirm), f"{case.name}: 确认后设计人员仍看到任务")
    _assert(case.target_row not in _row_numbers(superior_view_after_confirm), f"{case.name}: 勾选确认后上级列表未隐藏")

    snapshot_after_confirm = registry_hooks.get_task_snapshot(key)
    _assert(snapshot_after_confirm is not None, f"{case.name}: 确认后读取快照失败")
    _assert(snapshot_after_confirm["status"] == "confirmed", f"{case.name}: 确认后状态不是 confirmed")
    _assert(snapshot_after_confirm["display_status"] == "已审查", f"{case.name}: 确认后 display_status 不是 已审查")
    _assert(snapshot_after_confirm["confirmed_by"] == case.superior_name, f"{case.name}: confirmed_by 未写入")
    _assert(snapshot_after_confirm.get("archived_at") in (None, ""), f"{case.name}: 确认后不应立即归档")

    registry_hooks.on_unconfirmed_by_superior(key, user_name=case.superior_name)

    after_unconfirm_df = _process_file(case, work_path)
    superior_view_after_unconfirm = _make_view_df(
        after_unconfirm_df,
        work_path,
        case.file_type,
        case.project_id,
        case.superior_name,
        [case.superior_role],
    )
    _assert(case.target_row in _row_numbers(superior_view_after_unconfirm), f"{case.name}: 取消确认后任务未恢复待审查")

    snapshot_after_unconfirm = registry_hooks.get_task_snapshot(key)
    _assert(snapshot_after_unconfirm is not None, f"{case.name}: 取消确认后读取快照失败")
    _assert(snapshot_after_unconfirm["status"] == "completed", f"{case.name}: 取消确认后状态不是 completed")
    _assert(snapshot_after_unconfirm["display_status"] == "待审查", f"{case.name}: 取消确认后 display_status 不是 待审查")
    _assert(snapshot_after_unconfirm.get("confirmed_by") in (None, ""), f"{case.name}: 取消确认后 confirmed_by 未清空")

    reconfirm_time = NOW + timedelta(minutes=10)
    registry_hooks.on_confirmed_by_superior(
        case.file_type,
        str(work_path),
        case.target_row,
        case.superior_name,
        case.project_id,
        interface_id=interface_id,
        role=case.superior_role,
        now=reconfirm_time,
    )
    finalize_time = reconfirm_time + timedelta(days=8)
    registry_hooks.on_scan_finalize(f"smoke-{case.name}", now=finalize_time)

    archived_snapshot = registry_hooks.get_task_snapshot(key)
    _assert(archived_snapshot is not None, f"{case.name}: 归档后读取快照失败")
    _assert(archived_snapshot["status"] == "archived", f"{case.name}: 7天后未归档")

    extra_result = case.extra_verify(work_path, case.target_row) if case.extra_verify else {}
    if extra_result.get("expected_reply_status") is not None:
        _assert(
            extra_result["actual_reply_status"] == extra_result["expected_reply_status"],
            f"{case.name}: 文件6 M列回文状态错误",
        )

    return {
        "case": case.name,
        "row_index": case.target_row,
        "interface_id": interface_id,
        "response_cells": columns,
        "written_cells": {
            "response": str(written_cells["response"]),
            "time": str(written_cells["time"]),
            "name": str(written_cells["name"]),
        },
        "designer_hidden_after_response": True,
        "superior_visible_before_confirm": True,
        "superior_hidden_after_confirm": True,
        "restored_after_unconfirm": True,
        "archived_after_8_days": True,
        "registry_exact_tid": exact_after_response["exact_tid"],
        "registry_business_id": exact_after_response["business_id"],
        "source_column": case.source_column,
        "extra": extra_result,
    }


def main():
    if RUNTIME_DIR.exists():
        shutil.rmtree(RUNTIME_DIR)
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)

    cases = [
        SmokeCase(
            name="file1",
            file_type=1,
            source_name="1818按项目导出IDI手册2026-01-28-15_11_50.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file(path, NOW),
            response_number="SMK-F1-001",
            designer_name="申浩",
            superior_role="二室主任",
            superior_name="二室主任测试",
            target_row=23,
        ),
        SmokeCase(
            name="file2",
            file_type=2,
            source_name="内部接口信息单报表181820260128.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file2(path, NOW, "1818"),
            response_number="SMK-F2-001",
            designer_name="申浩",
            superior_role="二室主任",
            superior_name="二室主任测试",
            target_row=15558,
        ),
        SmokeCase(
            name="file3_m",
            file_type=3,
            source_name="外部接口ICM报表181820260128.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file3(path, NOW),
            response_number="SMK-F3M-001",
            designer_name="张亮",
            superior_role="一室主任",
            superior_name="一室主任测试",
            target_row=1069,
            prepare_copy=_prepare_file3_m,
            source_column="M",
        ),
        SmokeCase(
            name="file3_l",
            file_type=3,
            source_name="外部接口ICM报表181820260128.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file3(path, NOW),
            response_number="SMK-F3L-001",
            designer_name="张亮",
            superior_role="一室主任",
            superior_name="一室主任测试",
            target_row=2067,
            prepare_copy=_prepare_file3_l,
            source_column="L",
        ),
        SmokeCase(
            name="file4",
            file_type=4,
            source_name="外部接口单报表181820260128.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file4(path, NOW),
            response_number="SMK-F4-001",
            designer_name="周立欣",
            superior_role="一室主任",
            superior_name="一室主任测试",
            target_row=9506,
            prepare_copy=_prepare_file4,
        ),
        SmokeCase(
            name="file6",
            file_type=6,
            source_name="收发文清单1818.xlsx",
            project_id="1818",
            process_fn=lambda path: process_target_file6(path, NOW),
            response_number="SMK-F6-001",
            designer_name="张亮",
            superior_role="一室主任",
            superior_name="一室主任测试",
            target_row=470,
            extra_verify=_verify_file6_extra,
        ),
    ]

    requested = {arg.strip() for arg in sys.argv[1:] if arg.strip()}
    if requested:
        cases = [case for case in cases if case.name in requested]
        _assert(bool(cases), f"未找到指定 case: {sorted(requested)}")

    results = []
    for case in cases:
        print(f"[SMOKE] running {case.name}")
        results.append(_run_single_case(case))

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
