"""Generate example template xlsx files for file types 1/2/3/4/6."""

from __future__ import annotations

from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
EXAMPLE_DIR = REPO_ROOT / "example"


def _init_sheet(workbook: Workbook, title: str, max_cols: int) -> None:
    ws = workbook.active
    ws.title = title

    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=letter)
        ws.column_dimensions[letter].width = 14


def _write_row(ws, row_num: int, value_map: Dict[int, str]) -> None:
    for col_idx, value in value_map.items():
        ws.cell(row=row_num, column=col_idx + 1, value=value)


def build_file1_template() -> None:
    workbook = Workbook()
    _init_sheet(workbook, "文件1模板", 18)  # A-R
    ws = workbook.active

    # 关键列说明行
    _write_row(
        ws,
        2,
        {
            0: "接口号(A)",
            7: "科室编码(H)",
            10: "接口时间(K)",
            12: "完成列(M)",
            17: "责任人(R)",
        },
    )

    rows = [
        {0: "F1-001", 7: "25C1", 10: "2026.02.20", 12: "", 17: "张三"},
        {0: "F1-002", 7: "25C2", 10: "2026.12.30", 12: "", 17: "李四"},
        {0: "F1-003", 7: "25C3", 10: "2027.01.02", 12: "1125-001", 17: "王五"},
        {0: "F1-004", 7: "25C1", 10: "2026.02.01", 1: "作废", 12: "", 17: ""},
        {0: "F1-005", 7: "25C2", 10: "2026.03.05", 12: "", 17: "刘峰a"},
    ]
    for offset, row_map in enumerate(rows, start=3):
        _write_row(ws, offset, row_map)

    workbook.save(EXAMPLE_DIR / "待处理文件1_模板.xlsx")


def build_file2_template() -> None:
    workbook = Workbook()
    _init_sheet(workbook, "文件2模板", 39)  # A-AM
    ws = workbook.active
    _write_row(
        ws,
        2,
        {
            4: "版次(E)",
            8: "科室(I)",
            12: "接口时间(M)",
            13: "完成列(N)",
            17: "接口号(R)",
            27: "AB排除列",
            38: "责任人(AM)",
        },
    )
    rows = [
        {
            4: "A",
            8: "河北分公司-建筑结构所-结构一室",
            12: "2026.02.21",
            13: "",
            17: "F2-001",
            38: "张三",
        },
        {
            4: "B",
            8: "河北分公司-建筑结构所-结构二室",
            12: "2026.02.22",
            13: "",
            17: "F2-001",
            38: "张三,李四",
        },
        {
            4: "A",
            8: "25C3",
            12: "2027.01.03",
            13: "已回复",
            17: "F2-002",
            38: "",
        },
        {
            4: "C",
            8: "河北分公司-建筑结构所",
            12: "2026.03.01",
            13: "",
            17: "F2-003",
            5: "传递",
            27: "4444XXX",
            38: "王五a",
        },
    ]
    for offset, row_map in enumerate(rows, start=3):
        _write_row(ws, offset, row_map)
    workbook.save(EXAMPLE_DIR / "待处理文件2_模板.xlsx")


def build_file3_template() -> None:
    workbook = Workbook()
    _init_sheet(workbook, "文件3模板", 42)  # A-AP
    ws = workbook.active
    _write_row(
        ws,
        2,
        {
            2: "接口号(C)",
            8: "I列(B)",
            11: "时间L",
            12: "时间M",
            16: "完成Q",
            19: "完成T",
            28: "版次AC",
            37: "机构AL",
            40: "科室AO",
            41: "责任人AP",
        },
    )
    rows = [
        {
            2: "F3-001(设计人员)",
            8: "B",
            12: "2026.02.25",
            19: "",
            28: "A",
            37: "河北分公司-建筑结构所",
            40: "结构一室",
            41: "张三",
        },
        {
            2: "F3-001(设计人员)",
            8: "B",
            12: "2026.02.25",
            19: "",
            28: "B",
            37: "河北分公司-建筑结构所",
            40: "结构一室",
            41: "张三,李四",
        },
        {
            2: "F3-002",
            8: "B",
            11: "2026.03.02",
            16: "",
            28: "A",
            37: "河北分公司-建筑结构所",
            40: "",
            41: "请确认",
        },
        {
            2: "F3-003",
            8: "B",
            11: "2027.01.02",
            16: "已完成",
            28: "C",
            37: "河北分公司-建筑结构所",
            40: "结构二室",
            41: "王五a",
        },
    ]
    for offset, row_map in enumerate(rows, start=3):
        _write_row(ws, offset, row_map)
    workbook.save(EXAMPLE_DIR / "待处理文件3_模板.xlsx")


def build_file4_template() -> None:
    workbook = Workbook()
    _init_sheet(workbook, "文件4模板", 34)  # A-AH
    ws = workbook.active
    _write_row(
        ws,
        2,
        {
            4: "接口号(E)",
            8: "版次(I)",
            15: "P列(B)",
            18: "接口时间(S)",
            21: "完成(V)",
            31: "机构AF",
            32: "科室AG",
            33: "责任人(AH)",
        },
    )
    rows = [
        {
            4: "F4-001",
            8: "A",
            15: "B",
            18: "2026.02.24",
            21: "",
            31: "河北分公司-建筑结构所",
            32: "结构一室",
            33: "张三",
        },
        {
            4: "F4-001",
            8: "B",
            15: "",
            28: "B",
            18: "2026.02.24",
            21: "",
            31: "河北分公司-建筑结构所",
            32: "结构一室",
            33: "张三,李四",
        },
        {
            4: "F4-002",
            8: "A",
            15: "B",
            18: "2027.01.03",
            21: "完成",
            31: "河北分公司-建筑结构所",
            32: "",
            33: "",
        },
    ]
    for offset, row_map in enumerate(rows, start=3):
        _write_row(ws, offset, row_map)
    workbook.save(EXAMPLE_DIR / "待处理文件4_模板.xlsx")


def build_file6_template() -> None:
    workbook = Workbook()
    _init_sheet(workbook, "文件6模板", 29)  # A-AC
    ws = workbook.active
    _write_row(
        ws,
        2,
        {
            4: "接口号(E)",
            8: "接口时间(I)",
            9: "完成(J)",
            12: "状态(M)",
            21: "机构V",
            22: "主办室W",
            23: "责任人X",
            28: "版次AC",
        },
    )
    rows = [
        {
            4: "F6-001",
            8: "2026.02.20",
            9: "",
            12: "尚未回复",
            21: "河北分公司.建筑结构所",
            22: "结构一室",
            23: "张三,李四",
            28: "A",
        },
        {
            4: "F6-001",
            8: "2026.02.20",
            9: "",
            12: "尚未回复",
            21: "河北分公司.建筑结构所",
            22: "结构一室",
            23: "张三、王五a",
            28: "B",
        },
        {
            4: "F6-002",
            8: "2027.01.05",
            9: "",
            12: "超期未回复",
            21: "河北分公司.建筑结构所",
            22: "建筑总图室",
            23: "",
            28: "A",
        },
    ]
    for offset, row_map in enumerate(rows, start=3):
        _write_row(ws, offset, row_map)
    workbook.save(EXAMPLE_DIR / "待处理文件6_模板.xlsx")


def main() -> None:
    EXAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    build_file1_template()
    build_file2_template()
    build_file3_template()
    build_file4_template()
    build_file6_template()
    print(f"模板生成完成: {EXAMPLE_DIR}")


if __name__ == "__main__":
    main()
