#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回文单号输入处理模块
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import os

# 导入Registry模块
try:
    from registry import hooks as registry_hooks
except ImportError:
    print("警告: 未找到registry模块")
    registry_hooks = None


class InterfaceInputDialog(tk.Toplevel):
    """回文单号输入弹窗"""
    
    def __init__(self, parent, interface_id, file_type, file_path, row_index, 
                 user_name, project_id, source_column=None, file_manager=None, 
                 viewer=None, item_id=None, columns=None):
        """
        参数:
            parent: 父窗口
            interface_id: 接口号
            file_type: 文件类型(1-6)
            file_path: 原始Excel文件路径
            row_index: Excel行号
            user_name: 当前用户姓名
            project_id: 项目号
            source_column: 文件3专用，'M'或'L'，表示筛选来源
            file_manager: 文件管理器实例（用于自动勾选）
            viewer: Treeview控件（用于立即刷新显示）
            item_id: Treeview中的行ID（用于立即刷新显示）
            columns: 列名列表（用于查找"是否已完成"列索引）
        """
        super().__init__(parent)
        
        self.interface_id = interface_id
        self.file_type = file_type
        self.file_path = file_path
        self.row_index = row_index
        self.user_name = user_name
        self.project_id = project_id
        self.source_column = source_column
        self.file_manager = file_manager
        self.viewer = viewer  # 保存Treeview引用
        self.item_id = item_id  # 保存行ID
        self.columns = columns  # 保存列名
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置界面"""
        self.title("回文单号输入")
        self.geometry("400x200")
        self.resizable(False, False)
        
        # 居中显示
        self.transient(self.master)
        self.grab_set()
        
        # 标题
        title_label = ttk.Label(self, text=f"接口号: {self.interface_id}", 
                                font=('Arial', 12, 'bold'))
        title_label.pack(pady=10)
        
        # 输入框
        input_frame = ttk.Frame(self)
        input_frame.pack(pady=10, padx=20, fill='x')
        
        ttk.Label(input_frame, text="回文单号:").pack(side='left', padx=5)
        
        self.entry = ttk.Entry(input_frame, width=30)
        self.entry.pack(side='left', padx=5, fill='x', expand=True)
        self.entry.focus_set()
        
        # 按钮
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="确认", command=self.on_confirm).pack(side='left', padx=10)
        ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)
        
        # 绑定Enter键
        self.entry.bind('<Return>', lambda e: self.on_confirm())
    
    def on_confirm(self):
        """确认按钮回调"""
        response_number = self.entry.get().strip()
        
        if not response_number:
            messagebox.showwarning("警告", "请输入回文单号", parent=self)
            return
        
        # 写入Excel
        try:
            success = write_response_to_excel(
                self.file_path,
                self.file_type,
                self.row_index,
                response_number,
                self.user_name,
                self.project_id,
                self.source_column
            )
            
            if success:
                # 【关键】Excel写入成功后，才更新Registry
                if registry_hooks:
                    try:
                        # 【修复】去除接口号中的角色后缀，并提取角色信息
                        # 例如："S-SA---1JT-01-25C1-25E6(设计人员)" -> 接口号="S-SA---1JT-01-25C1-25E6", 角色="设计人员"
                        import re
                        clean_interface_id = re.sub(r'\([^)]*\)$', '', self.interface_id).strip()
                        role_match = re.search(r'\(([^)]+)\)$', self.interface_id)
                        role = role_match.group(1) if role_match else None
                        
                        registry_hooks.on_response_written(
                            file_type=self.file_type,
                            file_path=self.file_path,
                            row_index=self.row_index,
                            interface_id=clean_interface_id,  # 使用清理后的接口号
                            response_number=response_number,
                            user_name=self.user_name,
                            project_id=self.project_id,
                            source_column=self.source_column,
                            role=role  # 传递角色信息
                        )
                        print(f"[Registry] ✓ 已记录回文单号写入事件")
                    except Exception as e:
                        print(f"[Registry] 回文单号写入钩子调用失败: {e}")
                        # Registry更新失败不影响Excel写入，但要提示用户
                        import traceback
                        traceback.print_exc()
                
                # 【新增】清除该文件的缓存，强制下次重新处理以应用Registry逻辑
                if self.file_manager:
                    try:
                        self.file_manager.clear_file_cache(self.file_path)
                        print(f"[缓存] ✓ 已清除文件缓存，下次处理将应用Registry逻辑")
                    except Exception as e:
                        print(f"[缓存] 清除文件缓存失败: {e}")
                
                # 【自动勾选】设计人员回填后自动勾选"已完成"
                if self.file_manager:
                    try:
                        self.file_manager.set_row_completed(
                            self.file_path,
                            self.row_index,
                            True,
                            self.user_name
                        )
                        print(f"[自动勾选] 已为设计人员{self.user_name}自动勾选行{self.row_index}")
                    except Exception as e:
                        print(f"[自动勾选] 失败: {e}")
                
                # 【立即刷新显示】更新Treeview中的勾选框和状态列
                if self.viewer and self.item_id and self.columns:
                    try:
                        # 获取当前行的值
                        current_values = list(self.viewer.item(self.item_id, "values"))
                        
                        # 1. 更新"是否已完成"列（勾选框）
                        if "是否已完成" in self.columns:
                            checkbox_idx = self.columns.index("是否已完成")
                            if checkbox_idx < len(current_values):
                                current_values[checkbox_idx] = "☑"
                        
                        # 2. 更新"状态"列（显示"待审查"）
                        if "状态" in self.columns:
                            status_idx = self.columns.index("状态")
                            if status_idx < len(current_values):
                                # 查询该任务的display_status
                                try:
                                    from registry.util import make_task_id
                                    import re
                                    
                                    # 清理接口号
                                    clean_interface_id = re.sub(r'\([^)]*\)$', '', self.interface_id).strip()
                                    
                                    task_key = {
                                        'file_type': self.file_type,
                                        'project_id': self.project_id,
                                        'interface_id': clean_interface_id,
                                        'source_file': os.path.basename(self.file_path),
                                        'row_index': self.row_index,
                                        'interface_time': ''  # 时间信息可选
                                    }
                                    
                                    # 查询状态
                                    status_map = registry_hooks.get_display_status([task_key], current_user_roles_str='')
                                    tid = make_task_id(
                                        self.file_type,
                                        self.project_id,
                                        clean_interface_id,
                                        os.path.basename(self.file_path),
                                        self.row_index
                                    )
                                    
                                    # 更新状态列
                                    if tid in status_map and status_map[tid]:
                                        current_values[status_idx] = status_map[tid]
                                        print(f"[立即刷新] 状态列已更新为: {status_map[tid]}")
                                except Exception as e:
                                    print(f"[立即刷新] 查询状态失败: {e}")
                        
                        # 应用更新
                        self.viewer.item(self.item_id, values=current_values)
                        print(f"[立即刷新] ✓ Treeview显示已更新（勾选框+状态列）")
                        
                    except Exception as e:
                        print(f"[立即刷新] 失败: {e}")
                
                messagebox.showinfo("成功", "回文单号已保存", parent=self)
                self.destroy()
            else:
                messagebox.showerror("失败", "保存失败，请重试", parent=self)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {str(e)}", parent=self)


def write_response_to_excel(file_path, file_type, row_index, response_number, 
                             user_name, project_id, source_column=None):
    """
    写入回文单号到Excel文件
    
    参数:
        file_path: Excel文件路径
        file_type: 文件类型(1-6)
        row_index: Excel行号（从2开始，因为第1行是标题）
        response_number: 回文单号
        user_name: 用户姓名
        project_id: 项目号
        source_column: 文件3专用，'M'或'L'
    
    返回:
        bool: 成功返回True，失败返回False
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            return False
        
        # 文件锁定检测
        try:
            # 尝试以独占模式打开
            with open(file_path, 'r+b') as f:
                pass
        except PermissionError:
            messagebox.showerror("文件占用", "有其他用户占用该文件，请稍后再试")
            return False
        
        # 使用openpyxl打开
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 获取写入列位置
        columns = get_write_columns(file_type, row_index, ws, source_column)
        
        if not columns:
            print(f"无法确定写入列位置: file_type={file_type}")
            return False
        
        # 写入数据
        response_col = columns['response_col']
        time_col = columns['time_col']
        name_col = columns['name_col']
        
        ws[f"{response_col}{row_index}"] = response_number
        ws[f"{time_col}{row_index}"] = date.today().strftime('%Y-%m-%d')
        ws[f"{name_col}{row_index}"] = user_name
        
         # 【新增】文件6特殊逻辑：自动更新M列（回复状态列）
        if file_type == 6:
            try:
                # I列（索引8）是预期时间列
                expected_time_cell = ws.cell(row_index, 9)  # I列是第9列（A=1）
                expected_time = expected_time_cell.value
                
                # 比较当前日期和预期时间
                from datetime import datetime
                today = date.today()
                
                # 解析预期时间
                if expected_time:
                    try:
                        # 尝试解析为日期对象
                        if isinstance(expected_time, datetime):
                            expected_date = expected_time.date()
                        elif isinstance(expected_time, date):
                            expected_date = expected_time
                        else:
                            # 尝试字符串解析
                            import pandas as pd
                            parsed = pd.to_datetime(expected_time, errors='coerce')
                            if pd.notna(parsed):
                                expected_date = parsed.date()
                            else:
                                expected_date = None
                        
                        # 根据对比结果写入M列（第13列）
                        if expected_date:
                            if today <= expected_date:
                                reply_status = "按时回复"
                            else:
                                reply_status = "延期回复"
                            
                            ws.cell(row_index, 13, reply_status)  # M列是第13列
                            print(f"[文件6] 自动更新M列: {reply_status} (预期:{expected_date}, 实际:{today})")
                        else:
                            print(f"[文件6] 无法解析预期时间，跳过M列更新")
                    except Exception as parse_error:
                        print(f"[文件6] 解析预期时间失败: {parse_error}")
                else:
                    print(f"[文件6] I列预期时间为空，跳过M列更新")
            except Exception as e:
                print(f"[文件6] 更新M列失败: {e}")
                # 即使M列更新失败，也不影响回文单号写入
        
        # 保存
        try:
            wb.save(file_path)
            wb.close()
            
            # 【关键】验证写入是否成功：重新打开文件检查
            print(f"[验证] 开始验证Excel写入...")
            verify_wb = load_workbook(file_path, read_only=True)
            verify_ws = verify_wb.active
            
            # 验证回文单号列
            verify_response = verify_ws[f"{response_col}{row_index}"].value
            if str(verify_response).strip() != str(response_number).strip():
                verify_wb.close()
                raise Exception(f"验证失败：回文单号列写入不匹配。期望:{response_number}, 实际:{verify_response}")
            
            verify_wb.close()
            print(f"[验证] ✓ Excel写入验证成功")
            print(f"成功写入: {file_path}, 行{row_index}, 回文单号={response_number}")
            return True
            
        except Exception as save_error:
            print(f"[ERROR] Excel保存或验证失败: {save_error}")
            raise  # 重新抛出异常，让上层处理
        
    except Exception as e:
        print(f"[ERROR] 写入回文单号失败!")
        print(f"  文件路径: {file_path}")
        print(f"  文件类型: {file_type}")
        print(f"  行号: {row_index}")
        print(f"  回文单号: {response_number}")
        print(f"  错误信息: {e}")
        import traceback
        traceback.print_exc()
        from tkinter import messagebox
        messagebox.showerror("写入失败", f"无法写入回文单号到Excel文件\n\n错误：{str(e)}")
        return False


def get_write_columns(file_type, row_index, worksheet, source_column=None):
    """
    获取各文件类型的写入列位置
    
    参数:
        file_type: 文件类型(1-6)
        row_index: Excel行号
        worksheet: openpyxl工作表对象
        source_column: 文件3专用，'M'或'L'
    
    返回:
        dict: {'response_col': 'S', 'time_col': 'N', 'name_col': 'V'}
        或 None（如果无法确定）
    """
    # 文件类型1-2, 4-6的固定列位置
    column_map = {
        1: {'response_col': 'S', 'time_col': 'M', 'name_col': 'V'},  
        2: {'response_col': 'P', 'time_col': 'N', 'name_col': 'AL'},
        4: {'response_col': 'U', 'time_col': 'V', 'name_col': 'AT'},
        5: {'response_col': 'V', 'time_col': 'N', 'name_col': 'W'},
        6: {'response_col': 'L', 'time_col': 'J', 'name_col': 'N'},
    }
    
    if file_type in column_map:
        return column_map[file_type]
    
    # 文件3特殊逻辑：根据source_column判断
    if file_type == 3:
        if source_column == 'M':
            # M列筛选：V/T/BM
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif source_column == 'L':
            # L列筛选：S/Q/BM
            return {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 如果未指定，尝试自动判断
            return determine_file3_source_and_columns(row_index, worksheet)
    
    return None


def determine_file3_source_and_columns(row_index, worksheet):
    """
    判断文件3某行是因M列还是L列被筛选出
    
    参数:
        row_index: Excel行号
        worksheet: openpyxl工作表对象
    
    返回:
        dict: 写入列位置
    """
    try:
        # 读取M列和L列的值
        m_val = worksheet[f"M{row_index}"].value
        l_val = worksheet[f"L{row_index}"].value
        
        # 读取T列和Q列的值（回复时间列）
        t_val = worksheet[f"T{row_index}"].value
        q_val = worksheet[f"Q{row_index}"].value
        
        # 简化判断逻辑：
        # 如果M列有时间数据且T列为空，判断为M列来源
        # 如果L列有时间数据且Q列为空，判断为L列来源
        # 优先M列
        
        m_has_time = m_val is not None and str(m_val).strip() != ''
        t_is_empty = t_val is None or str(t_val).strip() == ''
        
        l_has_time = l_val is not None and str(l_val).strip() != ''
        q_is_empty = q_val is None or str(q_val).strip() == ''
        
        if m_has_time and t_is_empty:
            # M列来源
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif l_has_time and q_is_empty:
            # L列来源
            return {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 默认M列
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
    
    except Exception as e:
        print(f"判断文件3来源失败: {e}")
        # 默认返回M列
        return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}


# 测试代码
if __name__ == "__main__":
    # 测试get_write_columns
    columns_1 = get_write_columns(1, 5, None)
    print(f"文件1写入列: {columns_1}")
    
    columns_3_m = get_write_columns(3, 5, None, 'M')
    print(f"文件3(M列)写入列: {columns_3_m}")
    
    columns_3_l = get_write_columns(3, 5, None, 'L')
    print(f"文件3(L列)写入列: {columns_3_l}")

