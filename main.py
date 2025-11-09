#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据处理模块
此文件包含所有Excel文件的数据处理逻辑
"""

import pandas as pd
import numpy as np
import datetime
import os
from pathlib import Path
import warnings
import re
from copy import copy

# 忽略pandas警告
warnings.filterwarnings('ignore')

## 日期筛选逻辑开关：由上层程序设置
# False: 使用新逻辑（当年1月1日~当月末/次月末）
# True:  使用旧逻辑（当月或当月+次月；文件6为未来14天）
USE_OLD_DATE_LOGIC = False


def process_excel_files(excel_files, current_datetime):
    """
    处理Excel文件的主函数
    
    参数:
        excel_files (list): Excel文件路径列表
        current_datetime (datetime): 当前日期时间
    
    返回:
        pandas.DataFrame: 处理后的结果数据
    """
    print(f"开始处理 {len(excel_files)} 个Excel文件...")
    print(f"处理时间: {current_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 记录到监控器
    try:
        import Monitor
        Monitor.log_info(f"开始处理 {len(excel_files)} 个Excel文件")
        Monitor.log_info(f"处理时间: {current_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
    except:
        pass
    
    # 查找待处理文件1（特定格式的文件）
    target_file, project_id = find_target_file(excel_files)
    
    if target_file is None:
        error_msg = "未找到符合格式要求的待处理文件（四位数字+按项目导出IDI手册+日期格式）"
        print(error_msg)
        try:
            import Monitor
            Monitor.log_error(error_msg)
        except:
            pass
        return pd.DataFrame({'错误信息': ['未找到符合格式的文件']})
    
    print(f"找到待处理文件1: {os.path.basename(target_file)}")
    try:
        import Monitor
        Monitor.log_success(f"找到待处理文件1: {os.path.basename(target_file)}")
    except:
        pass
    
    try:
        # 处理特定文件
        result = process_target_file(target_file, current_datetime)
        
        if result is not None and not result.empty:
            print(f"待处理文件1处理完成，生成 {len(result)} 行完成处理数据")
            return result
        else:
            print("处理完成，但没有符合条件的数据")
            return pd.DataFrame({'信息': ['没有符合条件的数据']})
                
    except Exception as e:
        print(f"处理待处理文件1时发生错误: {str(e)}")
        return pd.DataFrame({'错误信息': [str(e)]})


def find_target_file(excel_files):
    """
    查找符合特定格式的待处理文件1（兼容性函数，返回第一个匹配的文件）
    格式：四位数字+按项目导出IDI手册+日期
    例如：2016按项目导出IDI手册2025-08-01-17_55_52
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files1(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files1(excel_files):
    """
    查找所有符合特定格式的待处理文件1
    格式：四位数字+按项目导出IDI手册+日期
    例如：2016按项目导出IDI手册2025-08-01-17_55_52
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^(\d{4})按项目导出IDI手册\d{4}-\d{2}-\d{2}.*\.(xlsx|xls)$'
    matched_files = []
    
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件1...")
    except:
        pass
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件1格式: {file_name}, 项目号: {project_id}")
            try:
                import Monitor
                Monitor.log_success(f"找到待处理文件1: 项目{project_id} - {file_name}")
            except:
                pass
    
    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件1")
        try:
            import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件1，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except:
            pass
    else:
        try:
            import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件1")
        except:
            pass
    
    return matched_files


def find_target_file2(excel_files):
    """
    查找符合特定格式的待处理文件2（兼容性函数，返回第一个匹配的文件）
    格式：内部接口信息单报表+12位数字，前4位为项目号
    例如：内部接口信息单报表201612345678
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files2(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files2(excel_files):
    """
    查找所有符合特定格式的待处理文件2
    格式：内部接口信息单报表+12位数字，前4位为项目号
    例如：内部接口信息单报表201612345678
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^内部接口信息单报表(\d{4})\d{8}\.(xlsx|xls)$'
    matched_files = []
    
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件2...")
    except:
        pass
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件2格式: {file_name}, 项目号: {project_id}")
            try:
                import Monitor
                Monitor.log_success(f"找到待处理文件2: 项目{project_id} - {file_name}")
            except:
                pass
    
    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件2")
        try:
            import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件2，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except:
            pass
    else:
        try:
            import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件2")
        except:
            pass
    
    return matched_files


def find_target_file3(excel_files):
    """
    查找符合特定格式的待处理文件3（兼容性函数，返回第一个匹配的文件）
    格式：外部接口ICM报表+四位数字（项目号）+日期（8位）
    例如：外部接口ICM报表201620250801.xlsx
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files3(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files3(excel_files):
    """
    查找所有符合特定格式的待处理文件3
    格式：外部接口ICM报表+四位数字（项目号）+日期（8位）
    例如：外部接口ICM报表201620250801.xlsx
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^外部接口ICM报表(\d{4})\d{8}\.(xlsx|xls)$'
    matched_files = []
    
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件3...")
    except:
        pass
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件3格式: {file_name}, 项目号: {project_id}")
            try:
                import Monitor
                Monitor.log_success(f"找到待处理文件3: 项目{project_id} - {file_name}")
            except:
                pass
    
    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件3")
        try:
            import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件3，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except:
            pass
    else:
        try:
            import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件3")
        except:
            pass
    
    return matched_files


def find_target_file4(excel_files):
    """
    查找符合特定格式的待处理文件4（兼容性函数，返回第一个匹配的文件）
    格式：外部接口单报表+四位数字（项目号）+日期（8位）
    例如：外部接口单报表201620250801.xlsx
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files4(excel_files)
    if all_files:
        return all_files[0]
    return None, None

def find_all_target_files4(excel_files):
    """
    查找所有符合特定格式的待处理文件4
    格式：外部接口单报表+四位数字（项目号）+日期（8位）
    例如：外部接口单报表201620250801.xlsx
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^外部接口单报表(\d{4})\d{8}\.(xlsx|xls)$'
    matched_files = []
    
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件4...")
    except:
        pass
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            print(f"匹配到待处理文件4格式: {file_name}, 项目号: {project_id}")
            try:
                import Monitor
                Monitor.log_success(f"找到待处理文件4: 项目{project_id} - {file_name}")
            except:
                pass
    
    if matched_files:
        print(f"总共找到 {len(matched_files)} 个待处理文件4")
        try:
            import Monitor
            project_ids = list(set([pid for _, pid in matched_files]))
            Monitor.log_success(f"批量识别完成: 找到{len(matched_files)}个待处理文件4，涉及{len(project_ids)}个项目({', '.join(sorted(project_ids))})")
        except:
            pass
    else:
        try:
            import Monitor
            Monitor.log_warning("未找到任何符合格式的待处理文件4")
        except:
            pass
    
    return matched_files


def process_target_file(file_path, current_datetime):
    """
    处理待处理文件1的主函数
    
    参数:
        file_path (str): 待处理文件1的路径
        current_datetime (datetime): 当前日期时间
    
    返回:
        pandas.DataFrame: 完成处理数据
    """
    print(f"开始处理待处理文件1: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件1: {os.path.basename(file_path)}")
    except:
        pass
    
    # 读取Excel文件的第一个工作表（不强制Sheet1）
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')
        
    if df.empty:
        print("文件为空")
        try:
            import Monitor
            Monitor.log_error("文件为空，无法处理")
        except:
            pass
            return pd.DataFrame()
        
    print(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    try:
        import Monitor
        Monitor.log_info(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    except:
        pass
    
    # 显示前几行的数据概览，确保数据正确加载
    print("数据概览（前3行）：")
    for i in range(min(3, len(df))):
        if i == 0:
            print(f"第{i+1}行（表头）: {list(df.iloc[i])[:5]}...")  # 只显示前5列
        else:
            print(f"第{i+1}行（数据）: {list(df.iloc[i])[:5]}...")  # 只显示前5列
    
    # 执行四个处理步骤
    process1_rows = execute_process1(df)  # H列25C1/25C2/25C3筛选
    process2_rows = execute_process2(df, current_datetime)  # K列日期筛选
    process3_rows = execute_process3(df)  # M列空值且A列非空筛选
    process4_rows = execute_process4(df)  # 作废数据筛选
    
    # 计算最终结果：满足处理1、处理2、处理3，但排除处理4
    final_rows = process1_rows & process2_rows & process3_rows - process4_rows
    
    print(f"筛选统计 - P1:{len(process1_rows)}行 P2:{len(process2_rows)}行 P3:{len(process3_rows)}行 P4(排除):{len(process4_rows)}行 → 结果:{len(final_rows)}行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    print(f"\n========== [Registry] 开始查询待审查任务 ==========")
    print(f"[Registry] 总行数: {len(df)}, 原始筛选结果: {len(final_rows)}行")
    
    try:
        from registry import hooks as registry_hooks
        from registry.util import extract_interface_id, extract_project_id, make_task_id, make_business_id
        from registry.db import get_connection
        from registry.config import load_config
        
        # 【修复】直接查询数据库中有display_status的任务，按business_id匹配
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        from registry.hooks import _cfg
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            
            # 查询该文件类型下所有有display_status的任务
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status, row_index, source_file
                FROM tasks
                WHERE file_type = 1
                  AND display_status IS NOT NULL
                  AND display_status != ''
                  AND status != 'confirmed'
                  AND status != 'archived'
            """)
            
            registry_tasks = cursor.fetchall()
            print(f"[Registry] 数据库中该文件类型有{len(registry_tasks)}个有状态的任务（包括待完成、待审查等）")
            
            # 在当前Excel中查找这些接口号
            pending_rows = set()
            
            # 【调试】输出前3个数据库任务的详细信息
            if len(registry_tasks) > 0:
                print(f"[Registry调试] 数据库任务示例（前3个）:")
                for i, task in enumerate(registry_tasks[:3]):
                    print(f"  {i+1}. 接口={task[0][:40]}, 项目={task[1]}, display_status={task[2]}")
            
            # 【优化】先建立Excel接口号索引，提升性能
            # 【修复】项目号从文件名提取，不从df_row中提取
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else ""
            
            excel_index = {}  # {(interface_id, project_id): [idx1, idx2, ...]}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 1)
                    
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            print(f"[Registry] Excel索引建立完成（项目{file_project_id}），共{len(excel_index)}个唯一接口")
            
            # 【修复】按索引查找，避免双重循环
            for reg_interface_id, reg_project_id, reg_display_status, _, _ in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                
                if key in excel_index:
                    # 找到匹配的行
                    matched_indices = excel_index[key]
                    for idx in matched_indices:
                        # 如果不在final_rows中，添加
                        if idx not in final_rows:
                            pending_rows.add(idx)
                            print(f"[Registry] ✓ 发现待审查任务：第{idx+2}行 接口{reg_interface_id[:30]} 状态:{reg_display_status}")
                        else:
                            # 已经在final_rows中（可能M列实际为空）
                            print(f"[Registry提示] 接口{reg_interface_id[:30]}已在原始筛选结果中，跳过")
                else:
                    print(f"[Registry警告] 数据库任务未在Excel中找到: 接口={reg_interface_id[:40]}")
            
            print(f"\n[Registry] 统计: 数据库中{len(registry_tasks)}个待审查，在Excel中匹配到{len(pending_rows)}行")
        
            if pending_rows:
                final_rows = final_rows | pending_rows
                print(f"[Registry] ✓ 合并{len(pending_rows)}条待审查任务到结果")
            else:
                print(f"[Registry] 未找到待审查任务")
        else:
            print(f"[Registry] 数据库不存在或未配置，跳过待审查任务查询")
        
    except Exception as e:
        print(f"[Registry] ❌ 查询待审查任务失败（不影响主流程）: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"========== [Registry] 查询完成 ==========\n")
    
    # 记录到监控器
    try:
        import Monitor
        Monitor.log_info(f"处理1符合条件: {len(process1_rows)} 行")
        Monitor.log_info(f"处理2符合条件: {len(process2_rows)} 行")
        Monitor.log_info(f"处理3符合条件: {len(process3_rows)} 行")
        Monitor.log_info(f"处理4需排除: {len(process4_rows)} 行")
        if len(final_rows) > 0:
            Monitor.log_success(f"最终完成处理数据: {len(final_rows)} 行")
        else:
            Monitor.log_warning("经过四步筛选后，无符合条件的数据")
    except:
        pass
    
    if not final_rows:
        return pd.DataFrame()
    
    # 获取最终结果数据（排除第一行标题行）
    final_indices = [i for i in final_rows if i > 0]  # 排除第一行
    
    result_df = df.iloc[final_indices].copy()
    
    # 添加行号信息 - 修正行号计算以匹配用户期望
    excel_row_numbers = [i + 2 for i in final_indices]  # pandas索引+2 = Excel行号
    result_df['原始行号'] = excel_row_numbers
    # 新增“科室”列（基于H列：25C1/25C2/25C3）
    try:
        department_values = []
        for idx in result_df.index:
            cell_str = str(df.iloc[idx, 7]) if 7 < len(df.columns) else ""
            if "25C1" in cell_str:
                department_values.append("结构一室")
            elif "25C2" in cell_str:
                department_values.append("结构二室")
            elif "25C3" in cell_str:
                department_values.append("建筑总图室")
            else:
                department_values.append("")
        result_df["科室"] = department_values
    except Exception:
        result_df["科室"] = ""

    # 新增“接口时间”列（基于K列：索引10），格式 mm.dd
    try:
        time_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 10] if 10 < len(df.columns) else None
            # 解析日期
            try:
                if isinstance(cell_val, str):
                    # 多格式尝试
                    parsed = pd.to_datetime(cell_val, errors='coerce')
                else:
                    parsed = pd.to_datetime(cell_val, errors='coerce')
                if pd.isna(parsed):
                    time_values.append("")
                else:
                    # 【修复】保留完整年份，支持跨年延期判断
                    time_values.append(parsed.strftime('%Y.%m.%d'))
            except Exception:
                time_values.append("")
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""

    # 新增"责任人"列（基于R列：索引17，提取中文）
    try:
        zh_pattern = re.compile(r"[\u4e00-\u9fa5]+")
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 17] if 17 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            found = zh_pattern.findall(s)
            owners.append("".join(found))
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = ""
    
    # 【新增】添加source_file列（用于回文单号输入时定位源文件）
    result_df['source_file'] = file_path

    return result_df


def execute_process1(df):
    """
    处理1：H列数据筛选，筛选出包含"25C1"、"25C2"、"25C3"的数据
    
    参数:
        df (pandas.DataFrame): 原始数据
    
    返回:
        set: 符合条件的行索引集合
    """
    result_rows = set()
    
    # 记录处理开始
    try:
        import Monitor
        Monitor.log_process("开始执行处理1：筛选H列数据（25C1、25C2、25C3）")
    except:
        pass
    
    # 检查H列是否存在（列索引7，因为从0开始）
    if len(df.columns) <= 7:
        warning_msg = "警告：数据列数不足，无H列"
        print(warning_msg)
        try:
            import Monitor
            Monitor.log_warning(warning_msg)
        except:
            pass
        return result_rows
    
    h_column = df.iloc[:, 7]  # H列是第8列（索引7）
    
    # 搜索包含指定字符串的行
    target_values = ["25C1", "25C2", "25C3"]
    
    for idx, cell_value in h_column.items():
        if idx == 0:  # 跳过第一行标题
            continue
            
        cell_str = str(cell_value) if cell_value is not None else ""
        
        for target in target_values:
            if target in cell_str:
                result_rows.add(idx)
                break
    
    print(f"处理1完成：共找到 {len(result_rows)} 行符合H列筛选条件")
    try:
        import Monitor
        Monitor.log_success(f"处理1完成：共找到 {len(result_rows)} 行符合H列筛选条件")
    except:
        pass
    return result_rows


def execute_process2(df, current_datetime):
    """
    处理2：K列日期筛选逻辑
    根据当前日期决定筛选范围：
    - 如果今天是1-19号：筛选同年同月数据
    - 如果今天是20-31号：筛选同年同月及次月数据
    
    参数:
        df (pandas.DataFrame): 原始数据
        current_datetime (datetime): 当前日期时间
    
    返回:
        set: 符合条件的行索引集合
    """
    result_rows = set()
    
    # 记录处理开始
    try:
        import Monitor
        Monitor.log_process("开始执行处理2：筛选K列日期数据")
    except:
        pass
    
    # 检查K列是否存在（列索引10，因为从0开始）
    if len(df.columns) <= 10:
        warning_msg = "警告：数据列数不足，无K列"
        print(warning_msg)
        try:
            import Monitor
            Monitor.log_warning(warning_msg)
        except:
            pass
        return result_rows
    
    k_column = df.iloc[:, 10]  # K列是第11列（索引10）
    
    # 获取当前日期信息
    current_day = current_datetime.day
    current_year = current_datetime.year
    current_month = current_datetime.month
    
    print(f"当前日期：{current_datetime.strftime('%Y-%m-%d')}，今天是{current_day}号")
    
    if USE_OLD_DATE_LOGIC:
        # 旧逻辑：
        # 1~19：当月
        # 20~31：当月+次月
        start_date = datetime.datetime(current_year, current_month, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    else:
        # 新逻辑：
        # 1~19：当年1月1日 ~ 当月末
        # 20~31：当年1月1日 ~ 次月末
        start_date = datetime.datetime(current_year, 1, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    print(f"当日为{current_day}号，筛选范围：{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")
    
    # 遍历K列数据进行筛选
    for idx, cell_value in k_column.items():
        if idx == 0:  # 跳过第一行标题
            continue
        
        if pd.isna(cell_value):
            continue
        
        try:
            # 尝试解析日期，支持多种格式
            if isinstance(cell_value, str):
                # 尝试多种日期格式
                date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']
                cell_date = None
                
                for fmt in date_formats:
                    try:
                        cell_date = pd.to_datetime(cell_value, format=fmt, errors='raise')
                        break
                    except:
                        continue
                
                # 如果固定格式都失败，使用智能解析
                if cell_date is None or pd.isna(cell_date):
                    cell_date = pd.to_datetime(cell_value, errors='coerce')
            else:
                cell_date = pd.to_datetime(cell_value, errors='coerce')
            
            if pd.isna(cell_date):
                continue
            
            # 检查日期是否在筛选范围内
            if start_date <= cell_date <= end_date:
                result_rows.add(idx)
                
        except Exception as e:
            print(f"处理2：第{idx+1}行K列日期解析失败: {cell_value}, 错误: {e}")
            continue
    
    print(f"处理2完成：共找到 {len(result_rows)} 行符合K列日期筛选条件")
    try:
        import Monitor
        Monitor.log_success(f"处理2完成：共找到 {len(result_rows)} 行符合K列日期筛选条件")
    except:
        pass
    return result_rows


def execute_process3(df):
    """
    【修复】处理3：M列空值且A列非空筛选 + Registry待确认任务合并
    
    筛选逻辑：
    1. 原始筛选：M列为空值，同时A列不为空值的数据
    2. Registry查询：M列已填充但在Registry中有display_status（待确认）的任务
    3. 合并两部分结果
    
    参数:
        df (pandas.DataFrame): 原始数据
    
    返回:
        set: 符合条件的行索引集合
    """
    result_rows = set()
    
    # 记录处理开始
    try:
        import Monitor
        Monitor.log_process("开始执行处理3：筛选M列空值且A列非空数据")
    except:
        pass
    
    # 检查A列和M列是否存在
    if len(df.columns) <= 0:
        warning_msg = "警告：数据列数不足，无A列"
        print(warning_msg)
        try:
            import Monitor
            Monitor.log_warning(warning_msg)
        except:
            pass
        return result_rows
    
    if len(df.columns) <= 12:
        warning_msg = "警告：数据列数不足，无M列"
        print(warning_msg)
        try:
            import Monitor
            Monitor.log_warning(warning_msg)
        except:
            pass
        return result_rows
    
    a_column = df.iloc[:, 0]   # A列是第1列（索引0）
    m_column = df.iloc[:, 12]  # M列是第13列（索引12）
    
    # 【阶段1】原始筛选：M列为空且A列不为空
    for idx in range(len(df)):
        if idx == 0:  # 跳过第一行标题
            continue
            
        a_value = a_column.iloc[idx]
        m_value = m_column.iloc[idx]
        
        # 检查A列不为空且M列为空
        a_not_empty = not (pd.isna(a_value) or str(a_value).strip() == "")
        m_is_empty = pd.isna(m_value) or str(m_value).strip() == ""
        
        if a_not_empty and m_is_empty:
            result_rows.add(idx)
    
    print(f"处理3完成（原始筛选）：共找到 {len(result_rows)} 行符合M列空值且A列非空条件")
    
    # 【阶段2】Registry查询：M列已填充但待确认的任务
    # 注意：这个函数在process_target_file中被调用，文件路径需要从上层传递
    # 由于这里无法直接获取文件路径，我们返回原始结果，让调用层处理Registry逻辑
    # 实际的Registry合并逻辑应该在process_target_file中实现
    
    try:
        import Monitor
        Monitor.log_success(f"处理3完成：共找到 {len(result_rows)} 行符合M列空值且A列非空条件")
    except:
        pass
    return result_rows


def execute_process4(df):
    """
    处理4：筛选"作废"数据
    仅检查B列中是否包含"作废"标记
    
    参数:
        df (pandas.DataFrame): 原始数据
    
    返回:
        set: 符合条件的行索引集合（需要排除的行）
    """
    result_rows = set()
    
    # 记录处理开始
    try:
        import Monitor
        Monitor.log_process("开始执行处理4：筛选B列作废数据")
    except:
        pass
    
    # 检查B列是否存在（列索引1，因为从0开始）
    if len(df.columns) <= 1:
        warning_msg = "警告：数据列数不足，无B列"
        print(warning_msg)
        try:
            import Monitor
            Monitor.log_warning(warning_msg)
        except:
            pass
        return result_rows
    
    # 仅检查B列查找"作废"标记
    b_column = df.iloc[:, 1]  # B列是第2列（索引1）
    
    for idx, cell_value in b_column.items():
        if idx == 0:  # 跳过第一行标题
            continue
            
        cell_str = str(cell_value) if cell_value is not None else ""
        
        if "作废" in cell_str:
            result_rows.add(idx)
    
    print(f"处理4完成：共找到 {len(result_rows)} 行B列包含作废标记（需要排除）")
    try:
        import Monitor
        if len(result_rows) > 0:
            Monitor.log_warning(f"处理4完成：共找到 {len(result_rows)} 行B列包含作废标记（需要排除）")
        else:
            Monitor.log_success("处理4完成：未发现作废数据")
    except:
        pass
    return result_rows


def export_result_to_excel(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出完成处理数据到Excel文件
    新建空白Excel文件，重命名Sheet1，复制表头格式和内容，写入符合条件的完整数据，设置列宽
    
    参数:
        df (pandas.DataFrame): 完成处理数据（包含原始行号）
        original_file_path (str): 原始文件路径  
        current_datetime (datetime): 当前日期时间
        output_dir (str): 输出目录
        project_id (str): 项目号，用于创建结果文件夹
    
    返回:
        str: 导出文件路径
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    
    try:
        # 根据项目号创建结果文件夹
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            
            # 如果文件夹不存在则创建
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
                print(f"创建结果文件夹: {result_folder_path}")
            
            # 使用结果文件夹作为输出目录
            final_output_dir = result_folder_path
        else:
            # 如果没有项目号，使用原始输出目录
            final_output_dir = output_dir
        
        # 生成输出文件名，处理重名文件
        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"内部需打开接口{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        
        # 处理重名文件，自动加序号
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1
        
        # 第一步：新建空白Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 第二步：重命名Sheet1为"内部需打开接口"
        ws.title = "内部需打开接口"
        
        # 第三步：读取原始文件用于复制格式和数据
        original_wb = load_workbook(original_file_path)
        # 使用第一个工作表
        original_ws = original_wb.worksheets[0]
        
        # 读取原始数据（用于获取行数据）
        if original_file_path.endswith('.xlsx'):
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)
        
        # 第四步：复制原Excel第一行的数据格式和内容（表头）
        if original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            print(f"导出表头：从第1列到第{max_col}列")
            
            # 复制第一行的所有单元格（包括格式和内容）
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 复制格式
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception as style_error:
                    print(f"样式复制失败，仅复制值: {style_error}")
            
            print("已复制表头（格式和内容）")
        
        # 第五步：复制处理结果后的原Excel行数据
        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            print(f"准备导出 {len(qualified_rows)} 行符合条件的数据")
            
            write_row = 2  # 从第二行开始写入数据
            max_col = max(original_ws.max_column, len(original_df.columns))
            
            for excel_row_num in sorted(qualified_rows):
                # 确保行号在有效范围内
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    # 复制整行数据和格式
                    for col_idx in range(1, max_col + 1):
                        source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        
                        # 复制值
                        target_cell.value = source_cell.value
                        
                        # 复制格式
                        try:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                        except Exception as style_error:
                            print(f"数据行样式复制失败，仅复制值: {style_error}")
                    
                    write_row += 1
            
            print(f"已写入 {write_row - 2} 行数据")
        else:
            print("没有符合条件的数据行需要导出")
        
        # 第六步：设置列宽（能完全显示1~4行数据，乘以1.2系数）
        max_col = max(original_ws.max_column, len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            # 计算该列1~4行数据的最大显示宽度
            max_width = 8  # 最小宽度
            
            for row_idx in range(1, min(5, ws.max_row + 1)):  # 检查1~4行
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    # 计算单元格内容的显示宽度
                    cell_width = len(str(cell.value)) * 1.2  # 粗略估算
                    max_width = max(max_width, cell_width)
            
            # 应用1.2系数并设置列宽
            final_width = max_width * 1.2
            # openpyxl列宽限制在1-255之间
            final_width = min(max(final_width, 8), 100)
            
            # 设置列宽
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width
        
        print("已设置列宽")
        
        # 保存文件
        original_wb.close()
        wb.save(output_path)
        wb.close()
        
        print(f"内部需打开接口导出完成！文件保存到: {output_path}")
        try:
            import Monitor
            Monitor.log_success(f"内部需打开接口导出完成！文件保存到: {output_path}")
        except:
            pass
        
        return output_path
    except Exception as e:
        print(f"导出数据时发生错误: {str(e)}")
        raise


# ===================== 待处理文件2（内部需回复接口）相关处理 =====================
def process_target_file2(file_path, current_datetime, project_id=None):
    """
    处理待处理文件2（内部需回复接口）的主函数
    返回：pandas.DataFrame，包含原始行号
    
    筛选逻辑根据项目号决定：
    - 1907和2016项目：final = P1 & P2 & P4
    - 其他项目：final = P1 & P2 & P4 - P3
    """
    print(f"开始处理待处理文件2: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件2: {os.path.basename(file_path)}")
    except:
        pass

    # 读取Excel文件的第一个工作表（不强制Sheet1）
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')

    if df.empty:
        print("文件为空")
        try:
            import Monitor
            Monitor.log_error("文件为空，无法处理")
        except:
            pass
        return pd.DataFrame()

    # 处理1
    process1_rows = execute2_process1(df)
    # 处理2
    process2_rows = execute2_process2(df, current_datetime)
    # 处理3（用于排除）
    process3_rows = execute2_process3(df)
    # 处理4
    process4_rows = execute2_process4(df)

    # 根据项目号决定筛选逻辑
    # 1907和2016使用现有逻辑，其他项目排除process3
    if project_id in ['1907', '2016']:
        final_rows = process1_rows & process2_rows & process4_rows
        print(f"项目{project_id}使用标准逻辑（不排除process3）")
    else:
        final_rows = process1_rows & process2_rows & process4_rows - process3_rows
        print(f"项目{project_id}使用扩展逻辑（排除process3：{len(process3_rows)}行）")
    
    print(f"最终完成处理数据（原始筛选）: {len(final_rows)} 行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    print(f"\n========== [Registry] 开始查询待审查任务（文件类型2） ==========")
    try:
        from registry.util import extract_interface_id, extract_project_id
        from registry.db import get_connection
        from registry.config import load_config
        
        # 【修复】直接查询数据库中有display_status的任务，按business_id匹配
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        from registry.hooks import _cfg
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        
        pending_rows = set()
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status
                FROM tasks
                WHERE file_type = 2
                  AND display_status IS NOT NULL
                  AND display_status != ''
                  AND status != 'confirmed'
                  AND status != 'archived'
            """)
            
            registry_tasks = cursor.fetchall()
            print(f"[Registry] 数据库中该文件类型有{len(registry_tasks)}个有状态的任务")
            
            # 【优化】从文件名提取项目号，建立Excel索引
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else project_id
            
            excel_index = {}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 2)
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            # 按索引查找
            for reg_interface_id, reg_project_id, reg_display_status in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                if key in excel_index:
                    matched_indices = excel_index[key]
                    for idx in matched_indices:
                        if idx not in final_rows:
                            pending_rows.add(idx)
        
        if pending_rows:
            final_rows = final_rows | pending_rows
        
    except Exception as e:
        print(f"[Registry] 查询待确认任务失败（不影响主流程）: {e}")
    
    print(f"最终完成处理数据（含待确认）: {len(final_rows)} 行")

    # 日志
    try:
        import Monitor
        Monitor.log_info(f"处理1符合条件: {len(process1_rows)} 行")
        Monitor.log_info(f"处理2符合条件: {len(process2_rows)} 行")
        Monitor.log_info(f"处理3(排除项)符合条件: {len(process3_rows)} 行")
        Monitor.log_info(f"处理4符合条件: {len(process4_rows)} 行")
        if len(final_rows) > 0:
            Monitor.log_success(f"最终完成处理数据: {len(final_rows)} 行")
        else:
            Monitor.log_warning("经过筛选后，无符合条件的数据")
    except:
        pass

    if not final_rows:
        return pd.DataFrame()

    final_indices = [i for i in final_rows if i > 0]
    excel_row_numbers = [i + 2 for i in final_indices]  # pandas索引+2=Excel行号
    result_df = df.iloc[final_indices].copy()
    result_df['原始行号'] = excel_row_numbers
    
    # 新增"责任人"列（基于AM列：索引38，提取中文）
    try:
        zh_pattern = re.compile(r"[\u4e00-\u9fa5]+")
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 38] if 38 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            found = zh_pattern.findall(s)
            owner_str = "".join(found)
            # 空值显示"无"
            owners.append(owner_str if owner_str else "无")
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = "无"
    
    # 【新增】添加source_file列
    result_df['source_file'] = file_path
    # 新增"科室"列（基于I列内容包含：结构一室/结构二室/建筑总图室）
    try:
        department_values = []
        for idx in result_df.index:
            cell_str = str(df.iloc[idx, 8]) if 8 < len(df.columns) else ""
            if "结构一室" in cell_str:
                department_values.append("结构一室")
            elif "结构二室" in cell_str:
                department_values.append("结构二室")
            elif "建筑总图室" in cell_str:
                department_values.append("建筑总图室")
            else:
                department_values.append("")
        result_df["科室"] = department_values
    except Exception:
        result_df["科室"] = ""

    # 新增"接口时间"列（基于M列：索引12），格式 yyyy.mm.dd
    try:
        time_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 12] if 12 < len(df.columns) else None
            try:
                parsed = pd.to_datetime(cell_val, errors='coerce')
                if pd.isna(parsed):
                    time_values.append("")
                else:
                    # 【修复】保留完整年份，支持跨年延期判断
                    time_values.append(parsed.strftime('%Y.%m.%d'))
            except Exception:
                time_values.append("")
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""
    return result_df

def execute2_process1(df):
    """I列包含“河北分公司-建筑结构所”或包含“25C1/25C2/25C3”"""
    result_rows = set()
    if len(df.columns) <= 8:
        return result_rows
    i_column = df.iloc[:, 8]
    for idx, val in i_column.items():
        if idx == 0:
            continue
        s = str(val)
        if ("河北分公司-建筑结构所" in s) or ("25C1" in s) or ("25C2" in s) or ("25C3" in s):
            result_rows.add(idx)
    return result_rows

def execute2_process2(df, current_datetime):
    """M列日期筛选，逻辑同文件1的K列"""
    result_rows = set()
    if len(df.columns) <= 12:
        return result_rows
    m_column = df.iloc[:, 12]
    current_day = current_datetime.day
    current_year = current_datetime.year
    current_month = current_datetime.month
    if USE_OLD_DATE_LOGIC:
        start_date = datetime.datetime(current_year, current_month, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    else:
        start_date = datetime.datetime(current_year, 1, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    for idx, val in m_column.items():
        if idx == 0:
            continue
        cell_date = None
        try:
            if isinstance(val, str):
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                    try:
                        cell_date = pd.to_datetime(val, format=fmt, errors='raise')
                        break
                    except:
                        continue
                if cell_date is None or pd.isna(cell_date):
                    cell_date = pd.to_datetime(val, errors='coerce')
            else:
                cell_date = pd.to_datetime(val, errors='coerce')
            if pd.isna(cell_date):
                continue
            if start_date <= cell_date <= end_date:
                result_rows.add(idx)
        except:
            continue
    return result_rows

def execute2_process3(df):
    """AB列以4444开头，且F列为“传递”"""
    result_rows = set()
    if len(df.columns) <= 27 or len(df.columns) <= 5:
        return result_rows
    ab_column = df.iloc[:, 27]
    f_column = df.iloc[:, 5]
    for idx in range(len(df)):
        if idx == 0:
            continue
        ab_val = ab_column.iloc[idx]
        f_val = f_column.iloc[idx]
        if str(ab_val).startswith("4444") and str(f_val) == "传递":
            result_rows.add(idx)
    return result_rows

def execute2_process4(df):
    """N列为空且A列不为空"""
    result_rows = set()
    if len(df.columns) <= 13 or len(df.columns) <= 0:
        return result_rows
    n_column = df.iloc[:, 13]
    a_column = df.iloc[:, 0]
    for idx in range(len(df)):
        if idx == 0:
            continue
        n_val = n_column.iloc[idx]
        a_val = a_column.iloc[idx]
        a_not_empty = not (pd.isna(a_val) or str(a_val).strip() == "")
        n_is_empty = pd.isna(n_val) or str(n_val).strip() == ""
        if a_not_empty and n_is_empty:
            result_rows.add(idx)
    return result_rows

def export_result_to_excel2(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出内部需回复接口处理结果到Excel文件
    新建空白Excel文件，重命名Sheet1，复制表头格式和内容，写入符合条件的完整数据，设置列宽
    
    参数:
        df (pandas.DataFrame): 完成处理数据（包含原始行号）
        original_file_path (str): 原始文件路径  
        current_datetime (datetime): 当前日期时间
        output_dir (str): 输出目录
        project_id (str): 项目号，用于创建结果文件夹
    
    返回:
        str: 导出文件路径
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    
    try:
        # 根据项目号创建结果文件夹
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            
            # 如果文件夹不存在则创建
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
                print(f"创建结果文件夹: {result_folder_path}")
            
            # 使用结果文件夹作为输出目录
            final_output_dir = result_folder_path
        else:
            # 如果没有项目号，使用原始输出目录
            final_output_dir = output_dir
        
        # 生成输出文件名，处理重名文件
        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"内部需回复接口{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1
        
        # 第一步：新建空白Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 第二步：重命名Sheet1为"内部需回复接口"
        ws.title = "内部需回复接口"
        
        # 第三步：读取原始文件用于复制格式和数据
        original_wb = load_workbook(original_file_path)
        # 使用第一个工作表
        original_ws = original_wb.worksheets[0]
        
        # 读取原始数据
        if original_file_path.endswith('.xlsx'):
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)
        
        # 第四步：复制原Excel第一行的数据格式和内容（表头）
        if original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            print(f"导出表头：从第1列到第{max_col}列")
            
            # 复制第一行的所有单元格（包括格式和内容）
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 复制格式
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception as style_error:
                    print(f"样式复制失败，仅复制值: {style_error}")
            
            print("已复制表头（格式和内容）")
        
        # 第五步：复制处理结果后的原Excel行数据
        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            print(f"准备导出 {len(qualified_rows)} 行符合条件的数据")
            
            write_row = 2  # 从第二行开始写入数据
            max_col = max(original_ws.max_column, len(original_df.columns))
            
            for excel_row_num in sorted(qualified_rows):
                # 确保行号在有效范围内
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    # 复制整行数据和格式
                    for col_idx in range(1, max_col + 1):
                        source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        
                        # 复制值
                        target_cell.value = source_cell.value
                        
                        # 复制格式
                        try:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                        except Exception as style_error:
                            print(f"数据行样式复制失败，仅复制值: {style_error}")
                    
                    write_row += 1
            
            print(f"已写入 {write_row - 2} 行数据")
        else:
            print("没有符合条件的数据行需要导出")
        
        # 第六步：设置列宽（能完全显示1~4行数据，乘以1.2系数）
        max_col = max(original_ws.max_column, len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            # 计算该列1~4行数据的最大显示宽度
            max_width = 8  # 最小宽度
            
            for row_idx in range(1, min(5, ws.max_row + 1)):  # 检查1~4行
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    # 计算单元格内容的显示宽度
                    cell_width = len(str(cell.value)) * 1.2  # 粗略估算
                    max_width = max(max_width, cell_width)
            
            # 应用1.2系数并设置列宽
            final_width = max_width * 1.2
            # openpyxl列宽限制在1-255之间
            final_width = min(max(final_width, 8), 100)
            
            # 设置列宽
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width
        
        print("已设置列宽")
        
        # 保存文件
        original_wb.close()
        wb.save(output_path)
        wb.close()
        
        print(f"内部需回复接口导出完成！文件保存到: {output_path}")
        try:
            import Monitor
            Monitor.log_success(f"内部需回复接口导出完成！文件保存到: {output_path}")
        except:
            pass
        
        return output_path
    except Exception as e:
        print(f"导出数据时发生错误: {str(e)}")
        raise


# ===================== 待处理文件3（外部需打开接口）相关处理 =====================
def process_target_file3(file_path, current_datetime):
    """
    处理待处理文件3（外部需打开接口）的主函数
    
    参数:
        file_path (str): 待处理文件3的路径
        current_datetime (datetime): 当前日期时间
    
    返回:
        pandas.DataFrame: 完成处理数据，包含原始行号
    """
    print(f"开始处理待处理文件3: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件3: {os.path.basename(file_path)}")
    except:
        pass
    
    # 读取Excel文件的第一个工作表（不强制Sheet1）
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')
        
    if df.empty:
        print("文件为空")
        try:
            import Monitor
            Monitor.log_error("文件为空，无法处理")
        except:
            pass
        return pd.DataFrame()
        
    print(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    try:
        import Monitor
        Monitor.log_info(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    except:
        pass
    
    # 执行六个处理步骤
    process1_rows = execute3_process1(df)  # I列为"B"的数据
    process2_rows = execute3_process2(df)  # AL列以"河北分公司-建筑结构所"开头的数据
    process3_rows = execute3_process3(df, current_datetime)  # M列时间数据筛选
    process4_rows = execute3_process4(df, current_datetime)  # L列时间数据筛选
    process5_rows = execute3_process5(df)  # Q列为空值的数据
    process6_rows = execute3_process6(df)  # T列为空值的数据
    
    # 最终汇总逻辑：
    # (处理1 AND 处理2 AND 处理3 AND 处理6) OR (处理1 AND 处理2 AND 处理4 AND 处理5)
    group1 = process1_rows & process2_rows & process3_rows & process6_rows
    group2 = process1_rows & process2_rows & process4_rows & process5_rows
    final_rows = group1 | group2  # 并集关系
    
    print(f"最终完成处理数据（原始筛选）: {len(final_rows)} 行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    print(f"\n========== [Registry] 开始查询待审查任务（文件类型3） ==========")
    try:
        from registry.util import extract_interface_id, extract_project_id
        from registry.db import get_connection
        from registry.config import load_config
        
        # 【修复】直接查询数据库中有display_status的任务，按business_id匹配
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        from registry.hooks import _cfg
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        
        pending_rows = set()
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status
                FROM tasks
                WHERE file_type = 3
                  AND display_status IS NOT NULL
                  AND display_status != ''
                  AND status != 'confirmed'
                  AND status != 'archived'
            """)
            
            registry_tasks = cursor.fetchall()
            print(f"[Registry] 数据库中该文件类型有{len(registry_tasks)}个有状态的任务")
            
            # 【优化】从文件名提取项目号，建立Excel索引
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else ""
            
            excel_index = {}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 3)
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            # 按索引查找
            for reg_interface_id, reg_project_id, reg_display_status in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                if key in excel_index:
                    for idx in excel_index[key]:
                        if idx not in final_rows:
                            pending_rows.add(idx)
        
        if pending_rows:
            final_rows = final_rows | pending_rows
        
    except Exception as e:
        print(f"[Registry] 查询待审查任务失败: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"最终完成处理数据（含待审查）: {len(final_rows)} 行")
    
    # 日志记录
    try:
        import Monitor
        Monitor.log_info(f"处理1(I列为B): {len(process1_rows)} 行")
        Monitor.log_info(f"处理2(AL列河北分公司-建筑结构所开头): {len(process2_rows)} 行")
        Monitor.log_info(f"处理3(M列时间筛选): {len(process3_rows)} 行")
        Monitor.log_info(f"处理4(L列时间筛选): {len(process4_rows)} 行")
        Monitor.log_info(f"处理5(Q列为空): {len(process5_rows)} 行")
        Monitor.log_info(f"处理6(T列为空): {len(process6_rows)} 行")
        Monitor.log_info(f"组1(1&2&3-6): {len(group1)} 行")
        Monitor.log_info(f"组2(1&2&4-5): {len(group2)} 行")
        if len(final_rows) > 0:
            Monitor.log_success(f"最终完成处理数据: {len(final_rows)} 行")
        else:
            Monitor.log_warning("经过筛选后，无符合条件的数据")
    except:
        pass
    
    if not final_rows:
        return pd.DataFrame()
    
    # 转换为最终结果DataFrame
    final_indices = [i for i in final_rows if i >= 0]
    excel_row_numbers = [i + 2 for i in final_indices]  # pandas索引+2=Excel行号
    result_df = df.iloc[final_indices].copy()
    result_df['原始行号'] = excel_row_numbers
    
    # 【新增】添加来源标记（用于回文单号输入时判断写入列）
    source_columns = []
    for idx in final_indices:
        if idx in group1 and idx not in group2:
            source_columns.append('M')  # M列筛选路径
        elif idx in group2 and idx not in group1:
            source_columns.append('L')  # L列筛选路径
        else:
            source_columns.append('M')  # 两者都匹配，优先M列
    result_df['_source_column'] = source_columns
    # 新增“科室”列（基于AO列：匹配三种科室，空值则“请室主任确认”，否则保留原值）
    try:
        department_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 40] if 40 < len(df.columns) else None
            cell_str = "" if (cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val))) else str(cell_val).strip()
            if cell_str == "":
                department_values.append("请室主任确认")
            elif "结构一室" in cell_str:
                department_values.append("结构一室")
            elif "结构二室" in cell_str:
                department_values.append("结构二室")
            elif "建筑总图室" in cell_str:
                department_values.append("建筑总图室")
            else:
                department_values.append(cell_str)
        result_df["科室"] = department_values
    except Exception:
        result_df["科室"] = "请室主任确认"

    # 新增"接口时间"列（优先M列索引12，其次L列索引11），格式 yyyy.mm.dd
    try:
        time_values = []
        for idx in result_df.index:
            m_val = df.iloc[idx, 12] if 12 < len(df.columns) else None
            l_val = df.iloc[idx, 11] if 11 < len(df.columns) else None
            parsed = None
            try:
                parsed = pd.to_datetime(m_val, errors='coerce')
            except Exception:
                parsed = None
            if parsed is None or pd.isna(parsed):
                try:
                    parsed = pd.to_datetime(l_val, errors='coerce')
                except Exception:
                    parsed = None
            if parsed is None or pd.isna(parsed):
                time_values.append("")
            else:
                # 【修复】保留完整年份，支持跨年延期判断
                time_values.append(parsed.strftime('%Y.%m.%d'))
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""
    # 新增"责任人"列（基于AP列：索引41，提取中文）
    try:
        zh_pattern = re.compile(r"[\u4e00-\u9fa5]+")
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 41] if 41 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            found = zh_pattern.findall(s)
            owners.append("".join(found))
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = ""
    # 【新增】添加source_file列
    result_df['source_file'] = file_path
    return result_df


def execute3_process1(df):
    """
    处理1：读取待处理文件3中的I列的数据，筛选这一列中为"B"的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理1：筛选I列为'B'的数据")
    try:
        import Monitor
        Monitor.log_process("处理1：筛选I列为'B'的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 8:  # I列索引为8
        print("警告：文件列数不足，无法访问I列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问I列")
        except:
            pass
        return qualified_rows
    
    try:
        # I列索引为8（从0开始）
        i_column = df.iloc[:, 8]
        
        for index, value in enumerate(i_column):
            if pd.notna(value) and str(value).strip() == "B":
                qualified_rows.add(index)
        
        print(f"处理1完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理1完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理1执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理1执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute3_process2(df):
    """
    处理2：读取待处理文件3中AL列的数据，筛选这一列中以"河北分公司-建筑结构所"开头的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理2：筛选AL列以'河北分公司-建筑结构所'开头的数据")
    try:
        import Monitor
        Monitor.log_process("处理2：筛选AL列以'河北分公司-建筑结构所'开头的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 37:  # AL列索引为37
        print("警告：文件列数不足，无法访问AL列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问AL列")
        except:
            pass
        return qualified_rows
    
    try:
        # AL列索引为37（从0开始）
        al_column = df.iloc[:, 37]
        target_prefix = "河北分公司-建筑结构所"
        
        for index, value in enumerate(al_column):
            if pd.notna(value) and str(value).strip().startswith(target_prefix):
                qualified_rows.add(index)
        
        print(f"处理2完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理2完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理2执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理2执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute3_process3(df, current_datetime):
    """
    处理3：读取处理文件3中M列的数据，根据当前日期进行时间筛选
    
    参数:
        df (pandas.DataFrame): 输入数据
        current_datetime (datetime): 当前日期时间
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理3：筛选M列时间数据")
    try:
        import Monitor
        Monitor.log_process("处理3：筛选M列时间数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 12:  # M列索引为12
        print("警告：文件列数不足，无法访问M列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问M列")
        except:
            pass
        return qualified_rows
    
    try:
        # M列索引为12（从0开始）
        m_column = df.iloc[:, 12]
        current_day = current_datetime.day
        current_year = current_datetime.year
        current_month = current_datetime.month
        
        if USE_OLD_DATE_LOGIC:
            start_date = datetime.datetime(current_year, current_month, 1)
            if 1 <= current_day <= 19:
                if current_month == 12:
                    end_date = datetime.datetime(current_year, 12, 31)
                else:
                    end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
            else:
                if current_month == 11:
                    end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
                elif current_month == 12:
                    end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
                else:
                    end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
        else:
            start_date = datetime.datetime(current_year, 1, 1)
            if 1 <= current_day <= 19:
                if current_month == 12:
                    end_date = datetime.datetime(current_year, 12, 31)
                else:
                    end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
            else:
                if current_month == 11:
                    end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
                elif current_month == 12:
                    end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
                else:
                    end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
        
        print(f"筛选日期范围: {start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")
        
        for index, value in enumerate(m_column):
            if pd.notna(value):
                try:
                    # 尝试解析日期，支持多种格式
                    date_value = None
                    value_str = str(value).strip()
                    
                    # 尝试不同的日期格式
                    date_formats = [
                        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
                        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'
                    ]
                    
                    for fmt in date_formats:
                        try:
                            date_value = datetime.datetime.strptime(value_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    # 如果是pandas的Timestamp对象
                    if date_value is None and hasattr(value, 'year'):
                        date_value = datetime.datetime(value.year, value.month, value.day)
                    
                    if date_value and start_date <= date_value <= end_date:
                        qualified_rows.add(index)
                        
                except Exception as parse_error:
                    print(f"日期解析失败（第{index+1}行）: {value} - {parse_error}")
                    continue
        
        print(f"处理3完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理3完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理3执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理3执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute3_process4(df, current_datetime):
    """
    处理4：读取处理文件3中L列的数据，进行时间筛选（包括4444开头的特殊处理）
    
    参数:
        df (pandas.DataFrame): 输入数据
        current_datetime (datetime): 当前日期时间
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理4：筛选L列时间数据（包括4444开头特殊处理）")
    try:
        import Monitor
        Monitor.log_process("处理4：筛选L列时间数据（包括4444开头特殊处理）")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 11:  # L列索引为11
        print("警告：文件列数不足，无法访问L列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问L列")
        except:
            pass
        return qualified_rows
    
    try:
        # L列索引为11（从0开始）
        l_column = df.iloc[:, 11]
        current_day = current_datetime.day
        current_year = current_datetime.year
        current_month = current_datetime.month
        
        if USE_OLD_DATE_LOGIC:
            start_date = datetime.datetime(current_year, current_month, 1)
            if 1 <= current_day <= 19:
                if current_month == 12:
                    end_date = datetime.datetime(current_year, 12, 31)
                else:
                    end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
            else:
                if current_month == 11:
                    end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
                elif current_month == 12:
                    end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
                else:
                    end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
        else:
            start_date = datetime.datetime(current_year, 1, 1)
            if 1 <= current_day <= 19:
                if current_month == 12:
                    end_date = datetime.datetime(current_year, 12, 31)
                else:
                    end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
            else:
                if current_month == 11:
                    end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
                elif current_month == 12:
                    end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
                else:
                    end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
        
        print(f"筛选日期范围: {start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")
        
        for index, value in enumerate(l_column):
            if pd.notna(value):
                try:
                    value_str = str(value).strip()
                    date_value = None
                    
                    # 特殊处理：4444开头的数据当做当年处理
                    if value_str.startswith('4444'):
                        # 将4444替换为当前年份
                        modified_value_str = value_str.replace('4444', str(current_year), 1)
                        value_str = modified_value_str
                    
                    # 尝试不同的日期格式
                    date_formats = [
                        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
                        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'
                    ]
                    
                    for fmt in date_formats:
                        try:
                            date_value = datetime.datetime.strptime(value_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    # 如果是pandas的Timestamp对象
                    if date_value is None and hasattr(value, 'year'):
                        date_value = datetime.datetime(value.year, value.month, value.day)
                    
                    if date_value and start_date <= date_value <= end_date:
                        qualified_rows.add(index)
                        
                except Exception as parse_error:
                    print(f"日期解析失败（第{index+1}行）: {value} - {parse_error}")
                    continue
        
        print(f"处理4完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理4完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理4执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理4执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute3_process5(df):
    """
    处理5：读取待处理文件3中Q列的数据，筛选这一列中为空值的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理5：筛选Q列为空值的数据")
    try:
        import Monitor
        Monitor.log_process("处理5：筛选Q列为空值的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 16:  # Q列索引为16
        print("警告：文件列数不足，无法访问Q列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问Q列")
        except:
            pass
        return qualified_rows
    
    try:
        # Q列索引为16（从0开始）
        q_column = df.iloc[:, 16]
        
        for index, value in enumerate(q_column):
            if pd.isna(value) or str(value).strip() == '':
                qualified_rows.add(index)
        
        print(f"处理5完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理5完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理5执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理5执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute3_process6(df):
    """
    处理6：读取待处理文件3中T列的数据，筛选这一列中为空值的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理6：筛选T列为空值的数据")
    try:
        import Monitor
        Monitor.log_process("处理6：筛选T列为空值的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 19:  # T列索引为19
        print("警告：文件列数不足，无法访问T列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问T列")
        except:
            pass
        return qualified_rows
    
    try:
        # T列索引为19（从0开始）
        t_column = df.iloc[:, 19]
        
        for index, value in enumerate(t_column):
            if pd.isna(value) or str(value).strip() == '':
                qualified_rows.add(index)
        
        print(f"处理6完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理6完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理6执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理6执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def export_result_to_excel3(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出外部需打开接口处理结果到Excel文件
    新建空白Excel文件，重命名Sheet1，复制表头格式和内容，写入符合条件的完整数据，设置列宽
    
    参数:
        df (pandas.DataFrame): 完成处理数据（包含原始行号）
        original_file_path (str): 原始文件路径  
        current_datetime (datetime): 当前日期时间
        output_dir (str): 输出目录
        project_id (str): 项目号，用于创建结果文件夹
    
    返回:
        str: 导出文件路径
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    
    try:
        # 根据项目号创建结果文件夹
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            
            # 如果文件夹不存在则创建
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
                print(f"创建结果文件夹: {result_folder_path}")
            
            # 使用结果文件夹作为输出目录
            final_output_dir = result_folder_path
        else:
            # 如果没有项目号，使用原始输出目录
            final_output_dir = output_dir
        
        # 生成输出文件名，处理重名文件
        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"外部需打开接口{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1
        
        # 第一步：新建空白Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 第二步：重命名Sheet1为"外部需打开接口"
        ws.title = "外部需打开接口"
        
        # 第三步：读取原始文件用于复制格式和数据
        original_wb = load_workbook(original_file_path)
        # 使用第一个工作表
        original_ws = original_wb.worksheets[0]
        
        # 读取原始数据
        if original_file_path.endswith('.xlsx'):
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)
        
        # 第四步：复制原Excel第一行的数据格式和内容（表头）
        if original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            print(f"导出表头：从第1列到第{max_col}列")
            
            # 复制第一行的所有单元格（包括格式和内容）
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 复制格式
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception as style_error:
                    print(f"样式复制失败，仅复制值: {style_error}")
            
            print("已复制表头（格式和内容）")
        
        # 第五步：复制处理结果后的原Excel行数据
        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            print(f"准备导出 {len(qualified_rows)} 行符合条件的数据")
            
            write_row = 2  # 从第二行开始写入数据
            max_col = max(original_ws.max_column, len(original_df.columns))
            
            for excel_row_num in sorted(qualified_rows):
                # 确保行号在有效范围内
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    # 复制整行数据和格式
                    for col_idx in range(1, max_col + 1):
                        source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        
                        # 复制值
                        target_cell.value = source_cell.value
                        
                        # 复制格式
                        try:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                        except Exception as style_error:
                            print(f"数据行样式复制失败，仅复制值: {style_error}")
                    
                    write_row += 1
            
            print(f"已写入 {write_row - 2} 行数据")
        else:
            print("没有符合条件的数据行需要导出")
        
        # 第六步：设置列宽（能完全显示1~4行数据，乘以1.2系数）
        max_col = max(original_ws.max_column, len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            # 计算该列1~4行数据的最大显示宽度
            max_width = 8  # 最小宽度
            
            for row_idx in range(1, min(5, ws.max_row + 1)):  # 检查1~4行
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    # 计算单元格内容的显示宽度
                    cell_width = len(str(cell.value)) * 1.2  # 粗略估算
                    max_width = max(max_width, cell_width)
            
            # 应用1.2系数并设置列宽
            final_width = max_width * 1.2
            # openpyxl列宽限制在1-255之间
            final_width = min(max(final_width, 8), 100)
            
            # 设置列宽
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width
        
        print("已设置列宽")
        
        # 保存文件
        original_wb.close()
        wb.save(output_path)
        wb.close()
        
        print(f"外部需打开接口导出完成！文件保存到: {output_path}")
        try:
            import Monitor
            Monitor.log_success(f"外部需打开接口导出完成！文件保存到: {output_path}")
        except:
            pass
        
        return output_path
        
    except Exception as e:
        print(f"导出外部需打开接口数据时发生错误: {str(e)}")
        try:
            import Monitor
            Monitor.log_error(f"导出外部需打开接口数据时发生错误: {str(e)}")
        except:
            pass
        raise


# ===================== 待处理文件4（外部需回复接口）相关处理 =====================
def process_target_file4(file_path, current_datetime):
    """
    处理待处理文件4（外部需回复接口）的主函数
    
    参数:
        file_path (str): 待处理文件4的路径
        current_datetime (datetime): 当前日期时间
    
    返回:
        pandas.DataFrame: 完成处理数据，包含原始行号
    """
    print(f"开始处理待处理文件4: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件4: {os.path.basename(file_path)}")
    except:
        pass
    
    # 读取Excel文件的Sheet1
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')
        
    if df.empty:
        print("文件为空")
        try:
            import Monitor
            Monitor.log_error("文件为空，无法处理")
        except:
            pass
        return pd.DataFrame()
        
    print(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    try:
        import Monitor
        Monitor.log_info(f"读取到数据：{len(df)} 行，{len(df.columns)} 列")
    except:
        pass
    
    # 执行四个处理步骤
    process1_rows = execute4_process1(df)  # AF列以"河北分公司-建筑结构所"开头的数据
    process2_rows = execute4_process2(df)  # P列为"B"的数据
    process3_rows = execute4_process3(df, current_datetime)  # S列时间数据筛选
    process4_rows = execute4_process4(df)  # V列为空值的数据
    
    # 最终汇总逻辑：满足处理1、2、3，4
    final_rows = process1_rows & process2_rows & process3_rows & process4_rows
    
    print(f"最终完成处理数据（原始筛选）: {len(final_rows)} 行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    try:
        from registry.util import extract_interface_id, extract_project_id
        from registry.db import get_connection
        from registry.hooks import _cfg
        
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        pending_rows = set()
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status
                FROM tasks
                WHERE file_type = 4
                  AND display_status IS NOT NULL AND display_status != ''
                  AND status != 'confirmed' AND status != 'archived'
            """)
            registry_tasks = cursor.fetchall()
            
            # Excel索引优化
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else ""
            
            excel_index = {}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 4)
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            for reg_interface_id, reg_project_id, _ in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                if key in excel_index:
                    for idx in excel_index[key]:
                        if idx not in final_rows:
                            pending_rows.add(idx)
        
        if pending_rows:
            final_rows = final_rows | pending_rows
        
    except Exception as e:
        print(f"[Registry] 查询待审查任务失败: {e}")
    
    print(f"最终完成处理数据（含待确认）: {len(final_rows)} 行")
    
    # 日志记录
    try:
        import Monitor
        Monitor.log_info(f"处理1(AF列河北分公司-建筑结构所开头): {len(process1_rows)} 行")
        Monitor.log_info(f"处理2(P列为B): {len(process2_rows)} 行")
        Monitor.log_info(f"处理3(S列时间筛选): {len(process3_rows)} 行")
        Monitor.log_info(f"处理4(V列为空): {len(process4_rows)} 行")
        if len(final_rows) > 0:
            Monitor.log_success(f"最终完成处理数据: {len(final_rows)} 行")
        else:
            Monitor.log_warning("经过筛选后，无符合条件的数据")
    except:
        pass
    
    if not final_rows:
        return pd.DataFrame()
    
    # 转换为最终结果DataFrame
    final_indices = [i for i in final_rows if i >= 0]
    excel_row_numbers = [i + 2 for i in final_indices]  # pandas索引+2=Excel行号
    result_df = df.iloc[final_indices].copy()
    result_df['原始行号'] = excel_row_numbers
    # 新增“科室”列（基于AG列：匹配三种科室，空值则“请室主任确认”，否则保留原值）
    try:
        department_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 32] if 32 < len(df.columns) else None
            cell_str = "" if (cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val))) else str(cell_val).strip()
            if cell_str == "":
                department_values.append("请室主任确认")
            elif "结构一室" in cell_str:
                department_values.append("结构一室")
            elif "结构二室" in cell_str:
                department_values.append("结构二室")
            elif "建筑总图室" in cell_str:
                department_values.append("建筑总图室")
            else:
                department_values.append(cell_str)
        result_df["科室"] = department_values
    except Exception:
        result_df["科室"] = "请室主任确认"

    # 新增"接口时间"列（基于S列：索引18），格式 yyyy.mm.dd
    try:
        time_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 18] if 18 < len(df.columns) else None
            try:
                parsed = pd.to_datetime(cell_val, errors='coerce')
                if pd.isna(parsed):
                    time_values.append("")
                else:
                    # 【修复】保留完整年份，支持跨年延期判断
                    time_values.append(parsed.strftime('%Y.%m.%d'))
            except Exception:
                time_values.append("")
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""
    # 新增"责任人"列（基于AH列：索引33，提取中文）
    try:
        zh_pattern = re.compile(r"[\u4e00-\u9fa5]+")
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 33] if 33 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            found = zh_pattern.findall(s)
            owners.append("".join(found))
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = ""
    # 【新增】添加source_file列
    result_df['source_file'] = file_path
    return result_df


def execute4_process1(df):
    """
    处理1：读取待处理文件4中的AF列的数据，筛选这一列中以"河北分公司-建筑结构所"开头的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理1：筛选AF列以'河北分公司-建筑结构所'开头的数据")
    try:
        import Monitor
        Monitor.log_process("处理1：筛选AF列以'河北分公司-建筑结构所'开头的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 31:  # AF列索引为31
        print("警告：文件列数不足，无法访问AF列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问AF列")
        except:
            pass
        return qualified_rows
    
    try:
        # AF列索引为31（从0开始）
        af_column = df.iloc[:, 31]
        target_prefix = "河北分公司-建筑结构所"
        
        for index, value in enumerate(af_column):
            if pd.notna(value) and str(value).strip().startswith(target_prefix):
                qualified_rows.add(index)
        
        print(f"处理1完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理1完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理1执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理1执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute4_process2(df):
    """
    处理2：读取待处理文件4中的P列的数据，筛选这一列中为"B"的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理2：筛选P列为'B'的数据")
    try:
        import Monitor
        Monitor.log_process("处理2：筛选P列为'B'的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 15:  # P列索引为15
        print("警告：文件列数不足，无法访问P列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问P列")
        except:
            pass
        return qualified_rows
    
    try:
        # P列索引为15（从0开始）
        p_column = df.iloc[:, 15]
        
        for index, value in enumerate(p_column):
            if pd.notna(value) and str(value).strip() == "B":
                qualified_rows.add(index)
        
        print(f"处理2完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理2完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理2执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理2执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute4_process3(df, current_datetime):
    """
    处理3：读取待处理文件4中S列的数据，根据当前日期进行时间筛选
    
    参数:
        df (pandas.DataFrame): 输入数据
        current_datetime (datetime): 当前日期时间
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理3：筛选S列时间数据")
    try:
        import Monitor
        Monitor.log_process("处理3：筛选S列时间数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 18:  # S列索引为18
        print("警告：文件列数不足，无法访问S列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问S列")
        except:
            pass
        return qualified_rows
    
    try:
        # S列索引为18（从0开始）
        s_column = df.iloc[:, 18]
        current_day = current_datetime.day
        current_year = current_datetime.year
        current_month = current_datetime.month
        
        # 新逻辑：1~19号 → 当年1月1日至当月末；20~31号 → 当年1月1日至次月月末
        start_date = datetime.datetime(current_year, 1, 1)
        if 1 <= current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            elif current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
        
        print(f"筛选日期范围: {start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")
        
        for index, value in enumerate(s_column):
            if pd.notna(value):
                try:
                    # 尝试解析日期，支持多种格式
                    date_value = None
                    value_str = str(value).strip()
                    
                    # 尝试不同的日期格式
                    date_formats = [
                        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
                        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'
                    ]
                    
                    for fmt in date_formats:
                        try:
                            date_value = datetime.datetime.strptime(value_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    # 如果是pandas的Timestamp对象
                    if date_value is None and hasattr(value, 'year'):
                        date_value = datetime.datetime(value.year, value.month, value.day)
                    
                    if date_value and start_date <= date_value <= end_date:
                        qualified_rows.add(index)
                        
                except Exception as parse_error:
                    print(f"日期解析失败（第{index+1}行）: {value} - {parse_error}")
                    continue
        
        print(f"处理3完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理3完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理3执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理3执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def execute4_process4(df):
    """
    处理4：读取待处理文件4中V列的数据，筛选这一列中为空值的数据
    
    参数:
        df (pandas.DataFrame): 输入数据
    
    返回:
        set: 符合条件的行索引集合
    """
    print("执行处理4：筛选V列为空值的数据")
    try:
        import Monitor
        Monitor.log_process("处理4：筛选V列为空值的数据")
    except:
        pass
    
    qualified_rows = set()
    
    if len(df.columns) <= 21:  # V列索引为21
        print("警告：文件列数不足，无法访问V列")
        try:
            import Monitor
            Monitor.log_warning("文件列数不足，无法访问V列")
        except:
            pass
        return qualified_rows
    
    try:
        # V列索引为21（从0开始）
        v_column = df.iloc[:, 21]
        
        for index, value in enumerate(v_column):
            if pd.isna(value) or str(value).strip() == '':
                qualified_rows.add(index)
        
        print(f"处理4完成：找到 {len(qualified_rows)} 行符合条件")
        try:
            import Monitor
            Monitor.log_info(f"处理4完成：找到 {len(qualified_rows)} 行符合条件")
        except:
            pass
            
    except Exception as e:
        print(f"处理4执行出错: {e}")
        try:
            import Monitor
            Monitor.log_error(f"处理4执行出错: {e}")
        except:
            pass
    
    return qualified_rows


def export_result_to_excel4(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出外部需回复接口处理结果到Excel文件
    新建空白Excel文件，重命名Sheet1，复制表头格式和内容，写入符合条件的完整数据，设置列宽
    
    参数:
        df (pandas.DataFrame): 完成处理数据（包含原始行号）
        original_file_path (str): 原始文件路径  
        current_datetime (datetime): 当前日期时间
        output_dir (str): 输出目录
        project_id (str): 项目号，用于创建结果文件夹
    
    返回:
        str: 导出文件路径
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    
    try:
        # 根据项目号创建结果文件夹
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            
            # 如果文件夹不存在则创建
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
                print(f"创建结果文件夹: {result_folder_path}")
            
            # 使用结果文件夹作为输出目录
            final_output_dir = result_folder_path
        else:
            # 如果没有项目号，使用原始输出目录
            final_output_dir = output_dir
        
        # 生成输出文件名，处理重名文件
        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"外部需回复接口{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1
        
        # 第一步：新建空白Excel文件
        wb = Workbook()
        ws = wb.active
        
        # 第二步：重命名Sheet1为"外部需回复接口"
        ws.title = "外部需回复接口"
        
        # 第三步：读取原始文件用于复制格式和数据
        original_wb = load_workbook(original_file_path)
        original_ws = original_wb.worksheets[0]
        
        # 读取原始数据
        if original_file_path.endswith('.xlsx'):
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)
        
        # 第四步：复制原Excel第一行的数据格式和内容（表头）
        if original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            print(f"导出表头：从第1列到第{max_col}列")
            
            # 复制第一行的所有单元格（包括格式和内容）
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 复制格式
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception as style_error:
                    print(f"样式复制失败，仅复制值: {style_error}")
            
            print("已复制表头（格式和内容）")
        
        # 第五步：复制处理结果后的原Excel行数据
        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            print(f"准备导出 {len(qualified_rows)} 行符合条件的数据")
            
            write_row = 2  # 从第二行开始写入数据
            max_col = max(original_ws.max_column, len(original_df.columns))
            
            for excel_row_num in sorted(qualified_rows):
                # 确保行号在有效范围内
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    # 复制整行数据和格式
                    for col_idx in range(1, max_col + 1):
                        source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        
                        # 复制值
                        target_cell.value = source_cell.value
                        
                        # 复制格式
                        try:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                        except Exception as style_error:
                            print(f"数据行样式复制失败，仅复制值: {style_error}")
                    
                    write_row += 1
            
            print(f"已写入 {write_row - 2} 行数据")
        else:
            print("没有符合条件的数据行需要导出")
        
        # 第六步：设置列宽（能完全显示1~4行数据，乘以1.2系数）
        max_col = max(original_ws.max_column, len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            # 计算该列1~4行数据的最大显示宽度
            max_width = 8  # 最小宽度
            
            for row_idx in range(1, min(5, ws.max_row + 1)):  # 检查1~4行
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    # 计算单元格内容的显示宽度
                    cell_width = len(str(cell.value)) * 1.2  # 粗略估算
                    max_width = max(max_width, cell_width)
            
            # 应用1.2系数并设置列宽
            final_width = max_width * 1.2
            # openpyxl列宽限制在1-255之间
            final_width = min(max(final_width, 8), 100)
            
            # 设置列宽
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width
        
        print("已设置列宽")
        
        # 保存文件
        original_wb.close()
        wb.save(output_path)
        wb.close()
        
        print(f"外部需回复接口导出完成！文件保存到: {output_path}")
        try:
            import Monitor
            Monitor.log_success(f"外部需回复接口导出完成！文件保存到: {output_path}")
        except:
            pass
        
        return output_path
        
    except Exception as e:
        print(f"导出外部需回复接口数据时发生错误: {str(e)}")
        try:
            import Monitor
            Monitor.log_error(f"导出外部需回复接口数据时发生错误: {str(e)}")
        except:
            pass
        raise


# ===================== 待处理文件5（三维提资接口）相关处理 =====================
def find_target_file5(excel_files):
    """
    查找符合特定格式的待处理文件5（兼容性函数，返回第一个匹配的文件）
    格式：四位数字+接口提资清单
    例如：2016接口提资清单.xlsx
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files5(excel_files)
    if all_files:
        return all_files[0]
    return None, None


def find_all_target_files5(excel_files):
    """
    查找所有符合特定格式的待处理文件5
    格式：四位数字+接口提资清单+任意后缀
    例如：2016接口提资清单.xlsx
    返回：[(文件路径, 项目号), ...] 列表
    """
    pattern = r'^(\d{4})接口提资清单.*\.(xlsx|xls)$'
    matched_files = []
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件5(三维提资接口)...")
    except:
        pass
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        m = re.match(pattern, file_name)
        if m:
            project_id = m.group(1)
            matched_files.append((file_path, project_id))
            try:
                import Monitor
                Monitor.log_success(f"找到待处理文件5: 项目{project_id} - {file_name}")
            except:
                pass
    return matched_files


def process_target_file5(file_path, current_datetime):
    """
    处理待处理文件5（三维提资接口）的主函数
    最终条件：处理1 & 处理2 & 处理3
    - 处理1：G列为 25C1/25C2/25C3
    - 处理2：L列日期筛选（同文件1的K列逻辑）
    - 处理3：N列为空值
    """
    print(f"开始处理待处理文件5: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件5: {os.path.basename(file_path)}")
    except:
        pass

    # 读取Excel文件的Sheet1
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')

    if df.empty:
        try:
            import Monitor
            Monitor.log_warning("文件5为空，无法处理")
        except:
            pass
        return pd.DataFrame()

    p1 = execute5_process1(df)
    p2 = execute5_process2(df, current_datetime)
    p3 = execute5_process3(df)

    final_rows = p1 & p2 & p3
    
    print(f"最终完成处理数据（原始筛选）: {len(final_rows)} 行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    try:
        from registry.util import extract_interface_id, extract_project_id
        from registry.db import get_connection
        from registry.hooks import _cfg
        
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        pending_rows = set()
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status
                FROM tasks
                WHERE file_type = 5
                  AND display_status IS NOT NULL AND display_status != ''
                  AND status != 'confirmed' AND status != 'archived'
            """)
            registry_tasks = cursor.fetchall()
            
            # Excel索引优化+从文件名提取项目号
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else ""
            
            excel_index = {}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 5)
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            for reg_interface_id, reg_project_id, _ in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                if key in excel_index:
                    for idx in excel_index[key]:
                        if idx not in final_rows:
                            pending_rows.add(idx)
        
        if pending_rows:
            final_rows = final_rows | pending_rows
        
    except Exception as e:
        print(f"[Registry] 查询待审查任务失败: {e}")
    
    print(f"最终完成处理数据（含待确认）: {len(final_rows)} 行")
    
    try:
        import Monitor
        Monitor.log_info(f"文件5处理1(G列25C1/25C2/25C3): {len(p1)} 行")
        Monitor.log_info(f"文件5处理2(L列日期): {len(p2)} 行")
        Monitor.log_info(f"文件5处理3(N列为空): {len(p3)} 行")
        Monitor.log_success(f"文件5最终完成处理数据: {len(final_rows)} 行")
    except:
        pass

    if not final_rows:
        return pd.DataFrame()

    final_indices = [i for i in final_rows if i > 0]
    excel_row_numbers = [i + 2 for i in final_indices]
    result_df = df.iloc[final_indices].copy()
    result_df['原始行号'] = excel_row_numbers

    # 新增“科室”列（基于G列：25C1/25C2/25C3）
    try:
        department_values = []
        for idx in result_df.index:
            cell_str = str(df.iloc[idx, 6]) if 6 < len(df.columns) else ""
            if "25C1" in cell_str:
                department_values.append("结构一室")
            elif "25C2" in cell_str:
                department_values.append("结构二室")
            elif "25C3" in cell_str:
                department_values.append("建筑总图室")
            else:
                department_values.append("")
        result_df["科室"] = department_values
    except Exception:
        result_df["科室"] = ""

    # 新增"接口时间"列（基于L列：索引11），格式 yyyy.mm.dd
    try:
        time_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 11] if 11 < len(df.columns) else None
            try:
                parsed = pd.to_datetime(cell_val, errors='coerce')
                if pd.isna(parsed):
                    time_values.append("")
                else:
                    # 【修复】保留完整年份，支持跨年延期判断
                    time_values.append(parsed.strftime('%Y.%m.%d'))
            except Exception:
                time_values.append("")
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""

    # 新增"责任人"列（基于K列：索引10，提取中文）
    try:
        zh_pattern = re.compile(r"[\u4e00-\u9fa5]+")
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 10] if 10 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            found = zh_pattern.findall(s)
            owners.append("".join(found))
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = ""
    
    # 【新增】添加source_file列
    result_df['source_file'] = file_path

    return result_df


def execute5_process1(df):
    """G列为 25C1/25C2/25C3"""
    result_rows = set()
    if len(df.columns) <= 6:
        return result_rows
    g_column = df.iloc[:, 6]
    for idx, val in g_column.items():
        if idx == 0:
            continue
        s = str(val) if val is not None else ""
        if ("25C1" in s) or ("25C2" in s) or ("25C3" in s):
            result_rows.add(idx)
    return result_rows


def execute5_process2(df, current_datetime):
    """L列日期筛选，逻辑同文件1的K列"""
    result_rows = set()
    if len(df.columns) <= 11:
        return result_rows
    l_column = df.iloc[:, 11]
    current_day = current_datetime.day
    current_year = current_datetime.year
    current_month = current_datetime.month
    if USE_OLD_DATE_LOGIC:
        start_date = datetime.datetime(current_year, current_month, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    else:
        start_date = datetime.datetime(current_year, 1, 1)
        if current_day <= 19:
            if current_month == 12:
                end_date = datetime.datetime(current_year, 12, 31)
            else:
                end_date = datetime.datetime(current_year, current_month + 1, 1) - datetime.timedelta(days=1)
        else:
            if current_month == 12:
                end_date = datetime.datetime(current_year + 1, 2, 1) - datetime.timedelta(days=1)
            elif current_month == 11:
                end_date = datetime.datetime(current_year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(current_year, current_month + 2, 1) - datetime.timedelta(days=1)
    for idx, val in l_column.items():
        if idx == 0:
            continue
        cell_date = None
        try:
            if isinstance(val, str):
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                    try:
                        cell_date = pd.to_datetime(val, format=fmt, errors='raise')
                        break
                    except:
                        continue
                if cell_date is None or pd.isna(cell_date):
                    cell_date = pd.to_datetime(val, errors='coerce')
            else:
                cell_date = pd.to_datetime(val, errors='coerce')
            if pd.isna(cell_date):
                continue
            if start_date <= cell_date <= end_date:
                result_rows.add(idx)
        except:
            continue
    return result_rows


def execute5_process3(df):
    """N列为空值"""
    result_rows = set()
    if len(df.columns) <= 13:
        return result_rows
    n_column = df.iloc[:, 13]
    for idx, val in n_column.items():
        if idx == 0:
            continue
        if pd.isna(val) or str(val).strip() == "":
            result_rows.add(idx)
    return result_rows


def export_result_to_excel5(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出三维提资接口处理结果到Excel文件
    结构与其他导出函数一致；当源为.xls时仅复制值，不复制样式
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    try:
        # 根据项目号创建结果文件夹
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
            final_output_dir = result_folder_path
        else:
            final_output_dir = output_dir

        # 生成输出文件路径（避免重名）
        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"三维提资接口{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1

        # 新建目标工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "三维提资接口"

        # 读取源文件（xlsx用openpyxl，xls回退仅复制值）
        if original_file_path.endswith('.xlsx'):
            original_wb = load_workbook(original_file_path)
            original_ws = original_wb.worksheets[0]
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_wb = None
            original_ws = None
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)

        # 复制表头
        if original_ws is not None and original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                target_cell.value = source_cell.value
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception:
                    pass
        else:
            max_col = len(original_df.columns)
            for col_idx in range(1, max_col + 1):
                ws.cell(row=1, column=col_idx).value = original_df.iat[0, col_idx - 1]

        # 复制数据行
        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            write_row = 2
            max_col = max((original_ws.max_column if original_ws is not None else 0), len(original_df.columns))
            for excel_row_num in sorted(qualified_rows):
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    for col_idx in range(1, max_col + 1):
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        if original_ws is not None:
                            source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                            target_cell.value = source_cell.value
                            try:
                                if source_cell.font:
                                    target_cell.font = copy(source_cell.font)
                                if source_cell.fill:
                                    target_cell.fill = copy(source_cell.fill)
                                if source_cell.border:
                                    target_cell.border = copy(source_cell.border)
                                if source_cell.alignment:
                                    target_cell.alignment = copy(source_cell.alignment)
                                if source_cell.number_format:
                                    target_cell.number_format = source_cell.number_format
                            except Exception:
                                pass
                        else:
                            # 从DataFrame复制纯值（xls场景）
                            val = original_df.iat[excel_row_num - 1, col_idx - 1] if col_idx - 1 < len(original_df.columns) else None
                            target_cell.value = val
                    write_row += 1

        # 设置列宽
        max_col = max((original_ws.max_column if original_ws is not None else 0), len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            max_width = 8
            for row_idx in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_width = len(str(cell.value)) * 1.2
                    max_width = max(max_width, cell_width)
            final_width = min(max(max_width * 1.2, 8), 100)
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width

        # 保存
        if original_wb is not None:
            original_wb.close()
        wb.save(output_path)
        wb.close()

        try:
            import Monitor
            Monitor.log_success(f"三维提资接口导出完成！文件保存到: {output_path}")
        except Exception:
            pass
        return output_path
    except Exception as e:
        print(f"导出三维提资接口数据时发生错误: {str(e)}")
        try:
            import Monitor
            Monitor.log_error(f"导出三维提资接口数据时发生错误: {str(e)}")
        except Exception:
            pass
        raise
# ===================== 待处理文件6（收发文函）相关处理 =====================
def find_target_file6(excel_files):
    """
    查找符合特定格式的待处理文件6（兼容性函数，返回第一个匹配的文件）
    规则：文件名中包含“收发文清单”，若其后紧随四位数字则作为项目号
         示例：收发文清单2016.xlsx → 项目号=2016；
         未紧随四位数字时，项目号为空字符串（兼容旧命名）
    返回：(文件路径, 项目号) 或 (None, None)
    """
    all_files = find_all_target_files6(excel_files)
    if all_files:
        return all_files[0]
    return None, None


def find_all_target_files6(excel_files):
    """
    批量查找所有“收发文清单”文件。
    优先提取紧随其后的四位数字作为项目号；若未匹配，则项目号为空字符串。
    """
    matched_files = []
    try:
        import Monitor
        Monitor.log_process("开始批量识别待处理文件6(收发文函)...")
    except Exception:
        pass
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        if (file_name.endswith('.xlsx') or file_name.endswith('.xls')) and ("收发文清单" in file_name):
            # 优先匹配 紧随“收发文清单”的四位数字 作为项目号
            try:
                # 紧随“收发文清单”的四位数字作为项目号，例如：收发文清单2016.xlsx
                m = re.search(r"收发文清单(\d{4})", file_name)
                project_id = m.group(1) if m else ""
            except Exception:
                project_id = ""
            matched_files.append((file_path, project_id))
            try:
                import Monitor
                if project_id:
                    Monitor.log_success(f"找到待处理文件6: 项目{project_id} - {file_name}")
                else:
                    Monitor.log_success(f"找到待处理文件6(未识别项目号): {file_name}")
            except Exception:
                pass
    return matched_files


def filter_valid_names(names_str, valid_names_set):
    """
    过滤责任人姓名，只保留在姓名角色表中存在的姓名
    
    Args:
        names_str: 逗号分隔的姓名字符串，如"刘峰a,张三,李四b"
        valid_names_set: 有效姓名集合
    
    Returns:
        str: 过滤后的姓名字符串
    
    规则：
        - "刘峰a" → 尝试匹配"刘峰"（去除尾部字母）
        - 只保留在姓名角色表中存在的姓名
    """
    if not names_str or not valid_names_set:
        return names_str
    
    tokens = [t.strip() for t in names_str.split(',') if t.strip()]
    filtered_names = []
    
    for name in tokens:
        # 首先尝试精确匹配
        if name in valid_names_set:
            filtered_names.append(name)
        else:
            # 尝试去除尾部字母后匹配（如"刘峰a" → "刘峰"）
            import re
            # 移除尾部的英文字母（一个或多个）
            cleaned_name = re.sub(r'[a-zA-Z]+$', '', name)
            if cleaned_name and cleaned_name in valid_names_set:
                filtered_names.append(cleaned_name)
            # 如果都不匹配，不添加该姓名（过滤掉）
    
    return ','.join(filtered_names)


def process_target_file6(file_path, current_datetime, skip_date_filter=False, valid_names_set=None):
    """
    处理待处理文件6（收发文函）
    
    Args:
        file_path: Excel文件路径
        current_datetime: 当前时间
        skip_date_filter: 是否跳过I列日期范围筛选（管理员/所领导模式为True）
        valid_names_set: 有效姓名集合（用于过滤责任人）
    
    筛选条件：
      p1) V列包含"河北分公司.建筑结构所"
      p_i) I列不为空且为有效日期
      p3) I列日期 ≤ 今天+14天（普通模式）
      p4) M列等于"尚未回复"或"超期未回复"
      
    最终结果：
      - 【普通模式】: p1 & p_i & p3 & p4
      - 【管理员/所领导模式】: p1 & p_i & p4（跳过日期范围限制，但仍需I列非空）
    
    附加字段：
      - 接口时间：I列按 mm.dd 提取
      - 责任人：X列分隔的姓名集合（用于角色过滤，稍后基于包含关系过滤）
        【新增】只保留在姓名角色表中存在的姓名
      - 科室：空值（待处理文件6科室空值）
    """
    print(f"开始处理待处理文件6: {os.path.basename(file_path)}")
    try:
        import Monitor
        Monitor.log_process(f"开始处理待处理文件6: {os.path.basename(file_path)}")
    except Exception:
        pass

    # 读取Excel文件的第一个工作表（不强制Sheet1）
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, sheet_name=0, engine='xlrd')

    if df.empty:
        try:
            import Monitor
            Monitor.log_warning("文件6为空，无法处理")
        except Exception:
            pass
        return pd.DataFrame()

    p1 = execute6_process1(df)
    p_i_not_empty = execute6_process_i_not_empty(df)  # 【新增】I列非空检查
    p4 = execute6_process4(df)
    
    # 根据skip_date_filter决定是否使用p3（I列日期筛选）
    if skip_date_filter:
        # 管理员模式：跳过I列日期范围筛选，但仍需检查I列非空
        final_rows = p1 & p_i_not_empty & p4
        print(f"最终完成处理数据（原始筛选，管理员模式）: {len(final_rows)} 行")
    else:
        # 普通模式：使用所有筛选条件（包括I列非空和日期范围）
        p3 = execute6_process3(df, current_datetime)
        final_rows = p1 & p_i_not_empty & p3 & p4
        print(f"最终完成处理数据（原始筛选，普通模式）: {len(final_rows)} 行")
    
    # 【新增】Registry查询：查找有display_status的待审查任务（使用business_id匹配）
    try:
        from registry.util import extract_interface_id, extract_project_id
        from registry.db import get_connection
        from registry.hooks import _cfg
        
        # 【修复】通过registry_hooks._cfg()获取正确的配置（包含data_folder）
        cfg = _cfg()
        db_path = cfg.get('registry_db_path')
        pending_rows = set()
        
        if db_path and os.path.exists(db_path):
            conn = get_connection(db_path, True)
            cursor = conn.execute("""
                SELECT interface_id, project_id, display_status
                FROM tasks
                WHERE file_type = 6
                  AND display_status IS NOT NULL AND display_status != ''
                  AND status != 'confirmed' AND status != 'archived'
            """)
            registry_tasks = cursor.fetchall()
            
            # Excel索引优化+从文件名提取项目号
            import re
            filename = os.path.basename(file_path)
            match = re.search(r'(\d{4})', filename)
            file_project_id = match.group(1) if match else ""
            
            excel_index = {}
            for idx in range(len(df)):
                if idx == 0:
                    continue
                try:
                    row_data = df.iloc[idx]
                    df_interface_id = extract_interface_id(row_data, 6)
                    if df_interface_id and file_project_id:
                        key = (df_interface_id, file_project_id)
                        if key not in excel_index:
                            excel_index[key] = []
                        excel_index[key].append(idx)
                except:
                    continue
            
            for reg_interface_id, reg_project_id, _ in registry_tasks:
                key = (reg_interface_id, reg_project_id)
                if key in excel_index:
                    for idx in excel_index[key]:
                        if idx not in final_rows:
                            pending_rows.add(idx)
        
        if pending_rows:
            final_rows = final_rows | pending_rows
        
    except Exception as e:
        print(f"[Registry] 查询待审查任务失败: {e}")
    
    print(f"最终完成处理数据（含待确认）: {len(final_rows)} 行")
    
    # 日志记录
    try:
        import Monitor
        if skip_date_filter:
            Monitor.log_info(f"文件6处理1(V列机构匹配): {len(p1)} 行")
            Monitor.log_info(f"文件6 I列非空检查: {len(p_i_not_empty)} 行")
            Monitor.log_info(f"文件6处理4(M列=尚未回复或超期未回复): {len(p4)} 行")
            Monitor.log_success(f"文件6最终完成处理数据(管理员模式): {len(final_rows)} 行")
        else:
            Monitor.log_info(f"文件6处理1(V列机构匹配): {len(p1)} 行")
            Monitor.log_info(f"文件6 I列非空检查: {len(p_i_not_empty)} 行")
            Monitor.log_info(f"文件6处理3(I列日期≤今天+14天): {len(p3)} 行")
            Monitor.log_info(f"文件6处理4(M列=尚未回复或超期未回复): {len(p4)} 行")
            Monitor.log_success(f"文件6最终完成处理数据: {len(final_rows)} 行")
    except Exception:
        pass

    if not final_rows:
        return pd.DataFrame()

    final_indices = [i for i in final_rows if i > 0]
    excel_row_numbers = [i + 2 for i in final_indices]
    result_df = df.iloc[final_indices].copy()
    result_df['原始行号'] = excel_row_numbers

    # 接口时间（I列：索引8）
    try:
        time_values = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 8] if 8 < len(df.columns) else None
            parsed = pd.to_datetime(cell_val, errors='coerce')
            if pd.isna(parsed):
                time_values.append("")
            else:
                # 【修复】保留完整年份，支持跨年延期判断
                time_values.append(parsed.strftime('%Y.%m.%d'))
        result_df["接口时间"] = time_values
    except Exception:
        result_df["接口时间"] = ""

    # 待处理文件6科室空值
    try:
        result_df["科室"] = ""
    except Exception:
        pass

    # 责任人：X列（索引23）按分隔符拆分并保留原始姓名集合（供过滤）
    try:
        owners = []
        for idx in result_df.index:
            cell_val = df.iloc[idx, 23] if 23 < len(df.columns) else None
            s = str(cell_val) if cell_val is not None else ""
            # 分隔符：, ， ; ； / 、
            for sep in [',', '，', ';', '；', '/', '、']:
                s = s.replace(sep, ',')
            tokens = [t.strip() for t in s.split(',') if t.strip()]
            names_str = ','.join(tokens)
            
            # 【新增】过滤责任人：只保留在姓名角色表中存在的姓名
            if valid_names_set:
                names_str = filter_valid_names(names_str, valid_names_set)
            
            owners.append(names_str)
        result_df['责任人'] = owners
    except Exception:
        result_df['责任人'] = ""
    
    # 【新增】添加source_file列
    result_df['source_file'] = file_path

    return result_df


def execute6_process1(df):
    """V列包含“河北分公司.建筑结构所”"""
    result_rows = set()
    if len(df.columns) <= 21:
        return result_rows
    v_column = df.iloc[:, 21]
    for idx, val in v_column.items():
        if idx == 0:
            continue
        s = str(val) if val is not None else ""
        if "河北分公司.建筑结构所" in s:
            result_rows.add(idx)
    return result_rows


def execute6_process2(df):
    """H列为“是”"""
    result_rows = set()
    if len(df.columns) <= 7:
        return result_rows
    h_column = df.iloc[:, 7]
    for idx, val in h_column.items():
        if idx == 0:
            continue
        if str(val).strip() == "是":
            result_rows.add(idx)
    return result_rows


def execute6_process_i_not_empty(df):
    """I列不为空（管理员模式和普通模式都需要）"""
    result_rows = set()
    if len(df.columns) <= 8:
        return result_rows
    i_column = df.iloc[:, 8]
    for idx, val in i_column.items():
        if idx == 0:
            continue
        # 检查I列是否为空
        if val is not None and str(val).strip() != '':
            # 尝试解析为日期，确保是有效日期
            try:
                parsed = pd.to_datetime(val, errors='coerce')
                if not pd.isna(parsed):
                    result_rows.add(idx)
            except Exception:
                continue
    return result_rows


def execute6_process3(df, current_datetime):
    """I列为日期，筛选当日及之前 + 未来14天内（即 delta <= 14）"""
    result_rows = set()
    if len(df.columns) <= 8:
        return result_rows
    i_column = df.iloc[:, 8]
    if USE_OLD_DATE_LOGIC:
        # 旧逻辑：当日及之前 + 未来14天（即日期 <= 今天+14天）
        today = current_datetime.date()
        for idx, val in i_column.items():
            if idx == 0:
                continue
            try:
                parsed = pd.to_datetime(val, errors='coerce')
                if pd.isna(parsed):
                    continue
                d = parsed.date()
                delta = (d - today).days
                # 修改：包含过去的日期（delta < 0）+ 今天和未来14天（0 <= delta <= 14）
                if delta <= 14:
                    result_rows.add(idx)
            except Exception:
                continue
    else:
        # 新逻辑：与旧逻辑相同，待处理文件6不使用月度范围，而是使用简单的日期窗口
        # 当日及之前 + 未来14天（即日期 <= 今天+14天）
        today = current_datetime.date()
        for idx, val in i_column.items():
            if idx == 0:
                continue
            try:
                parsed = pd.to_datetime(val, errors='coerce')
                if pd.isna(parsed):
                    continue
                d = parsed.date()
                delta = (d - today).days
                # 包含过去的日期（delta < 0）+ 今天和未来14天（0 <= delta <= 14）
                if delta <= 14:
                    result_rows.add(idx)
            except Exception:
                continue
    return result_rows


def execute6_process4(df):
    """M列为'尚未回复'或'超期未回复'"""
    result_rows = set()
    if len(df.columns) <= 12:
        return result_rows
    m_column = df.iloc[:, 12]
    for idx, val in m_column.items():
        if idx == 0:
            continue
        val_str = str(val).strip()
        if val_str in ["尚未回复", "超期未回复"]:
            result_rows.add(idx)
    return result_rows


def export_result_to_excel6(df, original_file_path, current_datetime, output_dir, project_id=None):
    """
    导出收发文函处理结果到Excel文件
    Sheet 名称与文件前缀：收发文函
    注：待处理文件6项目号空值 → 不新建项目号结果文件夹
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy
    try:
        # 根据项目号决定输出目录（项目号为空则直接用输出目录）
        if project_id:
            result_folder_name = f"{project_id}结果文件"
            result_folder_path = os.path.join(output_dir, result_folder_name)
            if not os.path.exists(result_folder_path):
                os.makedirs(result_folder_path)
            final_output_dir = result_folder_path
        else:
            final_output_dir = output_dir

        date_str = current_datetime.strftime('%Y-%m-%d')
        base_filename = f"收发文函{date_str}"
        output_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(final_output_dir, output_filename)
        counter = 1
        while os.path.exists(output_path):
            output_filename = f"{base_filename}({counter}).xlsx"
            output_path = os.path.join(final_output_dir, output_filename)
            counter += 1

        wb = Workbook()
        ws = wb.active
        ws.title = "收发文函"

        original_wb = load_workbook(original_file_path)
        original_ws = original_wb.worksheets[0]
        if original_file_path.endswith('.xlsx'):
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='openpyxl', header=None)
        else:
            original_df = pd.read_excel(original_file_path, sheet_name=0, engine='xlrd', header=None)

        if original_ws.max_row > 0:
            max_col = max(original_ws.max_column, len(original_df.columns))
            for col_idx in range(1, max_col + 1):
                source_cell = original_ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                target_cell.value = source_cell.value
                try:
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                except Exception:
                    pass

        if not df.empty and '原始行号' in df.columns:
            qualified_rows = df['原始行号'].tolist()
            write_row = 2
            max_col = max(original_ws.max_column, len(original_df.columns))
            for excel_row_num in sorted(qualified_rows):
                if excel_row_num > 1 and excel_row_num <= len(original_df):
                    for col_idx in range(1, max_col + 1):
                        source_cell = original_ws.cell(row=excel_row_num, column=col_idx)
                        target_cell = ws.cell(row=write_row, column=col_idx)
                        target_cell.value = source_cell.value
                        try:
                            if source_cell.font:
                                target_cell.font = copy(source_cell.font)
                            if source_cell.fill:
                                target_cell.fill = copy(source_cell.fill)
                            if source_cell.border:
                                target_cell.border = copy(source_cell.border)
                            if source_cell.alignment:
                                target_cell.alignment = copy(source_cell.alignment)
                            if source_cell.number_format:
                                target_cell.number_format = source_cell.number_format
                        except Exception:
                            pass
                    write_row += 1

        max_col = max(original_ws.max_column, len(original_df.columns))
        for col_idx in range(1, max_col + 1):
            max_width = 8
            for row_idx in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_width = len(str(cell.value)) * 1.2
                    max_width = max(max_width, cell_width)
            final_width = min(max(max_width * 1.2, 8), 100)
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = final_width

        original_wb.close()
        wb.save(output_path)
        wb.close()
        try:
            import Monitor
            Monitor.log_success(f"收发文函导出完成！文件保存到: {output_path}")
        except Exception:
            pass
        return output_path
    except Exception as e:
        print(f"导出收发文函数据时发生错误: {str(e)}")
        try:
            import Monitor
            Monitor.log_error(f"导出收发文函数据时发生错误: {str(e)}")
        except Exception:
            pass
        raise


if __name__ == "__main__":
    # 如果直接运行此文件，显示提示信息
    print("Excel数据处理模块已加载")
    print("请通过主程序(base.py)来使用此模块")