#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回文单号输入处理模块
"""

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import os


class InterfaceInputDialog(tk.Toplevel):
    """回文单号输入弹窗"""
    
    def __init__(self, parent, interface_id, file_type, file_path, row_index, 
                 user_name, project_id, source_column=None):
        """
        参数:
            parent: 父窗口
            interface_id: 接口号
            file_type: 文件类型(1-6)
            file_path: 原始Excel文件路径
            row_index: Excel行号
            user_name: 当前用户姓名
            project_id: 项目号
            source_column: 文件3专用，'M'或'L'，表示筛选来源
        """
        super().__init__(parent)
        
        self.interface_id = interface_id
        self.file_type = file_type
        self.file_path = file_path
        self.row_index = row_index
        self.user_name = user_name
        self.project_id = project_id
        self.source_column = source_column
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置界面"""
        self.title("回文单号输入")
        self.geometry("400x200")
        self.resizable(False, False)
        
        # 居中显示
        self.transient(self.master)
        self.grab_set()
        
        # 标题
        title_label = ttk.Label(self, text=f"接口号: {self.interface_id}", 
                                font=('Arial', 12, 'bold'))
        title_label.pack(pady=10)
        
        # 输入框
        input_frame = ttk.Frame(self)
        input_frame.pack(pady=10, padx=20, fill='x')
        
        ttk.Label(input_frame, text="回文单号:").pack(side='left', padx=5)
        
        self.entry = ttk.Entry(input_frame, width=30)
        self.entry.pack(side='left', padx=5, fill='x', expand=True)
        self.entry.focus_set()
        
        # 按钮
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="确认", command=self.on_confirm).pack(side='left', padx=10)
        ttk.Button(button_frame, text="取消", command=self.destroy).pack(side='left', padx=10)
        
        # 绑定Enter键
        self.entry.bind('<Return>', lambda e: self.on_confirm())
    
    def on_confirm(self):
        """确认按钮回调"""
        response_number = self.entry.get().strip()
        
        if not response_number:
            messagebox.showwarning("警告", "请输入回文单号", parent=self)
            return
        
        # 写入Excel
        try:
            success = write_response_to_excel(
                self.file_path,
                self.file_type,
                self.row_index,
                response_number,
                self.user_name,
                self.project_id,
                self.source_column
            )
            
            if success:
                messagebox.showinfo("成功", "回文单号已保存", parent=self)
                self.destroy()
            else:
                messagebox.showerror("失败", "保存失败，请重试", parent=self)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {str(e)}", parent=self)


def write_response_to_excel(file_path, file_type, row_index, response_number, 
                             user_name, project_id, source_column=None):
    """
    写入回文单号到Excel文件
    
    参数:
        file_path: Excel文件路径
        file_type: 文件类型(1-6)
        row_index: Excel行号（从2开始，因为第1行是标题）
        response_number: 回文单号
        user_name: 用户姓名
        project_id: 项目号
        source_column: 文件3专用，'M'或'L'
    
    返回:
        bool: 成功返回True，失败返回False
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            return False
        
        # 文件锁定检测
        try:
            # 尝试以独占模式打开
            with open(file_path, 'r+b') as f:
                pass
        except PermissionError:
            messagebox.showerror("文件占用", "有其他用户占用该文件，请稍后再试")
            return False
        
        # 使用openpyxl打开
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 获取写入列位置
        columns = get_write_columns(file_type, row_index, ws, source_column)
        
        if not columns:
            print(f"无法确定写入列位置: file_type={file_type}")
            return False
        
        # 写入数据
        response_col = columns['response_col']
        time_col = columns['time_col']
        name_col = columns['name_col']
        
        ws[f"{response_col}{row_index}"] = response_number
        ws[f"{time_col}{row_index}"] = date.today().strftime('%Y-%m-%d')
        ws[f"{name_col}{row_index}"] = user_name
        
        # 保存
        wb.save(file_path)
        wb.close()
        
        print(f"成功写入: {file_path}, 行{row_index}, 回文单号={response_number}")
        return True
        
    except Exception as e:
        print(f"写入失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_write_columns(file_type, row_index, worksheet, source_column=None):
    """
    获取各文件类型的写入列位置
    
    参数:
        file_type: 文件类型(1-6)
        row_index: Excel行号
        worksheet: openpyxl工作表对象
        source_column: 文件3专用，'M'或'L'
    
    返回:
        dict: {'response_col': 'S', 'time_col': 'N', 'name_col': 'V'}
        或 None（如果无法确定）
    """
    # 文件类型1-2, 4-6的固定列位置
    column_map = {
        1: {'response_col': 'S', 'time_col': 'M', 'name_col': 'V'},  # 【修改】时间列从N改为M
        2: {'response_col': 'P', 'time_col': 'N', 'name_col': 'AL'},
        4: {'response_col': 'U', 'time_col': 'V', 'name_col': 'AT'},
        5: {'response_col': 'V', 'time_col': 'N', 'name_col': 'W'},
        6: {'response_col': 'L', 'time_col': 'J', 'name_col': 'N'},
    }
    
    if file_type in column_map:
        return column_map[file_type]
    
    # 文件3特殊逻辑：根据source_column判断
    if file_type == 3:
        if source_column == 'M':
            # M列筛选：V/T/BM
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif source_column == 'L':
            # L列筛选：S/Q/BM
            return {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 如果未指定，尝试自动判断
            return determine_file3_source_and_columns(row_index, worksheet)
    
    return None


def determine_file3_source_and_columns(row_index, worksheet):
    """
    判断文件3某行是因M列还是L列被筛选出
    
    参数:
        row_index: Excel行号
        worksheet: openpyxl工作表对象
    
    返回:
        dict: 写入列位置
    """
    try:
        # 读取M列和L列的值
        m_val = worksheet[f"M{row_index}"].value
        l_val = worksheet[f"L{row_index}"].value
        
        # 读取T列和Q列的值（回复时间列）
        t_val = worksheet[f"T{row_index}"].value
        q_val = worksheet[f"Q{row_index}"].value
        
        # 简化判断逻辑：
        # 如果M列有时间数据且T列为空，判断为M列来源
        # 如果L列有时间数据且Q列为空，判断为L列来源
        # 优先M列
        
        m_has_time = m_val is not None and str(m_val).strip() != ''
        t_is_empty = t_val is None or str(t_val).strip() == ''
        
        l_has_time = l_val is not None and str(l_val).strip() != ''
        q_is_empty = q_val is None or str(q_val).strip() == ''
        
        if m_has_time and t_is_empty:
            # M列来源
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
        elif l_has_time and q_is_empty:
            # L列来源
            return {'response_col': 'S', 'time_col': 'Q', 'name_col': 'BM'}
        else:
            # 默认M列
            return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}
    
    except Exception as e:
        print(f"判断文件3来源失败: {e}")
        # 默认返回M列
        return {'response_col': 'V', 'time_col': 'T', 'name_col': 'BM'}


# 测试代码
if __name__ == "__main__":
    # 测试get_write_columns
    columns_1 = get_write_columns(1, 5, None)
    print(f"文件1写入列: {columns_1}")
    
    columns_3_m = get_write_columns(3, 5, None, 'M')
    print(f"文件3(M列)写入列: {columns_3_m}")
    
    columns_3_l = get_write_columns(3, 5, None, 'L')
    print(f"文件3(L列)写入列: {columns_3_l}")

